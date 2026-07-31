#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ONCONOVA CIRUGÍA ONCOLÓGICA - Sistema de Oncología
Punto de entrada principal de la aplicación - Migrado completamente a TTKBootstrap

Este script se encarga de:
1. Configurar la ruta del ejecutable de Tesseract OCR.
2. Iniciar la interfaz gráfica de usuario moderna (dashboard).
"""

import ttkbootstrap as ttk
from ttkbootstrap import Style
from ttkbootstrap.constants import *
import tkinter.ttk as ttk_std
from PIL import Image, ImageTk
from tkinter import filedialog, messagebox
import tkinter as tk
from tksheet import Sheet  # V5.3.8: Tabla virtualizada tipo Excel
import threading
import concurrent.futures  # V6.9.10 PARALELO: ThreadPoolExecutor para procesar chunks en simultaneo
import os
import re
import pandas as pd
import seaborn as sns
from matplotlib.figure import Figure
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from datetime import datetime
import numpy as np
import argparse
import sys
import logging
import traceback
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import configparser
import pytesseract
from pathlib import Path

from core.calendario import CalendarioInteligente
# LAZY IMPORT: huv_web_automation solo se importa cuando se necesita (evita error de webdriver_manager al iniciar)
# from core.huv_web_automation import automatizar_entrega_resultados, Credenciales
from core.enhanced_export_system import EnhancedExportSystem
from config.version_info import get_version_string, get_build_info, get_full_version_info, get_dependencies_actual

# === IMPORTS PARA SISTEMA DE AUDITORÍA IA ===
from core.debug_mapper import DebugMapper
from core.ventana_auditoria_ia import mostrar_ventana_auditoria
from core.database_manager import get_registro_by_peticion

# === IMPORTS DE UI HELPERS ===
from ui_helpers import ocr_helpers, database_helpers, export_helpers, chart_helpers

# ======================== CONSTANTES UI ========================
# Movido desde: config/huv_constants.py

# Meses en español para formateo de fechas
MESES_ES = {
    1: 'Enero', 2: 'Febrero', 3: 'Marzo', 4: 'Abril',
    5: 'Mayo', 6: 'Junio', 7: 'Julio', 8: 'Agosto',
    9: 'Septiembre', 10: 'Octubre', 11: 'Noviembre', 12: 'Diciembre'
}

# Configuración para detección de duplicados
DUPLICATE_DETECTION = {
    'COLORS': {
        'duplicado': '#ff4444',      # Rojo para archivos duplicados
        'nuevo': '#44ff44',          # Verde para archivos nuevos
        'procesando': '#ffaa44',     # Naranja para en proceso
        'error': '#ff0000'           # Rojo intenso para errores
    },
    'SCROLL_SPEED_MULTIPLIER': 4,    # 4x más rápido que el scroll normal
    'HORIZONTAL_SCROLL_UNITS': 3     # Unidades por scroll
}

# =========================
# Configuración de Tesseract OCR
# =========================

def configure_tesseract():
    """
    Lee el archivo config.ini para encontrar la ruta de Tesseract OCR
    y la configura para que Pytesseract pueda utilizarla.
    """
    try:
        config = configparser.ConfigParser(interpolation=None)

        # CORREGIDO: Detectar si estamos en un ejecutable empaquetado
        if getattr(sys, 'frozen', False):
            # Estamos ejecutando como .exe - buscar config.ini junto al .exe
            base_path = Path(sys.executable).parent
        else:
            # Estamos ejecutando como script Python
            base_path = Path(__file__).resolve().parent

        config_path = base_path / 'config' / 'config.ini'
        config.read(config_path, encoding='utf-8')

        tesseract_cmd = None
        if sys.platform.startswith("win"):
            tesseract_cmd = config.get('PATHS', 'WINDOWS_TESSERACT', fallback=None)
        elif sys.platform.startswith("darwin"):
            tesseract_cmd = config.get('PATHS', 'MACOS_TESSERACT', fallback=None)
        else: # Asumimos Linux/otro
            tesseract_cmd = config.get('PATHS', 'LINUX_TESSERACT', fallback=None)

        if tesseract_cmd and Path(tesseract_cmd).exists():
            pytesseract.pytesseract.tesseract_cmd = tesseract_cmd
            logging.info(f"Tesseract OCR configurado en: {tesseract_cmd}")
        else:
            logging.warning("No se encontro la ruta de Tesseract en config.ini o la ruta no es valida.")
            logging.warning("El sistema intentara usar la variable de entorno PATH.")

    except Exception as e:
        logging.error(f"Error al configurar Tesseract desde config.ini: {e}")
        logging.warning("Se continuara usando la configuracion por defecto de Pytesseract.")

# =========================
# Configuración de temas TTKBootstrap
# =========================

# Mapeo de argumentos de tema a temas TTKBootstrap
THEME_MAP = {
    "huv": "huv",  # V6.9.16 - Tema institucional HUV claro (#2d3e5e)
    "dark": "darkly",
    "light": "flatly",
    "blue": "cosmo",
    "professional": "litera",
    "medical": "pulse",
    "modern": "superhero",
    "classic": "journal",
    # Agregar más temas TTKBootstrap compatibles
    "darkly": "darkly",
    "flatly": "flatly",
    "cosmo": "cosmo",
    "litera": "litera",
    "pulse": "pulse",
    "superhero": "superhero",
    "journal": "journal",
    "cyborg": "cyborg",
    "solar": "solar",
    "minty": "minty",
    "sandstone": "sandstone",
    "united": "united",
    "morph": "morph",
    "vapor": "vapor",
    "yeti": "yeti",
    "lumen": "lumen",
    "simplex": "simplex",
    "zephyr": "zephyr"
}

# Paleta de colores base (se ajustará según el tema)
COLORS = {
    "accent": "#2b6cb0",
    "bg": "#ffffff", 
    "surface": "#f8f9fa",
    "text": "#212529",
    "muted": "#6c757d"
}

# ======================================================================
# V6.9.27 - ESTILO DE GRAFICOS THEME-AWARE (oscuro / claro)
# Antes habia dos bloques fijos (seaborn oscuro + matplotlib claro) y el
# claro pisaba al oscuro -> los graficos quedaban SIEMPRE blancos, lo que
# en tema oscuro daba un contraste horrible. Ahora una sola funcion aplica
# el estilo correcto segun el tema activo. Se llama al inicio y al cambiar
# de tema desde el menu.
# ======================================================================
DARK_THEMES = {"darkly", "superhero", "cyborg", "solar", "vapor"}

def aplicar_estilo_graficos(es_oscuro: bool):
    """Aplica rcParams de matplotlib/seaborn coherentes con el tema activo.
    es_oscuro=True  -> fondos gris oscuro + texto/ejes claros (combina con 'darkly').
    es_oscuro=False -> fondos blancos + titulos navy (tema 'huv' claro).
    """
    try:
        import matplotlib as _mpl
        if es_oscuro:
            sns.set_theme(style="darkgrid", rc={
                "axes.facecolor": "#2b2f36", "grid.color": "#3a3f47",
                "figure.facecolor": "#23262b", "text.color": "#e6e8ec",
                "xtick.color": "#c2c7d0", "ytick.color": "#c2c7d0",
                "axes.labelcolor": "#e6e8ec", "axes.titlecolor": "#eef1f6",
            })
            _mpl.rcParams.update({
                "figure.facecolor": "#23262b", "axes.facecolor": "#2b2f36",
                "savefig.facecolor": "#23262b", "axes.edgecolor": "#3a3f47",
                "axes.linewidth": 0.8, "axes.labelcolor": "#e6e8ec",
                "axes.titlecolor": "#eef1f6", "axes.titlesize": 11, "axes.titleweight": "bold",
                "axes.grid": True, "axes.axisbelow": True,
                "grid.color": "#3a3f47", "grid.linewidth": 0.8,
                "axes.spines.top": False, "axes.spines.right": False,
                "xtick.color": "#c2c7d0", "ytick.color": "#c2c7d0", "text.color": "#e6e8ec",
                "axes.prop_cycle": _mpl.cycler(
                    color=["#5b8def", "#4ecb8d", "#e0a458", "#d9647a", "#9aa6bf", "#7c8cff"]),
            })
        else:
            sns.set_theme(style="whitegrid", rc={
                "axes.facecolor": "#ffffff", "figure.facecolor": "#ffffff",
                "text.color": "#2a2f3a",
            })
            _mpl.rcParams.update({
                "figure.facecolor": "#ffffff", "axes.facecolor": "#ffffff",
                "savefig.facecolor": "#ffffff", "axes.edgecolor": "#d2d9e6",
                "axes.linewidth": 0.8, "axes.labelcolor": "#2a2f3a",
                "axes.titlecolor": "#2d3e5e", "axes.titlesize": 11, "axes.titleweight": "bold",
                "axes.grid": True, "axes.axisbelow": True,
                "grid.color": "#eef1f6", "grid.linewidth": 0.8,
                "axes.spines.top": False, "axes.spines.right": False,
                "xtick.color": "#5a6172", "ytick.color": "#5a6172", "text.color": "#2a2f3a",
                "axes.prop_cycle": _mpl.cycler(
                    color=["#2d3e5e", "#9aa6bf", "#4a6da7", "#2f8f6b", "#d99a4e", "#c75c6e"]),
            })
    except Exception as _e:
        logging.warning(f"[graficos] No se pudo aplicar estilo ({'oscuro' if es_oscuro else 'claro'}): {_e}")

# V6.9.28 - La app usa SOLO tema claro (navy HUV). Graficos siempre en claro.
aplicar_estilo_graficos(es_oscuro=False)

# Módulos del proyecto
# Importar módulo unificado de extractores refactorizados
import core.unified_extractor as procesador_ihq_biomarcadores
import core.database_manager as database_manager


# ======================================================================
# V6.9.16 - TEMA INSTITUCIONAL HUV (UI/UX)
# Paleta minimalista construida sobre el azul institucional #2d3e5e.
# Se registra como tema ttkbootstrap 'huv'. Al estar definido a nivel
# modulo, queda disponible ANTES de instanciar la ventana (App).
# Filosofia: limpio, ordenado, fondos claros, acentos desaturados.
# ======================================================================
try:
    from ttkbootstrap.themes.standard import STANDARD_THEMES as _STD_THEMES

    _HUV_COLORS = {
        "primary":   "#2d3e5e",  # Azul institucional HUV (botones/acentos/headers)
        "secondary": "#8a909c",  # Gris neutro
        "success":   "#2f8f6b",  # Verde apagado (estados OK)
        "info":      "#4a6da7",  # Azul medio, armonico con el primary
        "warning":   "#d99a4e",  # Ambar suave
        "danger":    "#c75c6e",  # Rojo apagado (no estridente)
        "light":     "#f4f6f9",  # Gris muy claro (fondos de tarjeta/seccion)
        "dark":      "#2d3e5e",  # Mismo navy para superficies oscuras
        "bg":        "#ffffff",  # Fondo base limpio
        "fg":        "#2a2f3a",  # Texto principal (gris oscuro, no negro puro)
        "selectbg":  "#2d3e5e",  # Seleccion = azul institucional
        "selectfg":  "#ffffff",
        "border":    "#e4e7ec",  # Bordes sutiles
        "inputfg":   "#2a2f3a",
        "inputbg":   "#ffffff",
        "active":    "#eef1f6",  # Hover/activo muy claro
    }
    # Registrar solo si no existe (idempotente ante reimports)
    if "huv" not in _STD_THEMES:
        _STD_THEMES["huv"] = {"type": "light", "colors": _HUV_COLORS}
except Exception as _e:
    logging.warning(f"[tema] No se pudo registrar el tema institucional 'huv': {_e}")


# V6.9.27 - El estilo de graficos matplotlib ahora lo maneja la funcion
# aplicar_estilo_graficos() (theme-aware, definida mas arriba), que ya fue
# invocada con el estilo inicial. No se fija un estilo claro fijo aqui para
# no pisar el oscuro.


class App(ttk.Window):
    def __init__(self, info_usuario=None, tema="huv"):
        # Inicializar TTKBootstrap Window con el tema
        super().__init__(themename=tema)
        
        self.title("ONCONOVA · Gestión Oncológica Inteligente")
        self._configurar_icono_app()  # Icono institucional ONCONOVA
        aplicar_estilo_graficos(es_oscuro=False)  # V6.9.28 - app SOLO en claro (navy HUV)
        self.state('zoomed')  # Maximizar ventana

        # Información del usuario
        self.info_usuario = info_usuario or {"nombre": "Invitado", "cargo": "N/A", "ruta_foto": "SIN_FOTO", "ruta_directorio_fotos": ""}
        
        # Fuentes estándar para consistencia - usando las mismas del proyecto estadístico
        self.FONT_TITULO = ("Segoe UI", 22, "bold")
        self.FONT_SUBTITULO = ("Segoe UI", 12)
        self.FONT_NORMAL = ("Segoe UI", 11)
        self.FONT_ETIQUETA = ("Segoe UI", 9, "italic")
        self.FONT_NOMBRE_PERFIL = ("Segoe UI", 16, "bold")
        self.FONT_BOTONES = ("Segoe UI", 12)
        
        # Configurar tema actual
        self.current_theme = tema
        self.temas_disponibles = [
            'huv',  # V6.9.16 - Tema institucional HUV (#2d3e5e), primero por defecto
            'superhero', 'flatly', 'cyborg', 'journal', 'solar', 'darkly',
            'minty', 'pulse', 'sandstone', 'united', 'morph', 'vapor',
            'yeti', 'cosmo', 'litera', 'lumen', 'simplex', 'zephyr'
        ]
        
        # Configurar iconos y foto del usuario
        self.iconos = self._cargar_iconos()
        self.foto_usuario = self._cargar_foto_usuario()

        # Estados del nuevo sistema de navegación flotante
        self.header_visible = True
        self.sidebar_visible = False  # Sidebar no se usa en el nuevo diseño flotante
        self.floating_menu_visible = False
        self.welcome_screen_active = True
        self.current_view = "welcome"  # welcome, database, visualizar, dashboard

        # NUEVO: Variables para tracking de importación y auditoría IA
        self._ultimos_registros_procesados = []  # Lista de IDs recién importados
        self.ultimos_resultados_ia = None  # Resultados de última auditoría IA

        # Crear la interfaz
        self._create_layout()

    def _create_layout(self):
        """Crear la interfaz principal usando TTKBootstrap"""
        # Crear layout principal similar al proyecto estadístico
        main_frame = ttk.Frame(self, padding=0)
        main_frame.pack(expand=True, fill=BOTH)

        # ===== Header institucional =====
        self._create_header(main_frame)

        # Variables necesarias ANTES de crear la interfaz
        self.master_df = pd.DataFrame()  # DataFrame maestro (fuente única de verdad)
        self._compare_controls = {}

        # Inicializar componentes que serán referenciados (para evitar AttributeError)
        self.cmb_servicio = None
        self.cmb_malig = None
        self.cmb_resp = None
        self.tree = None
        self._ultimas_filas_seleccionadas = []

        # Inicializar sistema de exportación mejorado
        self.export_system = EnhancedExportSystem(self)

        # Variables para paneles flotantes
        self.details_panel = None
        self.filters_panel = None

        # ===== Separador eliminado (V6.9.16 - diseno minimalista sin linea) =====
        # Se mantiene como Frame invisible (height=0) para no romper las
        # referencias de _show_header / _hide_header. NO se empaqueta al inicio.
        self.header_separator = ttk.Frame(main_frame, height=0)

        # ===== Contenido principal (sin sidebar tradicional) =====
        self._create_main_content(main_frame)
        
        # ===== Menú flotante =====
        self._create_floating_menu()
        
        # ===== Botón flotante =====
        self._create_floating_button()

        # ===== Atajo de teclado: Ctrl+B abre/cierra el menu de navegacion (V6.9.16) =====
        self.bind_all("<Control-b>", lambda e: self._toggle_floating_menu())
        self.bind_all("<Control-B>", lambda e: self._toggle_floating_menu())

        # Inicializar estilo de treeview
        self._init_treeview_style()
        
        # Crear y mostrar pantalla de bienvenida inmediatamente
        self._create_welcome_screen()
        self.after(50, self.show_welcome_screen)

    def _create_header(self, parent):
        """Header institucional minimalista (V6.9.16).
        Solo titulo en azul institucional + perfil limpio a la derecha.
        Sin logos cuadrados, sin subtitulo gris, sin linea separadora."""
        self.header = ttk.Frame(parent, padding=(28, 16))
        self.header.pack(fill=X)

        # --- Titulo (izquierda) en azul institucional #2d3e5e ---
        center = ttk.Frame(self.header)
        center.pack(side=LEFT, expand=True, anchor=W)
        # Isotipo ONCONOVA a la izquierda del título
        self._header_logo_ref = self._cargar_logo_onconova(target=34)
        if self._header_logo_ref is not None:
            ttk.Label(center, image=self._header_logo_ref).pack(side=LEFT, padx=(0, 12))
        ttk.Label(
            center,
            text="ONCONOVA CIRUGÍA ONCOLÓGICA",
            font=("Segoe UI Semibold", 19),
            bootstyle="primary",
            anchor=W
        ).pack(side=LEFT, anchor=W)

        # --- Perfil (derecha): version + datos de usuario, estilo claro ---
        right = ttk.Frame(self.header)
        right.pack(side=RIGHT)

        # Badge de version (pill outline)
        version_btn = ttk.Button(
            right,
            text=f"v{get_version_string().split('-')[0].replace('v', '')}",
            command=self._show_version_info,
            bootstyle="primary-outline",
            width=7
        )
        version_btn.pack(side=RIGHT, padx=(14, 0))

        # Datos del usuario (texto plano, sin tarjeta oscura)
        datos = ttk.Frame(right)
        datos.pack(side=RIGHT)
        ttk.Label(
            datos,
            text=self.info_usuario.get("nombre", "Invitado"),
            font=("Segoe UI Semibold", 12),
            bootstyle="dark",
            anchor=E
        ).pack(anchor=E)
        ttk.Label(
            datos,
            text=self.info_usuario.get("cargo", "N/A"),
            font=("Segoe UI", 9),
            bootstyle="secondary",
            anchor=E
        ).pack(anchor=E)

    def _create_main_content(self, parent):
        """Crear el contenido principal sin sidebar tradicional"""
        # Contenedor principal que ocupa toda la ventana
        self.content_container = ttk.Frame(parent, padding=0)
        self.content_container.pack(expand=True, fill=BOTH)

        # Crear los diferentes paneles de contenido
        self._create_content_panels()

    def _create_floating_menu(self):
        """Menu de navegacion lateral minimalista y cohesivo (V6.9.16).
        Items uniformes estilo sidebar moderno: el item activo se resalta en
        azul institucional y el hover es sutil. Sin colores de semaforo ni
        relieve 3D. Header con el logo del hospital."""
        # --- Estilos propios de navegacion (sidebar) ---
        st = self.style
        st.configure(
            "Nav.TButton", font=("Segoe UI", 11), anchor="w",
            foreground="#2a2f3a", background="#ffffff",
            bordercolor="#ffffff", borderwidth=0, focusthickness=0,
            focuscolor="", relief="flat", padding=(18, 12)
        )
        st.map(
            "Nav.TButton",
            background=[("active", "#eef1f6"), ("pressed", "#e4e9f1")],
            foreground=[("active", "#2d3e5e")],
            relief=[("pressed", "flat"), ("active", "flat")]
        )
        st.configure(
            "NavActive.TButton", font=("Segoe UI Semibold", 11), anchor="w",
            foreground="#ffffff", background="#2d3e5e",
            bordercolor="#2d3e5e", borderwidth=0, focusthickness=0,
            focuscolor="", relief="flat", padding=(18, 12)
        )
        st.map(
            "NavActive.TButton",
            background=[("active", "#34466b"), ("pressed", "#26344f")],
            foreground=[("active", "#ffffff")],
            relief=[("pressed", "flat"), ("active", "flat")]
        )

        # --- Panel del menu (limpio, borde sutil, sin relieve 3D) ---
        self.floating_menu = ttk.Frame(self, padding=0, relief="solid", borderwidth=1)
        self.floating_menu.place(x=-320, y=20, width=300, height=500)

        # --- Header: logo del hospital + nombre ---
        header = ttk.Frame(self.floating_menu, padding=(22, 24, 22, 18))
        header.pack(fill=X)
        self._menu_logo_ref = self._cargar_logo_bienvenida(target=42, tinte="#2d3e5e")
        if self._menu_logo_ref is not None:
            ttk.Label(header, image=self._menu_logo_ref).pack(anchor="w")
        ttk.Label(
            header, text="HUV ONCOLOGÍA",
            font=("Segoe UI Semibold", 15), bootstyle="primary", anchor="w"
        ).pack(anchor="w", pady=(10, 0))
        ttk.Label(
            header, text="Panel de navegación",
            font=("Segoe UI", 9), bootstyle="secondary", anchor="w"
        ).pack(anchor="w")

        # --- Items de navegacion (cohesivos) ---
        nav_frame = ttk.Frame(self.floating_menu, padding=(12, 6))
        nav_frame.pack(fill=BOTH, expand=True)

        self.nav_buttons = {}
        nav_items = [
            ("Inicio", "home", self._nav_to_welcome),
            ("Base de Datos", "database", self._nav_to_database),
            ("Dashboard", "dashboard", self._nav_to_dashboard),
            # V6.9.44: "Análisis IA" OCULTO del menú (auditoría por LLM paralela al
            # flujo real con el agente data-auditor). La sección, _nav_to_analisis_ia
            # y los botones de Auditoría siguen intactos; solo se quitó esta entrada.
            # Reversible: descomentar la línea de abajo para volver a mostrarla.
            # ("Análisis IA", "analisis", self._nav_to_analisis_ia),
            ("Interoperabilidad QHORTE", "web", self._nav_to_web_auto),
        ]
        for text, view_id, callback in nav_items:
            btn = ttk.Button(
                nav_frame, text=text, style="Nav.TButton",
                command=lambda v=view_id, c=callback: self._on_nav_click(v, c),
                cursor="hand2", takefocus=False
            )
            btn.pack(fill=X, pady=2)
            self.nav_buttons[view_id] = btn

        # --- Footer: cerrar discreto ---
        footer = ttk.Frame(self.floating_menu, padding=(16, 12, 16, 18))
        footer.pack(fill=X, side=BOTTOM)
        ttk.Button(
            footer, text="✕   Cerrar menú", bootstyle="secondary-link",
            command=self._toggle_floating_menu, cursor="hand2", takefocus=False
        ).pack(anchor="w")

        self.menu_is_open = False
        # Marcar la vista inicial (bienvenida) como activa
        self._set_active_nav("home")

    def _on_nav_click(self, view_id, callback):
        """Marca el item como activo y ejecuta la navegacion."""
        self._set_active_nav(view_id)
        try:
            callback()
        except Exception as e:
            logging.error(f"[menu] Error navegando a {view_id}: {e}")

    def _set_active_nav(self, view_id):
        """Resalta en azul el item de navegacion activo; el resto queda neutro."""
        if not hasattr(self, "nav_buttons"):
            return
        for vid, btn in self.nav_buttons.items():
            try:
                btn.configure(style="NavActive.TButton" if vid == view_id else "Nav.TButton")
            except Exception:
                pass

    def _on_menu_btn_hover(self, button, entering):
        """Efectos hover para botones del menú"""
        if entering:
            # Efecto al entrar - escalar ligeramente
            button.configure(cursor="hand2")
            # Aquí se podría agregar más efectos visuales
        else:
            # Efecto al salir - restaurar
            button.configure(cursor="")

    def _create_floating_button(self):
        """Boton flotante circular navy para abrir el menu (V6.9.16).
        Cohesivo con el menu: circulo azul institucional con icono de menu
        (hamburguesa) en blanco y hover suave. Sin relieve 3D, sin vibracion
        ni cambio a naranja."""
        size = 42
        self._float_btn_size = size
        # V6.9.16: FAB DISCRETO. En reposo, gris muy claro con icono gris y
        # borde sutil (casi se funde con el fondo). Al pasar el mouse se tine
        # de navy. La via principal para abrir el menu es el atajo Ctrl+B.
        self._float_btn_color = "#eef1f6"    # reposo: gris azulado muy claro
        self._float_btn_icon = "#6c757d"     # reposo: icono gris medio
        self._float_btn_outline = "#d2d9e6"  # reposo: borde sutil
        self._float_btn_hover = "#2d3e5e"    # hover: navy institucional
        self._float_btn_x = 20               # margen izquierdo
        self._float_btn_yoff = 80            # px desde el borde INFERIOR (FAB)

        self.floating_btn_container = ttk.Frame(self, borderwidth=0)
        # V6.9.16: posicion FIJA abajo-izquierda (anchor sw) para no tapar
        # titulos ni contenido en ninguna vista. Coherente con el menu lateral.
        self.floating_btn_container.place(
            x=self._float_btn_x, rely=1.0, y=-self._float_btn_yoff,
            anchor="sw", width=size, height=size
        )

        # V6.9.44: el Canvas de tkinter NO puede ser transparente -> las esquinas
        # fuera del círculo siempre muestran su 'bg'. Antes se usaba el gris del
        # tema (TFrame), que contrastaba con el fondo BLANCO de las listas/tablas
        # (Treeview y tksheet usan #ffffff) y dibujaba un "cuadro" alrededor del
        # círculo. Se iguala al blanco del contenido para que las esquinas se
        # fundan y solo quede visible el círculo del FAB.
        canvas_bg = "#ffffff"

        self.floating_btn = tk.Canvas(
            self.floating_btn_container, width=size, height=size,
            highlightthickness=0, bd=0, bg=canvas_bg, cursor="hand2"
        )
        self.floating_btn.pack(expand=True, fill=BOTH)
        self._draw_floating_button(self._float_btn_color,
                                   icon_color=self._float_btn_icon,
                                   outline=self._float_btn_outline)

        # Eventos: click alterna el menu; hover cambia el color del circulo
        self.floating_btn.bind("<Button-1>", lambda e: self._toggle_floating_menu())
        self.floating_btn.bind("<Enter>", self._on_btn_hover_enter)
        self.floating_btn.bind("<Leave>", self._on_btn_hover_leave)

        # Animacion flotante eliminada (el FAB queda fijo, sin sube-y-baja)
        self.floating_btn_base_y = 150
        self.floating_animation_running = False
        self.hover_animation_running = False

    def _draw_floating_button(self, color, icon_color="#ffffff", outline=""):
        """Dibuja el circulo del FAB con el icono de menu (3 lineas).
        color = relleno del circulo; icon_color = color de las lineas del icono;
        outline = color del borde (cadena vacia = sin borde)."""
        if not hasattr(self, "floating_btn"):
            return
        c = self.floating_btn
        s = self._float_btn_size
        c.delete("all")
        # V6.9.22: pad mínimo -> el círculo llena casi todo el canvas, reduciendo
        # el "cuadro" del fondo a esquinas casi imperceptibles.
        pad = 1
        c.create_oval(pad, pad, s - pad, s - pad, fill=color,
                      outline=outline, width=1)
        cx, cy = s // 2, s // 2
        for dy in (-6, 0, 6):
            c.create_line(cx - 8, cy + dy, cx + 8, cy + dy,
                          fill=icon_color, width=2, capstyle="round")

    def _start_floating_animation(self):
        """Iniciar animación flotante continua sutil"""
        if not self.floating_animation_running and not self.floating_menu_visible:
            self.floating_animation_running = True
            self._animate_floating(0)

    def _animate_floating(self, step):
        """Animación flotante sutil - movimiento vertical suave como flotando"""
        if self.floating_animation_running and not self.floating_menu_visible:
            import math
            # Movimiento sinusoidal más sutil para botón pequeño (±3 pixels)  
            offset = math.sin(step * 0.08) * 3
            new_y = self.floating_btn_base_y + offset
            
            # Actualizar posición del botón con dimensiones más pequeñas
            if hasattr(self, 'floating_btn_container'):
                self.floating_btn_container.place(x=15, y=int(new_y), width=50, height=50)
            
            # Continuar la animación con ritmo más lento para mayor fluidez
            self.after(60, lambda: self._animate_floating(step + 1))
        elif self.floating_menu_visible:
            # Detener animación si el menú está visible
            self.floating_animation_running = False

    def _on_btn_hover_enter(self, event):
        """Hover: el FAB discreto se tine de navy con icono blanco."""
        self._draw_floating_button(self._float_btn_hover, icon_color="#ffffff", outline="")

    def _on_btn_hover_leave(self, event):
        """Reposo: vuelve al estilo discreto (gris claro, icono gris, borde sutil)."""
        self._draw_floating_button(self._float_btn_color,
                                   icon_color=self._float_btn_icon,
                                   outline=self._float_btn_outline)

    def _start_hover_vibration(self):
        """Animación de vibración cuando el mouse está encima"""
        self._vibrate_button(0)

    def _vibrate_button(self, step):
        """Efecto de vibración muy sutil para botón pequeño"""
        if self.hover_animation_running and step < 12:  # Vibración más corta para botón pequeño
            import random
            # Vibración muy sutil para botón compacto (±1 pixel)
            offset_x = random.randint(-1, 1)
            offset_y = random.randint(-1, 1)
            
            base_x = 15 + offset_x
            base_y = self.floating_btn_base_y + offset_y
            
            if hasattr(self, 'floating_btn_container'):
                self.floating_btn_container.place(x=base_x, y=base_y, width=50, height=50)
            
            # Continuar vibración con ritmo más rápido para efecto de tembleque
            self.after(30, lambda: self._vibrate_button(step + 1))
        elif self.hover_animation_running:
            # Reiniciar vibración para efecto continuo mientras hay hover
            self.after(100, lambda: self._vibrate_button(0))

    def _hide_floating_button(self):
        """Oculta el boton flotante al instante (queda cubierto por el menu
        que entra). V6.9.16: sin animacion de escala que recortaba el circulo."""
        self.floating_animation_running = False
        self.hover_animation_running = False
        if hasattr(self, 'floating_btn_container'):
            self.floating_btn_container.place_forget()

    def _show_floating_button(self):
        """Muestra el FAB deslizandolo suavemente desde la izquierda (ease-out),
        en su posicion fija abajo-izquierda. Sin recorte del circulo."""
        if not hasattr(self, 'floating_btn_container'):
            return
        size = getattr(self, '_float_btn_size', 50)
        yoff = getattr(self, '_float_btn_yoff', 80)
        end_x = getattr(self, '_float_btn_x', 20)
        start_x = -(size + 12)

        def slide_in_btn(step=0):
            if step <= 12:
                progress = step / 12
                eased = 1 - (1 - progress) ** 3  # ease-out cubica
                x = int(start_x + (end_x - start_x) * eased)
                self.floating_btn_container.place(
                    x=x, rely=1.0, y=-yoff, anchor="sw", width=size, height=size
                )
                self.after(12, lambda: slide_in_btn(step + 1))
            else:
                self.floating_btn_container.place(
                    x=end_x, rely=1.0, y=-yoff, anchor="sw", width=size, height=size
                )

        slide_in_btn()

    def _toggle_floating_menu(self):
        """Alternar visibilidad del menú flotante con animación"""
        if self.floating_menu_visible:
            self._hide_floating_menu()
        else:
            self._show_floating_menu()

    def _show_floating_menu(self):
        """Mostrar el menú flotante con animación suave y ocultar botón"""
        self.floating_menu_visible = True
        
        # Ocultar el botón flotante con animación de desvanecimiento
        self._hide_floating_button()
        
        # Animación suave con easing (ease-out)
        def slide_in(step=0):
            if step <= 25:
                # Función de easing ease-out cuádrica
                progress = step / 25
                eased_progress = 1 - (1 - progress) ** 2
                x_pos = -300 + (eased_progress * 300)  # De -300 a 0
                
                self.floating_menu.place(x=int(x_pos), y=20, width=300, height=500)
                self.after(12, lambda: slide_in(step + 1))
        
        slide_in()

    def _hide_floating_menu(self):
        """Ocultar el menú flotante con animación suave y mostrar botón"""
        self.floating_menu_visible = False
        
        # Animación suave con easing (ease-in)
        def slide_out(step=0):
            if step <= 25:
                # Función de easing ease-in cuádrica
                progress = step / 25
                eased_progress = progress ** 2
                x_pos = 0 - (eased_progress * 300)  # De 0 a -300
                
                self.floating_menu.place(x=int(x_pos), y=20, width=300, height=500)
                self.after(12, lambda: slide_out(step + 1))
            else:
                # Cuando la animación termine, mostrar el botón flotante
                self._show_floating_button()
        
        slide_out()

    # Funciones de navegación modernas
    def _nav_to_welcome(self):
        """Navegar a la pantalla de bienvenida"""
        self._hide_floating_menu()
        self.current_view = "welcome"
        self._show_header()  # Mostrar header solo en bienvenida
        self.show_welcome_screen()

    def _nav_to_database(self):
        """Navegar a la sección de base de datos"""
        self._hide_floating_menu()
        self._hide_header_if_not_welcome()
        self.current_view = "database"
        self._show_panel(self.database_frame)

    def _nav_to_visualizar(self):
        """Navegar a la pestaña Visualizador de Datos dentro de Base de Datos"""
        self._hide_floating_menu()
        self._hide_header_if_not_welcome()
        self.current_view = "database"
        self._show_panel(self.database_frame)

        # Seleccionar la pestaña del visualizador (índice 1)
        if hasattr(self, 'enhanced_dashboard') and hasattr(self.enhanced_dashboard, 'notebook'):
            try:
                self.enhanced_dashboard.notebook.select(1)  # Pestaña "Visualizador de Datos"
                logging.info("📊 Navegando a pestaña Visualizador de Datos en Base de Datos")
            except Exception as e:
                logging.error(f"Error seleccionando pestaña visualizador: {e}")

    def _nav_to_dashboard(self):
        """Navegar a la sección de dashboard"""
        self._hide_floating_menu()
        self._hide_header_if_not_welcome()
        self.current_view = "dashboard"
        self._show_panel(self.dashboard_frame)

        # CORREGIDO: Auto-cargar datos si están vacíos y cargar dashboard
        try:
            if self.master_df.empty:
                from core.database_manager import init_db, get_all_records_as_dataframe
                # V6.2.0: Comentado - init_db() ya se llama en ihq_processor antes del guardado
                # Llamarlo aquí causa que el UPDATE de relleno sobrescriba valores recién insertados
                # init_db()
                self.master_df = get_all_records_as_dataframe()

                # Ordenar por número de caso automáticamente
                if self.master_df is not None and not self.master_df.empty and "Numero de caso" in self.master_df.columns:
                    self.master_df = self.master_df.sort_values(
                        by="Numero de caso",
                        ascending=True,
                        na_position='last'
                    ).reset_index(drop=True)
            self.cargar_dashboard()
        except Exception as e:
            logging.error(f"Error auto-cargando dashboard: {e}")
            # Aún cargar dashboard vacío para mostrar interfaz
            self.cargar_dashboard()

    def _nav_to_web_auto(self):
        """Navegar a la sección de automatización web"""
        self._hide_floating_menu()
        self._hide_header_if_not_welcome()
        messagebox.showinfo("Web Automation", "Función de automatización web - En desarrollo")

    def _nav_to_analisis_ia(self):
        """Navegar a sección de Análisis con IA"""
        self._hide_floating_menu()
        self._hide_header_if_not_welcome()
        self.current_view = "analisis_ia"
        self._show_panel(self.analisis_ia_frame)

    def _mostrar_selector_tipo_auditoria(self, tipo_auditoria, registros_incompletos=None):
        """
        Muestra ventana para seleccionar tipo de auditoría (Parcial o Completa)

        Args:
            tipo_auditoria: 'parcial' (valor predefinido desde ventana de resultados)
            registros_incompletos: Lista de registros incompletos
        """
        # Guardar registros incompletos para usar después
        self._registros_incompletos_temp = registros_incompletos

        from core.ventana_selector_auditoria import mostrar_selector_auditoria

        # Mostrar selector
        mostrar_selector_auditoria(
            parent=self,
            callback_seleccion=self._iniciar_auditoria_ia
        )

    def _iniciar_auditoria_ia(self, tipo_auditoria):
        """
        Callback cuando usuario elige auditar con IA

        Args:
            tipo_auditoria: 'parcial' o 'completa'
        """
        logging.info(f"Iniciando auditoria IA - Tipo: {tipo_auditoria}")

        # Guardar tipo de auditoría para usar en el callback de resultados
        self._tipo_auditoria_actual = tipo_auditoria

        # Recuperar registros incompletos guardados
        registros_incompletos = getattr(self, '_registros_incompletos_temp', None)

        if tipo_auditoria == 'parcial' and registros_incompletos:
            # Auditoría solo de registros incompletos
            try:
                from core.auditoria_parcial import auditar_registros_incompletos

                numeros_peticion = [r['numero_peticion'] for r in registros_incompletos]
                logging.info(f"Auditando {len(numeros_peticion)} registros incompletos")

                # Esto mostrará VentanaAuditoriaIA automáticamente
                auditar_registros_incompletos(
                    numeros_peticion=numeros_peticion,
                    parent=self,
                    callback_completado=self._mostrar_resultados_auditoria
                )
            except ImportError:
                # Fallback: usar auditoría completa si la parcial no está implementada
                logging.warning("Auditoria parcial no disponible, usando auditoria completa")
                from core.ventana_auditoria_ia import mostrar_ventana_auditoria

                mostrar_ventana_auditoria(
                    parent=self,
                    callback_completado=self._mostrar_resultados_auditoria
                )

        elif tipo_auditoria == 'completa':
            # Auditoría COMPLETA - Solo registros recién importados (igual que PARCIAL)
            logging.info("Auditando registros recien importados con analisis profundo")

            # Obtener registros recién importados
            ultimos_registros = getattr(self, '_ultimos_registros_procesados', [])

            if not ultimos_registros:
                messagebox.showwarning(
                    "Sin registros para auditar",
                    "No hay registros recién importados para auditar.\n\n"
                    "La auditoría COMPLETA solo procesa los casos que acabas de importar.\n\n"
                    "Para auditar casos específicos:\n"
                    "1. Ve a 'Visualizar datos'\n"
                    "2. Selecciona los casos que deseas auditar\n"
                    "3. (Funcionalidad en desarrollo)"
                )
                return

            logging.info(f"Registros recien importados: {len(ultimos_registros)}")

            try:
                from pathlib import Path
                import glob
                from core.debug_mapper import DebugMapper
                from core.database_manager import get_registro_by_peticion

                # Preparar casos para auditoría COMPLETA
                project_root = Path(__file__).parent
                casos_preparados = []

                for numero in ultimos_registros:
                    try:
                        # Obtener datos del registro de BD
                        registro_bd = get_registro_by_peticion(numero)
                        if not registro_bd:
                            logging.warning(f"No se encontro registro en BD para {numero}")
                            continue

                        # Cargar debug_map para tener el PDF completo
                        debug_maps_dir = project_root / "data" / "debug_maps"
                        pattern = str(debug_maps_dir / f"debug_map_{numero}_*.json")
                        debug_map_files = glob.glob(pattern)

                        if debug_map_files:
                            # Usar el más reciente
                            debug_map_path = Path(sorted(debug_map_files)[-1])
                            try:
                                debug_map = DebugMapper.cargar_mapa(debug_map_path)
                            except Exception as e:
                                logging.warning(f"Error cargando debug_map para {numero}: {e}")
                                debug_map = {}
                        else:
                            logging.warning(f"No se encontro debug_map para {numero}")
                            debug_map = {}

                        # Preparar caso para auditoría COMPLETA
                        caso = {
                            'numero_peticion': numero,
                            'datos_bd': registro_bd,
                            'debug_map': debug_map,
                            'modo': 'completa'
                        }

                        casos_preparados.append(caso)

                    except Exception as e:
                        logging.error(f"Error preparando {numero}: {e}")
                        continue

                if not casos_preparados:
                    logging.error("No se pudieron preparar casos para auditoria")
                    messagebox.showerror(
                        "Error",
                        "No se pudieron preparar los casos para auditoría.\n"
                        "Verifique que existan debug_maps en data/debug_maps/"
                    )
                    return

                logging.info(f"{len(casos_preparados)} casos preparados para auditoria COMPLETA")

                # Mostrar ventana de auditoría con casos recién importados
                from core.ventana_auditoria_ia import mostrar_ventana_auditoria

                mostrar_ventana_auditoria(
                    parent=self,
                    casos=casos_preparados,
                    modo='completa',
                    callback_completado=self._mostrar_resultados_auditoria
                )

            except Exception as e:
                logging.error(f"Error preparando auditoria completa: {e}", exc_info=True)
                messagebox.showerror(
                    "Error",
                    f"Error preparando auditoría completa:\n{str(e)}"
                )

    def _auditar_seleccion_parcial(self):
        """V3.2.4: Auditar el item seleccionado en modo PARCIAL"""
        self._auditar_item_seleccionado(modo='parcial')

    def _auditar_seleccion_completa(self):
        """V3.2.4: Auditar el item seleccionado en modo COMPLETA"""
        self._auditar_item_seleccionado(modo='completa')

    def _auditar_item_seleccionado(self, modo='completa'):
        """
        V3.2.4.2: Audita los items seleccionados en la tabla (1 o múltiples)

        Comportamiento inteligente:
        - 1 item seleccionado: Auditar individualmente
        - Múltiples items + COMPLETA: Auditar 1 por 1
        - Múltiples items + PARCIAL: Auditar en lotes de 3 (como auditoría masiva)

        Args:
            modo: 'parcial' o 'completa'
        """
        # Obtener selección
        selection = self.tree.selection()
        if not selection:
            messagebox.showwarning(
                "Seleccionar registro",
                "Por favor selecciona al menos UN registro para auditar."
            )
            return

        num_seleccionados = len(selection)
        logging.info(f"Auditando {num_seleccionados} caso(s) seleccionado(s) (modo: {modo.upper()})")

        # V3.2.4.2: FIX 10 - Filtrado inteligente según estado de auditoría
        from core.database_manager import get_estado_auditoria

        # Extraer números de petición y filtrar según estado
        # v6.0.12: CORREGIDO - Sheet no tiene atributo ["columns"], usar método alternativo
        try:
            headers = self.sheet.headers() if hasattr(self.sheet, 'headers') else []
            col_idx = headers.index("Numero de caso") if "Numero de caso" in headers else 0
        except:
            col_idx = 0  # Fallback: primera columna es "Numero de caso"

        numeros_peticion = []
        casos_omitidos_info = {
            'ya_parcial': [],
            'ya_completa': [],
            'error': []
        }

        for item_id in selection:
            # tksheet: Usar get_row_data() en lugar de .item()
            values = self.sheet.get_row_data(item_id)
            try:
                if not values or len(values) <= col_idx:
                    logging.warning(f"No se pudieron obtener valores para item_id={item_id}")
                    continue
                numero = values[col_idx]
                estado = get_estado_auditoria(numero)

                # Lógica de filtrado inteligente
                if modo == 'parcial':
                    # AUDITORÍA PARCIAL: Solo procesar casos SIN auditoría
                    if estado == "PARCIAL":
                        casos_omitidos_info['ya_parcial'].append(numero)
                        logging.info(f"Omitiendo {numero}: Ya tiene auditoria PARCIAL")
                        continue
                    elif estado == "COMPLETA":
                        casos_omitidos_info['ya_completa'].append(numero)
                        logging.info(f"Omitiendo {numero}: Ya tiene auditoria COMPLETA")
                        continue
                    else:
                        # NULL o sin auditar → Procesar
                        numeros_peticion.append(numero)

                elif modo == 'completa':
                    # AUDITORÍA COMPLETA: Procesar NULL y PARCIAL, omitir COMPLETA
                    if estado == "COMPLETA":
                        casos_omitidos_info['ya_completa'].append(numero)
                        logging.info(f"Omitiendo {numero}: Ya tiene auditoria COMPLETA")
                        continue
                    else:
                        # NULL o PARCIAL → Procesar (PARCIAL se upgrade a COMPLETA)
                        numeros_peticion.append(numero)
                        if estado == "PARCIAL":
                            logging.info(f"{numero}: Upgrade de PARCIAL a COMPLETA")

            except (ValueError, IndexError) as e:
                casos_omitidos_info['error'].append(numero if 'numero' in locals() else 'desconocido')
                logging.warning(f"Error obteniendo numero de peticion de item {item_id}: {e}")
                continue

        # Mostrar resumen de omisiones si hay
        total_omitidos = len(casos_omitidos_info['ya_parcial']) + len(casos_omitidos_info['ya_completa']) + len(casos_omitidos_info['error'])

        if total_omitidos > 0:
            mensaje_omisiones = f"De {num_seleccionados} casos seleccionados:\n\n"
            mensaje_omisiones += f"✅ Se procesarán: {len(numeros_peticion)}\n"
            mensaje_omisiones += f"⏭️ Se omitirán: {total_omitidos}\n\n"

            if casos_omitidos_info['ya_parcial']:
                mensaje_omisiones += f"• {len(casos_omitidos_info['ya_parcial'])} ya tienen PARCIAL\n"
            if casos_omitidos_info['ya_completa']:
                mensaje_omisiones += f"• {len(casos_omitidos_info['ya_completa'])} ya tienen COMPLETA\n"
            if casos_omitidos_info['error']:
                mensaje_omisiones += f"• {len(casos_omitidos_info['error'])} con errores\n"

            logging.info(f"Resumen de filtrado:")
            logging.info(f"A procesar: {len(numeros_peticion)}")
            logging.info(f"Omitidos: {total_omitidos}")

            # Mostrar mensaje informativo al usuario
            if len(numeros_peticion) > 0:
                # Hay casos para procesar, informar omisiones
                messagebox.showinfo(
                    "Filtrado Inteligente",
                    mensaje_omisiones + f"\n¿Deseas continuar con los {len(numeros_peticion)} casos?"
                )

        # Verificar si quedan casos para procesar
        if not numeros_peticion:
            # Mensaje personalizado según el motivo
            if casos_omitidos_info['ya_completa'] and modo == 'completa':
                messagebox.showinfo(
                    "Nada que procesar",
                    f"Todos los {num_seleccionados} caso(s) seleccionado(s) ya tienen auditoría COMPLETA.\n\n"
                    f"✅ No es necesario volver a auditarlos."
                )
            elif casos_omitidos_info['ya_parcial'] and casos_omitidos_info['ya_completa'] and modo == 'parcial':
                messagebox.showinfo(
                    "Nada que procesar",
                    f"Todos los {num_seleccionados} caso(s) seleccionado(s) ya tienen auditoría.\n\n"
                    f"• {len(casos_omitidos_info['ya_parcial'])} con PARCIAL\n"
                    f"• {len(casos_omitidos_info['ya_completa'])} con COMPLETA\n\n"
                    f"💡 Selecciona casos sin auditar para procesarlos."
                )
            else:
                messagebox.showerror(
                    "Error",
                    f"No se pudo obtener los números de petición de los registros seleccionados.\n\n"
                    f"Casos con error: {len(casos_omitidos_info['error'])}"
                )
            return

        # V3.2.4.2: Lógica inteligente según número de items a procesar
        try:
            from pathlib import Path
            import glob
            from core.debug_mapper import DebugMapper
            from core.database_manager import get_registro_by_peticion
            from core.ventana_auditoria_ia import mostrar_ventana_auditoria

            project_root = Path(__file__).parent
            debug_maps_dir = project_root / "data" / "debug_maps"

            # Campos protegidos (no auditar)
            CAMPOS_PROTEGIDOS = [
                "Primer nombre", "Segundo nombre", "Primer apellido", "Segundo apellido",
                "N. de identificación", "Edad", "Genero",
                "Fecha de ingreso (2. Fecha de la muestra)", "Tipo de documento",
                "Numero de caso"
            ]

            # Preparar TODOS los casos seleccionados
            casos_preparados = []
            casos_omitidos = 0

            for numero_peticion in numeros_peticion:
                logging.info(f"Preparando caso: {numero_peticion}")

                # Obtener datos del registro de BD
                registro_bd = get_registro_by_peticion(numero_peticion)
                if not registro_bd:
                    logging.warning(f"No se encontro en BD, omitiendo {numero_peticion}")
                    casos_omitidos += 1
                    continue

                # Cargar debug_map
                pattern = str(debug_maps_dir / f"debug_map_{numero_peticion}_*.json")
                debug_map_files = glob.glob(pattern)

                if debug_map_files:
                    debug_map_path = Path(sorted(debug_map_files)[-1])
                    try:
                        debug_map = DebugMapper.cargar_mapa(debug_map_path)
                        logging.info(f"Debug map cargado para {numero_peticion}")
                    except Exception as e:
                        logging.warning(f"Error cargando debug_map para {numero_peticion}: {e}")
                        debug_map = {}
                else:
                    logging.warning(f"No se encontro debug_map para {numero_peticion}")
                    debug_map = {}

                # Pre-check para modo PARCIAL (V5.1.2: Usar validation_checker)
                campos_vacios = []
                if modo == 'parcial':
                    # V5.1.2: FIX - Usar validation_checker en lugar de chequeo manual
                    # Esto filtra solo los campos REALMENTE faltantes (basándose en ESTUDIOS SOLICITADOS)
                    from core.validation_checker import verificar_completitud_registro

                    analisis = verificar_completitud_registro(numero_peticion)
                    campos_vacios = (
                        analisis.get('campos_faltantes', []) +
                        analisis.get('biomarcadores_faltantes', [])
                    )

                    logging.info(f"Campos realmente faltantes en {numero_peticion}: {len(campos_vacios)}")

                    # Si no hay campos faltantes, omitir este caso en modo PARCIAL
                    if len(campos_vacios) == 0:
                        logging.info(f"Caso {numero_peticion} completo, omitiendo en modo PARCIAL")
                        casos_omitidos += 1
                        continue

                # Preparar caso
                caso = {
                    'numero_peticion': numero_peticion,
                    'datos_bd': registro_bd,
                    'debug_map': debug_map
                }

                # Agregar campos_a_buscar en modo PARCIAL
                if modo == 'parcial':
                    caso['campos_a_buscar'] = campos_vacios
                    # V3.2.4.2: FIX 8 - Agregar batch_size para procesamiento por lotes
                    caso['batch_size'] = 3  # Lotes de 3 casos simultáneos

                casos_preparados.append(caso)

            # Verificar si hay casos para procesar
            if not casos_preparados:
                if modo == 'parcial' and casos_omitidos > 0:
                    messagebox.showinfo(
                        "Nada que auditar",
                        f"Los {casos_omitidos} caso(s) seleccionado(s) ya están completos.\n\n"
                        f"✅ No hay campos vacíos que completar en modo PARCIAL.\n\n"
                        f"💡 Si deseas verificar la calidad de los datos existentes,\n"
                        f"   usa 'Auditoría COMPLETA' en su lugar."
                    )
                else:
                    messagebox.showerror(
                        "Error",
                        "No se pudo preparar ningún caso para auditar"
                    )
                return

            num_casos = len(casos_preparados)
            logging.info(f"{num_casos} caso(s) preparado(s) para auditoria {modo.upper()}")
            if casos_omitidos > 0:
                logging.info(f"{casos_omitidos} caso(s) omitido(s)")

            # Guardar el modo para el callback
            self._modo_auditoria_seleccion = modo

            # DECISIÓN INTELIGENTE: ¿Procesamiento individual o por lotes?
            if num_casos == 1:
                # UN SOLO CASO: Procesar individualmente (más info en UI)
                logging.info(f"Procesamiento INDIVIDUAL (1 caso)")
                mostrar_ventana_auditoria(
                    parent=self,
                    casos=casos_preparados,
                    modo=modo,
                    callback_completado=self._callback_auditoria_seleccion
                )

            elif modo == 'completa':
                # MÚLTIPLES + COMPLETA: Procesar 1 por 1 (análisis profundo)
                logging.info(f"Procesamiento SECUENCIAL (modo COMPLETA, 1 por 1)")
                mostrar_ventana_auditoria(
                    parent=self,
                    casos=casos_preparados,
                    modo='completa',
                    callback_completado=self._callback_auditoria_seleccion
                )

            else:  # modo == 'parcial' and num_casos > 1
                # MÚLTIPLES + PARCIAL: Procesar en LOTES de 3 (como auditoría masiva)
                num_lotes = (num_casos + 2) // 3  # Redondear hacia arriba
                logging.info(f"Procesamiento POR LOTES (modo PARCIAL)")
                logging.info(f"{num_casos} casos = {num_lotes} lote(s) de 3")
                logging.info(f"Tiempo estimado: ~{num_lotes * 30}s")

                mostrar_ventana_auditoria(
                    parent=self,
                    casos=casos_preparados,
                    modo='parcial',
                    callback_completado=self._callback_auditoria_seleccion
                )

        except Exception as e:
            logging.error(f"Error preparando auditoria: {e}", exc_info=True)
            messagebox.showerror(
                "Error",
                f"Error preparando auditoría:\n{str(e)}"
            )

    def _callback_auditoria_seleccion(self, resultados):
        """
        V3.2.4: Callback específico para auditoría de selección
        V3.2.4.2: Simplificado - La actualización de estados ahora se hace en _mostrar_resultados_auditoria()
        """
        logging.info(f"Auditoria de seleccion completada")

        if not resultados:
            logging.warning(f"No hay resultados")
            return

        # V3.2.4.2: FIX 7 - La actualización de estados ahora es centralizada
        # _mostrar_resultados_auditoria() se encarga de actualizar BD y refrescar UI
        self._mostrar_resultados_auditoria(resultados)

    def _mostrar_resultados_auditoria(self, resultados):
        """
        Callback cuando termina la auditoría IA
        Genera reporte Markdown y navega a la sección de Análisis IA

        Args:
            resultados: Dict con resultados de auditoría
        """
        logging.info(f"Auditoria completada - Mostrando resultados")

        # Guardar resultados en variable de instancia
        self.ultimos_resultados_ia = resultados

        # V3.2.4.2: FIX 7 - Actualizar estados de auditoría en BD ANTES de refrescar UI
        # Determinar tipo de auditoría (intentar ambas variables para soportar ambos flujos)
        tipo_auditoria = getattr(self, '_modo_auditoria_seleccion', None) or getattr(self, '_tipo_auditoria_actual', 'completa')

        # Actualizar estado en BD para cada caso auditado exitosamente
        from core.database_manager import set_estado_auditoria

        if isinstance(resultados, list):
            # Formato: lista de resultados individuales
            for resultado in resultados:
                if resultado.get('exito'):
                    numero_peticion = resultado.get('numero_peticion')
                    if numero_peticion:
                        estado = "PARCIAL" if tipo_auditoria == 'parcial' else "COMPLETA"
                        set_estado_auditoria(numero_peticion, estado)
                        logging.info(f"Estado actualizado: {numero_peticion} -> {estado}")
        elif isinstance(resultados, dict) and 'resultados' in resultados:
            # Formato: dict con clave 'resultados'
            for resultado in resultados.get('resultados', []):
                if resultado.get('exito'):
                    numero_peticion = resultado.get('numero_peticion')
                    if numero_peticion:
                        estado = "PARCIAL" if tipo_auditoria == 'parcial' else "COMPLETA"
                        set_estado_auditoria(numero_peticion, estado)
                        logging.info(f"Estado actualizado: {numero_peticion} -> {estado}")

        # Refrescar tabla DESPUÉS de actualizar BD
        self.refresh_data_and_table()

        # Generar reporte Markdown
        logging.info(f"Generando reporte Markdown...")
        ruta_reporte = self._generar_reporte_ia(resultados, tipo_auditoria)

        if ruta_reporte:
            logging.info(f"Reporte generado exitosamente: {ruta_reporte}")

            # Navegar a la sección de Análisis IA
            logging.info(f"Navegando a seccion Analisis IA...")
            try:
                self._nav_to_analisis_ia()
                logging.info(f"Navegacion exitosa")
            except Exception as e:
                logging.error(f"Error navegando: {e}", exc_info=True)

            # Actualizar lista de reportes
            try:
                self._actualizar_lista_reportes()
                logging.info(f"Lista de reportes actualizada")
            except Exception as e:
                logging.warning(f"Error actualizando lista de reportes: {e}")
                # Continuar sin fallar

            # Seleccionar automáticamente el reporte recién generado
            try:
                self._seleccionar_ultimo_reporte()
                logging.info(f"Reporte seleccionado automaticamente")
            except Exception as e:
                logging.warning(f"Error seleccionando reporte: {e}")
                # Continuar sin fallar

            # V2.1.6: No mostrar mensaje automáticamente
            # El mensaje se mostrará cuando el usuario haga clic en "Ver Resultados"
            # Guardar ruta para mostrar después
            self._ruta_ultimo_reporte = ruta_reporte
        else:
            logging.error(f"Error generando reporte")
            messagebox.showerror(
                "Error",
                "La auditoría se completó pero hubo un error al generar el reporte."
            )
            # Navegar al visualizador como fallback
            self._nav_to_visualizar()

    def _show_version_info(self):
        """Mostrar información detallada de la versión del sistema"""
        self._hide_floating_menu()
        
        try:
            version_info = get_full_version_info()
            actual_deps = get_dependencies_actual()
            
            # Crear ventana modal
            version_window = ttk.Toplevel(self)
            version_window.title(f"Acerca de - {version_info['project']['name']}")
            version_window.resizable(True, True)

            # Configurar fondo gris claro
            version_window.configure(bg='#f0f0f0')

            # V6.9.44: ventana ADAPTABLE. Antes se abría maximizada ('zoomed') y no
            # encajaba bien. Ahora abre en un tamaño moderado, CENTRADA sobre la app y
            # redimensionable; el contenido está en pestañas (General + Sistema con
            # scroll), así que se ve completo aunque la ventana sea pequeña.
            win_w, win_h = 860, 640
            self.update_idletasks()
            sw = version_window.winfo_screenwidth()
            sh = version_window.winfo_screenheight()
            px = self.winfo_rootx() + max((self.winfo_width() - win_w) // 2, 0)
            py = self.winfo_rooty() + max((self.winfo_height() - win_h) // 2, 0)
            # No dejar que se salga de la pantalla
            px = max(0, min(px, sw - win_w))
            py = max(0, min(py, sh - win_h))
            version_window.geometry(f"{win_w}x{win_h}+{px}+{py}")
            version_window.minsize(560, 440)

            # Después de maximizar, configurar modal
            version_window.transient(self)
            version_window.grab_set()
            
            # Frame principal
            main_frame = ttk.Frame(version_window, padding=10)
            main_frame.pack(fill=BOTH, expand=True)
            
            # Header con información principal
            header_frame = ttk.Frame(main_frame, bootstyle="primary", padding=15)
            header_frame.pack(fill=X, pady=(0, 10))
            
            ttk.Label(
                header_frame,
                text=version_info['project']['name'],
                font=("Arial", 18, "bold"),
                bootstyle="inverse-primary"
            ).pack()
            
            ttk.Label(
                header_frame,
                text=f"{get_version_string()} | {get_build_info()}",
                font=("Arial", 12),
                bootstyle="inverse-primary"
            ).pack(pady=(5, 0))
            
            ttk.Label(
                header_frame,
                text=version_info['project']['description'],
                font=("Arial", 10),
                bootstyle="inverse-primary"
            ).pack(pady=(5, 0))
            
            # Notebook para las diferentes secciones
            notebook = ttk.Notebook(main_frame)
            notebook.pack(fill=BOTH, expand=True, pady=10)
            
            # Tab 1: Información General - Compacta, sin scroll
            info_frame = ttk.Frame(notebook, padding=15)
            notebook.add(info_frame, text="📋 General")

            # Frame para centrar el contenido
            info_center = ttk.Frame(info_frame)
            info_center.pack(expand=True)
            
            # Traducir tipo de release a español
            release_type_es = {
                'stable': 'Estable',
                'beta': 'Beta',
                'alpha': 'Alfa',
                'rc': 'Release Candidate'
            }.get(version_info['version']['release_type'].lower(), version_info['version']['release_type'])

            self._create_info_section(info_center, "Información del Proyecto", [
                ("Nombre Completo", version_info['project']['full_name']),
                ("Organización", version_info['project']['organization']),
                ("Versión", version_info['version']['version']),
                ("Nombre de Versión", version_info['version']['version_name']),
                ("Tipo de Release", release_type_es),
                ("Fecha de Build", version_info['version']['build_date']),
                ("Número de Build", version_info['version']['build_number'])
            ])
            
            # Tab 2: Sistema - Con scroll habilitado
            system_frame = ttk.Frame(notebook, padding=10)
            notebook.add(system_frame, text="💻 Sistema")

            # Crear canvas y scrollbar para permitir scroll
            system_canvas = tk.Canvas(system_frame, bg='#f0f0f0', highlightthickness=0)
            system_scrollbar = ttk.Scrollbar(system_frame, orient="vertical", command=system_canvas.yview)
            system_scrollable = ttk.Frame(system_canvas)

            system_scrollable.bind(
                "<Configure>",
                lambda e: system_canvas.configure(scrollregion=system_canvas.bbox("all"))
            )

            system_canvas.create_window((0, 0), window=system_scrollable, anchor="nw")
            system_canvas.configure(yscrollcommand=system_scrollbar.set)

            system_canvas.pack(side="left", fill="both", expand=True)
            system_scrollbar.pack(side="right", fill="y")

            # Configurar columnas para distribución 50/50
            system_scrollable.columnconfigure(0, weight=1)
            system_scrollable.columnconfigure(1, weight=1)

            # Variables para trackear filas
            left_row = 0
            right_row = 0

            # COLUMNA IZQUIERDA
            # Información básica del sistema
            basic_system_info = [
                ("Versión Python", version_info['system']['python_version'].split()[0]),
                ("Plataforma", version_info['system']['platform']),
                ("Sistema", version_info['system'].get('system', 'No disponible')),
                ("Release", version_info['system'].get('release', 'No disponible')),
                ("Arquitectura", version_info['system']['architecture']),
                ("Máquina", version_info['system'].get('machine', 'No disponible')),
                ("Nodo", version_info['system'].get('node', 'No disponible')),
                ("Procesador", version_info['system']['processor'] or "No disponible")
            ]
            self._create_info_section_grid(system_scrollable, "Información Básica", basic_system_info, row=left_row, column=0)
            left_row += 1

            # COLUMNA DERECHA
            # Información de memoria
            if 'memoria_total' in version_info['system']:
                memory_info = [
                    ("Memoria Total", version_info['system']['memoria_total']),
                    ("Memoria Disponible", version_info['system']['memoria_disponible']),
                    ("Memoria Usada", version_info['system']['memoria_usada']),
                    ("Porcentaje Usado", version_info['system']['memoria_porcentaje'])
                ]
                self._create_info_section_grid(system_scrollable, "Información de Memoria", memory_info, row=right_row, column=1)
                right_row += 1

            # Información de CPU
            if 'cpu_cores' in version_info['system']:
                cpu_info = [
                    ("Núcleos Físicos", str(version_info['system']['cpu_cores'])),
                    ("Hilos Lógicos", str(version_info['system']['cpu_threads'])),
                    ("Frecuencia Máxima", version_info['system']['cpu_frecuencia'])
                ]
                self._create_info_section_grid(system_scrollable, "Información del Procesador", cpu_info, row=right_row, column=1)
                right_row += 1

            # Información de hardware adicional
            hardware_info = []
            if 'tarjeta_grafica' in version_info['system']:
                gpus = version_info['system']['tarjeta_grafica']
                if isinstance(gpus, list):
                    for i, gpu in enumerate(gpus):
                        hardware_info.append((f"Tarjeta Gráfica {i+1}", gpu))
                else:
                    hardware_info.append(("Tarjeta Gráfica", str(gpus)))

            if 'placa_madre' in version_info['system']:
                hardware_info.append(("Placa Madre", version_info['system']['placa_madre']))

            if hardware_info:
                self._create_info_section_grid(system_scrollable, "Hardware", hardware_info, row=right_row, column=1)
                right_row += 1

            # Información de discos - ANCHO COMPLETO (debajo de ambas columnas)
            disk_row = max(left_row, right_row)  # Empezar después de la columna más larga
            if 'discos' in version_info['system'] and isinstance(version_info['system']['discos'], list):
                for i, disco in enumerate(version_info['system']['discos']):
                    if isinstance(disco, dict):
                        disk_info = [
                            ("Dispositivo", disco.get('dispositivo', 'No disponible')),
                            ("Punto de Montaje", disco.get('punto_montaje', 'No disponible')),
                            ("Sistema de Archivos", disco.get('sistema_archivos', 'No disponible')),
                            ("Espacio Total", disco.get('total', 'No disponible')),
                            ("Espacio Usado", disco.get('usado', 'No disponible')),
                            ("Espacio Libre", disco.get('libre', 'No disponible')),
                            ("Porcentaje Usado", disco.get('porcentaje', 'No disponible'))
                        ]
                        self._create_info_section_grid(system_scrollable, f"Disco {i+1}", disk_info, row=disk_row+i, column=0, columnspan=2)
            
            # Tab 3: Equipo de Desarrollo - Centrado, sin espacios en blanco
            team_frame = ttk.Frame(notebook, padding=20)
            notebook.add(team_frame, text="👥 Equipo")

            # Frame para centrar el contenido
            team_center = ttk.Frame(team_frame)
            team_center.pack(expand=True)
            
            # Información del equipo
            role_titles = {
                'desarrollador': '👨‍💻 Desarrollador',
                'lider_investigacion': '👨‍⚕️ Líder de Investigación y Proyección Oncológica',
                'jefe_gestion_informacion': '👨‍💼 Jefe de Gestión de la Información'
            }

            # FIX: role_info puede ser un dict (un miembro) o una lista de dicts
            # (p. ej. 'desarrolladores'). Antes se indexaba la lista con string ->
            # "list indices must be integers or slices, not str".
            for role_key, role_info in version_info['team'].items():
                personas = role_info if isinstance(role_info, list) else [role_info]
                for persona in personas:
                    if not isinstance(persona, dict):
                        continue
                    role_data = [
                        ("Nombre", persona.get('nombre', 'No disponible')),
                        ("Cargo", persona.get('cargo', 'No disponible')),
                        ("Departamento", persona.get('departamento', 'No disponible')),
                        ("Correo", persona.get('correo', 'No disponible'))
                    ]
                    title = role_titles.get(role_key, persona.get('cargo', role_key))
                    self._create_info_section(team_center, title, role_data)

            # Frame de botones
            buttons_frame = ttk.Frame(main_frame, padding=10)
            buttons_frame.pack(fill=X, pady=10)

            # Botón dinámico que cambia según la pestaña
            copy_button = ttk.Button(
                buttons_frame,
                text="📋 Copiar Información",
                bootstyle="info"
            )
            copy_button.pack(side=LEFT, padx=(0, 10))

            # Función para actualizar el botón según la pestaña activa
            def update_copy_button():
                current_tab = notebook.index(notebook.select())
                if current_tab == 0:  # General
                    copy_button.configure(
                        text="📋 Copiar Info General",
                        command=lambda: self._copy_general_info(version_info),
                        state="normal"
                    )
                elif current_tab == 1:  # Sistema
                    copy_button.configure(
                        text="📋 Copiar Info Sistema",
                        command=lambda: self._copy_system_info(version_info),
                        state="normal"
                    )
                elif current_tab == 2:  # Equipo
                    # Ocultar el botón en la pestaña Equipo
                    copy_button.pack_forget()
                    return
                else:
                    copy_button.configure(state="disabled")

                # Asegurar que el botón esté visible si no es la pestaña Equipo
                if not copy_button.winfo_ismapped():
                    copy_button.pack(side=LEFT, padx=(0, 10))

            # Bind para actualizar cuando cambie de pestaña
            notebook.bind("<<NotebookTabChanged>>", lambda e: update_copy_button())

            # Inicializar el botón
            update_copy_button()

            ttk.Button(
                buttons_frame,
                text="✅ Cerrar",
                command=version_window.destroy,
                bootstyle="success"
            ).pack(side=RIGHT)
            
            # Habilitar scroll con la rueda del mouse en la pestaña Sistema
            def _on_mousewheel(event):
                try:
                    # Obtener el notebook tab actual
                    current_tab = notebook.index(notebook.select())

                    # Scroll solo en pestaña Sistema (tab 1)
                    if current_tab == 1:
                        system_canvas.yview_scroll(int(-1*(event.delta/120)), "units")
                except:
                    pass

            # Bind a la ventana completa
            version_window.bind_all("<MouseWheel>", _on_mousewheel)

            # Cleanup al cerrar
            def on_closing():
                try:
                    version_window.unbind_all("<MouseWheel>")
                except:
                    pass
                version_window.destroy()

            version_window.protocol("WM_DELETE_WINDOW", on_closing)
            
        except Exception as e:
            messagebox.showerror(
                "Error",
                f"Error al mostrar información de versión:\n{str(e)}"
            )

    def _create_info_section(self, parent, title, info_items):
        """Crear una sección de información con título y elementos usando pack"""
        # Frame para la sección
        section_frame = ttk.LabelFrame(parent, text=title, padding=10)
        section_frame.pack(fill=X, pady=(0, 10))

        # Grid de información
        for i, (label, value) in enumerate(info_items):
            ttk.Label(
                section_frame,
                text=f"{label}:",
                font=("Arial", 9, "bold")
            ).grid(row=i, column=0, sticky=W, padx=(0, 10), pady=2)

            ttk.Label(
                section_frame,
                text=str(value),
                font=("Arial", 9)
            ).grid(row=i, column=1, sticky=W, pady=2)

    def _create_info_section_grid(self, parent, title, info_items, row=0, column=0, columnspan=1):
        """Crear una sección de información con título y elementos usando grid layout"""
        # Frame para la sección
        section_frame = ttk.LabelFrame(parent, text=title, padding=10)
        section_frame.grid(row=row, column=column, columnspan=columnspan, sticky=(N, S, E, W), padx=5, pady=5)

        # Grid de información dentro del frame
        for i, (label, value) in enumerate(info_items):
            ttk.Label(
                section_frame,
                text=f"{label}:",
                font=("Arial", 9, "bold")
            ).grid(row=i, column=0, sticky=W, padx=(0, 10), pady=2)

            ttk.Label(
                section_frame,
                text=str(value),
                font=("Arial", 9)
            ).grid(row=i, column=1, sticky=W, pady=2)
    
    def _copy_general_info(self, version_info):
        """Copiar información de la pestaña General al clipboard"""
        try:
            release_type_es = {
                'stable': 'Estable',
                'beta': 'Beta',
                'alpha': 'Alfa',
                'rc': 'Release Candidate'
            }.get(version_info['version']['release_type'].lower(), version_info['version']['release_type'])

            info_text = f"""ONCONOVA CIRUGÍA ONCOLÓGICA - Información General
=====================================
Nombre Completo: {version_info['project']['full_name']}
Organización: {version_info['project']['organization']}
Versión: {version_info['version']['version']}
Nombre de Versión: {version_info['version']['version_name']}
Tipo de Release: {release_type_es}
Fecha de Build: {version_info['version']['build_date']}
Número de Build: {version_info['version']['build_number']}
"""

            self.clipboard_clear()
            self.clipboard_append(info_text)
            self.update()

            messagebox.showinfo("Copiado", "Información general copiada al portapapeles")

        except Exception as e:
            messagebox.showerror("Error", f"Error al copiar:\n{str(e)}")

    def _copy_system_info(self, version_info):
        """Copiar información de la pestaña Sistema al clipboard"""
        try:
            info_text = f"""ONCONOVA CIRUGÍA ONCOLÓGICA - Información del Sistema
=====================================
Versión Python: {version_info['system']['python_version'].split()[0]}
Plataforma: {version_info['system']['platform']}
Sistema: {version_info['system'].get('system', 'No disponible')}
Release: {version_info['system'].get('release', 'No disponible')}
Arquitectura: {version_info['system']['architecture']}
Máquina: {version_info['system'].get('machine', 'No disponible')}
Nodo: {version_info['system'].get('node', 'No disponible')}
Procesador: {version_info['system']['processor'] or 'No disponible'}
"""

            # Memoria
            if 'memoria_total' in version_info['system']:
                info_text += f"""
Memoria:
- Total: {version_info['system']['memoria_total']}
- Disponible: {version_info['system']['memoria_disponible']}
- Usada: {version_info['system']['memoria_usada']}
- Porcentaje Usado: {version_info['system']['memoria_porcentaje']}
"""

            # CPU
            if 'cpu_cores' in version_info['system']:
                info_text += f"""
Procesador:
- Núcleos Físicos: {version_info['system']['cpu_cores']}
- Hilos Lógicos: {version_info['system']['cpu_threads']}
- Frecuencia Máxima: {version_info['system']['cpu_frecuencia']}
"""

            # Discos
            if 'discos' in version_info['system']:
                info_text += "\nDiscos:\n"
                for i, disco in enumerate(version_info['system']['discos'], 1):
                    if isinstance(disco, dict):
                        info_text += f"""
Disco {i}:
- Dispositivo: {disco.get('dispositivo', 'No disponible')}
- Punto de Montaje: {disco.get('punto_montaje', 'No disponible')}
- Sistema de Archivos: {disco.get('sistema_archivos', 'No disponible')}
- Espacio Total: {disco.get('total', 'No disponible')}
- Espacio Usado: {disco.get('usado', 'No disponible')}
- Espacio Libre: {disco.get('libre', 'No disponible')}
- Porcentaje Usado: {disco.get('porcentaje', 'No disponible')}
"""

            self.clipboard_clear()
            self.clipboard_append(info_text)
            self.update()

            messagebox.showinfo("Copiado", "Información del sistema copiada al portapapeles")

        except Exception as e:
            messagebox.showerror("Error", f"Error al copiar:\n{str(e)}")

    def _hide_header_if_not_welcome(self):
        """Ocultar header si no estamos en la pantalla de bienvenida"""
        if self.header_visible:
            self.header.pack_forget()
            self.header_separator.pack_forget()
            self.header_visible = False
            # Actualizar posición base del botón flotante sin header
            self.floating_btn_base_y = 20
            # Reposicionar inmediatamente el botón si existe
            if hasattr(self, 'floating_btn_container') and not self.floating_menu_visible:
                self.floating_btn_container.place(x=15, y=20, width=50, height=50)

    def _show_header(self):
        """Mostrar header (solo para pantalla de bienvenida)"""
        if not self.header_visible:
            self.header.pack(fill=X, before=self.content_container)
            self.header_visible = True
            # V6.9.16: el FAB tiene posicion fija abajo-izquierda; ya no se
            # reposiciona segun el header, y el separador fue eliminado del diseno.

    def _create_sidebar(self):
        """Crear la barra lateral de navegación"""
        # Header del sidebar elegante
        top = ttk.Frame(self.sidebar, padding=10)
        top.pack(fill=X)
        
        # Título del sidebar
        ttk.Label(
            top, 
            text="� NAVEGACIÓN", 
            font=("Segoe UI", 12, "bold"),
            anchor="center"
        ).pack(fill=X, pady=(0, 10))

        # Navegación
        nav = ttk.Frame(self.sidebar, padding=(10, 10))
        nav.pack(fill=BOTH, expand=True)

        self.nav_buttons = {}

        # Botones de navegación reorganizados
        nav_items = [
            ("🏠 Inicio", "home", "light", self.show_welcome_screen),
            ("🗄️ Base de Datos", "database", "primary", self.show_database_frame),
            ("📈 Análisis Gráfico", "dashboard", "info", self.show_dashboard_frame),
            ("🔗 Interoperabilidad QHORTE\n(Sistema de Entrega)", "web", "warning", self.open_web_auto_modal),
        ]

        for text, icon_key, style, callback in nav_items:
            btn = ttk.Button(
                nav, 
                text=text,
                image=self.iconos.get(icon_key),
                compound=LEFT,
                bootstyle=style,
                command=callback,
                width=20
            )
            btn.pack(fill=X, pady=2)
            self.nav_buttons[text] = btn

        # Botón de navegación (mostrar/ocultar menús)
        self.nav_toggle_btn = ttk.Button(
            nav, 
            text="◀ Ocultar Menús", 
            command=self._toggle_navigation_visibility, 
            bootstyle="secondary",
            width=20
        )
        self.nav_toggle_btn.pack(fill=X, pady=(20, 0))

        # Footer
        ttk.Label(
            self.sidebar, 
            text="HUV • ONCONOVA", 
            anchor=CENTER, 
            padding=(10, 8), 
            bootstyle="light"
        ).pack(side=BOTTOM, fill=X)

    def _toggle_sidebar(self):
        """Alternar visibilidad de la sidebar"""
        target = 0 if self.sidebar_expanded else self.sidebar_width
        step = -24 if self.sidebar_expanded else 24

        def animate(curr):
            if (step > 0 and curr < target) or (step < 0 and curr > target):
                curr += step
                self.sidebar.configure(width=max(0, curr))
                self.after(10, lambda: animate(curr))
            else:
                self.sidebar.configure(width=target)
                self.sidebar_expanded = not self.sidebar_expanded

        current = self.sidebar.winfo_width() or (self.sidebar_width if self.sidebar_expanded else 0)
        animate(current)

    def _create_content_panels(self):
        """Crear los paneles de contenido principal con scroll"""
        # Panel de base de datos SIN scroll externo (V6.9.16): el notebook y el
        # scroll interno de cada seccion ya gestionan el desplazamiento. Esto evita
        # el doble scroll y el espacio en blanco arriba de las secciones.
        self.database_frame = ttk.Frame(self.content_container)
        self.database_frame.scrollable_frame = self.database_frame  # compat (_create_database_content)
        self._create_database_content()

        # Panel de visualización con scroll
        self.visualizar_frame = self._create_scrollable_frame(self.content_container)
        self._create_visualizar_content()

        # Panel de analisis grafico SIN scroll externo (V6.9.16): los graficos
        # usan grid 2x2 responsive (se ajustan al espacio disponible), por lo que
        # el scroll externo solo generaba el espacio vacio arriba.
        self.dashboard_frame = ttk.Frame(self.content_container)
        self.dashboard_frame.scrollable_frame = self.dashboard_frame  # compat (_create_dashboard_content)
        self._create_dashboard_content()

        # Panel de análisis IA con scroll
        self.analisis_ia_frame = self._create_scrollable_frame(self.content_container)
        self._crear_analisis_ia_content()

        # Panel activo actual
        self.panel_activo = None

    def _create_scrollable_frame(self, parent):
        """Crear un frame con barra de desplazamiento"""
        # Frame contenedor principal que llena toda el área
        container = ttk.Frame(parent, padding=0)
        
        # Canvas para scroll sin bordes ni highlight
        canvas = tk.Canvas(container, highlightthickness=0, borderwidth=0, relief='flat')
        scrollbar = ttk.Scrollbar(container, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas, padding=10)
        
        # Configurar scroll
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        # Hacer que el scrollable_frame se expanda horizontalmente en el canvas
        def _configure_canvas(event):
            canvas.configure(scrollregion=canvas.bbox("all"))
            # Hacer que el frame interno sea del mismo ancho que el canvas
            canvas_width = event.width
            canvas.itemconfig(window_id, width=canvas_width)
            # Auto-mostrar/ocultar scrollbar según si el contenido excede el canvas
            _update_scrollbar_visibility()
        
        def _update_scrollbar_visibility():
            """Mostrar scrollbar solo cuando el contenido excede el área visible"""
            try:
                canvas.update_idletasks()
                bbox = canvas.bbox("all")
                if bbox:
                    content_height = bbox[3] - bbox[1]
                    canvas_height = canvas.winfo_height()
                    if content_height > canvas_height:
                        if not scrollbar.winfo_ismapped():
                            scrollbar.pack(side="right", fill="y")
                    else:
                        if scrollbar.winfo_ismapped():
                            scrollbar.pack_forget()
            except Exception:
                pass
        
        window_id = canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.bind('<Configure>', _configure_canvas)
        canvas.configure(yscrollcommand=scrollbar.set)
        
        # Variable para trackear el último widget con focus para scroll anidado
        container.last_scroll_target = None

        # Scroll con rueda del mouse mejorado - maneja scroll anidado
        def _on_mousewheel(event):
            # Obtener el widget bajo el cursor
            widget = event.widget

            # Buscar si el widget o algún padre tiene un canvas con scroll
            current = widget
            scrollable_canvas = None

            while current and current != container:
                if hasattr(current, 'master') and isinstance(current.master, tk.Canvas):
                    # Este widget está dentro de un canvas scrollable
                    scrollable_canvas = current.master
                    break
                current = current.master if hasattr(current, 'master') else None

            # Si encontramos un canvas scrollable anidado, hacer scroll en ese
            if scrollable_canvas and scrollable_canvas != canvas:
                try:
                    scrollable_canvas.yview_scroll(int(-1*(event.delta/120)), "units")
                    return "break"
                except:
                    pass

            # Si no, hacer scroll en el canvas principal
            canvas.yview_scroll(int(-1*(event.delta/120)), "units")
            return "break"

        # Vincular el evento al canvas específico
        canvas.bind("<MouseWheel>", _on_mousewheel)

        # También vincular a todos los widgets hijos del scrollable_frame
        def _bind_to_mousewheel(widget):
            # V6.9.49 PERF FIX: NO tocar el tksheet Sheet ni sus canvas internos.
            # tksheet maneja su propio scroll de rueda NATIVO (suave y optimizado).
            # Bindear aquí _on_mousewheel (que hace tree-walk + yview_scroll + "break")
            # SECUESTRABA el wheel del Sheet y rompía su scroll suave -> "al bajar se
            # siente lento". Saltar el Sheet (sin recursar en sus hijos) deja intacto
            # su manejo nativo de la rueda.
            try:
                if isinstance(widget, Sheet):
                    return
            except Exception:
                pass
            # No vincular scroll a widgets que ya manejan su propio scroll
            widget_type = widget.winfo_class()
            if widget_type not in ['Listbox', 'Treeview', 'Text']:
                widget.bind("<MouseWheel>", _on_mousewheel)

            for child in widget.winfo_children():
                _bind_to_mousewheel(child)
        
        # Empacar elementos para llenar toda el área
        canvas.pack(side="left", fill="both", expand=True)
        # scrollbar se muestra/oculta automáticamente según contenido
        
        # Aplicar el binding del scroll a todos los widgets después de un pequeño delay
        # para asegurar que todos los widgets hijos se han creado
        def _apply_scroll_binding():
            try:
                _bind_to_mousewheel(scrollable_frame)
            except:
                pass  # Si hay error, continuamos sin problemas
        
        container.after(100, _apply_scroll_binding)
        
        # Guardar referencias para uso posterior
        container.scrollable_frame = scrollable_frame
        container.canvas = canvas
        container._bind_to_mousewheel = _bind_to_mousewheel  # Guardar función para uso posterior
        
        return container

    def _create_database_content(self):
        """Crear contenido del panel de base de datos con dashboard mejorado de ancho completo"""
        # Usar el frame scrollable
        frame = self.database_frame.scrollable_frame

        # ELIMINADO: Título duplicado (el dashboard ya tiene su propio título)
        # ttk.Label(
        #     frame,
        #     text="🏥 Dashboard Base de Datos ONCONOVA CIRUGÍA ONCOLÓGICA",
        #     font=self.FONT_TITULO
        # ).pack(pady=(0, 20), anchor=W)

        # Dashboard mejorado que usa todo el ancho disponible
        dashboard_container = ttk.Frame(frame, padding=0)
        dashboard_container.pack(expand=True, fill=BOTH)

        # Crear dashboard mejorado con funcionalidad de importación integrada
        try:
            from core.enhanced_database_dashboard import EnhancedDatabaseDashboard
            self.enhanced_dashboard = EnhancedDatabaseDashboard(dashboard_container)

            # Conectar métodos de importación del dashboard con la UI principal
            self._connect_import_functionality()

            # Poblar la pestaña de visualizador en el dashboard
            self._populate_visualizar_tab_in_dashboard()

            # V6.9.73: vista agrupada por paciente (pestaña propia)
            self._populate_pacientes_tab_in_dashboard()

        except ImportError as e:
            # Fallback en caso de error
            ttk.Label(
                dashboard_container,
                text=f"Error: No se pudo cargar el dashboard mejorado: {e}",
                font=("Segoe UI", 12)
            ).pack(pady=50)

    def _connect_import_functionality(self):
        """Conectar la funcionalidad de importación del dashboard con los métodos de la UI principal"""
        if hasattr(self, 'enhanced_dashboard'):
            # Conectar los métodos de importación
            dashboard = self.enhanced_dashboard

            # Reasignar los métodos del dashboard a los de la UI principal
            dashboard.select_pdf_file = self._select_pdf_file
            dashboard.select_pdf_folder = self._select_pdf_folder
            dashboard.process_selected_files = self._process_selected_files
            dashboard.process_selected_files_ia = self._process_selected_files_ia
            dashboard.seleccionar_pendientes = self._seleccionar_pendientes   # V6.9.66

            # V6.9.25: referencia al ÁRBOL navegable del dashboard (carpetas/subcarpetas)
            if hasattr(dashboard, 'import_files_tree'):
                self.files_tree = dashboard.import_files_tree
                dashboard.refresh_files_list = self._refresh_files_list

                # Actualizar la lista de archivos después de conectar
                try:
                    self._refresh_files_list()
                except Exception as e:
                    logging.error(f"Error al actualizar lista de archivos inicial: {e}")

            # v6.0.12: Cargar datos del dashboard inmediatamente al inicializar
            try:
                dashboard.refresh_all_data()
                logging.info("✅ Dashboard inicializado con datos de la base de datos")
            except Exception as e:
                logging.error(f"❌ Error al cargar datos iniciales del dashboard: {e}")
                logging.error(f"Traceback: {traceback.format_exc()}")

    def _populate_visualizar_tab_in_dashboard(self):
        """Poblar la pestaña de visualizador en el dashboard con el contenido completo IGUAL al visualizador original"""
        if not hasattr(self, 'enhanced_dashboard'):
            return

        dashboard = self.enhanced_dashboard

        # Verificar que existe el tab
        if not hasattr(dashboard, 'visualizar_tab'):
            logging.warning("Dashboard no tiene visualizar_tab")
            return

        # Usar el visualizar_tab directamente como frame
        frame = dashboard.visualizar_tab

        # Limpiar todo el contenido previo
        for widget in frame.winfo_children():
            widget.destroy()

        # ===== CREAR EL MISMO CONTENIDO QUE _create_visualizar_content() =====

        # Título principal compacto
        title_frame = ttk.Frame(frame)
        title_frame.pack(fill=X, padx=10, pady=5)

        ttk.Label(
            title_frame,
            text="Visualizador de datos",
            font=("Segoe UI Semibold", 20),
            bootstyle="primary"
        ).pack(side=LEFT)

        # Botones de acción en el header
        actions_frame = ttk.Frame(title_frame)
        actions_frame.pack(side=RIGHT)

        # Botones de accion unificados (V6.9.16): paleta sobria navy/gris.
        # Accion principal en navy solido; el resto en outline. Sin colores dispares.
        # V6.9.44: botón "🔍 Filtros" ELIMINADO. Era un placeholder (mostraba "se
        # implementará en una versión futura"). El buscador superior cubre "encontrar
        # un caso" y lo agregado lo da el PDF de Estadísticas. El método
        # _toggle_advanced_filters queda en el código, sin exponerse en la UI.

        # Boton de detalles flotante
        # V6.9.44: arranca DESHABILITADO y se habilita al seleccionar una fila.
        self.details_btn_dashboard = ttk.Button(
            actions_frame,
            text="📋 Detalles",
            command=self._toggle_details_panel,
            bootstyle="primary-outline",
            state="disabled"
        )
        self.details_btn_dashboard.pack(side=RIGHT, padx=(0, 5))

        # Boton exportar seleccion (inicialmente deshabilitado)
        self.export_selection_btn_dashboard = ttk.Button(
            actions_frame,
            text="📤 Exportar Selección",
            command=self._export_selected_data,
            bootstyle="primary-outline",
            state="disabled"
        )
        self.export_selection_btn_dashboard.pack(side=RIGHT, padx=(0, 5))

        # Boton exportar toda la base de datos
        ttk.Button(
            actions_frame,
            text="💾 Exportar Todo",
            command=self._export_full_database,
            bootstyle="primary-outline"
        ).pack(side=RIGHT, padx=(0, 5))

        # V6.9.44: botón "📊 Resumen IA" ELIMINADO. Generaba un resumen NARRATIVO
        # con el LLM (lento y con riesgo de inventar cifras). El reporte fiable es
        # "Estadísticas Generales → Generar PDF" (determinista, sin IA). El método
        # _generar_resumen_ia queda en el código pero ya no se expone en la UI.

        # Accion principal: Actualizar datos (navy solido)
        ttk.Button(
            actions_frame,
            text="🔄 Actualizar Datos",
            command=self.refresh_data_and_table,
            bootstyle="primary"
        ).pack(side=RIGHT, padx=(0, 5))

        # V6.9.50: abrir el Visualizador en ventana Qt (QTableView) de alto rendimiento
        ttk.Button(
            actions_frame,
            text="⚡ Tabla Rápida (Qt)",
            command=self._abrir_visor_qt,
            bootstyle="info"
        ).pack(side=RIGHT, padx=(0, 5))

        # V6.9.44: botones "Auditoría PARCIAL/COMPLETA" (barra del dashboard)
        # ELIMINADOS. Misma razón que en la otra barra: auditoría por LLM que
        # alimentaba "Análisis IA" (ya oculta). Métodos _auditar_* intactos; el
        # manejo de estado usa hasattr -> inocuo.

        # Frame para tabla - FULL SCREEN
        table_frame = ttk.Frame(frame)
        table_frame.pack(expand=True, fill=BOTH, padx=10, pady=5)
        table_frame.grid_rowconfigure(1, weight=1)
        table_frame.grid_columnconfigure(0, weight=1)

        # Campo de búsqueda
        self._search_placeholder = "Buscar por N° Petición, Cédula, Nombre o Apellido..."
        self.search_var_dashboard = tk.StringVar()
        self.search_var_dashboard.trace_add("write", self._filter_tabla_debounced)
        self._search_entry_dashboard = ttk.Entry(
            table_frame,
            textvariable=self.search_var_dashboard,
            font=("Segoe UI", 11)
        )
        self._search_entry_dashboard.grid(row=0, column=0, sticky="ew", padx=10, pady=10)
        self._search_entry_dashboard.insert(0, self._search_placeholder)
        self._search_entry_dashboard.bind("<FocusIn>", lambda e: self._on_search_focus_in(self._search_entry_dashboard, self.search_var_dashboard))
        self._search_entry_dashboard.bind("<FocusOut>", lambda e: self._on_search_focus_out(self._search_entry_dashboard, self.search_var_dashboard))

        # V6.9.55: filtro segmentado Todos / IHQ / Coloración (a la derecha del buscador)
        self._crear_filtro_tipo_registro(table_frame).grid(
            row=0, column=1, sticky="e", padx=(0, 10), pady=10
        )

        # Crear Sheet en el dashboard (compartiremos la misma instancia para sincronización)
        # IMPORTANTE: Usamos self.sheet para que sea la MISMA instancia que el visualizador original
        from tksheet import Sheet

        self.sheet_dashboard = Sheet(
            table_frame,
            page_up_down_select_row=True,
            expand_sheet_if_paste_too_big=False,
            column_width=150,
            startup_select=(0, 0, "rows"),
            headers_height=30,
            default_row_height=25,
            show_horizontal_grid=True,
            show_vertical_grid=True,
            show_top_left=False,
            show_row_index=True,
            show_header=True,
            empty_horizontal=0,
            empty_vertical=0,
            header_font=("Segoe UI", 10, "bold"),
            font=("Segoe UI", 10, "normal"),
            header_bg="#e9edf3",                 # gris azulado claro (header sutil)
            header_fg="#2d3e5e",                 # texto NAVY legible (alto contraste)
            header_grid_fg="#d2d9e6",            # lineas de la cabecera sutiles
            header_selected_cells_bg="#2d3e5e",  # al seleccionar columna -> navy
            header_selected_cells_fg="#ffffff",
            header_selected_columns_bg="#2d3e5e",
            header_selected_columns_fg="#ffffff",
            table_bg="#ffffff",
            table_fg="#2a2f3a",
            table_selected_cells_bg="#d7deea",   # seleccion azul-gris suave
            table_selected_cells_fg="#2a2f3a",
            table_selected_rows_bg="#eef1f6",
            table_selected_rows_fg="#2a2f3a",
            top_left_bg="#e9edf3",
            index_bg="#f4f6f9",                  # indice gris muy claro
            index_fg="#5a6172",
            index_selected_cells_bg="#d7deea",
            index_selected_rows_bg="#dfe4ee"
        )
        self.sheet_dashboard.grid(row=1, column=0, columnspan=2, sticky="nsew", padx=10, pady=(0, 10))

        # Habilitar funcionalidades tipo Excel
        self.sheet_dashboard.enable_bindings(
            "all", "copy", "row_select", "column_select", "drag_select",
            "select_all", "rc_select", "arrowkeys", "single_select", "drag_and_drop", "move_columns"
        )

        # Deshabilitar edición
        self.sheet_dashboard.disable_bindings(
            "edit_cell", "cut", "paste", "delete", "undo",
            "column_width_resize", "double_click_column_resize",
            "row_width_resize", "column_height_resize"
        )

        # Evento de selección
        self.sheet_dashboard.bind("<<SheetSelect>>", self.mostrar_detalle_registro)

        # V6.9.51: hover -> popup con valor completo + doble clic -> ventana de detalle.
        # (Esta tabla del dashboard también necesita ambos, no solo self.sheet.)
        self._instalar_hover_tooltip(self.sheet_dashboard)
        try:
            self.sheet_dashboard.MT.bind(
                "<Double-Button-1>",
                lambda e: self._abrir_detalle_fila(e, self.sheet_dashboard), add="+")
        except Exception:
            pass

        # Agregar métodos de compatibilidad Treeview → Sheet
        def _sheet_selection_dashboard():
            try:
                rows = set()
                selected_rows = self.sheet_dashboard.get_selected_rows()
                if selected_rows:
                    rows.update(selected_rows)
                selected_cells = self.sheet_dashboard.get_selected_cells()
                if selected_cells:
                    rows.update([cell[0] for cell in selected_cells])
                return list(rows)
            except Exception as e:
                logging.error(f"Error en _sheet_selection_dashboard: {e}")
                return []

        self.sheet_dashboard.selection = _sheet_selection_dashboard

        # Marcar que el dashboard tiene sheet
        dashboard.has_sheet = True
        dashboard.sheet_dashboard = self.sheet_dashboard

        # Cargar datos iniciales
        self.refresh_data_and_table()

        logging.info("✅ Pestaña de visualizador COMPLETA poblada en el dashboard con tabla Sheet")

    def _refresh_files_list_for_dashboard(self):
        """V6.9.25: Reconstruye el árbol navegable de archivos del dashboard."""
        if hasattr(self, 'enhanced_dashboard') and hasattr(self.enhanced_dashboard, 'import_files_tree'):
            self.files_tree = self.enhanced_dashboard.import_files_tree
            self._refresh_files_list()

    def _create_visualizar_content(self):
        """Crear contenido del panel de visualización mejorado - FULL SCREEN"""
        # Usar el frame principal directamente (sin scroll externo)
        frame = self.visualizar_frame

        # Limpiar cualquier contenido previo
        for widget in frame.winfo_children():
            widget.destroy()

        # Título principal compacto
        title_frame = ttk.Frame(frame)
        title_frame.pack(fill=X, padx=10, pady=5)

        ttk.Label(
            title_frame,
            text="Visualizador de datos",
            font=("Segoe UI Semibold", 20),
            bootstyle="primary"
        ).pack(side=LEFT)

        # Botones de acción en el header
        actions_frame = ttk.Frame(title_frame)
        actions_frame.pack(side=RIGHT)

        # V6.9.44: botón "🔍 Filtros" ELIMINADO (placeholder sin función real). Ver
        # nota en la barra del dashboard. _toggle_advanced_filters queda inactivo.

        # Botón de detalles flotante
        # V6.9.44: "Detalles" requiere una fila seleccionada -> arranca DESHABILITADO
        # y se "enciende" (info) al seleccionar, igual que "Exportar Selección".
        self.details_btn = ttk.Button(
            actions_frame,
            text="📋 Detalles",
            command=self._toggle_details_panel,
            bootstyle="primary",
            state="disabled"
        )
        self.details_btn.pack(side=RIGHT, padx=(0, 5))

        # Botón exportar selección (inicialmente deshabilitado)
        self.export_selection_btn = ttk.Button(
            actions_frame,
            text="📤 Exportar Selección",
            command=self._export_selected_data,
            bootstyle="success",
            state="disabled"
        )
        self.export_selection_btn.pack(side=RIGHT, padx=(0, 5))

        # Botón exportar toda la base de datos
        ttk.Button(
            actions_frame,
            text="💾 Exportar Todo",
            command=self._export_full_database,
            bootstyle="warning"
        ).pack(side=RIGHT, padx=(0, 5))

        # V6.9.44: botón "📊 Resumen IA" ELIMINADO (ver nota en el otro toolbar).

        ttk.Button(
            actions_frame,
            text="🔄 Actualizar Datos",
            command=self.refresh_data_and_table,
            bootstyle="primary"
        ).pack(side=RIGHT, padx=(0, 5))

        # V6.9.50: abrir el Visualizador en ventana Qt (QTableView) de alto rendimiento
        ttk.Button(
            actions_frame,
            text="⚡ Tabla Rápida (Qt)",
            command=self._abrir_visor_qt,
            bootstyle="info"
        ).pack(side=RIGHT, padx=(0, 5))

        # V6.9.44: botones "Auditoría PARCIAL/COMPLETA" ELIMINADOS. Eran la
        # auditoría por LLM que alimentaba la sección "Análisis IA" (ya oculta del
        # menú). El flujo real de auditoría es el agente data-auditor sobre los
        # debug_maps. Los métodos _auditar_* quedan en el código y el manejo de
        # estado (más abajo) usa hasattr -> queda inocuo sin estos botones.

        # Frame para tabla - FULL SCREEN
        table_frame = ttk.Frame(frame)
        table_frame.pack(expand=True, fill=BOTH, padx=10, pady=5)
        table_frame.grid_rowconfigure(1, weight=1)  # Treeview
        table_frame.grid_columnconfigure(0, weight=1)

        # Campo de búsqueda
        self.search_var = tk.StringVar()
        self.search_var.trace_add("write", self._filter_tabla_debounced)
        self._search_entry = ttk.Entry(
            table_frame,
            textvariable=self.search_var,
            font=("Segoe UI", 11)
        )
        self._search_entry.grid(row=0, column=0, sticky="ew", padx=10, pady=10)
        self._search_entry.insert(0, self._search_placeholder)
        self._search_entry.bind("<FocusIn>", lambda e: self._on_search_focus_in(self._search_entry, self.search_var))
        self._search_entry.bind("<FocusOut>", lambda e: self._on_search_focus_out(self._search_entry, self.search_var))

        # V6.9.55: filtro segmentado Todos / IHQ / Coloración (a la derecha del buscador)
        self._crear_filtro_tipo_registro(table_frame).grid(
            row=0, column=1, sticky="e", padx=(0, 10), pady=10
        )

        # V5.3.8: Sheet virtualizado tipo Excel (reemplaza Treeview)
        # VENTAJAS: Virtualización nativa, rendimiento profesional, comportamiento Excel

        # Crear Sheet con scrollbars integrados
        self.sheet = Sheet(
            table_frame,
            page_up_down_select_row=True,
            expand_sheet_if_paste_too_big=False,
            column_width=150,
            startup_select=None,  # No auto-seleccionar fila 0
            headers_height=30,
            default_row_height=25,
            show_horizontal_grid=True,
            show_vertical_grid=True,
            show_top_left=False,
            show_row_index=True,
            show_header=True,
            empty_horizontal=0,
            empty_vertical=0,
            header_font=("Segoe UI", 10, "bold"),
            font=("Segoe UI", 10, "normal"),  # FIX: Agregar estilo "normal" (3 elementos)
            header_bg="#E8F5E9",  # Verde muy claro (profesional)
            header_fg="#1B5E20",  # Verde oscuro
            table_bg="white",
            table_fg="black",
            table_selected_cells_bg="#BBDEFB",  # Azul claro
            table_selected_cells_fg="black",
            table_selected_rows_bg="#E3F2FD",  # Azul muy claro
            table_selected_rows_fg="black",
            top_left_bg="#E8F5E9",
            index_bg="#F5F5F5",
            index_fg="#424242",
            index_selected_cells_bg="#CFD8DC",
            index_selected_rows_bg="#B0BEC5"
        )
        self.sheet.grid(row=1, column=0, columnspan=2, sticky="nsew", padx=10, pady=(0, 10))

        # Habilitar funcionalidades tipo Excel (NO usar "all" para evitar single_select)
        self.sheet.enable_bindings(
            "copy",  # Ctrl+C
            "row_select",  # Selección de filas
            "column_select",  # Selección de columnas
            "drag_select",  # Arrastrar para seleccionar
            "select_all",  # Ctrl+A
            "rc_select",  # Click derecho
            "arrowkeys",  # Navegación con flechas
            "toggle_select",  # Ctrl+Click / Shift+Click para multi-selección
        )

        # Evento de selección
        self.sheet.bind("<<SheetSelect>>", self.mostrar_detalle_registro)
        # IMPORTANTE: También enlazar a ButtonRelease para capturar clicks que no disparen <<SheetSelect>>
        self.sheet.bind("<ButtonRelease-1>", self.mostrar_detalle_registro, add="+")
        # Y navegación con teclado
        self.sheet.bind("<KeyRelease-Up>", self.mostrar_detalle_registro, add="+")
        self.sheet.bind("<KeyRelease-Down>", self.mostrar_detalle_registro, add="+")
        self.sheet.bind("<KeyRelease-Left>", self.mostrar_detalle_registro, add="+")
        self.sheet.bind("<KeyRelease-Right>", self.mostrar_detalle_registro, add="+")
        # V6.9.51: doble clic en una fila -> ventana de detalle con TODO el texto completo
        # (las celdas truncan campos largos como 'Descripcion macroscopica').
        # IMPORTANTE: distintas versiones de tksheet enrutan sheet.bind("<Double-Button-1>")
        # de forma distinta (extra_double_b1_func vs bind raw); para que funcione SIEMPRE,
        # se enlaza directo al widget interno (MainTable) con tkinter puro. double_b1 NO
        # retorna "break", así que con add="+" nuestro handler corre después de él.
        try:
            self.sheet.MT.bind("<Double-Button-1>", self._abrir_detalle_fila, add="+")
            logging.info("Doble clic (detalle) enlazado a sheet.MT")
        except Exception as _e:
            logging.warning(f"No se pudo enlazar a sheet.MT ({_e}); usando sheet.bind")
            self.sheet.bind("<Double-Button-1>", self._abrir_detalle_fila, add="+")
        # V6.9.51: hover -> popup con el valor completo de la celda (campos largos).
        self._instalar_hover_tooltip(self.sheet)

        # Binding permanente para ordenamiento por click en encabezado
        # IMPORTANTE: Se bindea UNA SOLA VEZ aquí (no en _populate_treeview)
        def _on_header_click(event):
            """Callback cuando se hace clic en un encabezado para ordenar columna"""
            import logging

            # Validación robusta: verificar que el evento tenga atributo 'region'
            if not hasattr(event, 'region'):
                logging.debug("_on_header_click: event sin atributo 'region', ignorando")
                return

            if event.region != "header":
                logging.debug(f"_on_header_click: region='{event.region}' (no es header), ignorando")
                return

            # Procesar click en encabezado para ordenar
            col_idx = event.column
            if col_idx is not None and hasattr(self, 'master_df') and not self.master_df.empty:
                # Obtener DataFrame actual (puede estar filtrado)
                df_display = self.master_df

                if col_idx < len(df_display.columns):
                    col_name = df_display.columns[col_idx]

                    # Alternar orden ascendente/descendente
                    if not hasattr(self, '_last_sorted_col') or self._last_sorted_col != col_name:
                        self._last_sorted_col = col_name
                        self._last_sorted_reverse = False
                    else:
                        self._last_sorted_reverse = not self._last_sorted_reverse

                    logging.info(f"Ordenando por columna '{col_name}' (reverso={self._last_sorted_reverse})")
                    self._sort_treeview(col_name, self._last_sorted_reverse)

        # Bindear el handler (add="+" para no reemplazar handlers existentes)
        self.sheet.bind("<ButtonRelease-1>", _on_header_click, add="+")
        logging.info("Handler de ordenamiento por encabezado bindeado correctamente")

        # V5.3.8: CAPA DE COMPATIBILIDAD TREEVIEW → SHEET
        # =================================================
        # Agregar métodos a Sheet para que se comporte como Treeview

        def _sheet_selection():
            """Emula tree.selection() - Retorna lista de índices de filas seleccionadas

            v6.0.15: Lee selection_boxes del Sheet + fallback a últimas filas guardadas.
            Cuando el usuario hace clic en un botón, el Sheet pierde selección interna,
            así que usamos self._ultimas_filas_seleccionadas como caché confiable.
            """
            try:
                rows = set()

                # ESTRATEGIA 1: Leer selection_boxes directamente del Sheet
                # Esto captura rangos de celdas/filas seleccionadas
                try:
                    boxes = self.sheet.get_all_selection_boxes()
                    if boxes:
                        for box in boxes:
                            # box es (from_r, from_c, upto_r, upto_c)
                            if len(box) >= 4:
                                from_r, _, upto_r, _ = box[0], box[1], box[2], box[3]
                                for r in range(from_r, upto_r):
                                    rows.add(r)
                        if rows:
                            logging.info(f"_sheet_selection: Estrategia 1 (selection_boxes): {sorted(rows)}")
                except Exception:
                    pass

                # ESTRATEGIA 2: get_selected_rows y get_selected_cells
                if not rows:
                    selected_rows = self.sheet.get_selected_rows()
                    if selected_rows:
                        rows.update(selected_rows)
                    selected_cells = self.sheet.get_selected_cells()
                    if selected_cells:
                        rows.update(row for row, col in selected_cells)
                    if rows:
                        logging.info(f"_sheet_selection: Estrategia 2 (rows/cells): {sorted(rows)}")

                # ESTRATEGIA 3: get_currently_selected
                if not rows:
                    selected = self.sheet.get_currently_selected()
                    if selected and hasattr(selected, 'row') and selected.row is not None:
                        rows.add(selected.row)
                        logging.info(f"_sheet_selection: Estrategia 3 (currently_selected): fila {selected.row}")

                # ESTRATEGIA 4: Fallback a últimas filas guardadas (crucial para multi-selección)
                # Cuando el usuario hace clic en un botón, el Sheet pierde la selección
                if not rows and hasattr(self, '_ultimas_filas_seleccionadas') and self._ultimas_filas_seleccionadas:
                    rows.update(self._ultimas_filas_seleccionadas)
                    logging.info(f"_sheet_selection: Estrategia 4 (caché): {sorted(rows)}")

                # Actualizar caché si obtuvimos filas por estrategias 1-3
                if rows and not (not rows and hasattr(self, '_ultimas_filas_seleccionadas')):
                    self._ultimas_filas_seleccionadas = list(rows)

                if rows:
                    result = sorted(rows)
                    logging.debug(f"_sheet_selection: RETORNANDO {len(result)} fila(s): {result}")
                    return result
                else:
                    logging.debug("_sheet_selection: Sin selección, retornando []")
                    return []

            except Exception as e:
                logging.error(f"_sheet_selection: Error obteniendo selección: {e}", exc_info=True)
                return []

        def _sheet_item(row_idx, option=None):
            """Emula tree.item(item_id, option) - Retorna datos de la fila"""
            if option == 'values' or option is None:
                try:
                    row_data = self.sheet.get_row_data(row_idx, return_copy=True)
                    return {'values': row_data} if option is None else row_data
                except:
                    return {'values': []} if option is None else []
            return {}

        def _sheet_get_children(item=""):
            """Emula tree.get_children() - Retorna lista de todos los índices de filas"""
            total_rows = self.sheet.get_total_rows()
            return list(range(total_rows))

        def _sheet_index(item_id):
            """Emula tree.index(item_id) - Retorna el índice de la fila"""
            return item_id  # En Sheet, item_id YA ES el índice

        # Agregar métodos de compatibilidad al objeto sheet
        self.sheet.selection = _sheet_selection
        self.sheet.item = _sheet_item
        self.sheet.get_children = _sheet_get_children
        self.sheet.index = _sheet_index

        # COMPATIBILIDAD: Mantener alias 'tree' para código legacy
        self.tree = self.sheet  # Ahora funciona como Treeview

        # NUEVO: Agregar tooltips al pasar mouse sobre celdas
        self._setup_cell_tooltips()

        # V6.9.49 PERF: bindings de "_update_selection_buttons" ELIMINADOS por
        # estar DUPLICADOS. mostrar_detalle_registro ya está bindeado a
        # <<SheetSelect>>, <ButtonRelease-1> y <KeyRelease-*>, y dentro YA llama a
        # _update_export_button_state() + _update_audit_buttons_state() en cada
        # selección. Tener además estos 4 bindings hacía que el estado de botones
        # (incl. la consulta de auditoría) se recalculara 3-4x POR CLIC -> clic
        # "sumamente lento". Como mostrar_detalle_registro cubre exactamente los
        # mismos eventos, el comportamiento es idéntico pero sin el trabajo repetido.
        logging.info("Eventos de selección bindeados correctamente a Sheet")

        # Cargar datos automáticamente al inicializar el visualizador
        self.after(100, lambda: self.refresh_data_and_table() if hasattr(self, 'refresh_data_and_table') else None)

        # V6.9.4 — Auto-refresh periódico del Visualizador cada 60 segundos.
        # Necesario para que las PCs cliente (que solo CONSULTAN la BD MySQL
        # central) vean en tiempo real los datos que el servidor procesa con
        # IA. Sin esto, los clientes tienen que cerrar/abrir la app o pulsar
        # "Refrescar" manualmente para ver cambios.
        # 60s es un balance: refresca a tiempo sin sobrecargar la BD ni la UI.
        self._auto_refresh_enabled = True
        self.after(60_000, self._auto_refresh_tick)

        # El panel de detalles ahora será flotante (se crea en demanda)

    def _auto_refresh_tick(self):
        """V6.9.4 — Tick periódico que refresca el Visualizador desde la BD.
        Se re-programa cada 60 segundos mientras la app esté abierta.

        Se desactiva automáticamente durante operaciones críticas (como
        procesamiento IA en curso) para evitar contención de UI. El propio
        worker IA dispara refresh por su cuenta cuando agrega registros.
        """
        try:
            if not getattr(self, '_auto_refresh_enabled', True):
                # Re-programar igual por si el flag se vuelve True luego
                self.after(60_000, self._auto_refresh_tick)
                return

            # No refrescar si hay un procesamiento IA activo (el worker
            # IA ya dispara refresh por chunk procesado)
            ia_activo = (
                hasattr(self, '_processing_result_ia')
                and self._processing_result_ia
                and not self._processing_result_ia.get('done', True)
            )
            if not ia_activo:
                # V6.9.47: saltar la recarga completa si la BD no cambió desde la última
                # carga (evita el freeze periódico de ~1 s cada 60 s sin datos nuevos).
                try:
                    from core.database_manager import get_db_fingerprint
                    fp = get_db_fingerprint()
                    if fp is not None and fp == getattr(self, '_last_db_fingerprint', None):
                        return  # nada cambió -> no recargar (finally re-programa el tick)
                except Exception:
                    pass
                if hasattr(self, 'refresh_data_and_table'):
                    logging.info("[auto-refresh] Refrescando Visualizador (tick 60s)")
                    self.refresh_data_and_table()
        except Exception as e:
            logging.warning(f"[auto-refresh] Error: {e}")
        finally:
            # Re-programar siempre (incluso si hubo error)
            try:
                self.after(60_000, self._auto_refresh_tick)
            except Exception:
                pass

    def _create_dashboard_content(self):
        """Crear contenido del panel de dashboard"""
        # Usar el frame scrollable
        frame = self.dashboard_frame.scrollable_frame
        
        # Título principal actualizado
        ttk.Label(frame, text="📈 Análisis Gráfico de la Base de Datos", font=self.FONT_TITULO).pack(pady=(0, 10), anchor=W)
        
        # Container principal
        main_container = ttk.Frame(frame)
        main_container.pack(expand=True, fill=BOTH)
        main_container.grid_rowconfigure(0, weight=1)
        main_container.grid_columnconfigure(0, weight=0)  # sidebar
        main_container.grid_columnconfigure(1, weight=1)  # main area

        # Sidebar de filtros (inicialmente oculto)
        self.db_filters = {
            "fecha_desde": tk.StringVar(value=""),
            "fecha_hasta": tk.StringVar(value=""),
            "servicio": tk.StringVar(value=""),
            "malignidad": tk.StringVar(value=""),
            "responsable": tk.StringVar(value=""),
        }
        self.db_sidebar_collapsed = True
        self.db_sidebar = ttk.Frame(main_container, padding=15, width=280)
        
        # Título del sidebar de filtros
        ttk.Label(self.db_sidebar, text="Filtros", font=("Segoe UI", 16, "bold")).grid(row=0, column=0, sticky="w", pady=(0, 10))
        
        # Campos de filtros
        ttk.Label(self.db_sidebar, text="Fecha desde (dd/mm/aaaa)").grid(row=1, column=0, sticky="w", pady=(5, 2))
        ttk.Entry(self.db_sidebar, textvariable=self.db_filters["fecha_desde"]).grid(row=2, column=0, sticky="ew", pady=(0, 10))
        
        ttk.Label(self.db_sidebar, text="Fecha hasta (dd/mm/aaaa)").grid(row=3, column=0, sticky="w", pady=(5, 2))
        ttk.Entry(self.db_sidebar, textvariable=self.db_filters["fecha_hasta"]).grid(row=4, column=0, sticky="ew", pady=(0, 10))
        
        ttk.Label(self.db_sidebar, text="Servicio").grid(row=5, column=0, sticky="w", pady=(5, 2))
        self.cmb_servicio = ttk.Combobox(self.db_sidebar, textvariable=self.db_filters["servicio"], values=[])
        self.cmb_servicio.grid(row=6, column=0, sticky="ew", pady=(0, 10))
        
        ttk.Label(self.db_sidebar, text="Malignidad").grid(row=7, column=0, sticky="w", pady=(5, 2))
        self.cmb_malig = ttk.Combobox(self.db_sidebar, textvariable=self.db_filters["malignidad"], values=["", "PRESENTE", "AUSENTE"])
        self.cmb_malig.grid(row=8, column=0, sticky="ew", pady=(0, 10))
        
        ttk.Label(self.db_sidebar, text="Responsable").grid(row=9, column=0, sticky="w", pady=(5, 2))
        self.cmb_resp = ttk.Combobox(self.db_sidebar, textvariable=self.db_filters["responsable"], values=[])
        self.cmb_resp.grid(row=10, column=0, sticky="ew", pady=(0, 10))
        
        # Botones del sidebar
        btns_frame = ttk.Frame(self.db_sidebar)
        btns_frame.grid(row=11, column=0, sticky="ew", pady=(10, 0))
        btns_frame.grid_columnconfigure(0, weight=1)
        btns_frame.grid_columnconfigure(1, weight=1)
        
        ttk.Button(btns_frame, text="Refrescar", command=self._refresh_dashboard).grid(row=0, column=0, sticky="ew", padx=(0, 5))
        ttk.Button(btns_frame, text="Limpiar", command=self._clear_filters).grid(row=0, column=1, sticky="ew", padx=(5, 0))
        
        # Configurar grid del sidebar
        self.db_sidebar.grid_columnconfigure(0, weight=1)

        # Área principal con toolbar + notebook
        main_area = ttk.Frame(main_container)
        main_area.grid(row=0, column=1, sticky="nsew")
        main_area.grid_rowconfigure(1, weight=1)
        main_area.grid_columnconfigure(0, weight=1)

        # Toolbar superior
        toolbar = ttk.Frame(main_area, padding=(5, 5))
        toolbar.grid(row=0, column=0, sticky="ew", pady=(0, 10))
        
        # V6.9.16: se conserva SOLO el modal 'Filtros…' (el panel lateral y su
        # boton '≡ Mostrar filtros' se removieron por redundantes). El db_sidebar
        # sigue creado pero oculto; sus combos alimentan los 'values' del modal.
        ttk.Button(toolbar, text="Filtros…", command=self._open_filters_sheet,
                   bootstyle="primary-outline").pack(side=LEFT)

        # Notebook con las pestañas del dashboard
        self.tabs = ttk.Notebook(main_area)
        self.tabs.grid(row=1, column=0, sticky="nsew")

        # Crear las pestañas con scroll
        self.tab_overview   = ttk.Frame(self.tabs)
        self.tab_biomarkers = ttk.Frame(self.tabs)
        self.tab_times      = ttk.Frame(self.tabs)
        self.tab_quality    = ttk.Frame(self.tabs)
        self.tab_compare    = ttk.Frame(self.tabs)

        # Configurar cada pestaña para ser responsive con grid 2x2
        for tab in [self.tab_overview, self.tab_biomarkers, self.tab_times, self.tab_quality, self.tab_compare]:
            tab.grid_columnconfigure(0, weight=1)
            tab.grid_columnconfigure(1, weight=1)
            tab.grid_rowconfigure(0, weight=1)
            tab.grid_rowconfigure(1, weight=1)

        self.tabs.add(self.tab_overview,   text="Overview")
        self.tabs.add(self.tab_biomarkers, text="Biomarcadores")
        self.tabs.add(self.tab_times,      text="Tiempos")
        self.tabs.add(self.tab_quality,    text="Calidad")
        self.tabs.add(self.tab_compare,    text="Comparador")

        self._dash_canvases = []

    def _crear_analisis_ia_content(self):
        """Crear contenido completo de la sección de Análisis IA"""
        # Usar el frame scrollable
        frame = self.analisis_ia_frame.scrollable_frame

        # Variables de instancia
        self.tipo_reporte_var = tk.StringVar(value="parcial")
        self.reporte_seleccionado = None

        # HEADER
        header_frame = ttk.Frame(frame, bootstyle="primary", padding=15)
        header_frame.pack(fill=X, pady=(0, 20))

        ttk.Label(
            header_frame,
            text="🤖 ONCONOVA CIRUGÍA ONCOLÓGICA - Análisis con IA",
            font=("Arial", 20, "bold"),
            bootstyle="inverse-primary"
        ).pack()

        # PANEL SELECTOR DE TIPO DE REPORTE
        selector_frame = ttk.LabelFrame(frame, text="Tipo de Reporte", padding=15, bootstyle="info")
        selector_frame.pack(fill=X, pady=(0, 10))

        radio_frame = ttk.Frame(selector_frame)
        radio_frame.pack(fill=X)

        ttk.Radiobutton(
            radio_frame,
            text="Análisis Parcial",
            variable=self.tipo_reporte_var,
            value="parcial",
            command=self._actualizar_lista_reportes,
            bootstyle="info"
        ).pack(side=LEFT, padx=10)

        ttk.Radiobutton(
            radio_frame,
            text="Análisis Completo",
            variable=self.tipo_reporte_var,
            value="completo",
            command=self._actualizar_lista_reportes,
            bootstyle="info"
        ).pack(side=LEFT, padx=10)

        # LISTA DE REPORTES DISPONIBLES
        reportes_frame = ttk.LabelFrame(frame, text="Reportes Disponibles", padding=15, bootstyle="success")
        reportes_frame.pack(fill=BOTH, expand=True, pady=(0, 10))

        # Treeview con scrollbar
        tree_container = ttk.Frame(reportes_frame)
        tree_container.pack(fill=BOTH, expand=True)

        scrollbar = ttk.Scrollbar(tree_container, orient="vertical")
        scrollbar.pack(side=RIGHT, fill=Y)

        self.reportes_tree = ttk.Treeview(
            tree_container,
            columns=("Fecha", "Tipo", "Casos", "Archivo"),
            show="headings",
            yscrollcommand=scrollbar.set,
            bootstyle="success"
        )
        scrollbar.config(command=self.reportes_tree.yview)

        # Configurar columnas
        self.reportes_tree.heading("Fecha", text="Fecha")
        self.reportes_tree.heading("Tipo", text="Tipo")
        self.reportes_tree.heading("Casos", text="Casos")
        self.reportes_tree.heading("Archivo", text="Archivo")

        self.reportes_tree.column("Fecha", width=150, anchor=W)
        self.reportes_tree.column("Tipo", width=120, anchor=CENTER)
        self.reportes_tree.column("Casos", width=200, anchor=W)
        self.reportes_tree.column("Archivo", width=400, anchor=W)

        self.reportes_tree.pack(fill=BOTH, expand=True)

        # Bind de selección
        self.reportes_tree.bind("<<TreeviewSelect>>", self._cargar_reporte_seleccionado)

        # VISUALIZADOR MARKDOWN
        visualizador_frame = ttk.LabelFrame(frame, text="Contenido del Reporte", padding=15, bootstyle="warning")
        visualizador_frame.pack(fill=BOTH, expand=True, pady=(0, 10))

        # Frame para el text widget con scrollbar
        text_container = ttk.Frame(visualizador_frame)
        text_container.pack(fill=BOTH, expand=True)

        text_scrollbar = ttk.Scrollbar(text_container, orient="vertical")
        text_scrollbar.pack(side=RIGHT, fill=Y)

        self.markdown_text = tk.Text(
            text_container,
            wrap=WORD,
            font=("Consolas", 10),
            yscrollcommand=text_scrollbar.set,
            state=DISABLED
        )
        text_scrollbar.config(command=self.markdown_text.yview)
        self.markdown_text.pack(fill=BOTH, expand=True)

        # Configurar tags para formato Markdown
        self.markdown_text.tag_config("h1", font=("Arial", 18, "bold"), foreground="#2c3e50")
        self.markdown_text.tag_config("h2", font=("Arial", 15, "bold"), foreground="#34495e")
        self.markdown_text.tag_config("h3", font=("Arial", 13, "bold"), foreground="#7f8c8d")
        self.markdown_text.tag_config("bold", font=("Arial", 10, "bold"))
        self.markdown_text.tag_config("italic", font=("Arial", 10, "italic"))
        self.markdown_text.tag_config("code", font=("Consolas", 9), background="#ecf0f1")
        self.markdown_text.tag_config("list", lmargin1=20, lmargin2=40)

        # Botón copiar
        btn_frame = ttk.Frame(visualizador_frame)
        btn_frame.pack(fill=X, pady=(10, 0))

        ttk.Button(
            btn_frame,
            text="📋 Copiar Contenido",
            command=self._copiar_contenido_reporte,
            bootstyle="warning"
        ).pack(side=LEFT)

        # Cargar reportes inicialmente
        self._actualizar_lista_reportes()

    def _actualizar_lista_reportes(self):
        """Actualizar la lista de reportes según el tipo seleccionado"""
        logging.info(f"Actualizando lista de reportes - Tipo: {self.tipo_reporte_var.get()}")

        # Limpiar treeview
        for item in self.reportes_tree.get_children():
            self.reportes_tree.delete(item)

        # Directorio de reportes
        reportes_dir = Path("data/reportes_ia")
        if not reportes_dir.exists():
            logging.warning(f"Directorio {reportes_dir} no existe, creandolo...")
            reportes_dir.mkdir(parents=True, exist_ok=True)
            return

        # Obtener tipo seleccionado
        tipo = self.tipo_reporte_var.get()
        patron = "PARCIAL" if tipo == "parcial" else "COMPLETA"

        # Buscar archivos
        reportes = []
        for archivo in reportes_dir.glob("*.md"):
            if patron in archivo.name:
                # Parsear nombre: YYYYMMDD_HHMMSS_TIPO_casos.md
                partes = archivo.name.split("_")
                if len(partes) >= 3:
                    fecha_str = partes[0]
                    hora_str = partes[1]

                    # Formatear fecha
                    try:
                        fecha_obj = datetime.strptime(f"{fecha_str} {hora_str}", "%Y%m%d %H%M%S")
                        fecha_display = fecha_obj.strftime("%d/%m/%Y %H:%M:%S")
                    except:
                        fecha_display = f"{fecha_str} {hora_str}"

                    # Extraer casos
                    casos_str = "_".join(partes[3:]).replace(".md", "")

                    reportes.append({
                        "fecha": fecha_obj if 'fecha_obj' in locals() else datetime.now(),
                        "fecha_display": fecha_display,
                        "tipo": "Parcial" if patron == "PARCIAL" else "Completo",
                        "casos": casos_str,
                        "archivo": archivo.name,
                        "ruta": str(archivo)
                    })

        # Ordenar por fecha (más recientes primero)
        reportes.sort(key=lambda x: x["fecha"], reverse=True)

        # Agregar a treeview
        for rep in reportes:
            self.reportes_tree.insert("", "end", values=(
                rep["fecha_display"],
                rep["tipo"],
                rep["casos"],
                rep["archivo"]
            ), tags=(rep["ruta"],))

        logging.info(f"{len(reportes)} reportes cargados")

    def _cargar_reporte_seleccionado(self, event):
        """Cargar y mostrar el reporte seleccionado"""
        selection = self.reportes_tree.selection()
        if not selection:
            return

        # Obtener ruta del archivo
        item = selection[0]
        tags = self.reportes_tree.item(item, "tags")
        if not tags:
            return

        ruta_archivo = tags[0]
        logging.info(f"Cargando reporte: {ruta_archivo}")

        try:
            # Leer contenido
            with open(ruta_archivo, 'r', encoding='utf-8') as f:
                contenido = f.read()

            # Renderizar en el text widget
            self._renderizar_markdown(contenido)
            self.reporte_seleccionado = ruta_archivo

        except Exception as e:
            logging.error(f"Error cargando reporte: {e}")
            messagebox.showerror("Error", f"No se pudo cargar el reporte:\n{e}")

    def _renderizar_markdown(self, contenido):
        """Renderizar contenido Markdown con formato en el Text widget"""
        self.markdown_text.config(state=NORMAL)
        self.markdown_text.delete(1.0, END)

        lineas = contenido.split('\n')

        for linea in lineas:
            # Headers
            if linea.startswith('# '):
                self.markdown_text.insert(END, linea[2:] + '\n', "h1")
            elif linea.startswith('## '):
                self.markdown_text.insert(END, linea[3:] + '\n', "h2")
            elif linea.startswith('### '):
                self.markdown_text.insert(END, linea[4:] + '\n', "h3")
            # Listas
            elif linea.strip().startswith(('-', '*', '•')):
                self.markdown_text.insert(END, linea + '\n', "list")
            # Línea normal - procesar bold, italic, code
            else:
                self._procesar_linea_con_formato(linea)

        self.markdown_text.config(state=DISABLED)

    def _procesar_linea_con_formato(self, linea):
        """Procesar una línea aplicando formato inline (bold, italic, code)"""
        pos = 0
        while pos < len(linea):
            # Bold **texto**
            if linea[pos:pos+2] == '**':
                cierre = linea.find('**', pos+2)
                if cierre != -1:
                    self.markdown_text.insert(END, linea[pos+2:cierre], "bold")
                    pos = cierre + 2
                    continue

            # Italic *texto*
            if linea[pos] == '*' and (pos == 0 or linea[pos-1] != '*'):
                cierre = linea.find('*', pos+1)
                if cierre != -1 and (cierre+1 >= len(linea) or linea[cierre+1] != '*'):
                    self.markdown_text.insert(END, linea[pos+1:cierre], "italic")
                    pos = cierre + 1
                    continue

            # Code `texto`
            if linea[pos] == '`':
                cierre = linea.find('`', pos+1)
                if cierre != -1:
                    self.markdown_text.insert(END, linea[pos+1:cierre], "code")
                    pos = cierre + 1
                    continue

            # Carácter normal
            self.markdown_text.insert(END, linea[pos])
            pos += 1

        self.markdown_text.insert(END, '\n')

    def _copiar_contenido_reporte(self):
        """Copiar el contenido del reporte al portapapeles"""
        contenido = self.markdown_text.get(1.0, END)
        self.clipboard_clear()
        self.clipboard_append(contenido)
        messagebox.showinfo("Copiado", "Contenido copiado al portapapeles")

    def _seleccionar_ultimo_reporte(self):
        """Seleccionar automáticamente el primer reporte (más reciente) en la lista"""
        # Obtener primer item del treeview
        items = self.reportes_tree.get_children()
        if items:
            # Seleccionar y hacer focus en el primer item
            primer_item = items[0]
            self.reportes_tree.selection_set(primer_item)
            self.reportes_tree.focus(primer_item)
            self.reportes_tree.see(primer_item)

            # Cargar el reporte manualmente
            tags = self.reportes_tree.item(primer_item, "tags")
            if tags:
                ruta_archivo = tags[0]
                logging.info(f"Auto-cargando reporte: {ruta_archivo}")

                try:
                    with open(ruta_archivo, 'r', encoding='utf-8') as f:
                        contenido = f.read()
                    self._renderizar_markdown(contenido)
                    self.reporte_seleccionado = ruta_archivo
                except Exception as e:
                    logging.error(f"Error auto-cargando reporte: {e}")

    def _generar_reporte_ia(self, resultados, tipo):
        """
        Generar un archivo de reporte Markdown a partir de los resultados de auditoría

        Args:
            resultados: Dict con resultados de auditoría IA O lista de resultados
            tipo: 'parcial' o 'completa'

        Returns:
            str: Ruta al archivo generado
        """
        logging.info(f"Generando reporte IA tipo: {tipo}")

        # Crear directorio si no existe
        reportes_dir = Path("data/reportes_ia")
        reportes_dir.mkdir(parents=True, exist_ok=True)

        # V2.1.6: Detectar formato de entrada (dict o lista)
        if isinstance(resultados, list):
            # Formato nuevo: lista plana de resultados
            registros = resultados
            casos_procesados = len(registros)
            casos_exitosos = sum(1 for r in registros if r.get("exito", False))
            casos_errores = casos_procesados - casos_exitosos
            total_correcciones = sum(r.get("correcciones_aplicadas", 0) for r in registros)
        else:
            # Formato viejo: dict con "registros"
            registros = resultados.get("registros", [])
            casos_procesados = len(registros)
            casos_exitosos = sum(1 for r in registros if r.get("estado") == "exitoso")
            casos_errores = casos_procesados - casos_exitosos
            total_correcciones = sum(len(r.get("correcciones", [])) for r in registros)

        # Determinar rango de IHQ (compatible con ambos formatos)
        ihq_numeros = []
        for r in registros:
            # V2.1.6: Soportar ambos formatos
            ihq = r.get("ihq_numero") or r.get("numero_peticion", "")
            if ihq and ihq.startswith("IHQ"):
                try:
                    num = int(ihq.replace("IHQ", ""))
                    ihq_numeros.append(num)
                except:
                    pass

        if ihq_numeros:
            ihq_min = min(ihq_numeros)
            ihq_max = max(ihq_numeros)
            casos_str = f"IHQ{ihq_min:05d}_IHQ{ihq_max:05d}" if ihq_min != ihq_max else f"IHQ{ihq_min:05d}"
        else:
            casos_str = f"{casos_procesados}_casos"

        # Generar nombre de archivo
        timestamp = datetime.now()
        fecha_str = timestamp.strftime("%Y%m%d")
        hora_str = timestamp.strftime("%H%M%S")
        tipo_str = "PARCIAL" if tipo == "parcial" else "COMPLETA"
        nombre_archivo = f"{fecha_str}_{hora_str}_{tipo_str}_{casos_str}.md"
        ruta_archivo = reportes_dir / nombre_archivo

        # Generar contenido Markdown
        contenido = f"""# 📊 Reporte de Auditoría {tipo_str.title()} - ONCONOVA CIRUGÍA ONCOLÓGICA

**Fecha**: {timestamp.strftime("%d/%m/%Y %H:%M:%S")}
**Tipo**: {"Parcial" if tipo == "parcial" else "Completa"}
**Casos procesados**: {casos_procesados}

## 📋 Resumen Ejecutivo

- ✅ Casos exitosos: {casos_exitosos}
- ❌ Casos con errores: {casos_errores}
- 🔧 Total correcciones: {total_correcciones}

## 📝 Detalle por Caso

"""

        # V3.2.4.2: Agregar detalles de cada caso CON DIFERENCIAS según tipo
        for registro in registros:
            # Formato nuevo vs viejo
            if isinstance(resultados, list):
                # Formato nuevo
                ihq = registro.get("numero_peticion", "N/A")
                estado = "EXITOSO" if registro.get("exito", False) else "ERROR"
                num_correcciones = registro.get("correcciones_aplicadas", 0)
                detalles = registro.get("detalles", [])
                error = registro.get("error", "")
                # V3.2.4.2: Nuevos campos
                analisis_profundo = registro.get("analisis_profundo", {})
                no_encontrados = registro.get("no_encontrados", [])
            else:
                # Formato viejo
                ihq = registro.get("ihq_numero", "N/A")
                estado = registro.get("estado", "desconocido").upper()
                detalles = registro.get("correcciones", [])
                num_correcciones = len(detalles)
                error = registro.get("errores", [])
                analisis_profundo = {}
                no_encontrados = []

            contenido += f"### Caso: {ihq}\n\n"
            contenido += f"**Estado**: {estado}  \n"
            contenido += f"**Correcciones aplicadas**: {num_correcciones}\n\n"

            # Correcciones aplicadas
            if detalles and any(d.get("aplicada") for d in detalles):
                contenido += "**Correcciones aplicadas:**\n\n"
                for i, detalle in enumerate(detalles, 1):
                    if not detalle.get("aplicada"):
                        continue
                    campo = detalle.get("campo", "N/A")
                    valor_anterior = detalle.get("valor_anterior", "")
                    valor_nuevo = detalle.get("valor_nuevo", "")
                    razon = detalle.get("razon", "")

                    contenido += f"{i}. **Campo**: `{campo}`\n"
                    contenido += f"   - *Valor anterior*: {valor_anterior or '(vacío)'}\n"
                    contenido += f"   - *Valor nuevo*: {valor_nuevo}\n"
                    contenido += f"   - *Razón*: {razon}\n\n"

            # V3.2.4.2: AUDITORÍA PARCIAL - Mostrar campos no encontrados
            if tipo == 'parcial' and no_encontrados:
                contenido += f"**Campos no encontrados en el PDF** ({len(no_encontrados)}):\n\n"
                contenido += "> Estos campos están vacíos en BD y no se encontraron datos en el PDF para completarlos.\n\n"
                for campo in no_encontrados[:10]:  # Mostrar máximo 10
                    if isinstance(campo, dict):
                        campo_nombre = campo.get("campo", str(campo))
                    else:
                        campo_nombre = str(campo)
                    contenido += f"- {campo_nombre}\n"
                if len(no_encontrados) > 10:
                    contenido += f"\n... y {len(no_encontrados) - 10} campos más.\n"
                contenido += "\n"

            # V3.2.4.2: AUDITORÍA COMPLETA - Mostrar análisis profundo
            if tipo == 'completa' and analisis_profundo:
                contenido += "## 📊 Análisis Profundo\n\n"

                # Veracidad
                veracidad = analisis_profundo.get("veracidad_porcentaje")
                if veracidad is not None:
                    contenido += f"**Veracidad**: {veracidad}% de coincidencia entre PDF y BD\n\n"

                # Problemas detectados
                problemas = analisis_profundo.get("problemas_detectados", [])
                if problemas:
                    contenido += f"**⚠️ Problemas detectados** ({len(problemas)}):\n\n"
                    for problema in problemas:
                        contenido += f"- {problema}\n"
                    contenido += "\n"

                # Sugerencias
                sugerencias = analisis_profundo.get("sugerencias", [])
                if sugerencias:
                    contenido += f"**💡 Sugerencias de mejora** ({len(sugerencias)}):\n\n"
                    for sugerencia in sugerencias:
                        contenido += f"- {sugerencia}\n"
                    contenido += "\n"

                # Biomarcadores no mapeados
                bio_no_mapeados = analisis_profundo.get("biomarcadores_no_mapeados", [])
                if bio_no_mapeados:
                    contenido += f"**🧬 Biomarcadores sin columna** ({len(bio_no_mapeados)}):\n\n"
                    for bio in bio_no_mapeados:
                        contenido += f"- {bio}\n"
                    contenido += "\n"

            # Errores
            if error:
                contenido += "**Errores encontrados:**\n\n"
                if isinstance(error, list):
                    for e in error:
                        contenido += f"- {e}\n"
                else:
                    contenido += f"- {error}\n"
                contenido += "\n"

            contenido += "---\n\n"

        # Footer
        contenido += f"""
---
*Generated by ONCONOVA CIRUGÍA ONCOLÓGICA*
*{timestamp.strftime("%d/%m/%Y %H:%M:%S")}*
"""

        # Guardar archivo
        try:
            with open(ruta_archivo, 'w', encoding='utf-8') as f:
                f.write(contenido)
            logging.info(f"Reporte generado: {ruta_archivo}")
            return str(ruta_archivo)
        except Exception as e:
            logging.error(f"Error guardando reporte: {e}")
            return None

    def _create_kpi_card(self, parent, title, value):
        """Crear una tarjeta KPI usando TTKBootstrap"""
        card = ttk.Frame(parent, padding=10, relief="solid", borderwidth=1)
        
        ttk.Label(card, text=title, font=("Segoe UI", 12)).pack(anchor=W, padx=4, pady=(2, 0))
        value_lbl = ttk.Label(card, text=value, font=("Segoe UI", 26, "bold")).pack(anchor=W, padx=4, pady=(0, 2))
        
        # Store reference for updating
        card.value_lbl = list(card.winfo_children())[-1] if card.winfo_children() else None
        
        return card

    def _draw_rounded_rect(self, canvas, x1, y1, x2, y2, radius, **kwargs):
        """Dibujar un rectángulo con esquinas redondeadas en un Canvas"""
        points = [
            x1 + radius, y1,
            x2 - radius, y1,
            x2, y1,
            x2, y1 + radius,
            x2, y2 - radius,
            x2, y2,
            x2 - radius, y2,
            x1 + radius, y2,
            x1, y2,
            x1, y2 - radius,
            x1, y1 + radius,
            x1, y1,
            x1 + radius, y1,
        ]
        return canvas.create_polygon(points, smooth=True, **kwargs)

    def _create_circle_icon(self, parent, icon_text, size=52, bg_color="#2d3075", fg_color="white"):
        """Crear un icono circular con fondo coloreado usando Canvas"""
        canvas = tk.Canvas(parent, width=size, height=size, highlightthickness=0, bg=parent["bg"] if isinstance(parent, tk.Canvas) else "#1e2152")
        # Dibujar círculo de fondo
        pad = 2
        canvas.create_oval(pad, pad, size - pad, size - pad, fill=bg_color, outline="")
        # Texto del icono centrado
        canvas.create_text(size // 2, size // 2, text=icon_text, font=("Segoe UI", 16), fill=fg_color)
        return canvas

    def _cargar_logo_bienvenida(self, target=150, tinte="#2d3e5e"):
        """Carga el logo del hospital (favicon.png) para la pantalla de bienvenida.
        Busca en imagenes/favicon.png y en la raiz del proyecto. Lo redimensiona
        a 'target' px de alto manteniendo proporcion.

        Si 'tinte' no es None, recolorea la silueta del logo a ese color usando
        el canal alpha como mascara (el favicon es blanco con transparencia y se
        perderia sobre el fondo claro; lo pasamos a azul institucional #2d3e5e).
        Devuelve ImageTk.PhotoImage o None si la imagen no existe todavia."""
        try:
            base = database_manager.get_base_path()
            candidatos = [
                base / "imagenes" / "favicon.png",
                base / "favicon.png",
            ]
            ruta = next((p for p in candidatos if p.exists()), None)
            if ruta is None:
                return None
            img = Image.open(ruta).convert("RGBA")
            if tinte:
                rgb = tuple(int(tinte.lstrip('#')[i:i+2], 16) for i in (0, 2, 4))
                solido = Image.new("RGBA", img.size, rgb + (0,))
                solido.putalpha(img.getchannel("A"))  # conserva la forma del logo
                img = solido
            w, h = img.size
            if h > 0:
                escala = target / float(h)
                img = img.resize((max(1, int(w * escala)), target), Image.LANCZOS)
            return ImageTk.PhotoImage(img)
        except Exception as e:
            logging.warning(f"[welcome] No se pudo cargar favicon.png: {e}")
            return None

    def _cargar_logo_onconova(self, target=160):
        """Carga el isotipo ONCONOVA a color (imagenes/branding/onconova_isotipo.png),
        redimensionado a 'target' px de alto manteniendo proporción.
        Devuelve ImageTk.PhotoImage o None si no existe."""
        try:
            ruta = self._get_path(os.path.join("imagenes", "branding", "onconova_isotipo.png"))
            if not os.path.exists(ruta):
                return None
            img = Image.open(ruta).convert("RGBA")
            w, h = img.size
            if h > 0:
                escala = target / float(h)
                img = img.resize((max(1, int(w * escala)), target), Image.LANCZOS)
            return ImageTk.PhotoImage(img)
        except Exception as e:
            logging.warning(f"[welcome] No se pudo cargar onconova_isotipo.png: {e}")
            return None

    def _create_welcome_screen(self):
        """Pantalla de bienvenida minimalista clara (V6.9.16).
        Sin tarjeta oscura 'pegada': contenido centrado que respira sobre el
        fondo del tema. Logo del hospital grande, titulo en azul institucional
        y texto de alto contraste."""
        self.welcome_frame = ttk.Frame(self.content_container, padding=40)

        # Contenedor centrado (pack expand sin fill -> se centra en ambos ejes)
        center = ttk.Frame(self.welcome_frame)
        center.pack(expand=True)

        # --- Logo ONCONOVA (isotipo a color) grande y centrado ---
        logo_img = self._cargar_logo_onconova(target=170)
        if logo_img is None:
            logo_img = self._cargar_logo_bienvenida(target=150)  # fallback: logo HUV
        if logo_img is not None:
            self._welcome_logo_ref = logo_img  # mantener referencia (evita GC)
            ttk.Label(center, image=logo_img).pack(pady=(0, 24))

        # --- Titulo en azul institucional ---
        ttk.Label(
            center,
            text="Bienvenido a ONCONOVA",
            font=("Segoe UI Semibold", 30),
            bootstyle="primary",
            anchor="center",
            justify="center"
        ).pack()

        # --- Subtitulo de alto contraste ---
        ttk.Label(
            center,
            text="Investigación y mejora del área de oncología\ndel Hospital Universitario del Valle",
            font=("Segoe UI", 13),
            bootstyle="secondary",
            anchor="center",
            justify="center"
        ).pack(pady=(12, 30))

        # --- Accion segun haya datos o no ---
        has_data = False
        try:
            from core.database_manager import get_all_records_as_dataframe
            df_check = get_all_records_as_dataframe()
            has_data = not df_check.empty
        except Exception as e:
            logging.error(f"Error verificando datos: {e}")

        if not has_data:
            ttk.Label(
                center,
                text="Aún no hay información en la base de datos",
                font=("Segoe UI", 12),
                bootstyle="secondary",
                anchor="center"
            ).pack(pady=(0, 16))
            ttk.Button(
                center,
                text="+   Agregar información a la base de datos",
                bootstyle="primary",
                padding=(22, 12),
                command=self._goto_import_data_tab
            ).pack()
        else:
            ttk.Label(
                center,
                text="Seleccioná una opción del menú para comenzar\na trabajar con los datos oncológicos",
                font=("Segoe UI", 12),
                bootstyle="secondary",
                anchor="center",
                justify="center"
            ).pack()

        # Aviso del atajo de teclado para abrir el menu (V6.9.16)
        ttk.Label(
            center,
            text="💡  Presioná   Ctrl + B   para abrir el menú de navegación",
            font=("Segoe UI", 10),
            bootstyle="secondary",
            anchor="center"
        ).pack(pady=(38, 0))

    def _goto_import_data_tab(self):
        """Navegar directamente a la pestaña de importar datos del dashboard"""
        # Navegar a base de datos
        self._nav_to_database()
        # Esperar un poco para que se cargue el dashboard
        self.after(100, self._select_import_tab)

    def _select_import_tab(self):
        """Seleccionar la pestaña de importar datos en el dashboard mejorado"""
        try:
            if hasattr(self, 'enhanced_dashboard') and hasattr(self.enhanced_dashboard, 'notebook'):
                # La pestaña de importar es la número 4 (índice 4)
                self.enhanced_dashboard.notebook.select(4)
        except Exception as e:
            logging.error(f"Error seleccionando pestana de importar: {e}")



    # Métodos de navegación actualizados
    def show_database_frame(self):
        """Mostrar panel de base de datos (función compatibilidad)"""
        self._nav_to_database()
        # Actualizar dashboard mejorado cuando se muestre el panel
        if hasattr(self, 'enhanced_dashboard'):
            self.enhanced_dashboard.refresh_dashboard()

    def show_visualizar_frame(self):
        """Mostrar panel de visualización (función compatibilidad)"""
        self._nav_to_visualizar()

    def show_dashboard_frame(self):
        """Mostrar panel de análisis gráfico (función compatibilidad)"""
        self._nav_to_dashboard()
        # CORREGIDO: Asegurar que los datos estén cargados antes de mostrar el dashboard
        try:
            # Si master_df está vacío, cargar datos de la base de datos
            if self.master_df.empty:
                from core.database_manager import init_db, get_all_records_as_dataframe
                # V6.2.0: Comentado - init_db() ya se llama en ihq_processor antes del guardado
                # Llamarlo aquí causa que el UPDATE de relleno sobrescriba valores recién insertados
                # init_db()
                self.master_df = get_all_records_as_dataframe()

                # Ordenar por número de caso automáticamente
                if self.master_df is not None and not self.master_df.empty and "Numero de caso" in self.master_df.columns:
                    self.master_df = self.master_df.sort_values(
                        by="Numero de caso",
                        ascending=True,
                        na_position='last'
                    ).reset_index(drop=True)

            # Ahora cargar el dashboard con datos disponibles
            self.cargar_dashboard()
        except Exception as e:
            logging.error(f"Error cargando datos para dashboard: {e}")
            # Cargar dashboard vacío para mostrar mensaje apropiado
            self.cargar_dashboard()
        
    def show_welcome_screen(self):
        """Mostrar pantalla de bienvenida con header visible y menú flotante oculto"""
        # Ocultar otros paneles
        if self.panel_activo:
            self.panel_activo.pack_forget()
        
        # Ocultar menú flotante si está visible
        if self.floating_menu_visible:
            self._hide_floating_menu()
        
        # Mostrar header (solo en pantalla de bienvenida)
        self._show_header()
        
        # Mostrar la pantalla de bienvenida
        self.welcome_frame.pack(fill=BOTH, expand=True)
        self.panel_activo = self.welcome_frame
        self.welcome_screen_active = True
        self.current_view = "welcome"

    def _hide_welcome_and_show_panel(self, panel):
        """Ocultar pantalla de bienvenida, animar menús y mostrar panel"""
        # Si es la primera navegación desde la pantalla de bienvenida
        if self.welcome_screen_active:
            self._animate_menus_hide()
            self.welcome_screen_active = False
        
        # Cambiar el panel
        self._show_panel(panel)
    
    def _show_panel(self, panel):
        """Mostrar un panel específico"""
        if self.panel_activo:
            self.panel_activo.pack_forget()
        
        self.panel_activo = panel
        panel.pack(fill=BOTH, expand=True)
    
    def _animate_menus_hide(self):
        """Animar el ocultamiento de header y sidebar"""
        # Animar header hacia arriba
        self._animate_header_hide()
        # Animar sidebar hacia la izquierda  
        self._animate_sidebar_hide()
        # Actualizar el botón de navegación
        self.nav_toggle_btn.configure(text="▶ Mostrar Menús")

    def _animate_header_hide(self):
        """Animar ocultamiento del header hacia arriba - ULTRA RÁPIDO"""
        def slide_up(steps_remaining):
            if steps_remaining > 0:
                current_height = self.header.winfo_height()
                new_height = max(0, current_height - 20)  # Aumentado de 10 a 20
                if new_height > 0:
                    self.after(5, lambda: slide_up(steps_remaining - 1))  # Reducido de 20ms a 5ms
                else:
                    self.header.pack_forget()
                    self.header_separator.pack_forget()
                    self.header_visible = False
            
        if self.header_visible:
            slide_up(5)  # Reducido de 10 a 5 pasos

    def _animate_sidebar_hide(self):
        """Animar ocultamiento del sidebar hacia la izquierda - ULTRA RÁPIDO"""
        if not hasattr(self, 'sidebar') or not self.sidebar_visible:
            self.sidebar_visible = False
            return
        def slide_left(steps_remaining):
            if steps_remaining > 0:
                current_width = self.sidebar.winfo_width()
                new_width = max(0, current_width - 50)  # Aumentado de 30 a 50
                self.sidebar.config(width=new_width)
                if new_width > 0:
                    self.after(5, lambda: slide_left(steps_remaining - 1))  # Reducido de 20ms a 5ms
                else:
                    self.sidebar.pack_forget()
                    self.sidebar_visible = False
            
        if self.sidebar_visible:
            slide_left(4)  # Reducido de 8 a 4 pasos

    def _toggle_navigation_visibility(self):
        """Alternar visibilidad de header y sidebar con animación"""
        if self.header_visible and self.sidebar_visible:
            # Ocultar menús
            self._animate_menus_hide()
        else:
            # Mostrar menús
            self._animate_menus_show()

    def _animate_menus_show(self):
        """Animar la aparición de header y sidebar"""
        # Mostrar header
        if not self.header_visible:
            # Obtener el contenedor padre correcto
            parent = self.content_container.master
            self.header.pack(fill=X, before=parent.children[list(parent.children.keys())[0]])
            self.header_visible = True
            
        # Mostrar sidebar
        if not self.sidebar_visible and hasattr(self, 'sidebar'):
            # Obtener el contenedor padre correcto (body frame)
            body_parent = self.content_container.master
            self.sidebar.pack(side=LEFT, fill=Y, before=self.content_container)
            self.sidebar.config(width=self.sidebar_width)
            self.sidebar_visible = True
            
        # Actualizar botón
        self.nav_toggle_btn.configure(text="◀ Ocultar Menús")
        
        # Si no estamos en la pantalla de bienvenida, salir del modo bienvenida
        if self.welcome_screen_active and self.panel_activo != self.welcome_frame:
            self.welcome_screen_active = False

    def open_web_auto_modal(self):
        """Abrir modal de automatización web"""
        messagebox.showinfo("Web Automation", "Función de automatización web - En desarrollo")



    # Métodos de utilidad (conservando la lógica de carga de archivos)
    def _get_path(self, relative_path):
        """Obtiene la ruta absoluta de un archivo, compatible con PyInstaller"""
        try:
            base_path = sys._MEIPASS
        except Exception:
            base_path = os.path.abspath(os.path.dirname(__file__))
        return os.path.join(base_path, relative_path)

    def _configurar_icono_app(self):
        """Establece el icono institucional ONCONOVA en la ventana y barra de tareas.

        Usa el .ico multi-resolucion en Windows; si falla, recurre al PNG via
        iconphoto (multiplataforma). Nunca interrumpe el arranque de la app.
        """
        # Windows: darle identidad propia en la barra de tareas para que use el
        # icono de la app y NO el de python.exe / PyInstaller por defecto.
        try:
            import ctypes
            ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID("HUV.Onconova.GestorOncologia")
        except Exception:
            pass
        try:
            ico = self._get_path(os.path.join("imagenes", "branding", "onconova.ico"))
            if os.path.exists(ico):
                self.iconbitmap(default=ico)
                return
        except Exception as e:
            logging.warning(f"No se pudo aplicar onconova.ico: {e}")
        try:
            png = self._get_path(os.path.join("imagenes", "branding", "onconova_icono_app.png"))
            if os.path.exists(png):
                # Se conserva la referencia para evitar que el GC libere la imagen.
                self._app_icon_img = ImageTk.PhotoImage(Image.open(png))
                self.iconphoto(True, self._app_icon_img)
        except Exception as e:
            logging.warning(f"No se pudo aplicar el icono PNG de ONCONOVA: {e}")
    
    def _crear_menu_tema(self):
        """V6.9.27 - Barra de menu minima con selector de tema (oscuro/claro)."""
        try:
            import tkinter as tk
            menubar = tk.Menu(self)
            m = tk.Menu(menubar, tearoff=0)
            m.add_command(label="\U0001f319 Oscuro (Onconova)", command=lambda: self._cambiar_tema("darkly"))
            m.add_command(label="☀ Claro (navy HUV)", command=lambda: self._cambiar_tema("huv"))
            menubar.add_cascade(label="\U0001f3a8 Tema", menu=m)
            self.config(menu=menubar)
        except Exception as e:
            logging.warning(f"[tema] No se pudo crear el menu de tema: {e}")

    def _cambiar_tema(self, nombre):
        """V6.9.27 - Cambia el tema ttkbootstrap en caliente + ajusta los graficos."""
        try:
            self.style.theme_use(nombre)
            self.current_theme = nombre
            aplicar_estilo_graficos(es_oscuro=(nombre in DARK_THEMES))
            try:
                self._init_treeview_style()  # re-aplica estilo de tablas al nuevo tema
            except Exception:
                pass
            try:
                from tkinter import messagebox
                messagebox.showinfo(
                    "Tema",
                    f"Tema aplicado: {nombre}.\n\nLos graficos del dashboard se repintan al usar 'Refrescar'.")
            except Exception:
                pass
        except Exception as e:
            logging.warning(f"[tema] No se pudo cambiar a {nombre}: {e}")

    def _cargar_foto_usuario(self):
        """Cargar foto del usuario desde la ruta especificada"""
        try:
            ruta_foto = self.info_usuario.get("ruta_foto", "SIN_FOTO")
            if ruta_foto and ruta_foto != "SIN_FOTO" and os.path.exists(ruta_foto):
                img = Image.open(ruta_foto).resize((50, 50), Image.Resampling.LANCZOS)
                return ImageTk.PhotoImage(img)
            return None
        except Exception as e:
            logging.error(f"Error al cargar la foto del usuario: {e}")
            return None

    def _cargar_iconos(self):
        """Carga todos los iconos necesarios para la interfaz"""
        iconos = {}
        icon_files = {
            "logo1": "logo1.png", 
            "logo2": "logo2.png", 
            "logo3": "logo3.png",
            "usuario": "usuario.png"
        }
        for name, filename in icon_files.items():
            try:
                path = self._get_path(os.path.join("imagenes", filename))
                if name == "logo1":
                    size = (55, 55)       # logo izquierdo header (compacto)
                elif name == "logo3":
                    size = (60, 60)       # logo derecho header (compacto)
                elif name == "logo2":
                    size = (140, 140)
                elif name == "usuario":
                    size = (48, 48)       # foto usuario por defecto
                else:
                    size = (32, 32)
                img = Image.open(path).resize(size, Image.Resampling.LANCZOS)
                iconos[name] = ImageTk.PhotoImage(img)
            except Exception as e:
                logging.warning(f"No se encontró el icono o logo '{filename}': {e}")
                iconos[name] = None
        return iconos

    # =========================
    # Helpers UI (KPI y Status)
    # =========================
    # =========================
    # Métodos de utilidad y renderizado
    # =========================
    def _render_kpis(self, df):
        """Actualizar los valores de los KPIs con datos del DataFrame"""
        # V6.9.45: los KPIs son SOLO de IHQ -> excluir filas de coloración (clave M…).
        if df is not None and not df.empty and "Numero de caso" in df.columns:
            df = df[~df["Numero de caso"].astype(str).str.match(r"^[Mm]\d", na=False)]
        if df is None or df.empty:
            # Resetear valores cuando no hay datos
            try:
                if hasattr(self, 'kpi_total') and hasattr(self.kpi_total, 'value_lbl'):
                    self.kpi_total.value_lbl.configure(text="0")
                if hasattr(self, 'kpi_rango') and hasattr(self.kpi_rango, 'value_lbl'):
                    self.kpi_rango.value_lbl.configure(text="—")
                if hasattr(self, 'kpi_ultimo') and hasattr(self.kpi_ultimo, 'value_lbl'):
                    self.kpi_ultimo.value_lbl.configure(text="—")
            except Exception as e:
                logging.error(f"Error al resetear KPIs: {e}")
            return

        total = len(df)

        # Rango de fechas: fecha mínima y máxima
        fecha_cols = [
            "Fecha Informe",
            "Fecha de informe",
            "Fecha de ingreso",
        ]
        fecha_min = None
        fecha_max = None
        
        for c in fecha_cols:
            if c in df.columns:
                try:
                    fechas_validas = pd.to_datetime(df[c], dayfirst=True, errors="coerce").dropna()
                    if not fechas_validas.empty:
                        if fecha_min is None or fechas_validas.min() < fecha_min:
                            fecha_min = fechas_validas.min()
                        if fecha_max is None or fechas_validas.max() > fecha_max:
                            fecha_max = fechas_validas.max()
                except Exception:
                    pass

        # Formatear rango de fechas
        if fecha_min and fecha_max:
            if fecha_min.date() == fecha_max.date():
                rango_txt = fecha_min.strftime("%d/%m/%Y")
            else:
                rango_txt = f"{fecha_min.strftime('%d/%m/%Y')} - {fecha_max.strftime('%d/%m/%Y')}"
        else:
            rango_txt = "—"

        # Última importación (fecha más reciente)
        ultimo_txt = "—" if (fecha_max is None or pd.isna(fecha_max)) else fecha_max.strftime("%d/%m/%Y")

        # Actualizar KPIs con los nuevos valores
        try:
            if hasattr(self, 'kpi_total') and hasattr(self.kpi_total, 'value_lbl'):
                self.kpi_total.value_lbl.configure(text=f"{total:,}".replace(",", "."))
            if hasattr(self, 'kpi_rango') and hasattr(self.kpi_rango, 'value_lbl'):
                self.kpi_rango.value_lbl.configure(text=rango_txt)
            if hasattr(self, 'kpi_ultimo') and hasattr(self.kpi_ultimo, 'value_lbl'):
                self.kpi_ultimo.value_lbl.configure(text=ultimo_txt)
        except Exception as e:
            logging.error(f"Error al actualizar KPIs: {e}")

    def set_status(self, text):
        """Actualiza el texto de la barra de estado"""
        try:
            # Crear una barra de estado temporal si no existe
            if not hasattr(self, 'status_label'):
                self.status_label = ttk.Label(self, text=text, bootstyle="secondary")
                self.status_label.pack(side="bottom", fill="x", padx=5, pady=2)
            else:
                self.status_label.configure(text=text)
        except Exception as e:
            logging.error(f"Error al actualizar status: {e}")
            logging.info(f"[STATUS] {text}")  # Fallback al console

    # =========================
    # Métodos de carga de recursos (conservados)
    # =========================
            pass

    # =========================
    # Navegación (métodos obsoletos removidos - usar los métodos nuevos)
    # =========================




    # ---------- Helpers Dashboard ----------

    def _clear_filters(self):
        for k in self.db_filters:
            self.db_filters[k].set("")
        self._refresh_dashboard()

    def _refresh_dashboard(self):
        try:
            self.set_status("Actualizando dashboard…")
            self.cargar_dashboard()
        finally:
            self.set_status("Dashboard actualizado.")

    def _get_filtered_df(self, df):
        dff = df.copy()
        # V6.9.45: el Dashboard analítico excluye las filas de coloración (clave M…),
        # que no son casos IHQ (sin Dx Principal/Malignidad). Se ven en el visualizador.
        if "Numero de caso" in dff.columns:
            dff = dff[~dff["Numero de caso"].astype(str).str.match(r"^[Mm]\d", na=False)]
        fd = self.db_filters["fecha_desde"].get().strip()
        fh = self.db_filters["fecha_hasta"].get().strip()
        if fd:
            d0 = pd.to_datetime(fd, dayfirst=True, errors="coerce")
            if pd.notna(d0):
                dff = dff[dff["_fecha_informe"] >= d0]
        if fh:
            d1 = pd.to_datetime(fh, dayfirst=True, errors="coerce")
            if pd.notna(d1):
                dff = dff[dff["_fecha_informe"] <= d1]

        srv = self.db_filters["servicio"].get().strip()
        if srv:
            dff = dff[dff.get("Servicio", "").astype(str).eq(srv)]
        mal = self.db_filters["malignidad"].get().strip()
        if mal:
            dff = dff[dff.get("Malignidad", "").astype(str).str.upper().eq(mal)]
        rsp = self.db_filters["responsable"].get().strip()
        if rsp:
            dff = dff[dff.get("Patologo", "").astype(str).eq(rsp)]
        return dff

    def _clear_dash_area(self):
        # Desmonta los canvases previos para liberar memoria
        for cv in getattr(self, "_dash_canvases", []):
            try:
                cv.get_tk_widget().destroy()
            except Exception:
                pass
        self._dash_canvases = []

        # Limpia frames hijos en cada pestaña
        for tab in [self.tab_overview, self.tab_biomarkers, self.tab_times, self.tab_quality, self.tab_compare]:
            for child in tab.grid_slaves():
                child.destroy()

    def _chart_in(self, tab, row, col, render_fn, title, dff):
        card = ttk.Frame(tab, padding=8, relief="solid", borderwidth=1)
        card.grid(row=row, column=col, padx=8, pady=8, sticky="nsew")
        
        # Configurar responsive grid
        tab.grid_rowconfigure(row, weight=1)
        tab.grid_columnconfigure(col, weight=1)

        header = ttk.Frame(card)
        header.pack(fill="x", pady=(0, 6))
        ttk.Label(header, text=title, font=("Segoe UI", 11, "bold")).pack(side="left")
        ttk.Button(header, text="⛶", 
                  command=lambda: self._open_fullscreen_figure(render_fn, title, dff),
                  bootstyle="outline").pack(side="right")

        try:
            fig = render_fn()
            if fig is None:
                ttk.Label(card, text="(sin datos)", font=("Segoe UI", 10)).pack(padx=10, pady=20)
                return
            
            # Ajustar tamaño de figura para mejor responsive
            fig.set_size_inches(4.5, 3.2)
            fig.set_dpi(90)
            
            canvas = FigureCanvasTkAgg(fig, master=card)
            canvas.draw()
            widget = canvas.get_tk_widget()
            widget.pack(fill="both", expand=True, padx=4, pady=4)
            widget.bind("<Double-Button-1>", lambda e: self._open_fullscreen_figure(render_fn, title, dff))
            self._dash_canvases.append(canvas)
        except Exception as e:
            ttk.Label(card, text=f"Error: {e}", font=("Segoe UI", 9), bootstyle="danger").pack(padx=10, pady=10)

    def _toggle_db_sidebar(self):
        # Muestra/oculta el sidebar, ajusta texto del botón y grid
        if self.db_sidebar_collapsed:
            self.db_sidebar.grid(row=0, column=0, sticky="ns", padx=(0, 12), pady=6)
            self.btn_toggle_sidebar.configure(text="✕ Ocultar filtros")
        else:
            self.db_sidebar.grid_forget()
            self.btn_toggle_sidebar.configure(text="≡ Mostrar filtros")
        self.db_sidebar_collapsed = not self.db_sidebar_collapsed

    def _open_filters_sheet(self):
        # Modal de filtros (para no robar ancho). V6.9.16: cada campo se crea
        # con su propia fila como parent (antes se usaba pack(in_=...) con master
        # distinto, lo que dejaba los campos ocultos detras del frame).
        top = tk.Toplevel(self)
        top.title("Filtros")
        top.geometry("460x380")
        top.grab_set()
        top.transient(self)

        wrap = ttk.Frame(top, padding=16)
        wrap.pack(fill="both", expand=True)

        # Opciones de los combos (tomadas de los combos del panel oculto)
        def _vals(combo, default=None):
            if combo is not None:
                try:
                    return list(combo.cget("values"))
                except Exception:
                    pass
            return default if default is not None else []
        servicio_vals = _vals(self.cmb_servicio)
        malig_vals = _vals(self.cmb_malig, ["", "PRESENTE", "AUSENTE"])
        resp_vals = _vals(self.cmb_resp)

        campos = [
            ("Fecha desde (dd/mm/aaaa)", "entry", "fecha_desde", None),
            ("Fecha hasta (dd/mm/aaaa)", "entry", "fecha_hasta", None),
            ("Servicio", "combo", "servicio", servicio_vals),
            ("Malignidad", "combo", "malignidad", malig_vals),
            ("Responsable", "combo", "responsable", resp_vals),
        ]
        for lbl, tipo, key, vals in campos:
            r = ttk.Frame(wrap)
            r.pack(fill="x", pady=7)
            ttk.Label(r, text=lbl, width=22, anchor="w").pack(side="left")
            if tipo == "entry":
                w = ttk.Entry(r, textvariable=self.db_filters[key])
            else:
                w = ttk.Combobox(r, values=vals, textvariable=self.db_filters[key])
            w.pack(side="left", fill="x", expand=True)

        btns = ttk.Frame(wrap)
        btns.pack(fill="x", pady=(18, 0))
        ttk.Button(btns, text="Aplicar", bootstyle="primary",
                   command=lambda: (self._refresh_dashboard(), top.destroy())).pack(
                   side="left", expand=True, fill="x", padx=(0, 6))
        ttk.Button(btns, text="Limpiar", bootstyle="secondary-outline",
                   command=self._clear_filters).pack(
                   side="left", expand=True, fill="x", padx=(6, 0))

    def _open_fullscreen_figure(self, render_fn, title, dff):
        # Ventana a pantalla completa con inspector lateral
        fs = tk.Toplevel(self)
        fs.title(title)
        try:
            fs.state('zoomed')
        except Exception:
            pass
        fs.grid_rowconfigure(0, weight=1)
        fs.grid_columnconfigure(0, weight=1)
        fs.grid_columnconfigure(1, weight=0)

        # Área de gráfico
        graph_area = ttk.Frame(fs, padding=10)
        graph_area.grid(row=0, column=0, sticky="nsew", padx=(10,6), pady=10)
        fig = render_fn()
        if fig is None:
            ttk.Label(graph_area, text="(sin datos)").pack(padx=12, pady=12)
        else:
            canv = FigureCanvasTkAgg(fig, master=graph_area)
            canv.draw()
            canv.get_tk_widget().pack(fill="both", expand=True, padx=10, pady=10)

        # Inspector lateral
        insp = ttk.Frame(fs, padding=10, width=300)
        insp.grid(row=0, column=1, sticky="ns", padx=(6,10), pady=10)
        ttk.Label(insp, text="Inspector", font=("Segoe UI", 16, "bold")).pack(anchor="w", pady=(0,6))
        self._build_inspector(insp, title, dff)

        # Barra superior simple (cerrar)
        topbar = ttk.Frame(graph_area)
        topbar.pack(fill="x", pady=(0,10))
        ttk.Label(topbar, text=title, font=("Segoe UI", 14, "bold")).pack(side="left")
        ttk.Button(topbar, text="Cerrar", command=fs.destroy).pack(side="right")

    def _build_inspector(self, parent, title, dff):
        # Datos generales
        n = len(dff)
        fmin = pd.to_datetime(dff.get("_fecha_informe"), errors="coerce").min()
        fmax = pd.to_datetime(dff.get("_fecha_informe"), errors="coerce").max()
        rng = f"{fmin:%d/%m/%Y} – {fmax:%d/%m/%Y}" if pd.notna(fmin) and pd.notna(fmax) else "—"

        def row(k, v):
            r = ttk.Frame(parent)
            r.pack(fill="x", padx=12, pady=4)
            ttk.Label(r, text=k).pack(side="left")
            ttk.Label(r, text=v).pack(side="right")

        row("Registros filtrados", f"{n:,}".replace(",", "."))
        row("Rango de fechas", rng)

        # Secciones condicionales útiles
        if "Malignidad" in dff.columns:
            ser = dff["Malignidad"].astype(str).str.upper().value_counts()
            box = ttk.Frame(parent, padding=8, relief="solid", borderwidth=1); box.pack(fill="x", padx=12, pady=(10,4))
            ttk.Label(box, text="Malignidad", font=("Segoe UI", 13, "bold")).pack(anchor="w", padx=10, pady=(8,2))
            for k,v in ser.items():
                rowtxt = ttk.Frame(box); rowtxt.pack(fill="x", padx=10, pady=2)
                ttk.Label(rowtxt, text=f"{k}").pack(side="left")
                ttk.Label(rowtxt, text=str(v)).pack(side="right")

        if "Organo" in dff.columns:
            top_org = dff["Organo"].astype(str).replace({"": "No especificado"}).value_counts().head(8)
        elif "IHQ_ORGANO" in dff.columns:
            top_org = dff["IHQ_ORGANO"].astype(str).replace({"": "No especificado"}).value_counts().head(8)
        else:
            top_org = None

        if top_org is not None and not top_org.empty:
            box2 = ttk.Frame(parent, padding=8, relief="solid", borderwidth=1); box2.pack(fill="x", padx=12, pady=(10,12))
            ttk.Label(box2, text="Top Órganos", font=("Segoe UI", 13, "bold")).pack(anchor="w", padx=10, pady=(8,2))
            for k,v in top_org.items():
                rowtxt = ttk.Frame(box2); rowtxt.pack(fill="x", padx=10, pady=2)
                ttk.Label(rowtxt, text=f"{k}").pack(side="left")
                ttk.Label(rowtxt, text=str(v)).pack(side="right")

    # ---------- Renderers: Overview ----------

    def _g_line_informes_por_mes(self, df):
        if df.empty or df["_fecha_informe"].isna().all():
            return None
        ser = df.dropna(subset=["_fecha_informe"]).set_index("_fecha_informe").resample("MS").size()
        fig = Figure(figsize=(5.6, 3.2), dpi=100)
        ax = fig.add_subplot(111)
        ax.plot(ser.index, ser.values, marker="o")
        ax.set_title("Informes por mes")
        ax.set_xlabel("Mes")
        ax.set_ylabel("Conteo")
        fig.tight_layout()
        return fig

    def _g_pie_malignidad(self, df):
        if "Malignidad" not in df.columns or df.empty:
            return None
        ser = df["Malignidad"].astype(str).str.upper().replace({"": "DESCONOCIDO"}).value_counts()
        if ser.empty: return None
        # V6.9.16: agrupar categorias pequenas (<2.5%) en "OTROS" para que las
        # etiquetas no se amontonen; leyenda lateral en vez de texto sobre el pie;
        # formato donut con la paleta navy.
        total = ser.sum()
        umbral = 0.025 * total
        grandes = ser[ser >= umbral]
        resto = int(ser[ser < umbral].sum())
        if resto > 0:
            grandes = pd.concat([grandes, pd.Series({"OTROS": resto})])
        labels = [self._trunc_label(x, 28) for x in grandes.index]
        colores = ["#2d3e5e", "#b08d57", "#4a6da7", "#2f8f6b", "#9aa6bf",
                   "#c75c6e", "#d99a4e", "#6c757d"]
        fig = Figure(figsize=(6.4, 3.6), dpi=100)
        ax = fig.add_subplot(111)
        wedges, _t, _autotexts = ax.pie(
            grandes.values, autopct="%1.1f%%", startangle=90, pctdistance=0.78,
            colors=colores[:len(grandes)],
            textprops={"fontsize": 8, "color": "white"},
            wedgeprops={"width": 0.45, "edgecolor": "white", "linewidth": 1},
        )
        ax.legend(wedges, labels, loc="center left", bbox_to_anchor=(0.98, 0.5),
                  fontsize=8, frameon=False)
        ax.set_title("Distribución de Malignidad")
        ax.set_aspect("equal")
        ax.grid(False)
        fig.tight_layout()
        return fig

    def _g_bar_top_servicio(self, df, top=12):
        if "Servicio" not in df.columns or df.empty:
            return None
        ser = df["Servicio"].astype(str).value_counts().head(top)
        if ser.empty: return None
        fig = Figure(figsize=(5.6, 3.2), dpi=100)
        ax = fig.add_subplot(111)
        ax.bar(ser.index, ser.values)
        ax.set_title(f"Top Servicios (n={ser.sum()})")
        ax.set_ylabel("Informes")
        ax.tick_params(axis="x", rotation=30, labelsize=8)
        fig.tight_layout()
        return fig

    def _g_bar_top_organo(self, df, top=12):
        # soporta tanto columna Excel como IHQ_ORGANO
        col = "Organo" if "Organo" in df.columns else ("IHQ_ORGANO" if "IHQ_ORGANO" in df.columns else None)
        if not col: return None
        ser = df[col].astype(str).replace({"": "No especificado"}).value_counts().head(top)
        if ser.empty: return None
        fig = Figure(figsize=(5.6, 3.2), dpi=100)
        ax = fig.add_subplot(111)
        ax.bar(ser.index, ser.values)
        ax.set_title("Top Órganos")
        ax.set_ylabel("Informes")
        ax.tick_params(axis="x", rotation=30, labelsize=8)
        fig.tight_layout()
        return fig

    # ---------- Renderers: Biomarcadores ----------

    def _g_hist_ki67(self, df):
        col = "IHQ_KI-67" if "IHQ_KI-67" in df.columns else None
        if not col: return None
        
        # Limpiar y convertir los datos
        raw_data = df[col].astype(str)  # Convertir todo a string primero
        # Remover espacios y reemplazar cadenas vacías con NaN
        clean_data = raw_data.str.strip().replace('', pd.NA)
        # Remover el símbolo % y convertir a numérico
        numeric_data = clean_data.str.replace('%', '', regex=False)
        s = pd.to_numeric(numeric_data, errors="coerce").dropna()
        
        if s.empty: 
            return None
            
        fig = Figure(figsize=(5.6, 3.2), dpi=100)
        ax = fig.add_subplot(111)
        ax.hist(s.values, bins=min(12, len(s.unique())))  # Ajustar bins si hay pocos datos únicos
        ax.set_title("Ki-67 (%)")
        ax.set_xlabel("%")
        ax.set_ylabel("Frecuencia")
        fig.tight_layout()
        return fig

    def _g_bar_her2(self, df):
        col = "IHQ_HER2" if "IHQ_HER2" in df.columns else None
        if not col: return None
        order = ["0", "1+", "2+", "3+", "NEGATIVO", "POSITIVO"]
        ser = df[col].astype(str).str.upper().value_counts()
        ser = ser.reindex(order, fill_value=0) if any(k in ser.index for k in order) else ser
        if ser.sum() == 0: return None
        fig = Figure(figsize=(5.6, 3.2), dpi=100); ax = fig.add_subplot(111)
        ax.bar(ser.index, ser.values)
        ax.set_title("HER2 (score)")
        ax.set_ylabel("Informes")
        fig.tight_layout()
        return fig

    @staticmethod
    def _trunc_label(s, n=42):
        """Trunca una etiqueta larga para que no desborde el grafico."""
        s = str(s)
        return s if len(s) <= n else s[:n - 1] + "…"

    def _g_bar_re_rp(self, df):
        cols = [c for c in ["IHQ_RECEPTOR_ESTROGENOS", "IHQ_RECEPTOR_PROGESTERONA"] if c in df.columns]
        if not cols: return None
        data = []
        labels = []
        for c in cols:
            ser = df[c].astype(str).str.upper().replace({"": "ND"}).value_counts()
            data.append(ser)
            labels.append(c.replace("IHQ_RECEPTOR_", "").title())
        # Top 8 categorias combinadas (evita amontonar valores sucios poco frecuentes)
        total = pd.concat(data, axis=1).fillna(0).sum(axis=1).sort_values(ascending=True)
        cats = list(total.tail(8).index)
        mat = np.array([[d.get(k, 0) for k in cats] for d in data])
        fig = Figure(figsize=(5.6, 3.6), dpi=100); ax = fig.add_subplot(111)
        y = np.arange(len(cats)); h = 0.38
        colores = ["#2d3e5e", "#9aa6bf"]
        for i, rowv in enumerate(mat):
            ax.barh(y + i * h, rowv, height=h, label=labels[i],
                    color=colores[i % len(colores)])
        ax.set_yticks(y + h / 2)
        ax.set_yticklabels([self._trunc_label(k, 30) for k in cats], fontsize=8)
        ax.set_title("RE / RP (estado)")
        ax.set_xlabel("Informes")
        ax.legend(fontsize=8)
        fig.tight_layout()
        return fig

    def _g_bar_pdl1(self, df):
        # intenta TPS o CPS
        for col in ["IHQ_PDL-1", "IHQ_PDL1_TPS", "IHQ_PDL1_CPS"]:
            if col in df.columns:
                s = df[col].astype(str)
                break
        else:
            return None
        ser = s.replace({"": "ND"}).value_counts().head(12).iloc[::-1]
        if ser.empty: return None
        fig = Figure(figsize=(5.6, 3.6), dpi=100); ax = fig.add_subplot(111)
        ax.barh([self._trunc_label(x, 38) for x in ser.index], ser.values, color="#2d3e5e")
        ax.set_title("PD-L1")
        ax.set_xlabel("Informes")
        ax.tick_params(axis="y", labelsize=8)
        fig.tight_layout()
        return fig

    # ---------- Renderers: Tiempos ----------

    def _g_box_tiempo_proceso(self, df):
        f_ing = pd.to_datetime(df.get("Fecha de ingreso (2. Fecha de la muestra)", ""), dayfirst=True, errors="coerce")
        f_inf = pd.to_datetime(df.get("Fecha Informe", df.get("Fecha de informe", "")), dayfirst=True, errors="coerce")
        dias = (f_inf - f_ing).dt.days.dropna()
        if dias.empty: return None
        fig = Figure(figsize=(5.6, 3.2), dpi=100); ax = fig.add_subplot(111)
        ax.boxplot(dias.values, vert=True)
        ax.set_title("Tiempo de proceso (días)")
        fig.tight_layout()
        return fig

    def _g_line_throughput_semana(self, df):
        if df.empty or df["_fecha_informe"].isna().all(): return None
        ser = df.dropna(subset=["_fecha_informe"]).set_index("_fecha_informe").resample("W-MON").size()
        fig = Figure(figsize=(5.6, 3.2), dpi=100); ax = fig.add_subplot(111)
        ax.plot(ser.index, ser.values, marker="o")
        ax.set_title("Throughput semanal")
        ax.set_xlabel("Semana")
        ax.set_ylabel("Informes")
        fig.tight_layout()
        return fig

    def _g_scatter_edad_ki67(self, df):
        if "Edad" not in df.columns: return None
        x = pd.to_numeric(df["Edad"], errors="coerce")
        
        # Limpiar los datos Ki-67 igual que en _g_hist_ki67
        if "IHQ_KI-67" in df.columns:
            raw_ki67 = df["IHQ_KI-67"].astype(str).str.strip().replace('', pd.NA)
            clean_ki67 = raw_ki67.str.replace('%', '', regex=False)
            y = pd.to_numeric(clean_ki67, errors="coerce")
        else:
            y = pd.Series(np.nan, index=df.index)
            
        m = x.notna() & y.notna()
        if not m.any(): 
            return None
            
        fig = Figure(figsize=(5.6, 3.2), dpi=100); ax = fig.add_subplot(111)
        ax.scatter(x[m], y[m], alpha=0.6)
        ax.set_title("Edad vs Ki-67")
        ax.set_xlabel("Edad")
        ax.set_ylabel("Ki-67 (%)")
        fig.tight_layout()
        return fig

    # ---------- Renderers: Calidad ----------

    def _g_bar_missingness(self, df):
        cols = [
            "Servicio", "Malignidad", "Patologo",
            "Organo", "IHQ_HER2", "IHQ_KI-67"
        ]
        present = [c for c in cols if c in df.columns]
        if not present: return None
        miss = df[present].isna().mean().sort_values(ascending=False)
        fig = Figure(figsize=(5.6, 3.2), dpi=100); ax = fig.add_subplot(111)
        ax.bar(miss.index, (miss.values*100.0))
        ax.set_title("Campos vacíos (%)")
        ax.set_ylabel("% vacío")
        ax.tick_params(axis="x", rotation=25)
        fig.tight_layout()
        return fig

    def _g_bar_top_responsables(self, df, top=10):
        col = "Patologo"
        if col not in df.columns: return None
        ser = df[col].astype(str).value_counts().head(top).iloc[::-1]
        if ser.empty: return None
        fig = Figure(figsize=(5.6, 3.6), dpi=100); ax = fig.add_subplot(111)
        ax.barh([self._trunc_label(x, 38) for x in ser.index], ser.values, color="#2d3e5e")
        ax.set_title("Productividad por responsable (Top)")
        ax.set_xlabel("Informes")
        ax.tick_params(axis="y", labelsize=8)
        fig.tight_layout()
        return fig

    def _g_bar_largos_texto(self, df):
        # V6.9.16: el nombre real de la columna es "Descripcion Diagnostico"
        # (antes se buscaba un nombre largo con parentesis que ya no existe).
        # Busqueda flexible para tolerar variaciones del nombre.
        col = next((c for c in df.columns
                    if c.strip().lower().startswith("descripcion diagnostico")), None)
        if col is None: return None
        s = df[col].astype(str).str.len()
        bins = [0, 50, 150, 300, 600, 1200, np.inf]
        ser = pd.cut(s, bins=bins, labels=["<50", "50–150", "150–300", "300–600", "600–1200", "1200+"], include_lowest=True).value_counts().sort_index()
        if ser.sum() == 0: return None
        fig = Figure(figsize=(5.6, 3.2), dpi=100); ax = fig.add_subplot(111)
        ax.bar(ser.index.astype(str), ser.values, color="#2d3e5e")
        ax.set_title("Longitud del diagnóstico (bins)")
        ax.set_ylabel("Informes")
        fig.tight_layout()
        return fig

    # ---------- Comparador parametrizable ----------

    def _build_comparator(self, tab, df):
        # Controles
        ctrl = ttk.Frame(tab)
        ctrl.grid(row=0, column=0, columnspan=2, padx=10, pady=(10,0), sticky="ew")

        dims = [c for c in ["Servicio", "Patologo", "Malignidad", "Organo"] if c in df.columns]
        mets = [c for c in ["IHQ_KI-67"] if c in df.columns]  # se pueden añadir más numéricas

        self._compare_controls["dim"] = tk.StringVar(value=dims[0] if dims else "")
        self._compare_controls["agg"] = tk.StringVar(value="conteo")
        self._compare_controls["met"] = tk.StringVar(value=mets[0] if mets else "")

        row = ttk.Frame(ctrl, padding=10, relief="solid", borderwidth=1)
        row.pack(fill="x", padx=4, pady=4)
        ttk.Label(row, text="Dimensión:").pack(side="left", padx=6)
        ttk.Combobox(row, values=dims or [""], textvariable=self._compare_controls["dim"]).pack(side="left", padx=6)
        ttk.Label(row, text="Agregador:").pack(side="left", padx=6)
        ttk.Combobox(row, values=["conteo", "promedio"], textvariable=self._compare_controls["agg"]).pack(side="left", padx=6)
        ttk.Label(row, text="Métrica:").pack(side="left", padx=6)
        ttk.Combobox(row, values=mets or [""], textvariable=self._compare_controls["met"]).pack(side="left", padx=6)
        ttk.Button(row, text="Aplicar", command=lambda: self._chart_in(tab, 1, 0, lambda: self._g_compare(df), "Comparación de Datos", df)).pack(side="left", padx=10)

        # Gráfico inicial
        self._chart_in(tab, 1, 0, lambda: self._g_compare(df), "Comparación de Datos", df)

    def _g_compare(self, df):
        dim = self._compare_controls["dim"].get()
        agg = self._compare_controls["agg"].get()
        met = self._compare_controls["met"].get()
        if not dim or df.empty: return None

        fig = Figure(figsize=(11.6, 5.6), dpi=100); ax = fig.add_subplot(111)

        if agg == "conteo":
            ser = df[dim].astype(str).value_counts().head(20).iloc[::-1]
            ax.barh([self._trunc_label(x, 48) for x in ser.index], ser.values, color="#2d3e5e")
            ax.set_title(f"Conteo por {dim} (top 20)")
            ax.set_xlabel("Informes")
        else:
            if not met or met not in df.columns:
                return None
            s = pd.to_numeric(df[met], errors="coerce")
            grp = df.assign(_metric=s).groupby(dim)["_metric"].mean().dropna().sort_values().tail(20)
            ax.barh([self._trunc_label(x, 48) for x in grp.index], grp.values, color="#2d3e5e")
            ax.set_title(f"Promedio de {met} por {dim} (top 20)")
            ax.set_xlabel(met)

        ax.tick_params(axis="y", labelsize=8)
        fig.tight_layout()
        return fig

    # =========================
    # Funcionalidad
    # =========================
    def log_to_widget(self, message):
        """Log seguro que funciona con y sin log_textbox"""
        try:
            if hasattr(self, 'log_textbox') and self.log_textbox:
                self.log_textbox.configure(state="normal")
                self.log_textbox.insert("end", f"[{datetime.now().strftime('%H:%M:%S')}] {message}\n")
                self.log_textbox.see("end")
                self.log_textbox.configure(state="disabled")
            else:
                # Fallback a logging normal si no hay widget
                import logging
                logging.info(message)
        except Exception as e:
            # Si falla, usar logging como último recurso
            import logging
            logging.info(f"{message} (log_to_widget error: {e})")

    def select_files(self):
        self.pdf_files = filedialog.askopenfilenames(title="Seleccione archivos PDF", filetypes=[("PDF", "*.pdf")])
        if self.pdf_files:
            self.log_to_widget(f"Seleccionados {len(self.pdf_files)} archivos.")
            self.start_button.configure(state="normal")
            self.set_status(f"{len(self.pdf_files)} archivos listos.")
        else:
            self.log_to_widget("Selección cancelada.")
            self.start_button.configure(state="disabled")
            self.set_status("Selección cancelada.")

    def start_processing(self):
        if not self.pdf_files:
            messagebox.showwarning("Advertencia", "Por favor, seleccione archivos PDF primero.")
            return

        self.start_button.configure(state="disabled")
        self.select_files_button.configure(state="disabled")
        self.log_to_widget("=" * 50)
        self.log_to_widget("INICIANDO PROCESAMIENTO...")
        self.set_status("Procesando… esto puede tardar según el tamaño de los PDFs.")

        threading.Thread(target=self.processing_thread, daemon=True).start()

    def processing_thread(self):
        try:
            # FORZAR RECARGA DEL MÓDULO para asegurar que tenemos las últimas modificaciones
            import importlib
            import core.unified_extractor
            importlib.reload(core.unified_extractor)

            # === USAR PROCESAMIENTO CON AUDITORÍA IA INTEGRADA ===
            self.log_to_widget("\n" + "="*60)
            self.log_to_widget("🔄 CARGANDO SISTEMA DE AUDITORÍA IA...")
            self.log_to_widget("="*60)

            from core.process_with_audit import process_ihq_paths_with_audit, crear_callback_auditoria_para_ui

            self.log_to_widget("✅ Módulos de auditoría importados correctamente")

            output_dir = os.path.dirname(self.pdf_files[0])

            # Crear callback para auditoría IA
            self.log_to_widget("🔧 Configurando callback de auditoría IA...")
            callback_auditoria = crear_callback_auditoria_para_ui(self)
            self.log_to_widget("✅ Callback configurado")

            # Procesar con auditoría integrada
            self.log_to_widget("\n🚀 INICIANDO PROCESAMIENTO CON AUDITORÍA INTEGRADA...")
            self.log_to_widget("="*60)

            num_records = process_ihq_paths_with_audit(
                pdf_paths=self.pdf_files,
                output_dir=output_dir,
                ui_callback_auditoria=callback_auditoria,
                log_callback=self.log_to_widget
            )

            # Nota: El mensaje de éxito y el refresh se manejan en el callback de auditoría
            # Solo mostramos mensaje si no hay casos para auditar
            if num_records == 0:
                self.log_to_widget(f"⚠️ No se procesaron registros.")
                logging.warning("No se encontraron casos IHQ válidos en los PDFs seleccionados.")
            else:
                # El mensaje final se mostrará después de la auditoría
                self.log_to_widget(f"\n✅ Procesamiento completado: {num_records} registros")
                logging.info("Procesamiento completado - Auditando con IA...")

        except Exception as e:
            import traceback
            error_msg = f"ERROR: {e}\n{traceback.format_exc()}"
            self.log_to_widget(error_msg)
            logging.error(f"Error durante el procesamiento: {e}")
        finally:
            # V4.2.1 FIX: Usar after() para operaciones de UI desde thread
            def _restore_buttons():
                try:
                    self.start_button.configure(state="normal")
                    self.select_files_button.configure(state="normal")
                except Exception:
                    pass
            self.after(0, _restore_buttons)

    def refresh_data_and_table(self):
        """Actualizar datos y tabla desde la base de datos"""
        try:
            # V5.3.9.3: Usar logging en lugar de print (stdout puede estar cerrado)
            logging.info("🔄 Iniciando refresh de datos...")

            from core.database_manager import init_db, get_all_records_as_dataframe

            # V6.2.0: Comentado - init_db() ya se llama en ihq_processor antes del guardado
            # Llamarlo aquí (en refresh_data) causa que el UPDATE de relleno sobrescriba
            # valores recién insertados con N/A
            # init_db()

            # Cargar datos
            self.master_df = get_all_records_as_dataframe()
            # V6.9.49: invalidar caché de completitud. Los datos recién recargados
            # pueden haber cambiado (procesamiento/reprocesamiento); _apply_row_colors
            # la repoblará bajo demanda solo para los casos mostrados.
            self._completitud_por_caso = {}

            # Ordenar por número de caso automáticamente
            if self.master_df is not None and not self.master_df.empty and "Numero de caso" in self.master_df.columns:
                self.master_df = self.master_df.sort_values(
                    by="Numero de caso",
                    ascending=True,
                    na_position='last'
                ).reset_index(drop=True)

            # V6.9.47: registrar la huella de la BD recién cargada para que el
            # auto-refresh de 60 s pueda saltarse recargas cuando nada cambió.
            try:
                from core.database_manager import get_db_fingerprint
                self._last_db_fingerprint = get_db_fingerprint()
            except Exception:
                self._last_db_fingerprint = None

            if self.master_df is not None and not self.master_df.empty:
                logging.info(f"📊 Datos cargados: {len(self.master_df)} registros")

                # Actualizar tabla solo si existe
                if hasattr(self, 'tree') and self.tree is not None:
                    self._populate_treeview(self.master_df)
                    logging.info("🗂️ Tabla actualizada")

                # Actualizar KPIs si existe el método
                try:
                    self._render_kpis(self.master_df)
                    logging.info("📈 KPIs actualizados")
                except:
                    pass

                # Actualizar estado con información inteligente
                footer_info = self._crear_footer_inteligente()
                self.set_status(footer_info)
                logging.info("✅ Refresh completado exitosamente")

            else:
                logging.warning("⚠️ No hay datos en la base de datos")
                self.set_status("⚠️ No hay datos en la base de datos")

        except Exception as e:
            error_msg = f"No se pudieron cargar los datos: {e}"
            logging.error(f"❌ Error en refresh: {error_msg}")

            # Solo mostrar error si no es un problema de UI
            try:
                if hasattr(self, 'tree'):  # Solo mostrar error si la UI está inicializada
                    messagebox.showerror("Error de Base de Datos", error_msg)
                self.set_status("❌ Error al cargar datos")
            except:
                logging.error("❌ Error mostrando mensaje de error")

    def _abrir_visor_qt(self):
        """V6.9.50: abre el Visualizador en una ventana Qt (QTableView) de alto
        rendimiento, como PROCESO APARTE. Tkinter y Qt no comparten bucle de eventos,
        por eso es un subproceso. Hereda el entorno actual -> lee la MISMA BD (respeta
        ONCONOVA_DB_OVERRIDE: prod o DEV). Usa el intérprete de env_qt (con PySide6)."""
        import os
        import subprocess
        try:
            proj = os.path.dirname(os.path.abspath(__file__))
            script = os.path.join(proj, "visor_datos_qt.py")
            # Preferir pythonw.exe (sin ventana de consola) del entorno Qt
            py_w = os.path.join(proj, "env_qt", "Scripts", "pythonw.exe")
            py_c = os.path.join(proj, "env_qt", "Scripts", "python.exe")
            py = py_w if os.path.exists(py_w) else py_c
            if not os.path.exists(py) or not os.path.exists(script):
                messagebox.showwarning(
                    "Visor Qt no disponible",
                    "No se encontró el entorno Qt (env_qt) o visor_datos_qt.py.\n\n"
                    f"Intérprete: {py}\nScript: {script}"
                )
                return
            creationflags = getattr(subprocess, "CREATE_NO_WINDOW", 0) if os.name == "nt" else 0
            subprocess.Popen(
                [py, script], cwd=proj, env=os.environ.copy(), creationflags=creationflags
            )
            try:
                self.set_status("🪟 Abriendo Visualizador Qt en ventana aparte…")
            except Exception:
                pass
            logging.info("Visor Qt lanzado como proceso aparte")
        except Exception as e:
            logging.error(f"Error abriendo visor Qt: {e}", exc_info=True)
            try:
                messagebox.showerror("Error", f"No se pudo abrir el visor Qt:\n{e}")
            except Exception:
                pass

    def _ocultar_m_redundantes(self, df):
        """V6.9.48: quita del DISPLAY las filas M de coloración cuyo PACIENTE ya tiene una
        fila IHQ que lleva la coloración en su columna 'Diagnostico Coloracion 2' (sea 1 dx,
        o los N concatenados si tiene varias). NO se borran de la BD (son la fuente de
        verdad); solo se ocultan para no duplicar. Las coloraciones de pacientes SIN IHQ se
        muestran como sus propias filas."""
        if df is None or df.empty or "Numero de caso" not in df.columns:
            return df
        col = "Diagnostico Coloracion 2"
        base = getattr(self, "master_df", None)
        if base is None or base.empty:
            base = df
        if col not in base.columns or "N. de identificación" not in base.columns:
            return df
        if col not in df.columns or "N. de identificación" not in df.columns:
            return df

        def _ced(serie):
            return serie.astype(str).str.replace(r"\D", "", regex=True)

        # Cédulas que tienen una fila IHQ con la coloración ya reflejada en su columna.
        b_nc = base["Numero de caso"].astype(str)
        b_esM = b_nc.str.match(r"^[Mm]\d", na=False)
        b_dx = base[col].astype(str).str.strip()
        b_real = ~b_dx.str.lower().isin(["", "nan", "none", "n/a"])
        b_ihq = (~b_esM) & b_real
        ceds_cubiertas = set(_ced(base.loc[b_ihq, "N. de identificación"]))
        if not ceds_cubiertas:
            return df

        d_nc = df["Numero de caso"].astype(str)
        d_esM = d_nc.str.match(r"^[Mm]\d", na=False)
        d_ced = _ced(df["N. de identificación"])
        redundante = d_esM.values & d_ced.isin(ceds_cubiertas).values
        return df[~redundante]

    def _populate_treeview(self, df_to_display):
        """
        V5.3.8: Población de Sheet virtualizado (antes era Treeview)
        VENTAJA: Carga COMPLETA en <100ms para 1000+ filas × 88 columnas
        """
        # Verificar que el sheet exista
        if not hasattr(self, 'sheet') or self.sheet is None:
            return

        if df_to_display.empty:
            self.sheet.set_sheet_data([[]])  # Limpiar sheet
            self.sheet.headers([])
            return

        # V6.9.55: filtro por TIPO DE REGISTRO (Todos / IHQ / Coloración) del control
        # segmentado del Visualizador. Coloración = filas de clave 'M…' (^[Mm]\d);
        # IHQ = el resto. Se aplica AQUÍ porque TODOS los caminos de poblado (refresh,
        # buscador, cambio de filtro) pasan por _populate_treeview -> ambas tablas
        # (self.sheet y self.sheet_dashboard) quedan filtradas de forma consistente.
        # Por defecto ('todos') no hace nada -> comportamiento idéntico al anterior.
        _mostrar_todas_las_m = False
        _tipo_reg = "todos"
        try:
            _tv = getattr(self, "_tipo_registro_var", None)
            if _tv is not None:
                _tipo_reg = (_tv.get() or "todos").strip().lower()
        except Exception:
            _tipo_reg = "todos"
        if _tipo_reg in ("ihq", "coloracion") and "Numero de caso" in df_to_display.columns:
            _es_m = df_to_display["Numero de caso"].astype(str).str.match(r"^[Mm]\d", na=False)
            if _tipo_reg == "coloracion":
                df_to_display = df_to_display[_es_m]
                # En modo Coloración se muestran TODAS las coloraciones como filas
                # propias (no se ocultan las "redundantes" ya reflejadas en un IHQ).
                _mostrar_todas_las_m = True
            else:  # ihq
                df_to_display = df_to_display[~_es_m]
            if df_to_display.empty:
                self.sheet.set_sheet_data([[]])
                self.sheet.headers([])
                return

        # V6.9.46: ocultar filas M de coloración REDUNDANTES (su dx ya está reflejado en
        # la fila IHQ del paciente -> single-merged). Las multi (la fila IHQ marca
        # "varias (N)") y las coloraciones de pacientes SIN IHQ sí se muestran.
        # V6.9.55: en modo "Coloración" se OMITE este ocultamiento (se quieren ver
        # todas las coloraciones como filas propias).
        if not _mostrar_todas_las_m:
            try:
                df_to_display = self._ocultar_m_redundantes(df_to_display)
            except Exception as e:
                logging.warning(f"⚠️ No se pudo filtrar filas M redundantes: {e}")
            if df_to_display.empty:
                self.sheet.set_sheet_data([[]])
                self.sheet.headers([])
                return

        # V5.3.9.3: NO sobrescribir "Nombre Completo" si ya existe (creada en database_manager.py sin N/A)
        # Crear columna de Nombre Completo solo si NO existe
        if "Nombre Completo" not in df_to_display.columns and all(col in df_to_display.columns for col in ["Primer nombre", "Segundo nombre", "Primer apellido", "Segundo apellido"]):
            # Fallback: Si por alguna razón no existe, crearla (pero normalmente ya existe)
            from core.unified_extractor import build_clean_full_name

            def crear_nombre_limpio_ui(row):
                try:
                    return build_clean_full_name(
                        str(row.get("Primer nombre", "")),
                        str(row.get("Segundo nombre", "")),
                        str(row.get("Primer apellido", "")),
                        str(row.get("Segundo apellido", ""))
                    )
                except:
                    return "N/A"

            df_to_display["Nombre Completo"] = df_to_display.apply(crear_nombre_limpio_ui, axis=1)

        # V5.3.7: Columnas reorganizadas para mejor rendimiento y visualización
        # Eliminadas: EPS (innecesaria), columnas duplicadas
        # Reordenadas: IHQ_ORGANO e IHQ_ESTUDIOS_SOLICITADOS antes de biomarcadores
        # V6.0.12: ELIMINADAS columnas sensibles (N. de identificación, Nombre Completo) por privacidad
        cols_to_show = [
            "Numero de caso",
            # V6.9.44: re-añadidas a pedido (identificar/buscar paciente por cédula o
            # nombre). OJO: datos sensibles (Habeas Data) -> visibles en tabla/capturas.
            "N. de identificación",
            "Nombre Completo",
            # "EPS",  # ELIMINADA - No relevante para investigación
            "Procedimiento",
            "Organo",
            "Malignidad",
            "Diagnostico Coloracion",  # v6.1.0: Diagnóstico del Estudio M (Coloración)
            "Diagnostico Coloracion 2",  # V6.9.45: Dx del PDF de Coloración (estudio M autónomo)
            "Diagnostico Principal",
            "Factor pronostico",
            # "Descripcion macroscopica",  # V5.3.8: ELIMINADA - Texto muy largo, poco útil en tabla
            # "Descripcion microscopica",  # V5.3.8: ELIMINADA - Texto muy largo, poco útil en tabla
            # V5.3.7: IHQ_ORGANO e IHQ_ESTUDIOS_SOLICITADOS ANTES de biomarcadores
            "IHQ_ORGANO",
            "IHQ_ESTUDIOS_SOLICITADOS",
            # Biomarcadores principales
            "IHQ_HER2",
            "IHQ_KI-67",
            "IHQ_RECEPTOR_ESTROGENOS",
            "IHQ_RECEPTOR_PROGESTERONA",
            "IHQ_PDL-1",
            "IHQ_P16_ESTADO",
            "IHQ_P16_PORCENTAJE",
            "IHQ_P40_ESTADO",
            "IHQ_E_CADHERINA",  # v6.0.3 - E-Cadherina
            # Biomarcadores adicionales v4.0/v4.1
            "IHQ_CK7",
            "IHQ_DESMINA",
            "IHQ_LCA",
            "IHQ_CD11",
            "IHQ_MIOGENINA",
            "IHQ_MAMAGLOBINA",
            "IHQ_TIROGLOBULINA",
            "IHQ_CK34BETAE12",
            "IHQ_CK34BETA12",
            "IHQ_OCT4",
            "IHQ_PODOPLANINA",
            "IHQ_IDH",
            "IHQ_GPC3",
            "IHQ_AFP",
            "IHQ_IGD",
            "IHQ_BETACATENINA",
            "IHQ_ACTINA_MUSCULO_ESPECIFICA",
            "IHQ_MIELOPEROXIDASA",
            "IHQ_CD7",
            "IHQ_EBER",
            # V6.4.25: IHQ_CALRRETININA (typo v4.0) eliminado - duplicado de IHQ_CALRETININA (línea 4378)
            "IHQ_SYNAPTOFISINA",
            "IHQ_CKAE1E3",
            "IHQ_SINAPTOFISINA",
            "IHQ_CROMOGRANINA",
            "IHQ_CK56",
            "IHQ_CAM5",
            "IHQ_GLICOFORINA",
            "IHQ_TDT",
            "IHQ_ATRX",
            "IHQ_IDH1",
            "IHQ_CMYC",
            "IHQ_IGG4",
            "IHQ_IGG",
            "IHQ_MAMOGLOBINA",
            "IHQ_HEPATOCITO",  # V6.0.16: Auto-agregado
            "IHQ_CK19",
            "IHQ_CK20",
            "IHQ_CDX2",
            "IHQ_EMA",
            "IHQ_GATA3",
            "IHQ_SOX10",
            "IHQ_P53",
            "IHQ_TTF1",
            "IHQ_S100",
            "IHQ_VIMENTINA",
            "IHQ_CHROMOGRANINA",
            "IHQ_SYNAPTOPHYSIN",
            "IHQ_MELAN_A",
            # Marcadores CD
            "IHQ_CD2",
            "IHQ_CD3",
            "IHQ_CD5",
            "IHQ_CD10",
            "IHQ_CD20",
            "IHQ_CD30",
            "IHQ_CD34",
            "IHQ_CD38",
            "IHQ_CD45",
            "IHQ_CD56",
            "IHQ_CD61",
            "IHQ_CD68",
            "IHQ_CD117",
            "IHQ_CD138",
            "IHQ_KAPPA",
            "IHQ_LAMBDA",
            # V6.0.13: Biomarcadores para linfomas y mielomas
            "IHQ_BCL2",
            "IHQ_BCL6",
            "IHQ_MUM1", "IHQ_MUC1", "IHQ_MUC2"  # V6.4.43: MUC1 agregado,
            "IHQ_CD15",
            "IHQ_CD79A",
            "IHQ_ALK",
            # NUEVOS BIOMARCADORES v5.0 - CRÍTICOS PARA CASOS COMPLEJOS
            "IHQ_CKAE1AE3",
            "IHQ_NAPSIN",
            "IHQ_CDK4",
            "IHQ_MDM2",
            "IHQ_PAX5",
            "IHQ_ACTIN",
            # BIOMARCADORES ADICIONALES v5.1 - COMPLETAR CON TODAS LAS COLUMNAS DE BD
            "IHQ_PAX8",
            "IHQ_GFAP",
            # V6.5.83: IHQ_CAM52 eliminado (obsoleto, migrado a IHQ_CAM5)
            "IHQ_DOG1",
            "IHQ_H_CALDESMON",  # V6.1.2: Biomarcador IHQ250997 (tumor maligno indiferenciado)
            "IHQ_AML",  # V6.1.2: Biomarcador IHQ250997 (tumor maligno indiferenciado)
            "IHQ_HHV8",
            "IHQ_NEUN",
            "IHQ_P63",
            # V6.1.3: Biomarcador celulas mioepiteliales (IHQ250999) - CK5_6 ya existe en V5.3
            "IHQ_CALPONINA",
            "IHQ_BER_EP4",  # V6.0.12.1: Agregado BER-EP4 (Ep-CAM) - FIX IHQ250991
            "IHQ_WT1",
            # MARCADORES MMR (Mismatch Repair) - CRÍTICOS PARA CÁNCER COLORRECTAL
            "IHQ_MLH1",
            "IHQ_MSH2",
            "IHQ_MSH6",
            "IHQ_PMS2",
            # V5.3 - NUEVOS BIOMARCADORES (28 adicionales detectados en producción)
            "IHQ_CD23",
            "IHQ_CD4",
            "IHQ_CD8",
            "IHQ_CD99",
            "IHQ_CD1A",
            "IHQ_C4D",
            "IHQ_LMP1",
            "IHQ_CITOMEGALOVIRUS",
            "IHQ_SV40",
            "IHQ_CEA",
            "IHQ_CA19_9",
            "IHQ_CALRETININA",
            "IHQ_CK34BE12",
            "IHQ_CK5_6",
            "IHQ_HEPAR",
            "IHQ_GLIPICAN",
            "IHQ_ARGINASA",
            "IHQ_HMB45",
            "IHQ_PSA",
            "IHQ_INHIBINA",  # V6.4.60: Biomarcador hormonal
            "IHQ_RACEMASA",
            "IHQ_34BETA",
            "IHQ_B2",
            # V6.0.16 - Biomarcadores para linfomas (IHQ250988)
            "IHQ_SALL4",
            "IHQ_ALK1",
            # V5.3.7: Columnas de sistema al final
            "Estado Auditoria IA",  # V3.2.4
            "Fecha Ingreso Base de Datos",
        ]

        # V6.9.50: usar la fuente ÚNICA de columnas compartida con el visor Qt
        # (core/columnas_visor.py) para garantizar paridad EXACTA entre la tabla
        # Tkinter y el visor Qt. Esto además corrige un typo histórico: faltaba una
        # coma tras "IHQ_MUC2" -> se concatenaba con "IHQ_CD15" ("IHQ_MUC2IHQ_CD15")
        # y NINGUNA de las dos columnas se mostraba. La lista literal de arriba queda
        # solo como fallback por si fallara el import.
        try:
            from core.columnas_visor import COLS_TO_SHOW as _COLS_VISOR
            cols_to_show = list(_COLS_VISOR)
        except Exception as _e_cols:
            logging.warning(f"No se pudo cargar core.columnas_visor; uso lista local: {_e_cols}")

        # Filtrar solo las columnas que existen en el DataFrame
        # V6.9.56: además, OCULTAR las columnas que NO APLICAN (todas en N/A) — con
        # ~130 biomarcadores la tabla se llenaba de "N/A" inútiles. Si algún paciente
        # mostrado sí tiene el biomarcador, la columna reaparece automáticamente.
        try:
            from core.columnas_visor import columnas_visibles as _cols_vis
            available_cols = _cols_vis(df_to_display, cols_to_show)
        except Exception as _e_cv:
            logging.warning(f"columnas_visibles no aplicado: {_e_cv}")
            available_cols = [c for c in cols_to_show if c in df_to_display.columns]
        df_display = df_to_display[available_cols].copy()

        # Guardar DataFrame actual para ordenamiento
        self.current_displayed_df = df_display.copy()

        # V5.3.8: CONFIGURACIÓN DE SHEET - Una sola carga, virtualización automática
        # =========================================================================

        # Preparar encabezados (simplificados para mejor visualización)
        # V6.9.50: usar simplificar_header (fuente única) para que el renombrado de
        # display aplique aquí también: "Diagnostico Coloracion" (extraído del IHQ) ->
        # "Diagnostico Coloracion IHQ"; "Diagnostico Coloracion 2" (extraído de los PDFs
        # de Coloración) -> "Diagnostico Coloracion". NO cambia el nombre real de la BD.
        try:
            from core.columnas_visor import simplificar_header as _simpl_header
            headers = [_simpl_header(col) for col in df_display.columns]
        except Exception:
            headers = [col.split("(")[0].strip() for col in df_display.columns]

        # Convertir DataFrame a lista de listas (formato Sheet)
        # V6.9.57: las celdas SIN DATO se muestran VACÍAS (no "N/A"). Ocultar la
        # columna solo sirve si NINGÚN paciente de la vista tiene el biomarcador; en
        # la vista completa la columna se queda y el resto de celdas quedaba llena de
        # "N/A". Solo display: la BD, la búsqueda, el orden y la exportación no cambian.
        try:
            from core.columnas_visor import filas_para_display as _filas_disp
            sheet_data = _filas_disp(df_display)
        except Exception as _e_fd:
            logging.warning(f"filas_para_display no aplicado: {_e_fd}")
            sheet_data = df_display.fillna("").astype(str).values.tolist()

        # PASO 1: Cargar TODOS los datos de una sola vez (ultra rápido)
        self.sheet.set_sheet_data(data=sheet_data, reset_col_positions=True, reset_row_positions=True, redraw=False)
        self.sheet.headers(newheaders=headers, index=None, reset_col_positions=False, show_headers_if_not_sheet=True, redraw=False)

        # PASO 1.5: Ordenamiento por encabezado YA está configurado en __init__ (línea 2357-2391)
        # NO bindear aquí para evitar acumulación de handlers duplicados

        # PASO 2: Configurar anchos de columnas
        column_widths = {}
        for idx, col in enumerate(df_display.columns):
            if "Numero de caso" in col:
                width = 120
            # V6.0.12: Columnas eliminadas (privacidad)
            # elif "Nombre Completo" in col:
            #     width = 250
            # elif "N. de identificación" in col:
            #     width = 120
            elif "Fecha" in col:
                width = 120
            elif "Procedimiento" in col:
                width = 200
            elif "Organo" in col:
                width = 200
            elif "Malignidad" in col:
                width = 100
            elif "Diagnostico Coloracion" in col:
                # V6.9.50: 360 (antes 300). El encabezado renombrado "Diagnostico
                # Coloracion IHQ" (negrita) mide ~338px y a 300 se cortaba el "IHQ".
                width = 360
            elif "Diagnostico Principal" in col:
                width = 300
            elif "Factor pronostico" in col:
                width = 200
            elif "Descripcion" in col:
                width = 350
            elif col.startswith("IHQ_"):
                width = 150
            elif "Estado Auditoria IA" in col:
                width = 150
            elif "Fecha Ingreso" in col:
                width = 180
            else:
                width = 150  # Default

            self.sheet.column_width(column=idx, width=width, only_set_if_too_small=False, redraw=False)

        # PASO 3: DIBUJAR DATOS INMEDIATAMENTE (no bloquear UI)
        # ====================================================
        # V4.2.1 FIX: Mostrar datos ANTES de calcular colores para evitar tabla en blanco
        self.sheet.refresh()

        # PASO 4: Aplicar colores de fila de forma diferida (no bloquea UI)
        # ================================================================
        self.after(100, lambda: self._apply_row_colors(df_display, sheet_data))

        # PASO 7: Actualizar también sheet_dashboard si existe
        if hasattr(self, 'sheet_dashboard') and self.sheet_dashboard is not None:
            try:
                self.sheet_dashboard.set_sheet_data(data=sheet_data, reset_col_positions=True, reset_row_positions=True, redraw=False)
                self.sheet_dashboard.headers(newheaders=headers, index=None, reset_col_positions=False, show_headers_if_not_sheet=True, redraw=False)

                # Aplicar anchos de columna
                for col_idx, width in column_widths.items():
                    self.sheet_dashboard.column_width(column=col_idx, width=width, only_set_if_too_small=False, redraw=False)

                self.sheet_dashboard.refresh()
                logging.debug("✅ sheet_dashboard actualizado")
            except Exception as e:
                logging.error(f"Error actualizando sheet_dashboard: {e}")

        # Actualizar KPIs en base a lo mostrado
        try:
            self._render_kpis(df_display)
        except Exception:
            pass

    def _apply_row_colors(self, df_display, sheet_data):
        """V4.2.1: Aplicar colores de fila de forma diferida para no bloquear UI.
        Se ejecuta via after() DESPUÉS de que los datos ya se muestran en la tabla."""
        try:
            estado_col_idx = None
            peticion_col_idx = None
            if "Estado Auditoria IA" in df_display.columns:
                estado_col_idx = list(df_display.columns).index("Estado Auditoria IA")
            if "Numero de caso" in df_display.columns:
                peticion_col_idx = list(df_display.columns).index("Numero de caso")

            # V6.9.28 PERF: completitud en UNA sola consulta (en memoria, sin SELECT
            # por caso). V6.9.49 PERF: además se CACHEA a nivel de instancia. La
            # completitud de un caso NO cambia al filtrar/ordenar (mismos datos), así
            # que se calcula UNA vez por caso y se reutiliza. refresh_data_and_table
            # limpia self._completitud_por_caso al recargar la BD (datos nuevos).
            # Antes se recalculaban ~8.000 casos en CADA repoblado (cada tecla del
            # buscador, cada orden, cada auto-refresh) -> repoblado lento.
            completitud_cache = getattr(self, "_completitud_por_caso", None)
            if completitud_cache is None:
                completitud_cache = {}
                self._completitud_por_caso = completitud_cache
            if peticion_col_idx is not None:
                try:
                    from core.validation_checker import verificar_completitud_registro
                    numeros_peticion = set(df_display["Numero de caso"].dropna().unique())
                    # V6.9.44 FIX: la completitud se calcula sobre los DATOS EN VIVO
                    # (self.master_df -> BD configurada/MySQL), ya en memoria y con
                    # TODAS las columnas requeridas (sin consultas extra a la BD).
                    # V6.9.47 PERF: registros construidos de forma VECTORIZADA
                    # (drop_duplicates + to_dict). SOLO se evalúan filas IHQ: las filas
                    # M de coloración no son informes IHQ y no deben pintarse incompletas.
                    registros_por_caso = {}
                    numeros_ihq = {n for n in numeros_peticion
                                   if not re.match(r'^[Mm]\d', str(n))}
                    # V6.9.49: calcular SOLO los casos que aún no están cacheados.
                    numeros_faltantes = {n for n in numeros_ihq
                                         if n not in completitud_cache}
                    _mdf = getattr(self, "master_df", None)
                    if (_mdf is not None and not _mdf.empty
                            and "Numero de caso" in _mdf.columns and numeros_faltantes):
                        _sub = (_mdf[_mdf["Numero de caso"].isin(numeros_faltantes)]
                                .drop_duplicates("Numero de caso")
                                .fillna(""))
                        # to_dict('records') CONSERVA todas las columnas (incl. 'Numero
                        # de caso'); set_index la quitaría y el verificador la contaría
                        # como faltante -> todos los IHQ saldrían "incompletos" (rojos).
                        registros_por_caso = {r["Numero de caso"]: r
                                              for r in _sub.to_dict("records")}
                    for numero in numeros_faltantes:
                        reg = registros_por_caso.get(numero)
                        if reg is None:
                            # Sin datos en memoria para verificar -> NO marcar rojo
                            # (evita el falso positivo masivo del SQLite desincronizado).
                            completitud_cache[numero] = True
                            continue
                        try:
                            analisis = verificar_completitud_registro(numero, registro=reg)
                            completitud_cache[numero] = analisis.get('completo', True)
                        except Exception:
                            completitud_cache[numero] = True
                except Exception as e:
                    logging.warning(f"Error calculando completitud: {e}")

            # Clasificar filas
            rows_auditoria_parcial = []
            rows_auditoria_completa = []
            rows_incompletos = []

            # V6.9.47 PERF: clasificar sobre arrays (.values), no con iterrows
            # (~700 ms -> ~15 ms con 8.369 filas). El índice posicional i coincide con
            # la fila del Sheet (set_sheet_data usó el mismo orden del DataFrame).
            estado_vals = (df_display.iloc[:, estado_col_idx].values
                           if estado_col_idx is not None else None)
            num_vals = (df_display.iloc[:, peticion_col_idx].values
                        if peticion_col_idx is not None else None)
            for i in range(len(df_display)):
                if estado_vals is not None:
                    estado = estado_vals[i]
                    if estado == "PARCIAL":
                        rows_auditoria_parcial.append(i)
                        continue
                    elif estado == "COMPLETA":
                        rows_auditoria_completa.append(i)
                        continue
                if num_vals is not None and completitud_cache.get(num_vals[i]) is False:
                    rows_incompletos.append(i)

            # Aplicar highlighting
            if rows_auditoria_parcial:
                self.sheet.highlight_rows(rows=rows_auditoria_parcial, bg="#FFF3CD", fg="#856404", redraw=False)
            if rows_auditoria_completa:
                self.sheet.highlight_rows(rows=rows_auditoria_completa, bg="#D4EDDA", fg="#155724", redraw=False)
            if rows_incompletos:
                self.sheet.highlight_rows(rows=rows_incompletos, bg="#FFE5E5", fg="#721C24", redraw=False)

            self.sheet.refresh()

            # Aplicar también al dashboard sheet
            if hasattr(self, 'sheet_dashboard') and self.sheet_dashboard is not None:
                if rows_incompletos:
                    self.sheet_dashboard.highlight_rows(rows=rows_incompletos, bg="#FFE5E5", fg="#721C24", redraw=False)
                    self.sheet_dashboard.refresh()

        except Exception as e:
            logging.warning(f"⚠️ Error aplicando colores de fila: {e}")

    def _sort_treeview(self, col, reverse):
        """
        V5.3.8: Ordenamiento optimizado para Sheet
        Ordena el DataFrame y recarga el Sheet (mega rápido con virtualización)
        """
        if not hasattr(self, 'current_displayed_df') or self.current_displayed_df is None:
            return

        from datetime import datetime as _dt

        def _sort_key(val):
            """Función de ordenamiento inteligente"""
            s = str(val).strip()
            # Intentar número
            try:
                return (0, float(s.replace(",", ".")))
            except:
                pass
            # Intentar fecha
            for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
                try:
                    return (1, _dt.strptime(s, fmt))
                except:
                    pass
            # Texto
            return (2, s.lower())

        # Ordenar DataFrame
        try:
            df_sorted = self.current_displayed_df.sort_values(
                by=col,
                ascending=not reverse,
                key=lambda x: x.map(_sort_key),
                na_position='last'
            )
            # Recargar Sheet con datos ordenados
            self._populate_treeview(df_sorted)
        except Exception as e:
            logging.warning(f"Error ordenando por {col}: {e}")

    def _filter_tabla_debounced(self, *args):
        """V6.9.49 PERF: debounce del buscador. trace_add("write") dispara en CADA
        tecla; sin debounce, escribir "garcia" reconstruía la tabla entera
        (~8.000x163 + coloreo) 6 veces seguidas. Ahora se espera ~300 ms tras la
        última tecla y se filtra UNA sola vez. Aplica a ambos buscadores
        (Visualizador y Dashboard), que comparten esta lógica."""
        try:
            job = getattr(self, '_filter_job', None)
            if job:
                self.after_cancel(job)
        except Exception:
            pass
        self._filter_job = self.after(300, self.filter_tabla)

    def filter_tabla(self, *args):
        # Guardia: no filtrar si aún no hay datos cargados
        if not hasattr(self, 'master_df') or self.master_df is None or self.master_df.empty:
            return
        # Leer de ambas search vars (dashboard y visualizador)
        query = ""
        placeholder = getattr(self, '_search_placeholder', '').lower()
        if hasattr(self, 'search_var_dashboard'):
            q = self.search_var_dashboard.get().strip().lower()
            if q and q != placeholder:
                query = q
        if not query and hasattr(self, 'search_var'):
            q = self.search_var.get().strip().lower()
            if q and q != placeholder:
                query = q
        if not query:
            self._populate_treeview(self.master_df)
            return

        df = self.master_df.copy()
        if df.empty:
            self._populate_treeview(df)
            return
            
        # V6.9.44: búsqueda ampliada -> también por CÉDULA/identificación y por nombre
        # y apellido completos (antes solo N° caso + primer nombre/apellido).
        search_cols = ["Numero de caso", "N. de identificación", "Nombre Completo",
                       "Primer nombre", "Segundo nombre",
                       "Primer apellido", "Segundo apellido"]
        # V6.9.44: búsqueda por TOKENS sobre un "haystack" combinado (todas las cols
        # de búsqueda unidas en un solo texto por fila). Permite buscar el NOMBRE
        # COMPLETO o varias palabras en cualquier orden ("diego garcia"), insensible
        # a mayúsculas y acentos. Antes hacía contains columna-por-columna, por eso un
        # nombre completo nunca coincidía con una sola columna y no devolvía nada.
        import unicodedata as _ud
        def _strip(_s):
            _s = str(_s).lower()
            return ''.join(_c for _c in _ud.normalize('NFKD', _s) if not _ud.combining(_c))

        _present = [c for c in search_cols if c in df.columns]
        _haystack = df[_present].fillna('').astype(str).agg(' '.join, axis=1).map(_strip)

        _tokens = [_strip(t) for t in query.split() if t.strip()]
        mask = pd.Series([True] * len(df), index=df.index)
        for _tok in _tokens:
            mask &= _haystack.str.contains(_tok, na=False, regex=False)

        self._populate_treeview(df[mask])

    def _crear_filtro_tipo_registro(self, parent):
        """V6.9.55: control segmentado para filtrar la tabla del Visualizador por
        TIPO DE REGISTRO: Todos / IHQ / Coloración.

        Las filas de coloración son las de clave 'M…' (^[Mm]\\d) en 'Numero de caso';
        las IHQ son el resto (mismo criterio que _get_filtered_df y _ocultar_m_redundantes).
        El estado (self._tipo_registro_var) se COMPARTE entre el Visualizador y el
        Dashboard, así ambas tablas y buscadores quedan sincronizados. Al cambiar,
        re-dispara filter_tabla (respeta la búsqueda activa); el filtrado real se
        aplica en _populate_treeview, por donde pasan TODOS los caminos de poblado."""
        if not hasattr(self, '_tipo_registro_var'):
            self._tipo_registro_var = tk.StringVar(value="todos")
        cont = ttk.Frame(parent)
        ttk.Label(cont, text="Ver:", font=("Segoe UI", 10)).pack(side=LEFT, padx=(0, 6))
        for _val, _txt in (("todos", "Todos"), ("ihq", "IHQ"), ("coloracion", "Coloración")):
            ttk.Radiobutton(
                cont,
                text=_txt,
                value=_val,
                variable=self._tipo_registro_var,
                command=self.filter_tabla,
                bootstyle="primary-outline-toolbutton",
            ).pack(side=LEFT)
        return cont

    def _on_search_focus_in(self, entry_widget, string_var):
        """Limpiar placeholder al enfocar el campo de búsqueda"""
        if string_var.get() == self._search_placeholder:
            string_var.set("")

    def _on_search_focus_out(self, entry_widget, string_var):
        """Restaurar placeholder si el campo queda vacío"""
        if not string_var.get().strip():
            string_var.set(self._search_placeholder)

    def mostrar_detalle_registro(self, event):
        # v6.0.15: Extraer TODAS las filas seleccionadas del evento y del Sheet
        # V6.9.49 PERF: logs degradados a DEBUG. Esta función corre en CADA selección
        # (clic y teclado); con nivel INFO escribía decenas de líneas a consola por
        # interacción (I/O sincrónica) y sumaba a la lentitud percibida del clic.
        logging.debug("mostrar_detalle_registro: EVENTO DISPARADO")

        # Extraer filas de selection_boxes del evento (captura multi-selección)
        event_rows = set()
        if isinstance(event, dict):
            # Extraer de selection_boxes (contiene TODOS los rangos seleccionados)
            sel_boxes = event.get('selection_boxes', {})
            for box in sel_boxes:
                # box es Box_nt(from_r, from_c, upto_r, upto_c)
                if hasattr(box, 'from_r') and hasattr(box, 'upto_r'):
                    for r in range(box.from_r, box.upto_r):
                        event_rows.add(r)
                elif isinstance(box, (tuple, list)) and len(box) >= 4:
                    for r in range(box[0], box[2]):
                        event_rows.add(r)

            # Fallback: extraer de 'selected'
            if not event_rows:
                selected_info = event.get('selected', None)
                if selected_info and hasattr(selected_info, 'row') and selected_info.row is not None:
                    event_rows.add(selected_info.row)

        if event_rows:
            logging.info(f"mostrar_detalle_registro: {len(event_rows)} fila(s) desde evento: {sorted(event_rows)}")

        # Obtener selección del Sheet (puede tener más filas si toggle_select)
        try:
            selection = self.tree.selection()
            logging.info(f"self.tree.selection() retornó: {selection}")

            # Combinar: si el evento tiene filas Y selection también, usar la unión
            if event_rows and selection:
                combined = set(selection) | event_rows
                selection = sorted(combined)
            elif event_rows and not selection:
                selection = sorted(event_rows)

            # Guardar TODAS las filas seleccionadas (para multi-selección en auditoría)
            if selection and len(selection) > 0:
                self.ultima_seleccion = selection[0]
                self._ultimas_filas_seleccionadas = list(selection)
                logging.info(f"💾 Guardada selección: {len(selection)} fila(s): {selection}")
        except Exception as e:
            logging.error(f"ERROR al llamar self.tree.selection(): {e}", exc_info=True)

        # v6.0.12: Actualizar estado de TODOS los botones cuando hay selección
        self._update_export_button_state()
        self._update_audit_buttons_state()

        # NUEVO: Si el panel de detalles está abierto, actualizarlo con el nuevo registro
        try:
            if (hasattr(self.export_system, 'details_panel') and
                self.export_system.details_panel.winfo_exists() and
                selection and len(selection) > 0):
                logging.info("Panel de detalles abierto - actualizando contenido...")
                self.export_system._update_details_panel_content(selection[0])
        except Exception as e:
            logging.error(f"Error al actualizar panel de detalles: {e}")

        # El panel de detalles ahora es flotante y se maneja en el export_system
        # Aquí solo manejamos la selección

    # ================================================================
    #  V6.9.51 — Detalle de fila (doble clic): ver TODO el texto completo
    # ================================================================
    def _abrir_detalle_fila(self, event=None, sheet=None):
        """Doble clic en una fila -> ventana con TODOS los campos y su TEXTO COMPLETO.
        Las celdas de la tabla truncan los campos largos (p. ej. 'Descripcion
        macroscopica'); esta ventana los muestra completos y permite copiarlos.
        'sheet' permite usarlo tanto en self.sheet como en self.sheet_dashboard.

        Nota: tksheet invoca este callback vía try_binding(), que TRAGA excepciones;
        por eso envolvemos todo y mostramos el error al usuario en vez de fallar mudo."""
        sheet = sheet or self.sheet
        try:
            logging.info("_abrir_detalle_fila: doble clic recibido")
            # Anti-rebote: evita abrir dos ventanas si el evento se dispara duplicado.
            if getattr(self, "_detalle_abriendo", False):
                return
            self._detalle_abriendo = True
            self.after(450, lambda: setattr(self, "_detalle_abriendo", False))
            # No abrir si el doble clic fue sobre el encabezado/índice (ordenar/redimensionar).
            try:
                if sheet.identify_region(event) in ("header", "index", "top left"):
                    return
            except Exception:
                pass

            row_idx = self._fila_desde_evento(event, sheet)
            if row_idx is None:
                logging.info("_abrir_detalle_fila: no se pudo determinar la fila")
                return

            headers = list(sheet.headers())
            valores = list(sheet.get_row_data(row_idx))
            registro = list(zip(headers, valores))
            # V6.9.58: abrir la FICHA DEL PACIENTE — agrupa TODOS sus estudios (IHQ +
            # Coloraciones) en una sola ventana. Si no se puede resolver el paciente
            # (sin cédula / sin master_df), cae al detalle del registro de siempre.
            num = next((str(v) for h, v in registro
                        if h == "Numero de caso" and str(v).strip()), "")
            if not self._mostrar_ficha_paciente(num):
                self._mostrar_ventana_detalle(registro)
        except Exception as e:
            logging.exception("_abrir_detalle_fila falló")
            try:
                messagebox.showerror("Detalle del registro",
                                     f"No se pudo abrir el detalle:\n{e}")
            except Exception:
                pass

    def _fila_desde_evento(self, event, sheet=None):
        """Índice de fila del doble clic, robusto entre versiones de tksheet."""
        sheet = sheet or self.sheet
        # 1) identify_row desde el evento (lo más preciso).
        try:
            r = sheet.identify_row(event, allow_end=False)
            if r is not None:
                return int(r)
        except Exception:
            pass
        # 2) Celda actualmente seleccionada (double_b1 la selecciona antes del hook).
        try:
            cur = sheet.get_currently_selected()
            if cur:
                if getattr(cur, "row", None) is not None:
                    return int(cur.row)
                if isinstance(cur, (tuple, list)) and cur and cur[0] is not None:
                    return int(cur[0])
        except Exception:
            pass
        # 3) Filas seleccionadas.
        try:
            rows = sheet.get_selected_rows()
            if rows:
                return sorted(rows)[0]
        except Exception:
            pass
        # 4) Celdas seleccionadas.
        try:
            cells = sheet.get_selected_cells()
            if cells:
                return sorted(cells)[0][0]
        except Exception:
            pass
        return None

    def _mostrar_ventana_detalle(self, registro):
        """Ventana modal con cada campo (etiqueta + valor completo). Oculta los vacíos,
        resalta los campos narrativos largos y permite seleccionar/copiar el texto."""
        campos_largos = {
            "Descripcion macroscopica", "Descripcion microscopica",
            "Diagnostico Coloracion", "Diagnostico Coloracion IHQ",
            "Diagnostico Principal", "Factor pronostico", "Datos Clinicos",
        }
        num = next((str(v) for h, v in registro
                    if h == "Numero de caso" and str(v).strip()), "")

        win = tk.Toplevel(self)
        win.title(f"Detalle del registro — {num}" if num else "Detalle del registro")
        win.geometry("840x660")
        win.transient(self)

        cont = ttk.Frame(win, padding=10)
        cont.pack(fill="both", expand=True)

        # Texto con scroll nativo (maneja rueda del mouse y campos muy largos).
        txt = tk.Text(cont, wrap="word", font=("Segoe UI", 10),
                      padx=12, pady=10, relief="flat", background="#ffffff",
                      foreground="#1b1b1b", cursor="arrow")
        scroll = ttk.Scrollbar(cont, orient="vertical", command=txt.yview)
        txt.configure(yscrollcommand=scroll.set)
        txt.pack(side="left", fill="both", expand=True)
        scroll.pack(side="right", fill="y")

        txt.tag_configure("campo", font=("Segoe UI", 9, "bold"),
                          foreground="#1B5E20", spacing1=10, spacing3=2)
        txt.tag_configure("valor", font=("Segoe UI", 10), foreground="#1b1b1b",
                          lmargin1=6, lmargin2=6, spacing3=4)
        txt.tag_configure("valor_largo", font=("Segoe UI", 10), foreground="#1b1b1b",
                          lmargin1=8, lmargin2=8, spacing1=2, spacing3=6,
                          background="#F1F8E9")

        lineas_copia = []
        mostrados = 0
        for h, val in registro:
            val = str(val or "").strip()
            if not val:
                continue  # ocultar vacíos -> tarjeta compacta (no ~140 columnas IHQ vacías)
            mostrados += 1
            lineas_copia.append(f"{h}: {val}")
            txt.insert("end", f"{h}\n", "campo")
            txt.insert("end", f"{val}\n", "valor_largo" if h in campos_largos else "valor")
        if mostrados == 0:
            txt.insert("end", "(Sin datos en este registro)")

        # Solo lectura pero SELECCIONABLE/COPIABLE (bloquea edición, permite Ctrl+C/Ctrl+A).
        def _solo_lectura(e):
            if (e.state & 0x4) and e.keysym.lower() in ("c", "a"):
                return
            if e.keysym in ("Left", "Right", "Up", "Down", "Home", "End",
                            "Prior", "Next"):
                return
            return "break"
        txt.bind("<Key>", _solo_lectura)

        # Barra inferior: copiar todo + cerrar.
        barra = ttk.Frame(win, padding=(10, 6))
        barra.pack(fill="x")

        def _copiar_todo():
            try:
                self.clipboard_clear()
                self.clipboard_append("\n".join(lineas_copia))
            except Exception:
                pass

        ttk.Button(barra, text="📋 Copiar todo", command=_copiar_todo,
                   bootstyle="secondary").pack(side="left")
        ttk.Button(barra, text="Cerrar", command=win.destroy,
                   bootstyle="primary").pack(side="right")

        win.bind("<Escape>", lambda e: win.destroy())
        # Centrar sobre la ventana principal y FORZAR al frente (si no, puede abrirse
        # detrás de la app maximizada y parecer que "no pasa nada").
        try:
            win.update_idletasks()
            x = self.winfo_x() + (self.winfo_width() - win.winfo_width()) // 2
            y = self.winfo_y() + (self.winfo_height() - win.winfo_height()) // 2
            win.geometry(f"+{max(x, 0)}+{max(y, 0)}")
        except Exception:
            pass
        try:
            win.lift()
            win.focus_force()
            win.attributes("-topmost", True)
            win.after(300, lambda: win.attributes("-topmost", False))
        except Exception:
            pass

    # ================================================================
    #  V6.9.73 — VISTA "POR PACIENTE" (pestaña propia)
    # ================================================================
    #  El Visualizador es una fila por ESTUDIO: correcto para estadística y
    #  exportación, pero deja los estudios de un mismo paciente dispersos por
    #  la lista. Aquí se agrupan: una fila por paciente, desplegable en sus
    #  estudios.
    #
    #  Es una vista APARTE, con su propia hoja: no toca la tabla principal ni
    #  el dato. Cada estudio sigue siendo su propia fila en la BD.
    #
    #  Se agrupa SOLO por cédula (99,7% de las filas la tienen). Agrupar por
    #  nombre fusionaría homónimos, y mezclar la historia clínica de dos
    #  pacientes distintos es un error grave, no un detalle de presentación.
    #  Las filas sin cédula fiable se muestran sueltas, nunca fusionadas.
    # ================================================================
    # El NOMBRE del paciente (y el nº de caso en los hijos) NO va como columna: va
    # como texto del árbol, que tksheet dibuja en la columna índice junto con la
    # flecha de desplegar. Por eso aquí solo están las columnas de datos.
    _PAC_COLS = ["Cédula", "Estudios", "Órgano", "Diagnóstico", "Biomarcadores", "Fecha"]
    _PAC_PREFIJO = "P#"
    # No son biomarcadores: son metadatos del estudio.
    _PAC_NO_BIO = {"IHQ_ORGANO", "IHQ_ESTUDIOS_SOLICITADOS"}

    def _pac_es_coloracion(self, num) -> bool:
        return bool(re.match(r"^[Mm]\d", str(num or "").strip()))

    # Campos de fecha en orden de preferencia (mismos que usa la ficha)
    _PAC_FECHAS = ("Fecha de ingreso (2. Fecha de la muestra)", "Fecha Ingreso",
                   "Fecha Informe", "Fecha Ingreso Base de Datos")

    def _pac_construir_filas(self, df, solo_multi=False, filtro=""):
        """Devuelve (filas_para_tree_build, n_pacientes, n_estudios).
        Cada fila lleva 2 columnas auxiliares al final: iid y parent.

        Trabaja sobre LISTAS de Python, no sobre el DataFrame fila a fila:
        acceder con df.loc[i] 22.500 veces crea una Serie por fila y tardaba
        10 s. Con las columnas volcadas a listas baja a menos de 1 s.
        """
        import collections

        def col(nombre):
            if nombre in df.columns:
                return df[nombre].fillna("").astype(str).tolist()
            return [""] * len(df)

        ced = [re.sub(r"\D", "", s) for s in col("N. de identificación")]
        nom = col("Nombre Completo")
        num = [s.strip() for s in col("Numero de caso")]
        organo = col("Organo")
        dx_ihq = col("Diagnostico Principal")
        dx_col = col("Diagnostico Coloracion 2")
        dx_alt = col("Diagnostico Coloracion")
        fechas = [col(c) for c in self._PAC_FECHAS]
        NA = self._FICHA_NA

        # Biomarcadores: en vez de 125 columnas casi siempre vacías, UNA columna con
        # los que ESE estudio tiene con resultado. Se prepara una sola vez, en listas.
        bio_cols = [c for c in df.columns
                    if str(c).upper().startswith("IHQ_") and str(c) not in self._PAC_NO_BIO]
        bio_vals = [(str(c)[4:], col(c)) for c in bio_cols]

        def _val(*candidatos):
            for v in candidatos:
                v = (v or "").strip()
                if v and v.upper() not in NA:
                    return v
            return ""

        cuenta = collections.Counter(c for c in ced if len(c) >= 4)

        f = filtro.strip().lower()
        # V6.9.86: primero se AGRUPA y despues se filtra el grupo ENTERO.
        # Filtrando fila a fila se partía al paciente: hay 9 con el nombre escrito
        # distinto entre sus propias filas ('LUCELLY' / 'LUCENY'), así que buscar
        # una grafía mostraba solo parte de su historia — y con "solo varios
        # estudios" encendido aparecían con un único estudio, justo lo que ese
        # interruptor promete que no pasa. Buscar una errata debe traer al
        # paciente completo.
        grupos, sueltos = {}, []
        for k in range(len(num)):
            if len(ced[k]) >= 4:
                grupos.setdefault(ced[k], []).append(k)
            else:
                sueltos.append(k)
        if f:
            grupos = {c: ks for c, ks in grupos.items()
                      if f in c or any(f in nom[k].lower() for k in ks)}
            sueltos = [k for k in sueltos
                       if f in nom[k].lower() or f in ced[k]]

        def _bio(k):
            """Biomarcadores de ese estudio, en una línea.
            "NO MENCIONADO" se CONSERVA (significa que se solicitó pero el informe
            no lo reporta: es una señal real de calidad del dato), pero se manda al
            final para que lo primero que se lea sean los que sí tienen resultado."""
            res = [(v[k].strip().upper() == "NO MENCIONADO", f"{nom_b}: {v[k].strip()}")
                   for nom_b, v in bio_vals
                   if v[k].strip() and v[k].strip().upper() not in NA]
            res.sort(key=lambda x: x[0])
            return "   ·   ".join(t for _, t in res)

        def _dx(k):
            """El diagnóstico vive en un campo distinto según el tipo de estudio."""
            if self._pac_es_coloracion(num[k]):
                return _val(dx_col[k], dx_alt[k])
            return _val(dx_ihq[k], dx_alt[k])

        def _fecha(k):
            return _val(*(fc[k] for fc in fechas))

        def _orden_fecha(s):
            """DD/MM/AAAA -> AAAAMMDD para poder comparar. Lo que no case va al
            final, nunca delante de una fecha real."""
            m = re.match(r"^\s*(\d{2})/(\d{2})/(\d{4})", str(s or ""))
            return (m.group(3) + m.group(2) + m.group(1)) if m else ""

        # [Cédula, Estudios, Órgano, Diagnóstico, Biomarcadores, Fecha] + [texto, iid, parent]
        def _fila_estudio(k, pid):
            es_col = self._pac_es_coloracion(num[k])
            return ["", "Coloración" if es_col else "IHQ",
                    _val(organo[k]), _dx(k), _bio(k), _fecha(k),
                    ("🎨  " if es_col else "🔬  ") + num[k],
                    num[k] or f"_e{k}", pid]

        orden = sorted(grupos.items(), key=lambda kv: (nom[kv[1][0]].upper(), kv[0]))
        filas, n_pac, n_est = [], 0, 0
        for c, idxs in orden:
            # V6.9.86: se cuenta sobre el grupo YA construido, no sobre el
            # recuento global. Ahora que el filtro no parte pacientes son lo
            # mismo, y así el interruptor no puede volver a contradecirse.
            if solo_multi and len(idxs) < 2:
                continue
            n_pac += 1
            estudios = sorted(idxs, key=lambda k: (self._pac_es_coloracion(num[k]), num[k]))
            n_col = sum(1 for k in estudios if self._pac_es_coloracion(num[k]))
            n_ihq = len(estudios) - n_col
            partes = []
            if n_ihq:
                partes.append(f"{n_ihq} IHQ")
            if n_col:
                partes.append("1 coloración" if n_col == 1 else f"{n_col} coloraciones")
            pid = self._PAC_PREFIJO + c
            # Fila del PACIENTE. V6.9.85: deja de ir en blanco, pero SIN elegir un
            # estudio para "representar" al paciente —eso afirmaría algo que el
            # informe no dice, y era la razón de vaciarla—. Solo se muestra lo que
            # es cierto del paciente entero:
            #   · Órgano      el CONJUNTO de sus órganos, no uno.
            #   · Diagnóstico solo si TODOS sus estudios dicen lo mismo; si
            #                 difieren, se deja vacío (hay que abrir y mirar).
            #   · Fecha       la más reciente, que es un agregado honesto.
            # Biomarcadores sigue vacío: son de cada estudio y juntarlos daría una
            # línea ilegible que además mezclaría resultados de muestras distintas.
            orgs = []
            for k in estudios:
                o = _val(organo[k])
                if o and o not in orgs:
                    orgs.append(o)
            org_pac = " · ".join(orgs[:2])
            if len(orgs) > 2:
                org_pac += f"  +{len(orgs) - 2}"
            dxs = {d for d in (_dx(k) for k in estudios) if d}
            dx_pac = dxs.pop() if len(dxs) == 1 else ""
            fec_pac = max((_fecha(k) for k in estudios), key=_orden_fecha, default="")
            # V6.9.86: si el paciente tiene UN solo estudio, la fila no agrega
            # nada — ES ese estudio, así que puede mostrar sus biomarcadores sin
            # mezclar muestras. Con varios se deja vacío: unir los marcadores de
            # muestras distintas daría una línea contradictoria (el mismo
            # marcador con dos resultados) además de ilegible.
            # El 90 % de los pacientes son solo coloraciones y ahí queda vacío
            # con razón: una tinción básica no lleva biomarcadores.
            # `_bio` recorre 146 columnas; no se llama para coloraciones, que no
            # llevan biomarcadores. Son el 90 % de los pacientes.
            bio_pac = ""
            if len(estudios) == 1 and not self._pac_es_coloracion(num[estudios[0]]):
                bio_pac = _bio(estudios[0])
            filas.append([c, f"{len(estudios)}  ·  " + " + ".join(partes),
                          org_pac, dx_pac, bio_pac, fec_pac,
                          nom[estudios[0]] or "(sin nombre)", pid, ""])
            n_est += len(estudios)
            filas.extend(_fila_estudio(k, pid) for k in estudios)

        # sin cédula fiable: nunca se fusionan (juntarlos por nombre mezclaría
        # homónimos), así que van sueltos al final como estudio individual
        if not solo_multi:
            for k in sueltos:
                n_pac += 1
                n_est += 1
                fila = _fila_estudio(k, "")
                fila[0] = "(sin cédula)"
                fila[6] = f"{nom[k] or '(sin nombre)'}  ·  {num[k]}"
                filas.append(fila)
        return filas, n_pac, n_est

    def _populate_pacientes_tab_in_dashboard(self):
        """Construye la pestaña 'Por Paciente'. Hoja NUEVA creada en modo árbol
        desde el principio (no se reutiliza la del Visualizador)."""
        try:
            dash = getattr(self, "enhanced_dashboard", None)
            if dash is None or not hasattr(dash, "paciente_tab"):
                return
            frame = dash.paciente_tab
            for w in frame.winfo_children():
                w.destroy()

            barra = ttk.Frame(frame)
            barra.pack(fill=X, padx=10, pady=(8, 4))
            ttk.Label(barra, text="👤 Pacientes", font=("Segoe UI", 14, "bold")).pack(side=LEFT)

            self._pac_resumen = ttk.Label(barra, text="", font=("Segoe UI", 9),
                                          foreground="#5f6472")
            self._pac_resumen.pack(side=LEFT, padx=(12, 0))

            ttk.Button(barra, text="↻ Actualizar", bootstyle="secondary",
                       command=self._pac_refrescar).pack(side=RIGHT)
            self._pac_solo_multi = tk.BooleanVar(value=False)
            ttk.Checkbutton(barra, text="Solo con varios estudios",
                            variable=self._pac_solo_multi, bootstyle="round-toggle",
                            command=self._pac_refrescar).pack(side=RIGHT, padx=10)

            fbuscar = ttk.Frame(frame)
            fbuscar.pack(fill=X, padx=10, pady=(0, 6))
            self._pac_buscar = tk.StringVar()
            e = ttk.Entry(fbuscar, textvariable=self._pac_buscar, font=("Segoe UI", 10))
            e.pack(fill=X)
            e.insert(0, "")
            ttk.Label(fbuscar, text="Buscar por nombre o cédula…", font=("Segoe UI", 8),
                      foreground="#8a8f98").pack(anchor="w")
            self._pac_buscar.trace_add("write", lambda *_: self._pac_refrescar_debounce())

            cont = ttk.Frame(frame)
            cont.pack(expand=True, fill=BOTH, padx=10, pady=(0, 10))
            self.sheet_pac = Sheet(
                cont,
                treeview=True,
                headers=list(self._PAC_COLS),
                # El índice ES la columna del árbol: ahí van la flecha de desplegar,
                # la sangría y el nombre del paciente / nº de caso. Sin él no hay
                # forma de expandir nada.
                show_row_index=True,
                index_width=330,
                treeview_indent=24,
                headers_height=30,
                default_row_height=25,
                header_font=("Segoe UI", 10, "bold"),
                font=("Segoe UI", 10, "normal"),
                header_bg="#E8F5E9", header_fg="#1B5E20",
                table_bg="white", table_fg="black",
                table_selected_cells_bg="#BBDEFB", table_selected_cells_fg="black",
                startup_select=None,
                empty_horizontal=0, empty_vertical=0,
            )
            self.sheet_pac.pack(expand=True, fill=BOTH)
            self.sheet_pac.enable_bindings("copy", "row_select", "drag_select",
                                           "arrowkeys", "rc_select")
            self.sheet_pac.bind("<Double-Button-1>", self._pac_abrir_ficha, add="+")
            self._pac_refrescar()
        except Exception:
            logging.exception("No se pudo construir la pestaña 'Por Paciente'")

    def _pac_refrescar_debounce(self):
        """Evita reconstruir el árbol en cada tecla."""
        try:
            if getattr(self, "_pac_after", None):
                self.after_cancel(self._pac_after)
        except Exception:
            pass
        self._pac_after = self.after(350, self._pac_refrescar)

    def _pac_refrescar(self):
        try:
            sheet = getattr(self, "sheet_pac", None)
            if sheet is None:
                return
            df = getattr(self, "master_df", None)
            if df is None or getattr(df, "empty", True) \
                    or "N. de identificación" not in df.columns:
                return
            filas, n_pac, n_est = self._pac_construir_filas(
                df,
                solo_multi=bool(getattr(self, "_pac_solo_multi", None)
                                and self._pac_solo_multi.get()),
                filtro=(getattr(self, "_pac_buscar", None).get()
                        if getattr(self, "_pac_buscar", None) else ""))
            n = len(self._PAC_COLS)
            if not filas:
                sheet.set_sheet_data([[""] * n], redraw=True)
            else:
                # Las 3 últimas columnas (texto del árbol, iid y padre) son de
                # servicio: tree_build las consume y con include_*_column=False no
                # entran en la tabla. Si no, aparecían como dos columnas sueltas
                # "G" y "H" llenas de "P#12345678".
                sheet.tree_build(data=filas, iid_column=n + 1, parent_column=n + 2,
                                 text_column=n, open_ids=[], ncols=n,
                                 include_text_column=False, include_iid_column=False,
                                 include_parent_column=False)
                for i, ancho in enumerate((110, 175, 165, 430, 430, 95)):
                    try:
                        sheet.column_width(column=i, width=ancho, redraw=False)
                    except Exception:
                        pass
                sheet.redraw()
            if getattr(self, "_pac_resumen", None) is not None:
                self._pac_resumen.config(
                    text=f"{n_pac:,} pacientes  ·  {n_est:,} estudios".replace(",", "."))
        except Exception:
            logging.exception("No se pudo refrescar la vista por paciente")

    def _pac_abrir_ficha(self, event=None):
        """Doble clic -> ficha completa del paciente (la misma de V6.9.58)."""
        try:
            sheet = self.sheet_pac
            r = sheet.identify_row(event, allow_end=False)
            if r is None:
                return
            # identify_row da la fila MOSTRADA; con los pacientes colapsados eso NO
            # coincide con la fila de datos (hay miles de estudios ocultos entre
            # medias). Sin traducir el índice se abría la ficha de otro paciente.
            filas_vis = sheet.MT.displayed_rows
            if not sheet.MT.all_rows_displayed and filas_vis is not None:
                if r >= len(filas_vis):
                    return
                r = filas_vis[r]
            # El iid del nodo ES el identificador: nº de caso en los estudios y
            # "P#<cédula>" en la fila del paciente. Se lee de ahí en vez de
            # parsear el texto que se dibuja.
            iid = str(getattr(sheet.MT._row_index[r], "iid", "")) \
                if r < len(sheet.MT._row_index) else ""
            num = (self._primer_caso_de_cedula(iid[len(self._PAC_PREFIJO):])
                   if iid.startswith(self._PAC_PREFIJO) else iid)
            if num:
                self._mostrar_ficha_paciente(num)
        except Exception:
            logging.exception("No se pudo abrir la ficha desde la vista por paciente")

    def _primer_caso_de_cedula(self, ced) -> str:
        """Nº de caso de cualquier estudio de esa cédula (para abrir su ficha)."""
        try:
            ced = re.sub(r"\D", "", str(ced or ""))
            df = getattr(self, "master_df", None)
            if not ced or df is None or getattr(df, "empty", True):
                return ""
            m = (df["N. de identificación"].fillna("").astype(str)
                 .str.replace(r"\D", "", regex=True)) == ced
            if not m.any():
                return ""
            return str(df.loc[m, "Numero de caso"].iloc[0]).strip()
        except Exception:
            return ""

    # ================================================================
    #  V6.9.58 — FICHA DEL PACIENTE (agrupa IHQ + Coloraciones)
    # ================================================================
    #  El DATO no se toca: cada estudio (IHQ###### o M######) sigue siendo su
    #  propia fila. Aquí solo se AGRUPAN en la VISTA por paciente (cédula), que
    #  es lo que faltaba: un paciente puede tener varios IHQ y varias
    #  coloraciones (hasta 9 estudios) y en la tabla plana quedaban dispersos.
    # ================================================================
    _FICHA_NA = {"", "N/A", "NA", "NAN", "NONE", "NULL", "NO APLICA", "-", "--"}

    def _ficha_es_coloracion(self, num) -> bool:
        return bool(re.match(r"^[Mm]\d", str(num or "").strip()))

    def _ficha_fecha(self, fila):
        """Mejor fecha disponible del estudio (para ordenar y mostrar)."""
        for c in ("Fecha de ingreso (2. Fecha de la muestra)", "Fecha Ingreso",
                  "Fecha Informe", "Fecha Ingreso Base de Datos"):
            v = str(fila.get(c, "") or "").strip()
            if v and v.upper() not in self._FICHA_NA:
                return v
        return ""

    def _mostrar_ficha_paciente(self, num_click) -> bool:
        """Ventana con TODOS los estudios del paciente del registro clicado,
        agrupados y en orden. Devuelve False si no se pudo resolver el paciente
        (el llamador entonces abre el detalle simple del registro)."""
        try:
            df = getattr(self, "master_df", None)
            if df is None or getattr(df, "empty", True):
                return False
            if "Numero de caso" not in df.columns or "N. de identificación" not in df.columns:
                return False

            nums = df["Numero de caso"].astype(str).str.strip()
            fila = df[nums == str(num_click).strip()]
            if fila.empty:
                return False
            ced = re.sub(r"\D", "", str(fila.iloc[0].get("N. de identificación", "")))
            if len(ced) < 4:
                return False  # sin cédula fiable -> no se puede agrupar

            ceds = df["N. de identificación"].astype(str).str.replace(r"\D", "", regex=True)
            estudios = df[ceds == ced].copy()
            if estudios.empty:
                return False

            # Orden: cronológico (fecha) y, a igualdad, por número de caso.
            estudios["_f"] = estudios.apply(self._ficha_fecha, axis=1)
            def _clave(v):
                m = re.match(r"(\d{1,2})[/-](\d{1,2})[/-](\d{2,4})", str(v))
                if m:
                    d, mo, a = m.groups()
                    a = ("20" + a) if len(a) == 2 else a
                    return f"{a}{int(mo):02d}{int(d):02d}"
                return "0"
            estudios["_k"] = estudios["_f"].map(_clave)
            estudios = estudios.sort_values(["_k", "Numero de caso"])

            nombre = str(fila.iloc[0].get("Nombre Completo", "") or "").strip()
            n_ihq = sum(1 for n in estudios["Numero de caso"] if not self._ficha_es_coloracion(n))
            n_col = len(estudios) - n_ihq
            self._render_ficha(nombre, ced, estudios, str(num_click).strip(), n_ihq, n_col)
            return True
        except Exception:
            logging.exception("_mostrar_ficha_paciente falló")
            return False

    def _render_ficha(self, nombre, ced, estudios, num_click, n_ihq, n_col):
        win = tk.Toplevel(self)
        win.title(f"Ficha del paciente — {nombre or ced}")
        win.geometry("900x720")
        win.transient(self)

        cont = ttk.Frame(win, padding=10)
        cont.pack(fill="both", expand=True)
        txt = tk.Text(cont, wrap="word", font=("Segoe UI", 10), padx=14, pady=10,
                      relief="flat", background="#ffffff", foreground="#1b1b1b",
                      cursor="arrow")
        scroll = ttk.Scrollbar(cont, orient="vertical", command=txt.yview)
        txt.configure(yscrollcommand=scroll.set)
        txt.pack(side="left", fill="both", expand=True)
        scroll.pack(side="right", fill="y")

        txt.tag_configure("paciente", font=("Segoe UI", 13, "bold"), foreground="#0D47A1",
                          spacing1=4, spacing3=2)
        txt.tag_configure("sub", font=("Segoe UI", 9), foreground="#5f6368", spacing3=10)
        txt.tag_configure("h_ihq", font=("Segoe UI", 11, "bold"), foreground="#1B5E20",
                          background="#E8F5E9", spacing1=12, spacing3=4)
        txt.tag_configure("h_col", font=("Segoe UI", 11, "bold"), foreground="#4A148C",
                          background="#F3E5F5", spacing1=12, spacing3=4)
        txt.tag_configure("actual", font=("Segoe UI", 9, "bold"), foreground="#E65100")
        txt.tag_configure("campo", font=("Segoe UI", 9, "bold"), foreground="#37474F",
                          spacing1=6, spacing3=1)
        txt.tag_configure("valor", font=("Segoe UI", 10), lmargin1=8, lmargin2=8, spacing3=3)
        txt.tag_configure("largo", font=("Segoe UI", 10), lmargin1=10, lmargin2=10,
                          background="#F1F8E9", spacing1=2, spacing3=6)
        txt.tag_configure("bio", font=("Consolas", 9), foreground="#263238",
                          lmargin1=10, lmargin2=10, background="#FAFAFA", spacing3=4)
        # V6.9.74: datos de contexto — presentes pero secundarios frente al diagnóstico
        txt.tag_configure("sub2", font=("Segoe UI", 9), foreground="#5f6368",
                          lmargin1=10, lmargin2=10, spacing3=2)

        copia = []
        txt.insert("end", f"{nombre or '(sin nombre)'}\n", "paciente")
        resumen = f"CC {ced}   ·   {len(estudios)} estudio(s):  {n_ihq} IHQ  ·  {n_col} Coloración(es)"
        # V6.9.74: la cabecera solo traía nombre y cédula. Edad, género y documento
        # son del PACIENTE (no de cada estudio), así que van aquí una sola vez.
        _demo = []
        for _c, _et in (("Edad", "años"), ("Genero", ""), ("Tipo de documento", "doc.")):
            _v = next((str(estudios.iloc[i].get(_c, "") or "").strip()
                       for i in range(len(estudios))
                       if str(estudios.iloc[i].get(_c, "") or "").strip().upper()
                       not in self._FICHA_NA), "")
            if not _v:
                continue
            if _c == "Edad":
                _demo.append(f"{_v} años")
            elif _c == "Genero":
                _demo.append(_v.capitalize())
            else:
                # las siglas de documento (CC, TI, RC…) van en MAYÚSCULAS, no "Ti"
                _demo.append(f"doc. {_v.upper()}")
        if _demo:
            resumen += "   ·   " + "  ·  ".join(_demo)
        txt.insert("end", f"{resumen}\n", "sub")
        copia.append(f"{nombre} — CC {ced} — {resumen}")

        CLAVE = ["Organo", "Procedimiento", "Malignidad", "Servicio", "Médico tratante"]
        # V6.9.74: datos del estudio que la ficha no mostraba y sí tienen dato
        # (patólogo, dónde y cuándo se hizo, quién lo cubre).
        CONTEXTO = ["Patologo", "Tipo de examen", "Especialidad", "Sede", "EPS",
                    "Hospitalizado", "Departamento", "Municipio", "CUPS",
                    "Fecha de toma (1. Fecha de la toma)",
                    "Fecha de ingreso (2. Fecha de la muestra)", "Fecha Informe"]
        # El diagnóstico vive en un campo DISTINTO según el tipo de estudio:
        #   · Coloración (fila M): su dx está en "Diagnostico Coloracion 2".
        #   · IHQ:                 su dx está en "Diagnostico Principal".
        # En la fila IHQ, "Diagnostico Coloracion 2" contiene el texto CONCATENADO de
        # las coloraciones del paciente -> aquí NO se muestra: esas coloraciones ya
        # aparecen como secciones propias (evita el duplicado que confunde).
        DX_IHQ = [("Diagnostico Principal", "Diagnóstico"),
                  ("Diagnostico Coloracion", "Diagnóstico citado en el informe IHQ"),
                  ("Factor pronostico", "Factor pronóstico")]
        DX_COL = [("Diagnostico Coloracion 2", "Diagnóstico")]
        LARGOS = ["Descripcion macroscopica", "Descripcion microscopica",
                  "Descripcion macroscopica Coloracion", "Descripcion microscopica Coloracion"]

        def _ok(v):
            v = str(v or "").strip()
            return v if v.upper() not in self._FICHA_NA else ""

        for _, est in estudios.iterrows():
            num = str(est.get("Numero de caso", "")).strip()
            es_col = self._ficha_es_coloracion(num)
            fecha = self._ficha_fecha(est)
            tag = "h_col" if es_col else "h_ihq"
            icono = "🎨" if es_col else "🔬"
            tipo = "COLORACIÓN" if es_col else "ESTUDIO IHQ"
            cab = f"{icono}  {num}   ·   {tipo}" + (f"   ·   {fecha}" if fecha else "")
            txt.insert("end", f"\n{cab}\n", tag)
            if num == num_click:
                txt.insert("end", "   ▲ el registro que abriste\n", "actual")
            copia.append(f"\n=== {cab} ===")

            # Datos clave en una línea compacta
            partes = [f"{c}: {_ok(est.get(c))}" for c in CLAVE if _ok(est.get(c))]
            if partes:
                txt.insert("end", "   " + "   ·   ".join(partes) + "\n", "valor")
                copia.append("  " + " · ".join(partes))

            # V6.9.74: contexto del estudio (patólogo, sede, EPS, fechas…)
            ctx = [f"{c.split(' (')[0]}: {_ok(est.get(c))}" for c in CONTEXTO if _ok(est.get(c))]
            if ctx:
                txt.insert("end", "   " + "   ·   ".join(ctx) + "\n", "sub2")
                copia.append("  " + " · ".join(ctx))

            # Diagnósticos (el campo correcto según el tipo de estudio)
            for c, etiqueta in (DX_COL if es_col else DX_IHQ):
                v = _ok(est.get(c))
                if not v:
                    continue
                txt.insert("end", f"{etiqueta}\n", "campo")
                txt.insert("end", f"{v}\n", "largo")
                copia.append(f"  {etiqueta}: {v}")

            # Panel solicitado (contexto: qué marcadores pidió el patólogo)
            solic = _ok(est.get("IHQ_ESTUDIOS_SOLICITADOS"))
            if solic:
                txt.insert("end", "Biomarcadores solicitados\n", "campo")
                txt.insert("end", f"   {solic}\n", "valor")
                copia.append(f"  Solicitados: {solic}")

            # Biomarcadores: SOLO los que este estudio tiene con RESULTADO.
            # IHQ_ORGANO / IHQ_ESTUDIOS_SOLICITADOS NO son biomarcadores (son metadatos).
            _NO_BIO = {"IHQ_ORGANO", "IHQ_ESTUDIOS_SOLICITADOS"}
            bios = [(c.replace("IHQ_", ""), _ok(est.get(c)))
                    for c in estudios.columns
                    if str(c).startswith("IHQ_") and str(c) not in _NO_BIO]
            bios = [(k, v) for k, v in bios if v]
            if bios:
                txt.insert("end", f"Resultados de biomarcadores ({len(bios)})\n", "campo")
                txt.insert("end", "   " + "   ·   ".join(f"{k}: {v}" for k, v in bios) + "\n", "bio")
                copia.append("  Biomarcadores: " + " · ".join(f"{k}: {v}" for k, v in bios))
            else:
                # V6.9.74: antes esta sección simplemente NO aparecía, y eso se leía
                # como "faltan datos". Ahora se dice por qué no hay: una coloración es
                # una tinción básica, no lleva biomarcadores.
                _razon = ("las coloraciones son tinciones básicas y no llevan biomarcadores"
                          if es_col else "este estudio no reporta ningún biomarcador")
                txt.insert("end", "Resultados de biomarcadores\n", "campo")
                txt.insert("end", f"   Sin biomarcadores — {_razon}.\n", "sub2")
                copia.append(f"  Biomarcadores: ninguno ({_razon})")

            # Descripciones largas
            for c in LARGOS:
                v = _ok(est.get(c))
                if not v:
                    continue
                txt.insert("end", f"{c}\n", "campo")
                txt.insert("end", f"{v}\n", "largo")
                copia.append(f"  {c}: {v}")

            # V6.9.74: CAJÓN FINAL — cualquier campo con dato que no se haya mostrado
            # arriba. Garantiza que la ficha no esconda nada: si el informe lo trae,
            # aquí sale. Se excluyen los ya mostrados y los de identidad del paciente
            # (ya están en la cabecera) para no repetir.
            _ya = set(CLAVE) | set(CONTEXTO) | set(LARGOS) | {
                "Numero de caso", "Nombre Completo", "N. de identificación", "Edad",
                "Genero", "Tipo de documento", "Primer nombre", "Segundo nombre",
                "Primer apellido", "Segundo apellido", "IHQ_ESTUDIOS_SOLICITADOS",
                "Diagnostico Principal", "Diagnostico Coloracion",
                "Diagnostico Coloracion 2", "Factor pronostico"}
            # "_f"/"_k" son columnas auxiliares que _mostrar_ficha_paciente añade para
            # ordenar los estudios; no son datos del informe y no deben salir aquí.
            resto = [(c, _ok(est.get(c))) for c in estudios.columns
                     if c not in _ya and not str(c).startswith(("IHQ_", "_"))
                     and _ok(est.get(c))]
            if resto:
                txt.insert("end", f"Otros datos del informe ({len(resto)})\n", "campo")
                for c, v in resto:
                    txt.insert("end", f"   {c}: {v}\n", "sub2")
                copia.append("  Otros: " + " · ".join(f"{c}: {v}" for c, v in resto))

        # Solo lectura pero copiable
        def _solo_lectura(e):
            if (e.state & 0x4) and e.keysym.lower() in ("c", "a"):
                return
            if e.keysym in ("Left", "Right", "Up", "Down", "Home", "End", "Prior", "Next"):
                return
            return "break"
        txt.bind("<Key>", _solo_lectura)

        barra = ttk.Frame(win, padding=(10, 6))
        barra.pack(fill="x")

        def _copiar():
            try:
                self.clipboard_clear()
                self.clipboard_append("\n".join(copia))
            except Exception:
                pass

        ttk.Button(barra, text="📋 Copiar ficha", command=_copiar,
                   bootstyle="secondary").pack(side="left")
        ttk.Button(barra, text="Cerrar", command=win.destroy,
                   bootstyle="primary").pack(side="right")
        win.bind("<Escape>", lambda e: win.destroy())
        try:
            win.update_idletasks()
            x = self.winfo_x() + (self.winfo_width() - win.winfo_width()) // 2
            y = self.winfo_y() + (self.winfo_height() - win.winfo_height()) // 2
            win.geometry(f"+{max(x, 0)}+{max(y, 0)}")
            win.lift()
            win.focus_force()
            win.attributes("-topmost", True)
            win.after(300, lambda: win.attributes("-topmost", False))
        except Exception:
            pass

    # ================================================================
    #  V6.9.51 — Hover: popup con el VALOR COMPLETO de la celda
    # ================================================================
    def _instalar_hover_tooltip(self, sheet):
        """Instala en un Sheet el popup que muestra el TEXTO COMPLETO de la celda al
        pasar el mouse (las celdas truncan campos largos como 'Descripcion macroscopica').
        Se enlaza al widget interno (MainTable) con tkinter puro -> funciona en cualquier
        versión de tksheet."""
        if sheet is None:
            return
        try:
            sheet.MT.bind("<Motion>", lambda e, s=sheet: self._hover_tip_motion(e, s), add="+")
            sheet.MT.bind("<Leave>", lambda e: self._hover_tip_hide(), add="+")
            logging.info("Hover-tooltip de celda instalado en un Sheet")
        except Exception as e:
            logging.warning(f"No se pudo instalar hover-tooltip: {e}")

    def _hover_tip_motion(self, event, sheet):
        """Detecta la celda bajo el cursor y programa el popup (con breve retardo)."""
        try:
            r = sheet.MT.identify_row(y=event.y, allow_end=False)
            c = sheet.MT.identify_col(x=event.x, allow_end=False)
        except Exception:
            r = c = None
        if r is None or c is None:
            self._hover_tip_hide()
            self._tip_cell = None
            return
        celda = (id(sheet), r, c)
        if getattr(self, "_tip_cell", None) == celda:
            return  # misma celda: no recrear
        self._tip_cell = celda
        self._hover_tip_hide()
        if getattr(self, "_tip_after", None):
            try:
                self.after_cancel(self._tip_after)
            except Exception:
                pass
        xr, yr = event.x_root, event.y_root
        self._tip_after = self.after(
            450, lambda: self._hover_tip_show(sheet, r, c, xr, yr))

    def _hover_tip_show(self, sheet, r, c, x_root, y_root):
        """Crea el popup con el valor completo de la celda (r, c)."""
        try:
            val = sheet.get_cell_data(r, c)
        except Exception:
            val = None
        val = "" if val is None else str(val).strip()
        if not val:
            return
        self._hover_tip_hide()
        try:
            tip = tk.Toplevel(self)
            tip.wm_overrideredirect(True)  # sin barra de título
            try:
                tip.attributes("-topmost", True)
            except Exception:
                pass
            tk.Label(
                tip, text=val, justify="left", wraplength=560,
                background="#FFFFE0", foreground="#1b1b1b",
                relief="solid", borderwidth=1, font=("Segoe UI", 10),
                padx=8, pady=6,
            ).pack()
            tip.update_idletasks()
            # Ajustar posición para no salir de la pantalla.
            sw, sh = self.winfo_screenwidth(), self.winfo_screenheight()
            w, h = tip.winfo_reqwidth(), tip.winfo_reqheight()
            x, y = x_root + 16, y_root + 18
            if x + w > sw:
                x = max(0, sw - w - 8)
            if y + h > sh:
                y = max(0, y_root - h - 12)
            tip.geometry(f"+{x}+{y}")
            self._tip = tip
        except Exception as e:
            logging.debug(f"_hover_tip_show error: {e}")

    def _hover_tip_hide(self):
        tip = getattr(self, "_tip", None)
        if tip is not None:
            try:
                tip.destroy()
            except Exception:
                pass
        self._tip = None

    def _export_full_database(self):
        """Exportar toda la base de datos usando el sistema mejorado"""
        try:
            self.export_system.export_full_database()
        except Exception as e:
            messagebox.showerror("Error de Exportación", f"Error al exportar la base de datos:\n{str(e)}")

    # ================================================================
    #  RESUMEN IA — Análisis profesional de la base de datos
    # ================================================================

    def _generar_resumen_ia(self):
        """Genera un resumen profesional de la BD usando IA (en hilo aparte)."""
        if not hasattr(self, 'master_df') or self.master_df is None or self.master_df.empty:
            messagebox.showwarning("Sin Datos", "No hay datos en la base de datos para analizar.")
            return

        # Mostrar overlay mientras se genera
        self._resumen_ia_overlay = tk.Toplevel(self)
        overlay = self._resumen_ia_overlay
        overlay.title("Generando Resumen IA...")
        overlay.transient(self)
        overlay.grab_set()
        overlay.resizable(False, False)
        overlay.protocol("WM_DELETE_WINDOW", lambda: None)

        w, h = 420, 140
        x = self.winfo_x() + (self.winfo_width() - w) // 2
        y = self.winfo_y() + (self.winfo_height() - h) // 2
        overlay.geometry(f"{w}x{h}+{x}+{y}")

        frame = ttk.Frame(overlay, padding=25)
        frame.pack(fill=BOTH, expand=True)
        ttk.Label(frame, text="📊 Generando resumen con IA…", font=("Segoe UI", 13, "bold")).pack(pady=(0, 8))
        ttk.Label(frame, text="Esto puede tomar varios minutos con modelos locales.", font=("Segoe UI", 10)).pack()
        pb = ttk.Progressbar(frame, mode="indeterminate", bootstyle="info-striped")
        pb.pack(fill=X, pady=(10, 0))
        pb.start(15)

        # Resultado compartido entre hilos
        self._resumen_ia_result = {"done": False, "texto": "", "error": None}

        thread = threading.Thread(target=self._resumen_ia_worker, daemon=True)
        thread.start()
        self.after(500, self._poll_resumen_ia)

    def _compilar_estadisticas(self):
        """Compila estadísticas locales del DataFrame para enviar a la IA."""
        df = self.master_df.copy()
        stats = {}

        stats["total_casos"] = len(df)
        stats["nota"] = "Todos los casos son estudios de Inmunohistoquímica (IHQ)"

        # Rango de fechas
        for col in ["Fecha Informe", "Fecha de informe", "Fecha de ingreso"]:
            if col in df.columns:
                fechas = pd.to_datetime(df[col], dayfirst=True, errors="coerce").dropna()
                if not fechas.empty:
                    stats["fecha_min"] = fechas.min().strftime("%d/%m/%Y")
                    stats["fecha_max"] = fechas.max().strftime("%d/%m/%Y")
                    break

        # Distribución de malignidad
        if "Malignidad" in df.columns:
            stats["malignidad"] = df["Malignidad"].fillna("SIN DATO").value_counts().head(10).to_dict()

        # Distribución por órgano — NORMALIZACIÓN CANÓNICA
        # Usa core/normalizador_organos para agrupar por categoría anatómica
        # real (MAMA, COLON, MEDULA OSEA, etc.) en lugar de strings literales.
        # Prefiere IHQ_ORGANO (más limpio) sobre Organo.
        from core.normalizador_organos import (
            normalizar_organo,
            elegir_columna_organo,
        )
        col_organo = elegir_columna_organo(df.columns)
        if col_organo is not None:
            serie_norm = df[col_organo].apply(normalizar_organo)
            total_validos = int((serie_norm != "SIN DATO").sum())
            top = serie_norm.value_counts()
            # Excluir SIN DATO del ranking principal
            top_sin_nulos = top[top.index != "SIN DATO"]
            stats["organos_normalizados"] = top_sin_nulos.head(20).to_dict()
            stats["organos_total_con_dato"] = total_validos
            stats["organos_columna_fuente"] = col_organo
            stats["organos_categorias_distintas"] = int(top_sin_nulos.shape[0])
        # Distribución por procedimiento — filtrar "INMUNOHISTOQUIMICA" (no es procedimiento quirúrgico)
        # V6.6.8: Agregar totales y agrupación por tipo (biopsia/cirugía/etc.)
        if "Procedimiento" in df.columns:
            procs = df["Procedimiento"].fillna("SIN DATO").astype(str).str.upper().str.strip()
            procs = procs[~procs.isin(["INMUNOHISTOQUIMICA", "INMUNOHISTOQUÍMICA"])]
            procs_validos = procs[procs != "SIN DATO"]
            stats["procedimientos"] = procs_validos.value_counts().head(10).to_dict()
            stats["procedimientos_total_con_dato"] = int(len(procs_validos))
            stats["procedimientos_categorias_distintas"] = int(procs_validos.nunique())
            # Agrupar por tipo clínico: BIOPSIA / CIRUGIA / PUNCION
            def _clasificar_proc(p: str) -> str:
                p = p.upper()
                if any(k in p for k in ["BIOPSIA", "BX ", "BIOP"]):
                    return "BIOPSIA"
                if any(k in p for k in ["ECTOMIA", "ECTOMÍA", "RESECCION", "RESECCIÓN",
                                         "CIRUG", "MASTECTOM", "CUADRANTECT",
                                         "HEMICOLECT", "APENDICECT", "NEFRECT",
                                         "HISTERECT", "TIROIDECT", "PROSTATECT",
                                         "GASTRECT", "SIGMOIDECT", "HEPATECT",
                                         "SALPINGOOFOR", "ESPLENECT", "LARINGECT"]):
                    return "CIRUGIA"
                if any(k in p for k in ["PUNCION", "PUNCIÓN", "ASPIRADO", "PAAF"]):
                    return "PUNCION/ASPIRADO"
                if "LEGRADO" in p or "CURETAJE" in p:
                    return "LEGRADO"
                return "OTRO"
            tipos = procs_validos.apply(_clasificar_proc)
            stats["procedimientos_por_tipo"] = tipos.value_counts().to_dict()

        # Diagnósticos — filtrar entradas genéricas que no son diagnósticos reales
        filtro_diag = ["ESTUDIO DE INMUNOHISTOQUÍMICA", "INMUNOHISTOQUÍMICA",
                       "ESTUDIO DE INMUNOHISTOQUIMICA", "INMUNOHISTOQUIMICA"]
        for diag_col in ["Diagnostico Principal", "Diagnostico Coloracion"]:
            if diag_col in df.columns:
                vals = df[diag_col].dropna().astype(str).str.strip()
                vals = vals[(vals != "") & (~vals.str.upper().isin([f.upper() for f in filtro_diag]))]
                if not vals.empty:
                    stats[f"top_{diag_col.lower().replace(' ', '_')}"] = vals.value_counts().head(10).to_dict()

        # Diagnósticos CATEGORIZADOS — agrupa los 883 diagnósticos literales
        # distintos en categorías clínicas (CARCINOMA DUCTAL DE MAMA,
        # ADENOCARCINOMA, LINFOMA, etc.). Esto refleja el volumen real
        # del HUV en lugar de strings literales fragmentados.
        # V6.6.8: Usar categorizar_diagnostico_con_organo para refinar diagnósticos
        # genéricos ("ADENOCARCINOMA SIN ORIGEN", "CARCINOMA OTRO") usando el campo
        # Organo del caso. Esto resuelve el feedback clínico: "Adenocarcinoma debe
        # especificar ubicación", "Carcinoma escamocelular debe aclarar órgano".
        from core.normalizador_diagnosticos import (
            categorizar_diagnostico,
            categorizar_diagnostico_con_organo,
        )
        from core.normalizador_organos import normalizar_organo as _norm_organo
        if "Diagnostico Principal" in df.columns:
            # Pasar diagnóstico + órgano canónico para inferencia contextual
            organo_col = col_organo if 'col_organo' in dir() and col_organo else None
            if organo_col is not None and organo_col in df.columns:
                organos_norm = df[organo_col].apply(_norm_organo)
                cat_serie = df.apply(
                    lambda row: categorizar_diagnostico_con_organo(
                        row["Diagnostico Principal"],
                        organos_norm.loc[row.name] if row.name in organos_norm.index else None
                    ),
                    axis=1
                )
            else:
                cat_serie = df["Diagnostico Principal"].apply(categorizar_diagnostico)
            cat_top = cat_serie.value_counts()

            # V6.6.9 FIX feedback clínico: separar DIAGNÓSTICOS ONCOLÓGICOS reales
            # (neoplasias benignas o malignas) de HALLAZGOS NO-NEOPLÁSICOS
            # (negativo para malignidad, muestra no representativa, gliosis,
            # rechazo trasplante, malformaciones, etc.). Antes todas estas
            # categorías aparecían mezcladas en "Diagnósticos Principales", lo
            # cual el patólogo correctamente señaló como incorrecto: "Negativo
            # para malignidad NO es un diagnóstico oncológico".
            CATEGORIAS_NO_NEOPLASICAS = {
                "NEGATIVO PARA MALIGNIDAD",
                "MUESTRA NO REPRESENTATIVA / NO DIAGNOSTICA",
                "HALLAZGO HISTOLOGICO NORMAL / NO PATOLOGICO",
                "RESULTADO IHQ (SIN DIAGNOSTICO ESPECIFICO)",
                "ESTUDIO IHQ (SIN DIAGNOSTICO ESPECIFICO)",
                "GLIOSIS / LESION REACTIVA SNC",
                "RECHAZO DE TRASPLANTE",
                "MALFORMACION DEL DESARROLLO / HETEROTOPIA SNC",
                # V6.9.23: nuevas categorías NO oncológicas del categorizador
                "PROCESO INFLAMATORIO / INFECCIOSO (NO NEOPLASICO)",
                "HALLAZGO NO NEOPLASICO / NEGATIVO (OTRO)",
                "ESTUDIO DE MEDULA OSEA (MORFOLOGIA)",
                "MUESTRA INSUFICIENTE / LIMITADA (OTRO)",
                "SIN DIAGNOSTICO EN TEXTO / REVISAR (EXTRACCION)",
                "ENFERMEDAD DE HIRSCHSPRUNG / CELULAS GANGLIONARES",
            }

            EXCLUIR_DEL_TOP = (CATEGORIAS_NO_NEOPLASICAS | {"SIN DATO", "OTRO / NO CATEGORIZADO"})

            # Diagnósticos oncológicos reales (para gráfico/tabla principal)
            cat_oncologico = cat_top[~cat_top.index.isin(EXCLUIR_DEL_TOP)]
            stats["diagnosticos_categorizados"] = cat_oncologico.head(20).to_dict()
            stats["diagnosticos_total_categorizado"] = int(cat_oncologico.sum())

            # Hallazgos NO neoplásicos (separados, para sección aparte)
            hallazgos_no_neo = cat_top[cat_top.index.isin(CATEGORIAS_NO_NEOPLASICAS)]
            stats["hallazgos_no_neoplasicos"] = hallazgos_no_neo.to_dict()
            stats["hallazgos_no_neoplasicos_total"] = int(hallazgos_no_neo.sum())

            # Métricas de control
            stats["diagnosticos_otro_no_categorizado"] = int(cat_top.get("OTRO / NO CATEGORIZADO", 0))
            stats["diagnosticos_estudio_ihq_sin_dx"] = int(cat_top.get("ESTUDIO IHQ (SIN DIAGNOSTICO ESPECIFICO)", 0))
            stats["diagnosticos_sin_dato"] = int(cat_top.get("SIN DATO", 0))

        # Biomarcadores principales — top 15, valores compactos
        bio_cols = [c for c in df.columns if c.startswith("IHQ_")]
        # Excluir columnas auxiliares (no son biomarcadores clínicos)
        excluir = {"IHQ_ORGANO", "IHQ_ESTUDIOS_SOLICITADOS"}
        bio_cols = [c for c in bio_cols if c not in excluir]
        bio_counts = {}
        for bc in bio_cols:
            serie = df[bc].dropna().astype(str).str.strip()
            serie = serie[(serie != "") & (serie.str.upper() != "N/A") & (serie.str.upper() != "NO MENCIONADO")]
            if len(serie) > 0:
                bio_counts[bc] = len(serie)
        top_bio = sorted(bio_counts.items(), key=lambda x: x[1], reverse=True)[:15]
        bio_summary = {}
        for bc, count in top_bio:
            serie = df[bc].dropna().astype(str).str.strip()
            serie = serie[(serie != "") & (serie.str.upper() != "N/A") & (serie.str.upper() != "NO MENCIONADO")]
            bio_summary[bc] = {
                "n": count,
                "top": serie.value_counts().head(3).to_dict()
            }
        if bio_summary:
            stats["biomarcadores_top15"] = bio_summary
        stats["total_biomarcadores_distintos"] = len(bio_counts)

        # Paneles más solicitados (estudios solicitados)
        if "IHQ_ESTUDIOS_SOLICITADOS" in df.columns:
            estudios = df["IHQ_ESTUDIOS_SOLICITADOS"].dropna().astype(str).str.strip()
            estudios = estudios[(estudios != "") & (estudios.str.upper() != "N/A")]
            if not estudios.empty:
                stats["paneles_ihq_solicitados"] = estudios.value_counts().head(10).to_dict()

        # Servicio solicitante — filtrar N/A y SIN DATO
        # V6.6.8 FIX: Agregar totales para que la IA sepa cuántos casos hay en
        # total y cuántos servicios distintos. Antes solo enviaba top 10 sin
        # contexto, generando reportes que aparentaban no sumar al total.
        if "Servicio" in df.columns:
            servicios = df["Servicio"].fillna("").astype(str).str.strip()
            servicios = servicios[(servicios != "") & (servicios.str.upper() != "N/A") & (servicios.str.upper() != "SIN DATO")]
            if not servicios.empty:
                stats["servicios"] = servicios.value_counts().head(10).to_dict()
                stats["servicios_total_con_dato"] = int(len(servicios))
                stats["servicios_categorias_distintas"] = int(servicios.nunique())

        return stats

    def _construir_resumen_factual(self, stats: dict) -> str:
        """V6.9.19: Secciones FACTUALES del resumen generadas de forma DETERMINISTA
        desde las estadísticas (cifras y nombres EXACTOS, sin IA). Evita que un modelo
        pequeño invente cifras o copie las instrucciones del prompt. La IA solo redacta
        el 'Resumen Ejecutivo' y las 'Observaciones clínicas'."""
        total = stats.get("total_casos", 0) or 0
        def lst(d, n=None):
            items = list((d or {}).items())
            if n:
                items = items[:n]
            return [f"- {k}: {v}" for k, v in items]
        L = []
        L.append("# Volumen y Temporalidad")
        L.append(f"- Total de casos (IHQ): {total}")
        if stats.get("fecha_min"):
            L.append(f"- Periodo: {stats.get('fecha_min')} a {stats.get('fecha_max', '?')}")
        mal = stats.get("malignidad") or {}
        if mal:
            L.append("")
            L.append("# Malignidad")
            for k, v in mal.items():
                p = f"{(100.0 * v / total):.1f}%" if total else "0.0%"
                L.append(f"- {k}: {v} ({p})")
        org = stats.get("organos_normalizados") or {}
        if org:
            L.append("")
            L.append("# Distribución Anatómica")
            L.append(f"- (Total con dato: {stats.get('organos_total_con_dato', '?')} · categorías distintas: {stats.get('organos_categorias_distintas', '?')})")
            L += lst(org)
        dx = stats.get("diagnosticos_categorizados") or {}
        if dx:
            L.append("")
            L.append("# Diagnósticos Oncológicos Principales (neoplasias)")
            L.append(f"- (Total categorizado como oncológico: {stats.get('diagnosticos_total_categorizado', '?')})")
            L += lst(dx)
        hno = stats.get("hallazgos_no_neoplasicos") or {}
        if hno:
            L.append("")
            L.append("# Hallazgos No-Neoplásicos (sección separada — no son neoplasias)")
            L.append(f"- (Total: {stats.get('hallazgos_no_neoplasicos_total', '?')})")
            L += lst(hno)
        bios = stats.get("biomarcadores_top15") or {}
        if bios:
            L.append("")
            L.append("# Biomarcadores (top 15 por N evaluados)")
            L.append(f"- (Biomarcadores distintos con dato: {stats.get('total_biomarcadores_distintos', '?')})")
            for marc, info in sorted(bios.items(), key=lambda x: x[1].get("n", 0), reverse=True):
                top = info.get("top") or {}
                if top:
                    val, n = next(iter(top.items()))
                    pred = f"{val} (N={n})"
                else:
                    pred = "—"
                L.append(f"- {str(marc).replace('IHQ_', '')}: N={info.get('n', 0)}, predominante {pred}")
        serv = stats.get("servicios") or {}
        if serv:
            L.append("")
            L.append("# Servicios Solicitantes (top 10)")
            L.append(f"- (Total con dato: {stats.get('servicios_total_con_dato', '?')} · servicios distintos: {stats.get('servicios_categorias_distintas', '?')})")
            L += lst(serv)
        ptipo = stats.get("procedimientos_por_tipo") or {}
        if ptipo:
            L.append("")
            L.append("# Procedimientos por Tipo")
            L += lst(ptipo)
            ptop = stats.get("procedimientos") or {}
            if ptop:
                L.append("- Específicos más frecuentes:")
                L += [f"  - {k}: {v}" for k, v in list(ptop.items())[:10]]
        L.append("")
        L.append("# Control de Calidad de Categorización")
        L.append(f"- Sin categorizar (otro): {stats.get('diagnosticos_otro_no_categorizado', 0)}")
        L.append(f"- Estudio IHQ sin diagnóstico específico: {stats.get('diagnosticos_estudio_ihq_sin_dx', 0)}")
        L.append(f"- Sin dato: {stats.get('diagnosticos_sin_dato', 0)}")
        return "\n".join(L)

    def _resumen_ia_worker(self):
        """Hilo: compila estadísticas y arma el resumen.
        V6.9.19 HÍBRIDO: secciones factuales deterministas + IA solo para prosa."""
        import json
        import re
        try:
            stats = self._compilar_estadisticas()
            # Guardamos las estadísticas crudas para el dashboard de gráficos
            self._resumen_ia_result["stats"] = stats
            # 1) Secciones FACTUALES deterministas (fuente de verdad: cifras exactas)
            factual = self._construir_resumen_factual(stats)

            # 2) Contexto compacto para la IA (solo cifras clave -> prompt pequeño)
            total = stats.get("total_casos", 0)
            def _top(d, n=5):
                return ", ".join(f"{k} ({v})" for k, v in list((d or {}).items())[:n]) or "no disponible"
            contexto = (
                f"Total casos IHQ: {total}\n"
                f"Periodo: {stats.get('fecha_min', '?')} a {stats.get('fecha_max', '?')}\n"
                f"Malignidad: {_top(stats.get('malignidad'), 6)}\n"
                f"Top organos: {_top(stats.get('organos_normalizados'))}\n"
                f"Top diagnosticos oncologicos: {_top(stats.get('diagnosticos_categorizados'))}\n"
                f"Diagnosticos oncologicos totales: {stats.get('diagnosticos_total_categorizado', '?')}\n"
                f"Hallazgos no-neoplasicos totales: {stats.get('hallazgos_no_neoplasicos_total', '?')}\n"
                f"Biomarcadores distintos: {stats.get('total_biomarcadores_distintos', '?')}\n"
            )
            system_prompt = (
                "Eres analista clínico-oncológico del Hospital Universitario del Valle. "
                "Te entrego cifras YA CALCULADAS y EXACTAS de un periodo de patología "
                "(inmunohistoquímica). Escribe en español, SIN inventar ni recalcular "
                "cifras y SIN repetir estas instrucciones. Devuelve EXACTAMENTE estos dos "
                "bloques en Markdown y NADA MÁS:\n\n"
                "# Resumen Ejecutivo\n"
                "- (3 a 4 viñetas con lo más relevante: volumen y periodo, malignidad, "
                "órganos y diagnósticos predominantes)\n\n"
                "# Observaciones clínicas\n"
                "- (2 a 3 frases prudentes de interpretación general, sin sobre-interpretar "
                "ni diagnosticar casos individuales)\n\n"
                "No agregues otras secciones: el resto del informe ya está hecho."
            )

            exec_block, obs_block = "", ""
            try:
                from core.llm_client import LMStudioClient
                client = LMStudioClient(timeout=900)
                resultado = client.completar(
                    prompt=f"Cifras del periodo:\n{contexto}",
                    system_prompt=system_prompt,
                    temperature=0.3,
                    max_tokens=700,
                )
                if resultado.get("exito"):
                    llm = (resultado.get("respuesta") or "").strip()
                    partes = re.split(r'(?im)^\s*#+\s*Observaciones', llm, maxsplit=1)
                    exec_block = partes[0].strip()
                    if len(partes) > 1:
                        obs_block = "# Observaciones" + partes[1].rstrip()
                    if exec_block and not re.match(r'(?im)^\s*#', exec_block):
                        exec_block = "# Resumen Ejecutivo\n" + exec_block
            except Exception:
                pass  # la IA es opcional: si falla, entregamos el informe factual

            # 3) Fallback determinista del Resumen Ejecutivo si la IA no respondió
            if not exec_block:
                org0 = next(iter((stats.get("organos_normalizados", {}) or {}).items()), ("no disponible", 0))
                dx0 = next(iter((stats.get("diagnosticos_categorizados", {}) or {}).items()), ("no disponible", 0))
                exec_block = "\n".join([
                    "# Resumen Ejecutivo",
                    f"- Volumen: {total} casos de inmunohistoquímica ({stats.get('fecha_min', '?')} a {stats.get('fecha_max', '?')}).",
                    f"- Diagnósticos categorizados como oncológicos: {stats.get('diagnosticos_total_categorizado', '?')}; hallazgos no-neoplásicos: {stats.get('hallazgos_no_neoplasicos_total', '?')}.",
                    f"- Órgano más frecuente: {org0[0]} ({org0[1]}). Diagnóstico oncológico más frecuente: {dx0[0]} ({dx0[1]}).",
                ])

            # 4) Ensamblar: Resumen Ejecutivo (IA) -> Secciones factuales -> Observaciones (IA)
            secciones = [exec_block, factual]
            if obs_block:
                secciones.append(obs_block)
            self._resumen_ia_result["texto"] = "\n\n".join(s for s in secciones if s)

        except Exception as e:
            self._resumen_ia_result["error"] = str(e)
        finally:
            self._resumen_ia_result["done"] = True

    def _poll_resumen_ia(self):
        """Polling main thread: espera a que el hilo de resumen IA termine."""
        if not self._resumen_ia_result["done"]:
            self.after(500, self._poll_resumen_ia)
            return

        # Cerrar overlay
        if hasattr(self, '_resumen_ia_overlay') and self._resumen_ia_overlay:
            self._resumen_ia_overlay.grab_release()
            self._resumen_ia_overlay.destroy()
            self._resumen_ia_overlay = None

        error = self._resumen_ia_result.get("error")
        if error:
            messagebox.showerror("Error Resumen IA", f"No se pudo generar el resumen:\n{error}")
            return

        texto = self._resumen_ia_result["texto"]
        stats = self._resumen_ia_result.get("stats", {})
        self._mostrar_ventana_resumen_ia(texto, stats)

    def _mostrar_ventana_resumen_ia(self, texto: str, stats: dict | None = None):
        """Muestra el resumen IA con dashboard: Informe + Gráficos + Tablas."""
        stats = stats or {}
        win = tk.Toplevel(self)
        win.title("📊 Resumen IA — Base de Datos IHQ")
        win.geometry("1200x800")
        win.transient(self)

        # ── Cabecera ──────────────────────────────────────────────
        top_frame = ttk.Frame(win, padding=10)
        top_frame.pack(fill=X)

        ttk.Label(
            top_frame,
            text="📊 Resumen Profesional generado por IA",
            font=("Segoe UI", 14, "bold"),
        ).pack(side=LEFT)

        ttk.Button(
            top_frame,
            text="💾 Exportar",
            command=lambda: self._exportar_resumen_ia(texto, stats, win),
            bootstyle="success",
        ).pack(side=RIGHT, padx=5)

        ttk.Button(
            top_frame,
            text="📋 Copiar",
            command=lambda: self._copiar_resumen_ia(texto, win),
            bootstyle="info",
        ).pack(side=RIGHT, padx=5)

        # ── KPI cards ─────────────────────────────────────────────
        kpi_frame = ttk.Frame(win, padding=(10, 0, 10, 10))
        kpi_frame.pack(fill=X)
        self._render_kpi_cards(kpi_frame, stats)

        # ── Notebook con pestañas ─────────────────────────────────
        notebook = ttk.Notebook(win)
        notebook.pack(fill=BOTH, expand=True, padx=10, pady=(0, 10))

        # Tab 1 — Informe IA (markdown con formato)
        tab_informe = ttk.Frame(notebook)
        notebook.add(tab_informe, text="📝 Informe IA")
        self._render_informe_markdown(tab_informe, texto)

        # Tab 2 — Gráficos (matplotlib)
        tab_graficos = ttk.Frame(notebook)
        notebook.add(tab_graficos, text="📊 Gráficos")
        self._render_graficos_dashboard(tab_graficos, stats)

        # Tab 3 — Tablas (Treeviews limpios)
        tab_tablas = ttk.Frame(notebook)
        notebook.add(tab_tablas, text="📋 Tablas")
        self._render_tablas_dashboard(tab_tablas, stats)

    # ──────────────────────────────────────────────────────────────
    # Helpers de presentación del Resumen IA
    # ──────────────────────────────────────────────────────────────

    def _render_kpi_cards(self, parent, stats: dict):
        """Tarjetas con métricas clave en la parte superior."""
        if not stats:
            return

        total = stats.get("total_casos", 0)
        malig = stats.get("malignidad", {}) or {}
        n_malig = sum(v for k, v in malig.items() if "MALIGN" in str(k).upper() and "BENIG" not in str(k).upper())
        # Si no se detectó por keyword, usa la primera clave
        if n_malig == 0 and malig:
            primera = next(iter(malig.keys()))
            if "MALIGN" in primera.upper() and "BENIG" not in primera.upper():
                n_malig = malig[primera]
        pct_malig = (n_malig / total * 100) if total else 0

        organos_dist = stats.get("organos_categorias_distintas", 0)
        bio_total = stats.get("total_biomarcadores_distintos", 0)
        dx_cat = stats.get("diagnosticos_total_categorizado", 0)

        cards = [
            ("Total casos", f"{total:,}", "primary"),
            ("% Malignos", f"{pct_malig:.1f}%", "danger"),
            ("Categorías anatómicas", f"{organos_dist}", "info"),
            ("Biomarcadores distintos", f"{bio_total}", "warning"),
            ("Diagnósticos categorizados", f"{dx_cat:,}", "success"),
        ]

        for titulo, valor, estilo in cards:
            card = ttk.Frame(parent, padding=10, bootstyle=estilo)
            card.pack(side=LEFT, fill=BOTH, expand=True, padx=4)
            ttk.Label(
                card, text=titulo, font=("Segoe UI", 9),
                bootstyle=f"inverse-{estilo}",
            ).pack(anchor="w")
            ttk.Label(
                card, text=valor, font=("Segoe UI", 18, "bold"),
                bootstyle=f"inverse-{estilo}",
            ).pack(anchor="w")

    def _render_informe_markdown(self, parent, texto: str):
        """Renderiza el markdown con tipografía y colores básicos."""
        text_frame = ttk.Frame(parent, padding=10)
        text_frame.pack(fill=BOTH, expand=True)

        text_widget = tk.Text(
            text_frame,
            wrap=tk.WORD,
            font=("Segoe UI", 11),
            padx=20,
            pady=15,
            relief="flat",
            borderwidth=0,
            spacing1=4,
            spacing3=4,
        )
        scrollbar = ttk.Scrollbar(text_frame, orient=tk.VERTICAL, command=text_widget.yview)
        text_widget.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side=RIGHT, fill=Y)
        text_widget.pack(fill=BOTH, expand=True)

        # Tags de estilo
        text_widget.tag_configure("h1", font=("Segoe UI", 18, "bold"), foreground="#0d6efd",
                                  spacing1=12, spacing3=8)
        text_widget.tag_configure("h2", font=("Segoe UI", 14, "bold"), foreground="#198754",
                                  spacing1=10, spacing3=6)
        text_widget.tag_configure("h3", font=("Segoe UI", 12, "bold"), foreground="#6c757d",
                                  spacing1=8, spacing3=4)
        text_widget.tag_configure("bold", font=("Segoe UI", 11, "bold"))
        text_widget.tag_configure("bullet", lmargin1=20, lmargin2=40)
        text_widget.tag_configure("table", font=("Consolas", 10), foreground="#212529")
        text_widget.tag_configure("hr", foreground="#dee2e6")

        # Parser sencillo de markdown línea a línea
        import re as _re
        for raw in texto.splitlines():
            linea = raw.rstrip()
            if linea.startswith("# "):
                text_widget.insert("end", linea[2:] + "\n", "h1")
            elif linea.startswith("## "):
                text_widget.insert("end", linea[3:] + "\n", "h2")
            elif linea.startswith("### "):
                text_widget.insert("end", linea[4:] + "\n", "h3")
            elif linea.strip() in {"---", "***"}:
                text_widget.insert("end", "─" * 80 + "\n", "hr")
            elif linea.lstrip().startswith(("* ", "- ")):
                text_widget.insert("end", "  • " + linea.lstrip()[2:] + "\n", "bullet")
            elif linea.startswith("|"):
                text_widget.insert("end", linea + "\n", "table")
            else:
                # Sustituye **negritas** por tag
                pos = 0
                for m in _re.finditer(r"\*\*(.+?)\*\*", linea):
                    text_widget.insert("end", linea[pos:m.start()])
                    text_widget.insert("end", m.group(1), "bold")
                    pos = m.end()
                text_widget.insert("end", linea[pos:] + "\n")

        text_widget.configure(state="disabled")

    def _render_graficos_dashboard(self, parent, stats: dict):
        """Tablero matplotlib con 4 gráficos principales."""
        if not stats:
            ttk.Label(parent, text="No hay estadísticas disponibles para graficar.",
                      font=("Segoe UI", 11)).pack(pady=40)
            return

        try:
            import matplotlib
            matplotlib.use("TkAgg")
            from matplotlib.figure import Figure
            from matplotlib.backends.backend_tkagg import (
                FigureCanvasTkAgg, NavigationToolbar2Tk,
            )
        except Exception as e:  # matplotlib no disponible
            ttk.Label(
                parent,
                text=f"Matplotlib no disponible: {e}",
                font=("Segoe UI", 10),
            ).pack(pady=40)
            return

        fig = Figure(figsize=(12, 9), dpi=100, tight_layout=True)
        fig.patch.set_facecolor("#ffffff")

        # 1) Malignidad — pie (V6.9.20: agrupa rebanadas pequeñas (<4%) en OTROS
        # para evitar etiquetas montadas; antes el top-5 incluía slices <2% solapados)
        ax1 = fig.add_subplot(2, 2, 1)
        malig = stats.get("malignidad", {}) or {}
        if malig:
            total_m = sum(malig.values()) or 1
            grandes, otros = {}, 0
            for k, v in sorted(malig.items(), key=lambda x: x[1], reverse=True):
                if v / total_m >= 0.04:
                    grandes[k] = v
                else:
                    otros += v
            if otros:
                grandes["OTROS"] = otros
            etiquetas = list(grandes.keys())
            valores = list(grandes.values())
            paleta = ["#dc3545", "#198754", "#ffc107", "#6c757d", "#0dcaf0", "#2d3e5e"]
            colores = [paleta[i % len(paleta)] for i in range(len(etiquetas))]
            ax1.pie(valores, labels=etiquetas, autopct="%1.1f%%",
                    colors=colores, startangle=90, pctdistance=0.78,
                    labeldistance=1.06, textprops={"fontsize": 8})
            ax1.set_title("Malignidad", fontsize=12, fontweight="bold", pad=12)
        else:
            ax1.text(0.5, 0.5, "Sin datos", ha="center", va="center")
            ax1.axis("off")

        # 2) Top órganos — barra horizontal
        ax2 = fig.add_subplot(2, 2, 2)
        organos = stats.get("organos_normalizados", {}) or {}
        if organos:
            top = list(organos.items())[:10]
            top.reverse()
            etq = [k for k, _ in top]
            val = [v for _, v in top]
            ax2.barh(etq, val, color="#0d6efd")
            ax2.set_title("Top 10 órganos (canónicos)", fontsize=12, fontweight="bold")
            ax2.tick_params(axis="y", labelsize=9)
            for i, v in enumerate(val):
                ax2.text(v, i, f" {v}", va="center", fontsize=8)
        else:
            ax2.text(0.5, 0.5, "Sin datos", ha="center", va="center")
            ax2.axis("off")

        # 3) Top diagnósticos oncológicos (V6.6.9: solo neoplasias, sin hallazgos
        # no-neoplásicos como "negativo para malignidad")
        ax3 = fig.add_subplot(2, 2, 3)
        dx = stats.get("diagnosticos_categorizados", {}) or {}
        # Excluir "OTRO / NO CATEGORIZADO" del gráfico para ver categorías clínicas
        dx = {k: v for k, v in dx.items() if k != "OTRO / NO CATEGORIZADO"}
        if dx:
            top = list(dx.items())[:10]
            top.reverse()
            etq = [k[:40] + ("…" if len(k) > 40 else "") for k, _ in top]
            val = [v for _, v in top]
            ax3.barh(etq, val, color="#198754")
            ax3.set_title("Top 10 diagnósticos oncológicos", fontsize=12, fontweight="bold")
            ax3.tick_params(axis="y", labelsize=8)
            for i, v in enumerate(val):
                ax3.text(v, i, f" {v}", va="center", fontsize=8)
        else:
            ax3.text(0.5, 0.5, "Sin datos", ha="center", va="center")
            ax3.axis("off")

        # 4) Top biomarcadores — barra horizontal
        ax4 = fig.add_subplot(2, 2, 4)
        bios = stats.get("biomarcadores_top15", {}) or {}
        if bios:
            items = sorted(bios.items(), key=lambda x: x[1].get("n", 0), reverse=True)[:10]
            items.reverse()
            etq = [k.replace("IHQ_", "")[:25] for k, _ in items]
            val = [v.get("n", 0) for _, v in items]
            ax4.barh(etq, val, color="#fd7e14")
            ax4.set_title("Top 10 biomarcadores (N evaluados)", fontsize=12, fontweight="bold")
            ax4.tick_params(axis="y", labelsize=9)
            for i, v in enumerate(val):
                ax4.text(v, i, f" {v}", va="center", fontsize=8)
        else:
            ax4.text(0.5, 0.5, "Sin datos", ha="center", va="center")
            ax4.axis("off")

        canvas = FigureCanvasTkAgg(fig, master=parent)
        canvas.draw()
        canvas.get_tk_widget().pack(fill=BOTH, expand=True, padx=8, pady=8)

        toolbar_frame = ttk.Frame(parent)
        toolbar_frame.pack(fill=X, padx=8)
        toolbar = NavigationToolbar2Tk(canvas, toolbar_frame)
        toolbar.update()

    def _copiar_tabla_resumen(self, tv, headers, btn=None):
        """V6.9.19: Copia el contenido de un Treeview como TSV (pegable en Excel/Word).
        Incluye encabezados; da feedback breve en el botón."""
        lineas = ["\t".join(str(h) for h in headers)]
        for iid in tv.get_children():
            lineas.append("\t".join(str(v) for v in tv.item(iid, "values")))
        try:
            self.clipboard_clear()
            self.clipboard_append("\n".join(lineas))
        except Exception:
            pass
        if btn is not None:
            try:
                _orig = btn.cget("text")
                btn.configure(text="✓ Copiado")
                btn.after(1300, lambda: btn.configure(text=_orig))
            except Exception:
                pass

    def _render_tablas_dashboard(self, parent, stats: dict):
        """Tablas claras con scroll para cada bloque del informe."""
        if not stats:
            ttk.Label(parent, text="No hay estadísticas disponibles.",
                      font=("Segoe UI", 11)).pack(pady=40)
            return

        # Scroll global
        canvas = tk.Canvas(parent, highlightthickness=0)
        vbar = ttk.Scrollbar(parent, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=vbar.set)
        vbar.pack(side=RIGHT, fill=Y)
        canvas.pack(side=LEFT, fill=BOTH, expand=True)

        inner = ttk.Frame(canvas, padding=10)
        canvas_window = canvas.create_window((0, 0), window=inner, anchor="nw")

        def _on_inner_config(_e=None):
            canvas.configure(scrollregion=canvas.bbox("all"))

        def _on_canvas_config(e):
            canvas.itemconfigure(canvas_window, width=e.width)

        inner.bind("<Configure>", _on_inner_config)
        canvas.bind("<Configure>", _on_canvas_config)

        bloques = [
            ("🧬 Distribución anatómica", stats.get("organos_normalizados", {}),
             ("Órgano", "Casos")),
            ("🩺 Diagnósticos oncológicos (neoplasias)",
             stats.get("diagnosticos_categorizados", {}),
             ("Categoría neoplásica", "Casos")),
            ("📋 Hallazgos no-neoplásicos (separados)",
             stats.get("hallazgos_no_neoplasicos", {}),
             ("Categoría", "Casos")),
            ("⚕️ Procedimientos", stats.get("procedimientos", {}),
             ("Procedimiento", "Casos")),
            ("🏥 Servicios solicitantes", stats.get("servicios", {}),
             ("Servicio", "Casos")),
        ]

        for titulo, data, (col1, col2) in bloques:
            if not data:
                continue
            hdr = ttk.Frame(inner)
            hdr.pack(fill=X, pady=(10, 4))
            ttk.Label(hdr, text=titulo, font=("Segoe UI", 12, "bold"),
                      bootstyle="primary").pack(side=LEFT, anchor="w")

            tv = ttk.Treeview(
                inner, columns=("c1", "c2"), show="headings",
                height=min(12, len(data)),
                bootstyle="primary",
            )
            tv.heading("c1", text=col1)
            tv.heading("c2", text=col2)
            tv.column("c1", anchor="w", width=400)
            tv.column("c2", anchor="e", width=100)
            for k, v in data.items():
                tv.insert("", "end", values=(k, v))
            # V6.9.22: fila TOTAL (suma de la columna de casos)
            try:
                tv.insert("", "end", values=("TOTAL", sum(int(x) for x in data.values())), tags=("tot",))
                tv.tag_configure("tot", font=("Segoe UI Semibold", 10))
            except Exception:
                pass

            btn_cp = ttk.Button(hdr, text="📋 Copiar", bootstyle="secondary-outline", width=12)
            btn_cp.configure(command=lambda t=tv, h=(col1, col2), b=btn_cp: self._copiar_tabla_resumen(t, h, b))
            btn_cp.pack(side=RIGHT)

            tv.pack(fill=X, pady=(0, 6))

        # Bloque biomarcadores con N y resultado predominante
        bios = stats.get("biomarcadores_top15", {}) or {}
        if bios:
            hdr_bio = ttk.Frame(inner)
            hdr_bio.pack(fill=X, pady=(10, 4))
            ttk.Label(hdr_bio, text="🔬 Biomarcadores (top 15)",
                      font=("Segoe UI", 12, "bold"),
                      bootstyle="primary").pack(side=LEFT, anchor="w")
            tv = ttk.Treeview(
                inner, columns=("m", "n", "top"),
                show="headings",
                height=min(15, len(bios)),
                bootstyle="primary",
            )
            tv.heading("m", text="Biomarcador")
            tv.heading("n", text="N evaluados")
            tv.heading("top", text="Resultado predominante")
            tv.column("m", anchor="w", width=240)
            tv.column("n", anchor="e", width=110)
            tv.column("top", anchor="w", width=420)
            for marcador, info in sorted(bios.items(), key=lambda x: x[1].get("n", 0), reverse=True):
                top = info.get("top", {}) or {}
                if top:
                    primer_valor, primer_n = next(iter(top.items()))
                    resumen = f"{primer_valor} (N={primer_n})"
                else:
                    resumen = "—"
                tv.insert("", "end", values=(marcador, info.get("n", 0), resumen))
            # V6.9.22: fila TOTAL (suma de N evaluados de los biomarcadores mostrados)
            try:
                tv.insert("", "end", values=("TOTAL", sum(int(i.get("n", 0)) for i in bios.values()), ""), tags=("tot",))
                tv.tag_configure("tot", font=("Segoe UI Semibold", 10))
            except Exception:
                pass
            btn_bio = ttk.Button(hdr_bio, text="📋 Copiar", bootstyle="secondary-outline", width=12)
            btn_bio.configure(command=lambda t=tv, b=btn_bio: self._copiar_tabla_resumen(
                t, ("Biomarcador", "N evaluados", "Resultado predominante"), b))
            btn_bio.pack(side=RIGHT)
            tv.pack(fill=X, pady=(0, 6))

    def _exportar_resumen_ia(self, texto: str, stats: dict | None = None, parent=None):
        """Permite elegir entre Excel formateado o PDF profesional."""
        stats = stats or {}
        dlg = tk.Toplevel(parent or self)
        dlg.title("Exportar Resumen IA")
        dlg.geometry("420x200")
        dlg.transient(parent or self)
        dlg.grab_set()
        dlg.resizable(False, False)

        ttk.Label(
            dlg, text="¿En qué formato deseas exportar el resumen?",
            font=("Segoe UI", 11, "bold"), padding=15,
        ).pack()

        btn_frame = ttk.Frame(dlg, padding=10)
        btn_frame.pack(fill=BOTH, expand=True)

        def _hacer_excel():
            dlg.destroy()
            self._exportar_resumen_excel(texto, stats)

        def _hacer_pdf():
            dlg.destroy()
            self._exportar_resumen_pdf(texto, stats)

        ttk.Button(
            btn_frame, text="📊  Excel (.xlsx)\ncon tablas y gráficos",
            bootstyle="success", command=_hacer_excel, width=22,
        ).pack(side=LEFT, expand=True, padx=8, ipady=8)

        ttk.Button(
            btn_frame, text="📄  PDF\nInforme profesional",
            bootstyle="danger", command=_hacer_pdf, width=22,
        ).pack(side=LEFT, expand=True, padx=8, ipady=8)

    # ── Excel ─────────────────────────────────────────────────────────
    def _exportar_resumen_excel(self, texto: str, stats: dict):
        """Genera un .xlsx con KPIs, tablas formateadas y gráficos nativos."""
        fecha_str = datetime.now().strftime("%Y-%m-%d_%H%M%S")
        filepath = filedialog.asksaveasfilename(
            title="Exportar Resumen IA — Excel",
            defaultextension=".xlsx",
            initialfile=f"Resumen_IA_HUV_{fecha_str}.xlsx",
            filetypes=[("Excel", "*.xlsx")],
        )
        if not filepath:
            return

        try:
            from openpyxl import Workbook
            from openpyxl.styles import (
                Font, PatternFill, Alignment, Border, Side,
            )
            from openpyxl.chart import BarChart, PieChart, Reference
            from openpyxl.utils import get_column_letter
        except Exception as e:
            messagebox.showerror("Error", f"openpyxl no disponible: {e}")
            return

        wb = Workbook()

        # Estilos reutilizables
        azul_huv = "0D6EFD"
        verde_huv = "198754"
        gris_claro = "F1F3F5"
        thin = Side(style="thin", color="DEE2E6")
        borde = Border(top=thin, bottom=thin, left=thin, right=thin)
        font_titulo = Font(name="Calibri", size=18, bold=True, color="FFFFFF")
        fill_titulo = PatternFill("solid", fgColor=azul_huv)
        font_h2 = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
        fill_h2 = PatternFill("solid", fgColor=verde_huv)
        font_th = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        fill_th = PatternFill("solid", fgColor=azul_huv)
        fill_alt = PatternFill("solid", fgColor=gris_claro)
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left = Alignment(horizontal="left", vertical="center", wrap_text=True)
        right = Alignment(horizontal="right", vertical="center")

        def _aplicar_tabla(ws, fila_ini, col_ini, headers, filas):
            for i, h in enumerate(headers):
                c = ws.cell(row=fila_ini, column=col_ini + i, value=h)
                c.font = font_th
                c.fill = fill_th
                c.alignment = center
                c.border = borde
            for r, row in enumerate(filas, start=1):
                for j, val in enumerate(row):
                    c = ws.cell(row=fila_ini + r, column=col_ini + j, value=val)
                    c.alignment = right if isinstance(val, (int, float)) else left
                    c.border = borde
                    if r % 2 == 0:
                        c.fill = fill_alt
            return fila_ini + len(filas)

        # ── Hoja 1: Resumen Ejecutivo (KPIs) ──────────────────────
        ws1 = wb.active
        ws1.title = "Resumen Ejecutivo"
        ws1.merge_cells("A1:F1")
        c = ws1["A1"]
        c.value = "📊 RESUMEN IA — BASE DE DATOS IHQ HUV"
        c.font = font_titulo
        c.fill = fill_titulo
        c.alignment = center
        ws1.row_dimensions[1].height = 36

        ws1["A2"] = "Generado:"
        ws1["B2"] = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
        ws1["A3"] = "Periodo cubierto:"
        ws1["B3"] = f"{stats.get('fecha_min', '—')} → {stats.get('fecha_max', '—')}"
        for r in (2, 3):
            ws1.cell(row=r, column=1).font = Font(bold=True)

        # KPIs
        total = stats.get("total_casos", 0)
        malig = stats.get("malignidad", {}) or {}
        n_malig = next(
            (v for k, v in malig.items()
             if "MALIGN" in str(k).upper() and "BENIG" not in str(k).upper()),
            0,
        )
        pct_malig = (n_malig / total * 100) if total else 0
        kpis = [
            ("Total casos", total),
            ("% Malignos", f"{pct_malig:.1f}%"),
            ("Categorías anatómicas", stats.get("organos_categorias_distintas", 0)),
            ("Biomarcadores distintos", stats.get("total_biomarcadores_distintos", 0)),
            ("Diagnósticos categorizados", stats.get("diagnosticos_total_categorizado", 0)),
        ]
        ws1.cell(row=5, column=1, value="Indicadores Clave").font = font_h2
        ws1.cell(row=5, column=1).fill = fill_h2
        ws1.cell(row=5, column=1).alignment = center
        ws1.merge_cells("A5:F5")
        ws1.row_dimensions[5].height = 24

        for i, (k, v) in enumerate(kpis):
            row = 6 + i
            a = ws1.cell(row=row, column=1, value=k)
            b = ws1.cell(row=row, column=2, value=v)
            a.font = Font(bold=True)
            a.fill = fill_alt
            a.border = borde
            b.border = borde
            b.alignment = right
            ws1.row_dimensions[row].height = 20

        for col in range(1, 7):
            ws1.column_dimensions[get_column_letter(col)].width = 24

        # ── Hoja 2: Distribución Anatómica ─────────────────────────
        ws2 = wb.create_sheet("Distribución Anatómica")
        ws2.merge_cells("A1:C1")
        ws2["A1"] = "🧬 Top categorías anatómicas (canónicas)"
        ws2["A1"].font = font_titulo
        ws2["A1"].fill = fill_titulo
        ws2["A1"].alignment = center
        ws2.row_dimensions[1].height = 30
        organos = stats.get("organos_normalizados", {}) or {}
        filas = [(k, v) for k, v in organos.items()]
        end_row = _aplicar_tabla(ws2, 3, 1, ["Órgano", "Casos"], filas)
        ws2.column_dimensions["A"].width = 38
        ws2.column_dimensions["B"].width = 14

        if filas:
            chart = BarChart()
            chart.type = "bar"
            chart.style = 11
            chart.title = "Top órganos"
            chart.y_axis.title = "Órgano"
            chart.x_axis.title = "Casos"
            data = Reference(ws2, min_col=2, min_row=3, max_row=end_row, max_col=2)
            cats = Reference(ws2, min_col=1, min_row=4, max_row=end_row)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            chart.height = 12
            chart.width = 20
            ws2.add_chart(chart, "E3")

        # ── Hoja 3: Diagnósticos ──────────────────────────────────
        ws3 = wb.create_sheet("Diagnósticos")
        ws3.merge_cells("A1:C1")
        ws3["A1"] = "🩺 Diagnósticos por categoría clínica"
        ws3["A1"].font = font_titulo
        ws3["A1"].fill = fill_titulo
        ws3["A1"].alignment = center
        ws3.row_dimensions[1].height = 30
        dx = stats.get("diagnosticos_categorizados", {}) or {}
        filas = [(k, v) for k, v in dx.items() if k != "OTRO / NO CATEGORIZADO"]
        end_row = _aplicar_tabla(ws3, 3, 1, ["Categoría clínica", "Casos"], filas)
        ws3.column_dimensions["A"].width = 50
        ws3.column_dimensions["B"].width = 14

        # Notas adicionales
        nota_row = end_row + 2
        ws3.cell(row=nota_row, column=1, value="Sin diagnóstico específico (Estudio IHQ):").font = Font(bold=True, italic=True)
        ws3.cell(row=nota_row, column=2, value=stats.get("diagnosticos_estudio_ihq_sin_dx", 0))
        ws3.cell(row=nota_row + 1, column=1, value="Otros (no categorizados):").font = Font(bold=True, italic=True)
        ws3.cell(row=nota_row + 1, column=2, value=stats.get("diagnosticos_otro_no_categorizado", 0))
        ws3.cell(row=nota_row + 2, column=1, value="Sin dato:").font = Font(bold=True, italic=True)
        ws3.cell(row=nota_row + 2, column=2, value=stats.get("diagnosticos_sin_dato", 0))

        if filas:
            chart = BarChart()
            chart.type = "bar"
            chart.style = 12
            chart.title = "Top diagnósticos clínicos"
            data = Reference(ws3, min_col=2, min_row=3, max_row=end_row, max_col=2)
            cats = Reference(ws3, min_col=1, min_row=4, max_row=end_row)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            chart.height = 12
            chart.width = 22
            ws3.add_chart(chart, "E3")

        # ── Hoja 4: Biomarcadores ─────────────────────────────────
        ws4 = wb.create_sheet("Biomarcadores")
        ws4.merge_cells("A1:D1")
        ws4["A1"] = "🔬 Top biomarcadores"
        ws4["A1"].font = font_titulo
        ws4["A1"].fill = fill_titulo
        ws4["A1"].alignment = center
        ws4.row_dimensions[1].height = 30
        bios = stats.get("biomarcadores_top15", {}) or {}
        filas = []
        for marcador, info in sorted(bios.items(), key=lambda x: x[1].get("n", 0), reverse=True):
            top = info.get("top", {}) or {}
            if top:
                primer_valor, primer_n = next(iter(top.items()))
                resumen = f"{primer_valor} (N={primer_n})"
            else:
                resumen = "—"
            filas.append((marcador, info.get("n", 0), resumen))
        end_row = _aplicar_tabla(
            ws4, 3, 1,
            ["Biomarcador", "N evaluados", "Resultado predominante"], filas,
        )
        ws4.column_dimensions["A"].width = 32
        ws4.column_dimensions["B"].width = 14
        ws4.column_dimensions["C"].width = 42

        if filas:
            chart = BarChart()
            chart.type = "bar"
            chart.style = 13
            chart.title = "Biomarcadores por N evaluados"
            data = Reference(ws4, min_col=2, min_row=3, max_row=end_row, max_col=2)
            cats = Reference(ws4, min_col=1, min_row=4, max_row=end_row)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            chart.height = 12
            chart.width = 22
            ws4.add_chart(chart, "E3")

        # ── Hoja 5: Procedimientos & Servicios ────────────────────
        ws5 = wb.create_sheet("Procedimientos & Servicios")
        ws5.merge_cells("A1:E1")
        ws5["A1"] = "⚕️ Procedimientos y servicios solicitantes"
        ws5["A1"].font = font_titulo
        ws5["A1"].fill = fill_titulo
        ws5["A1"].alignment = center
        ws5.row_dimensions[1].height = 30
        procs = stats.get("procedimientos", {}) or {}
        servs = stats.get("servicios", {}) or {}
        _aplicar_tabla(ws5, 3, 1, ["Procedimiento", "Casos"],
                       [(k, v) for k, v in procs.items()])
        _aplicar_tabla(ws5, 3, 4, ["Servicio", "Casos"],
                       [(k, v) for k, v in servs.items()])
        ws5.column_dimensions["A"].width = 32
        ws5.column_dimensions["B"].width = 12
        ws5.column_dimensions["D"].width = 38
        ws5.column_dimensions["E"].width = 12

        # Pie de malignidad
        if malig:
            ws5.cell(row=20, column=1, value="Distribución de malignidad").font = font_h2
            ws5.cell(row=20, column=1).fill = fill_h2
            ws5.cell(row=20, column=1).alignment = center
            ws5.merge_cells("A20:B20")
            mfilas = [(k, v) for k, v in malig.items()]
            mend = _aplicar_tabla(ws5, 21, 1, ["Categoría", "Casos"], mfilas)
            chart = PieChart()
            chart.title = "Malignidad"
            chart.style = 10
            data = Reference(ws5, min_col=2, min_row=21, max_row=mend)
            cats = Reference(ws5, min_col=1, min_row=22, max_row=mend)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            chart.height = 9
            chart.width = 13
            ws5.add_chart(chart, "D20")

        # ── Hoja 6: Informe IA en texto ───────────────────────────
        ws6 = wb.create_sheet("Informe IA")
        ws6.merge_cells("A1:H1")
        ws6["A1"] = "📝 Informe redactado por IA"
        ws6["A1"].font = font_titulo
        ws6["A1"].fill = fill_titulo
        ws6["A1"].alignment = center
        ws6.row_dimensions[1].height = 30
        ws6.column_dimensions["A"].width = 110

        for i, linea in enumerate(texto.splitlines(), start=3):
            cell = ws6.cell(row=i, column=1, value=linea)
            if linea.startswith("# "):
                cell.font = Font(size=14, bold=True, color=azul_huv)
                cell.value = linea[2:]
            elif linea.startswith("## "):
                cell.font = Font(size=12, bold=True, color=verde_huv)
                cell.value = linea[3:]
            elif linea.startswith("### "):
                cell.font = Font(size=11, bold=True)
                cell.value = linea[4:]
            else:
                cell.font = Font(size=10)
            cell.alignment = Alignment(wrap_text=True, vertical="top")

        try:
            wb.save(filepath)
            messagebox.showinfo(
                "Exportación exitosa",
                f"Excel guardado en:\n{filepath}",
            )
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo guardar el Excel:\n{e}")

    # ── PDF ───────────────────────────────────────────────────────────
    def _exportar_resumen_pdf(self, texto: str, stats: dict):
        """Genera un PDF profesional con KPIs, tablas y gráficos matplotlib."""
        fecha_str = datetime.now().strftime("%Y-%m-%d_%H%M%S")
        filepath = filedialog.asksaveasfilename(
            title="Exportar Resumen IA — PDF",
            defaultextension=".pdf",
            initialfile=f"Resumen_IA_HUV_{fecha_str}.pdf",
            filetypes=[("PDF", "*.pdf")],
        )
        if not filepath:
            return

        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.lib import colors
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import cm, mm
            from reportlab.platypus import (
                SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
                PageBreak, Image,
            )
        except Exception as e:
            messagebox.showerror("Error", f"reportlab no disponible: {e}")
            return

        # Generar imágenes de gráficos con matplotlib
        chart_paths = []
        try:
            import tempfile
            import matplotlib
            matplotlib.use("Agg")
            import matplotlib.pyplot as plt

            tmpdir = tempfile.mkdtemp(prefix="huv_pdf_")

            def _save_bar(data: dict, titulo: str, color: str, fname: str, top_n=10):
                if not data:
                    return None
                items = list(data.items())[:top_n]
                items.reverse()
                etq = [str(k)[:42] for k, _ in items]
                val = [v for _, v in items]
                fig, ax = plt.subplots(figsize=(7, 4.2), dpi=150)
                ax.barh(etq, val, color=color)
                ax.set_title(titulo, fontsize=12, fontweight="bold")
                ax.tick_params(axis="y", labelsize=9)
                for i, v in enumerate(val):
                    ax.text(v, i, f" {v}", va="center", fontsize=8)
                fig.tight_layout()
                p = f"{tmpdir}/{fname}.png"
                fig.savefig(p, bbox_inches="tight")
                plt.close(fig)
                return p

            def _save_pie(data: dict, titulo: str, fname: str):
                if not data:
                    return None
                # V6.9.20: agrupar rebanadas pequeñas (<4%) en OTROS (evita etiquetas montadas)
                total_p = sum(data.values()) or 1
                grandes, otros = {}, 0
                for k, v in sorted(data.items(), key=lambda x: x[1], reverse=True):
                    if v / total_p >= 0.04:
                        grandes[k] = v
                    else:
                        otros += v
                if otros:
                    grandes["OTROS"] = otros
                etq = list(grandes.keys())
                val = list(grandes.values())
                paleta = ["#dc3545", "#198754", "#ffc107", "#6c757d", "#0dcaf0", "#2d3e5e"]
                fig, ax = plt.subplots(figsize=(5, 4.2), dpi=150)
                ax.pie(val, labels=etq, autopct="%1.1f%%",
                       colors=[paleta[i % len(paleta)] for i in range(len(etq))],
                       startangle=90, pctdistance=0.78, labeldistance=1.06,
                       textprops={"fontsize": 9})
                ax.set_title(titulo, fontsize=12, fontweight="bold", pad=12)
                fig.tight_layout()
                p = f"{tmpdir}/{fname}.png"
                fig.savefig(p, bbox_inches="tight")
                plt.close(fig)
                return p

            chart_paths.append(_save_pie(stats.get("malignidad", {}), "Malignidad", "malig"))
            chart_paths.append(_save_bar(stats.get("organos_normalizados", {}),
                                         "Top 10 órganos", "#0d6efd", "organos"))
            dx = {k: v for k, v in (stats.get("diagnosticos_categorizados", {}) or {}).items()
                  if k != "OTRO / NO CATEGORIZADO"}
            chart_paths.append(_save_bar(dx, "Top 10 diagnósticos clínicos", "#198754", "dx"))
            bios_top = {k: v.get("n", 0) for k, v in
                        sorted((stats.get("biomarcadores_top15", {}) or {}).items(),
                               key=lambda x: x[1].get("n", 0), reverse=True)}
            chart_paths.append(_save_bar(bios_top, "Top 10 biomarcadores (N evaluados)",
                                         "#fd7e14", "bios"))
        except Exception as e:
            logging.warning(f"No se pudieron generar gráficos para PDF: {e}")

        # Documento
        doc = SimpleDocTemplate(
            filepath, pagesize=A4,
            leftMargin=2*cm, rightMargin=2*cm,
            topMargin=2*cm, bottomMargin=2*cm,
            title="Resumen IA HUV", author="ONCONOVA",
        )
        styles = getSampleStyleSheet()
        h1 = ParagraphStyle("h1", parent=styles["Heading1"],
                            fontSize=20, textColor=colors.HexColor("#0D6EFD"),
                            spaceAfter=12)
        h2 = ParagraphStyle("h2", parent=styles["Heading2"],
                            fontSize=14, textColor=colors.HexColor("#198754"),
                            spaceBefore=12, spaceAfter=8)
        h3 = ParagraphStyle("h3", parent=styles["Heading3"],
                            fontSize=11, textColor=colors.HexColor("#6c757d"),
                            spaceBefore=6, spaceAfter=4)
        body = ParagraphStyle("body", parent=styles["Normal"],
                              fontSize=10, leading=14, spaceAfter=6)

        story = []

        # Portada
        story.append(Paragraph("📊 Resumen IA — Base de Datos IHQ", h1))
        story.append(Paragraph(
            "Hospital Universitario del Valle — ONCONOVA Gestor Oncológico",
            body))
        story.append(Spacer(1, 6))
        story.append(Paragraph(
            f"<b>Generado:</b> {datetime.now().strftime('%d/%m/%Y %H:%M')}<br/>"
            f"<b>Periodo cubierto:</b> {stats.get('fecha_min', '—')} → "
            f"{stats.get('fecha_max', '—')}",
            body))
        story.append(Spacer(1, 12))

        # KPIs
        total = stats.get("total_casos", 0)
        malig = stats.get("malignidad", {}) or {}
        n_malig = next(
            (v for k, v in malig.items()
             if "MALIGN" in str(k).upper() and "BENIG" not in str(k).upper()),
            0,
        )
        pct_malig = (n_malig / total * 100) if total else 0

        kpi_data = [
            ["Total casos", "% Malignos", "Cat. anatómicas",
             "Biomarcadores", "Dx categorizados"],
            [str(total), f"{pct_malig:.1f}%",
             str(stats.get("organos_categorias_distintas", 0)),
             str(stats.get("total_biomarcadores_distintos", 0)),
             str(stats.get("diagnosticos_total_categorizado", 0))],
        ]
        kpi_tbl = Table(kpi_data, colWidths=[3.4*cm]*5)
        kpi_tbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0D6EFD")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 9),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("BACKGROUND", (0, 1), (-1, 1), colors.HexColor("#F1F3F5")),
            ("FONTSIZE", (0, 1), (-1, 1), 14),
            ("FONTNAME", (0, 1), (-1, 1), "Helvetica-Bold"),
            ("TEXTCOLOR", (0, 1), (-1, 1), colors.HexColor("#212529")),
            ("ROWBACKGROUNDS", (0, 1), (-1, 1), [colors.HexColor("#F1F3F5")]),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#DEE2E6")),
            ("TOPPADDING", (0, 0), (-1, -1), 8),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ]))
        story.append(kpi_tbl)
        story.append(Spacer(1, 16))

        # Helper para generar tablas de datos
        def _tabla_datos(titulo, data, col1, col2, max_filas=15):
            if not data:
                return
            story.append(Paragraph(titulo, h2))
            rows = [[col1, col2]]
            for k, v in list(data.items())[:max_filas]:
                rows.append([str(k), str(v)])
            t = Table(rows, colWidths=[12*cm, 3*cm])
            t.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0D6EFD")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("ALIGN", (1, 0), (1, -1), "RIGHT"),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#DEE2E6")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1),
                 [colors.white, colors.HexColor("#F8F9FA")]),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ]))
            story.append(t)
            story.append(Spacer(1, 10))

        # Sección gráficos
        story.append(Paragraph("Indicadores visuales", h2))
        # Pares de gráficos en 2 columnas
        valid_charts = [p for p in chart_paths if p]
        for i in range(0, len(valid_charts), 2):
            par = valid_charts[i:i+2]
            row = []
            for p in par:
                row.append(Image(p, width=8*cm, height=5.5*cm))
            if len(row) == 1:
                row.append("")
            t = Table([row], colWidths=[8.5*cm, 8.5*cm])
            t.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "MIDDLE")]))
            story.append(t)
            story.append(Spacer(1, 8))

        story.append(PageBreak())

        # Tablas detalle
        _tabla_datos("🧬 Top órganos (canónicos)",
                     stats.get("organos_normalizados", {}), "Órgano", "Casos")
        dx_clin = {k: v for k, v in (stats.get("diagnosticos_categorizados", {}) or {}).items()
                   if k != "OTRO / NO CATEGORIZADO"}
        _tabla_datos("🩺 Diagnósticos (categorías clínicas)",
                     dx_clin, "Categoría", "Casos")
        story.append(Paragraph(
            f"<i>Sin diagnóstico específico (Estudio IHQ): {stats.get('diagnosticos_estudio_ihq_sin_dx', 0)} · "
            f"Otros no categorizados: {stats.get('diagnosticos_otro_no_categorizado', 0)} · "
            f"Sin dato: {stats.get('diagnosticos_sin_dato', 0)}</i>",
            body))
        story.append(Spacer(1, 8))

        bios = stats.get("biomarcadores_top15", {}) or {}
        if bios:
            story.append(Paragraph("🔬 Top biomarcadores", h2))
            rows = [["Biomarcador", "N evaluados", "Resultado predominante"]]
            for marcador, info in sorted(bios.items(),
                                         key=lambda x: x[1].get("n", 0), reverse=True):
                top = info.get("top", {}) or {}
                if top:
                    primer_valor, primer_n = next(iter(top.items()))
                    resumen = f"{primer_valor} (N={primer_n})"
                else:
                    resumen = "—"
                rows.append([marcador, str(info.get("n", 0)), resumen])
            t = Table(rows, colWidths=[5.5*cm, 2.5*cm, 9*cm])
            t.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0D6EFD")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("ALIGN", (1, 0), (1, -1), "RIGHT"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#DEE2E6")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1),
                 [colors.white, colors.HexColor("#F8F9FA")]),
            ]))
            story.append(t)
            story.append(Spacer(1, 8))

        _tabla_datos("⚕️ Procedimientos", stats.get("procedimientos", {}),
                     "Procedimiento", "Casos")
        _tabla_datos("🏥 Servicios solicitantes", stats.get("servicios", {}),
                     "Servicio", "Casos")

        story.append(PageBreak())

        # Informe IA narrativo
        story.append(Paragraph("📝 Informe redactado por IA", h1))
        import re as _re
        for raw in texto.splitlines():
            linea = raw.rstrip()
            if not linea:
                story.append(Spacer(1, 4))
                continue
            if linea.startswith("# "):
                story.append(Paragraph(linea[2:], h1))
            elif linea.startswith("## "):
                story.append(Paragraph(linea[3:], h2))
            elif linea.startswith("### "):
                story.append(Paragraph(linea[4:], h3))
            elif linea.strip() in {"---", "***"}:
                story.append(Spacer(1, 6))
            elif linea.startswith("|"):
                # Líneas de tabla markdown — las dejamos como código simple
                story.append(Paragraph(
                    f"<font face='Courier' size='8'>{linea.replace('<', '&lt;').replace('>', '&gt;')}</font>",
                    body,
                ))
            else:
                # Convertir **negritas**
                txt = _re.sub(r"\*\*(.+?)\*\*", r"<b>\1</b>", linea)
                story.append(Paragraph(txt, body))

        try:
            doc.build(story)
            messagebox.showinfo(
                "Exportación exitosa",
                f"PDF guardado en:\n{filepath}",
            )
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo generar el PDF:\n{e}")

    def _copiar_resumen_ia(self, texto: str, ventana: tk.Toplevel):
        """Copia el resumen IA al portapapeles."""
        ventana.clipboard_clear()
        ventana.clipboard_append(texto)
        messagebox.showinfo("Copiado", "Resumen copiado al portapapeles.")

    def _export_selected_data(self):
        """Exportar datos seleccionados usando el sistema mejorado (v6.0.12: con logging mejorado)"""
        try:
            import pandas as pd

            # v6.0.12: Logging inicial para debugging
            logging.debug("_export_selected_data: Iniciando exportación de selección")

            # Obtener elementos seleccionados del treeview
            selected_items = self.tree.selection()
            logging.debug(f"_export_selected_data: selection() retornó {len(selected_items) if selected_items else 0} items")

            if not selected_items:
                logging.warning("_export_selected_data: No hay selección para exportar")
                messagebox.showwarning("Sin Selección", "No hay elementos seleccionados para exportar")
                return

            logging.info(f"Exportar Seleccion: {len(selected_items)} elementos seleccionados")

            # Validar que master_df existe
            if not hasattr(self, 'master_df') or self.master_df is None or self.master_df.empty:
                logging.error("master_df no existe o esta vacio")
                messagebox.showerror(
                    "Error",
                    "Los datos no están cargados en memoria.\n"
                    "Recarga la base de datos e intenta de nuevo."
                )
                return

            # Obtener datos de las filas seleccionadas (enfoque simple que funciona)
            selected_rows_data = []
            for item in selected_items:
                # tksheet: Usar get_row_data() en lugar de .item()
                values = self.sheet.get_row_data(item)
                logging.info(f"DEBUG item={item}: values={values}")

                if values:
                    # El primer valor es el número de petición/caso
                    numero_peticion = values[0]

                    # Buscar el registro correspondiente en master_df usando la primera columna
                    # (esto es lo que funciona en _export_selected_data_professional)
                    matching_row = self.master_df[self.master_df.iloc[:, 0] == numero_peticion]

                    if not matching_row.empty:
                        selected_rows_data.append(matching_row.iloc[0])
                        logging.info(f"Fila '{numero_peticion}' encontrada y agregada")
                    else:
                        logging.warning(f"No se encontro la fila '{numero_peticion}' en master_df")
                else:
                    logging.warning(f"Item sin valores: {item}")

            if not selected_rows_data:
                logging.error("No se pudieron obtener datos de la seleccion")
                messagebox.showwarning(
                    "Sin Datos",
                    "No se pudieron obtener los datos de la selección.\n"
                    "Intenta recargar la base de datos."
                )
                return

            logging.info(f"Filas obtenidas exitosamente: {len(selected_rows_data)}")

            # Crear DataFrame con los registros seleccionados
            selected_df = pd.DataFrame(selected_rows_data)

            # Exportar usando el sistema mejorado
            self.export_system.export_selected_data(selected_df)

        except Exception as e:
            logging.error(f"Error en _export_selected_data: {e}", exc_info=True)
            messagebox.showerror("Error de Exportación", f"Error al exportar la selección:\n{str(e)}")

    def _toggle_details_panel(self):
        """Mostrar/ocultar panel flotante de detalles"""
        try:
            self.export_system.toggle_floating_details_panel()
        except Exception as e:
            messagebox.showerror("Error", f"Error al mostrar panel de detalles:\n{str(e)}")

    def _toggle_advanced_filters(self):
        """Mostrar/ocultar panel de filtros avanzados"""
        try:
            self.export_system.toggle_advanced_filters_panel()
        except Exception as e:
            messagebox.showerror("Error", f"Error al mostrar filtros avanzados:\n{str(e)}")

    def _update_export_button_state(self):
        """Actualizar estado del botón de exportar selección según la selección"""
        try:
            logging.debug("_update_export_button_state: INICIANDO")

            # Obtener selección del sheet apropiado
            selected_items = []
            if hasattr(self, 'sheet') and self.sheet is not None:
                try:
                    selected_items = self.sheet.selection()
                except:
                    pass
            elif hasattr(self, 'sheet_dashboard') and self.sheet_dashboard is not None:
                try:
                    selected_items = self.sheet_dashboard.selection()
                except:
                    pass
            elif hasattr(self, 'tree') and self.tree is not None:
                selected_items = self.tree.selection()

            logging.debug(f"_update_export_button_state: selection() = {selected_items}")
            has_selection = bool(selected_items)

            # Actualizar botón original
            if hasattr(self, 'export_selection_btn'):
                logging.info("_update_export_button_state: export_selection_btn existe")
                if has_selection:
                    logging.info("_update_export_button_state: HAY SELECCIÓN -> Habilitando botón")
                    self.export_selection_btn.configure(state="normal")
                else:
                    logging.info("_update_export_button_state: SIN SELECCIÓN -> Deshabilitando botón")
                    self.export_selection_btn.configure(state="disabled")

            # Actualizar botón del dashboard
            if hasattr(self, 'export_selection_btn_dashboard'):
                if has_selection:
                    self.export_selection_btn_dashboard.configure(state="normal")
                else:
                    self.export_selection_btn_dashboard.configure(state="disabled")

            # V6.9.44: "Detalles" también requiere selección -> mismo estado que
            # "Exportar Selección" (se "enciende" al seleccionar una fila).
            _estado_det = "normal" if has_selection else "disabled"
            for _attr in ("details_btn", "details_btn_dashboard"):
                _btn = getattr(self, _attr, None)
                if _btn is not None:
                    try:
                        _btn.configure(state=_estado_det)
                    except Exception:
                        pass

        except Exception as e:
            logging.error(f"Error actualizando estado del boton: {e}", exc_info=True)

    def _update_audit_buttons_state(self):
        """v6.0.14: Actualizar estado de botones de auditoría según la selección"""
        try:
            logging.debug("_update_audit_buttons_state: INICIANDO")

            # Obtener selección del sheet apropiado
            selection = []
            if hasattr(self, 'sheet') and self.sheet is not None:
                try:
                    selection = self.sheet.selection()
                except:
                    pass
            elif hasattr(self, 'sheet_dashboard') and self.sheet_dashboard is not None:
                try:
                    selection = self.sheet_dashboard.selection()
                except:
                    pass
            elif hasattr(self, 'tree') and self.tree is not None:
                selection = self.tree.selection()

            logging.debug(f"_update_audit_buttons_state: selection() = {selection}")
            has_selection = bool(selection)
            logging.debug(f"_update_audit_buttons_state: has_selection = {has_selection}")

            if not has_selection:
                logging.info(f"_update_audit_buttons_state: CHECKPOINT 2 - Dentro del if not has_selection")
                # Sin selección → deshabilitar botones
                logging.info("_update_audit_buttons_state: SIN SELECCIÓN -> Deshabilitando botones")

                # Deshabilitar botones originales
                if hasattr(self, 'audit_parcial_btn'):
                    logging.info("  -> Deshabilitando audit_parcial_btn")
                    self.audit_parcial_btn.configure(state="disabled")
                if hasattr(self, 'audit_completa_btn'):
                    logging.info("  -> Deshabilitando audit_completa_btn")
                    self.audit_completa_btn.configure(state="disabled")

                # Deshabilitar botones del dashboard
                if hasattr(self, 'audit_parcial_btn_dashboard'):
                    self.audit_parcial_btn_dashboard.configure(state="disabled")
                if hasattr(self, 'audit_completa_btn_dashboard'):
                    self.audit_completa_btn_dashboard.configure(state="disabled")
            else:
                # Hay selección → determinar estados
                logging.info("_update_audit_buttons_state: CHECKPOINT 3 - Dentro del else (HAY SELECCIÓN)")
                logging.info("_update_audit_buttons_state: HAY SELECCIÓN -> Determinando estados...")
                from core.database_manager import get_estado_auditoria

                try:
                    # Obtener el sheet correcto (dashboard o original)
                    active_sheet = None
                    if hasattr(self, 'sheet_dashboard') and self.sheet_dashboard is not None:
                        try:
                            if self.sheet_dashboard.selection():
                                active_sheet = self.sheet_dashboard
                        except:
                            pass
                    if active_sheet is None and hasattr(self, 'sheet') and self.sheet is not None:
                        active_sheet = self.sheet

                    if active_sheet is None:
                        logging.error("No hay sheet activo disponible")
                        return

                    # Obtener índices de columnas necesarias
                    try:
                        headers = active_sheet.headers() if hasattr(active_sheet, 'headers') else []
                        col_idx = headers.index("Numero de caso") if "Numero de caso" in headers else 0
                        # V6.9.49 PERF: índice de la columna de estado YA visible en la
                        # tabla. Permite leer el estado de la fila en memoria en vez de
                        # consultar get_estado_auditoria() a la BD (MySQL, a veces central)
                        # en CADA caso seleccionado y en CADA clic -> era la causa #1 del
                        # clic "sumamente lento". El valor es el mismo que ve el usuario.
                        estado_idx = headers.index("Estado Auditoria IA") if "Estado Auditoria IA" in headers else None
                        logging.debug(f"_update_audit_buttons_state: col_idx={col_idx}, estado_idx={estado_idx}")
                    except Exception as e:
                        logging.error(f"_update_audit_buttons_state: ERROR obteniendo col_idx: {e}")
                        col_idx = 0
                        estado_idx = None

                    # Obtener estados de todos los items seleccionados (sin tocar la BD)
                    estados = []
                    for item_id in selection:
                        # tksheet: Usar get_row_data() en lugar de .item()
                        values = active_sheet.get_row_data(item_id)
                        if values and len(values) > col_idx:
                            numero_peticion = values[col_idx]
                            if estado_idx is not None and len(values) > estado_idx:
                                # Leer estado directamente de la fila (en memoria)
                                estado = values[estado_idx]
                                estado = "" if estado is None else str(estado).strip()
                                if estado.upper() in ("N/A", "NONE", "NAN", "NULL"):
                                    estado = ""
                            else:
                                # Fallback: la columna de estado no está en el Sheet -> BD
                                estado = get_estado_auditoria(numero_peticion)
                            estados.append(estado)
                        else:
                            logging.warning(f"_update_audit_buttons_state: No se pudieron obtener valores para item_id={item_id}")

                    # Lógica de habilitación basada en estados
                    logging.debug(f"_update_audit_buttons_state: estados recolectados = {estados}")

                    if all(e == "COMPLETA" for e in estados):
                        # TODOS tienen auditoría COMPLETA → Bloquear ambos
                        logging.info("_update_audit_buttons_state: TODOS COMPLETA -> Bloqueando ambos")
                        if hasattr(self, 'audit_parcial_btn'):
                            self.audit_parcial_btn.configure(state="disabled")
                        if hasattr(self, 'audit_completa_btn'):
                            self.audit_completa_btn.configure(state="disabled")
                        if hasattr(self, 'audit_parcial_btn_dashboard'):
                            self.audit_parcial_btn_dashboard.configure(state="disabled")
                        if hasattr(self, 'audit_completa_btn_dashboard'):
                            self.audit_completa_btn_dashboard.configure(state="disabled")

                    elif all(e == "PARCIAL" for e in estados):
                        # TODOS tienen auditoría PARCIAL → Solo permitir COMPLETA
                        logging.info("_update_audit_buttons_state: TODOS PARCIAL -> Solo COMPLETA habilitada")
                        if hasattr(self, 'audit_parcial_btn'):
                            self.audit_parcial_btn.configure(state="disabled")
                        if hasattr(self, 'audit_completa_btn'):
                            self.audit_completa_btn.configure(state="normal")
                        if hasattr(self, 'audit_parcial_btn_dashboard'):
                            self.audit_parcial_btn_dashboard.configure(state="disabled")
                        if hasattr(self, 'audit_completa_btn_dashboard'):
                            self.audit_completa_btn_dashboard.configure(state="normal")

                    elif all(e in [None, "NULL", ""] for e in estados):
                        # TODOS sin auditoría → Permitir ambas
                        logging.info("_update_audit_buttons_state: TODOS SIN AUDITORIA -> Habilitando ambos")
                        if hasattr(self, 'audit_parcial_btn'):
                            logging.info("  -> Habilitando audit_parcial_btn")
                            self.audit_parcial_btn.configure(state="normal")
                        if hasattr(self, 'audit_completa_btn'):
                            logging.info("  -> Habilitando audit_completa_btn")
                            self.audit_completa_btn.configure(state="normal")
                        if hasattr(self, 'audit_parcial_btn_dashboard'):
                            self.audit_parcial_btn_dashboard.configure(state="normal")
                        if hasattr(self, 'audit_completa_btn_dashboard'):
                            self.audit_completa_btn_dashboard.configure(state="normal")

                    else:
                        # Mezcla de estados → Permitir ambas
                        logging.info("_update_audit_buttons_state: MEZCLA DE ESTADOS -> Habilitando ambos")
                        if hasattr(self, 'audit_parcial_btn'):
                            logging.info("  -> Habilitando audit_parcial_btn")
                            self.audit_parcial_btn.configure(state="normal")
                        if hasattr(self, 'audit_completa_btn'):
                            logging.info("  -> Habilitando audit_completa_btn")
                            self.audit_completa_btn.configure(state="normal")
                        if hasattr(self, 'audit_parcial_btn_dashboard'):
                            self.audit_parcial_btn_dashboard.configure(state="normal")
                        if hasattr(self, 'audit_completa_btn_dashboard'):
                            self.audit_completa_btn_dashboard.configure(state="normal")

                except Exception as e:
                    logging.error(f"Error en lógica de auditoría: {e}")
                    # Si hay error, deshabilitar botones
                    if hasattr(self, 'audit_parcial_btn'):
                        self.audit_parcial_btn.configure(state="disabled")
                    if hasattr(self, 'audit_completa_btn'):
                        self.audit_completa_btn.configure(state="disabled")
                    if hasattr(self, 'audit_parcial_btn_dashboard'):
                        self.audit_parcial_btn_dashboard.configure(state="disabled")
                    if hasattr(self, 'audit_completa_btn_dashboard'):
                        self.audit_completa_btn_dashboard.configure(state="disabled")

        except Exception as e:
            logging.error(f"Error actualizando botones de auditoría: {e}", exc_info=True)
    
    def _setup_cell_tooltips(self):
        """
        V5.3.8: Configurar tooltips emergentes al pasar el mouse sobre celdas del Sheet
        Muestra el contenido COMPLETO de la celda cuando es muy largo
        """
        # Crear tooltip widget (inicialmente oculto)
        self.tooltip = None
        self.tooltip_job = None
        self._last_tooltip_cell = None  # Rastrear última celda con tooltip

        def show_tooltip(event):
            """Mostrar tooltip con el contenido completo de la celda"""
            # Cancelar tooltip anterior si existe
            if self.tooltip_job:
                self.after_cancel(self.tooltip_job)
                self.tooltip_job = None

            try:
                # V5.3.8: Obtener celda bajo el cursor usando métodos correctos de Sheet
                # Método 1: get_cell_at_position
                cell_info = None
                try:
                    # Convertir coordenadas de evento a coordenadas del canvas
                    x = self.sheet.canvasx(event.x)
                    y = self.sheet.canvasy(event.y)

                    # Intentar obtener la celda en esa posición
                    # tksheet usa diferentes métodos dependiendo de la versión
                    if hasattr(self.sheet, 'get_cell_at_position'):
                        cell_info = self.sheet.get_cell_at_position(x, y)
                    elif hasattr(self.sheet.MT, 'identify_row') and hasattr(self.sheet.MT, 'identify_col'):
                        # Acceder al MainTable interno
                        row = self.sheet.MT.identify_row(y=event.y)
                        col = self.sheet.MT.identify_col(x=event.x)
                        if row is not None and col is not None:
                            cell_info = {'row': row, 'column': col}
                except:
                    pass

                # Si no se pudo obtener la celda, salir
                if not cell_info:
                    if self.tooltip:
                        self.tooltip.destroy()
                        self.tooltip = None
                        self._last_tooltip_cell = None
                    return

                # Extraer fila y columna del resultado
                row = cell_info.get('row') if isinstance(cell_info, dict) else getattr(cell_info, 'row', None)
                col = cell_info.get('column') if isinstance(cell_info, dict) else getattr(cell_info, 'column', None)

                if row is None or col is None:
                    return

                # Evitar recrear tooltip para la misma celda
                if self._last_tooltip_cell == (row, col):
                    return

                # Destruir tooltip anterior
                if self.tooltip:
                    self.tooltip.destroy()
                    self.tooltip = None

                # Obtener valor de la celda
                try:
                    cell_value = self.sheet.get_cell_data(row, col, return_copy=True)
                except:
                    return

                if not cell_value:
                    return

                cell_value = str(cell_value).strip()

                # Solo mostrar tooltip si el valor es suficientemente largo (>30 chars)
                # o contiene saltos de línea
                if len(cell_value) < 30 and '\n' not in cell_value:
                    return

                # Evitar tooltips para valores vacíos o inútiles
                if cell_value in ['', 'N/A', 'nan', 'None', 'null']:
                    return

                # Crear tooltip después de un pequeño delay
                def create_tooltip():
                    try:
                        self.tooltip = tk.Toplevel(self.sheet)
                        self.tooltip.wm_overrideredirect(True)

                        # Posicionar tooltip cerca del cursor
                        x_pos = event.x_root + 15
                        y_pos = event.y_root + 10

                        # Ajustar si está muy cerca del borde derecho
                        screen_width = self.tooltip.winfo_screenwidth()
                        if x_pos + 450 > screen_width:
                            x_pos = screen_width - 460

                        self.tooltip.wm_geometry(f"+{x_pos}+{y_pos}")

                        # Frame con borde y sombra
                        frame = tk.Frame(
                            self.tooltip,
                            relief="solid",
                            borderwidth=2,
                            background="#2C3E50",  # Borde azul oscuro profesional
                            padx=1,
                            pady=1
                        )
                        frame.pack()

                        inner_frame = tk.Frame(
                            frame,
                            background="#FFFEF0",  # Fondo crema claro
                            padx=10,
                            pady=8
                        )
                        inner_frame.pack()

                        # Texto del tooltip (máximo 1000 caracteres para auditoría)
                        display_text = cell_value[:1000] + "..." if len(cell_value) > 1000 else cell_value

                        label = tk.Label(
                            inner_frame,
                            text=display_text,
                            background="#FFFEF0",
                            foreground="#1A1A1A",
                            font=("Segoe UI", 9),
                            wraplength=450,  # Ancho máximo del tooltip
                            justify=tk.LEFT,
                            anchor="w"
                        )
                        label.pack()

                        # Agregar longitud del texto si es muy largo
                        if len(cell_value) > 100:
                            length_label = tk.Label(
                                inner_frame,
                                text=f"({len(cell_value)} caracteres)",
                                background="#FFFEF0",
                                foreground="#7F8C8D",
                                font=("Segoe UI", 8, "italic")
                            )
                            length_label.pack(anchor="e", pady=(5, 0))

                        # Guardar celda actual
                        self._last_tooltip_cell = (row, col)

                    except Exception as e:
                        logging.warning(f"Error creando tooltip: {e}")
                        if self.tooltip:
                            self.tooltip.destroy()
                            self.tooltip = None

                # V6.9.49: 150ms (antes 400). El throttle de <Motion> ya difiere el
                # cómputo de celda; este delay restante solo retarda crear el Toplevel.
                self.tooltip_job = self.after(150, create_tooltip)

            except Exception as e:
                # Silenciar errores de tooltips para no interrumpir la UI
                pass

        def hide_tooltip(event=None):
            """Ocultar tooltip"""
            if self.tooltip_job:
                self.after_cancel(self.tooltip_job)
                self.tooltip_job = None
            # V6.9.49: cancelar también el cómputo diferido pendiente por <Motion>
            _mj = getattr(self, '_tooltip_motion_job', None)
            if _mj:
                self.after_cancel(_mj)
                self._tooltip_motion_job = None

            if self.tooltip:
                self.tooltip.destroy()
                self.tooltip = None

            self._last_tooltip_cell = None

        # V6.9.49 PERF: throttle de <Motion>. El cómputo de la celda bajo el cursor
        # dentro de show_tooltip (canvasx/canvasy + identify_row/identify_col +
        # get_cell_data) es CARO y antes corría en CADA evento <Motion> -> mover el
        # mouse o scrollear se sentía pesado. Ahora <Motion> solo guarda el evento y
        # reprograma show_tooltip para cuando el cursor se detiene (~250ms sin
        # moverse): mientras el mouse se mueve no se hace prácticamente nada.
        self._tooltip_motion_job = None

        def _on_motion_throttled(event):
            if getattr(self, '_tooltip_motion_job', None):
                self.after_cancel(self._tooltip_motion_job)
            self._tooltip_motion_job = self.after(250, lambda e=event: show_tooltip(e))

        # Vincular eventos al Sheet
        self.sheet.bind("<Motion>", _on_motion_throttled, add="+")
        self.sheet.bind("<Leave>", hide_tooltip, add="+")
        self.sheet.bind("<Button-1>", hide_tooltip, add="+")  # Ocultar al hacer clic

    def _delayed_refresh_after_processing(self):
        """Refresh retardado después del procesamiento para asegurar actualización"""
        try:
            # V5.3.9.3: Usar logging en lugar de print (stdout puede estar cerrado)
            logging.info("🔄 Ejecutando refresh automático después del procesamiento...")

            # Forzar refresh de datos
            self.refresh_data_and_table()

            # Si estamos en otra vista, cambiar automáticamente al visualizador
            if hasattr(self, 'current_view') and self.current_view != "visualizar":
                logging.info("📊 Cambiando automáticamente al Visualizador de Datos...")
                self.show_visualizar_frame()

            # Actualizar estado de los botones de exportación
            self._update_export_button_state()

            logging.info("✅ Refresh automático completado")

        except Exception as e:
            logging.error(f"❌ Error en refresh automático: {e}")
            # Mostrar mensaje al usuario si falla
            try:
                messagebox.showwarning(
                    "Actualización automática",
                    f"Los datos se procesaron correctamente, pero no se pudo actualizar la vista automáticamente.\n\n"
                    f"Por favor, ve al Visualizador y haz clic en 'Actualizar Datos'.\n\nError: {e}"
                )
            except:
                pass

    def cargar_dashboard(self):
        # 1) Preparar DF y combos de filtros
        df = self.master_df.copy()
        if df is None or df.empty:
            self._render_kpis(df)
            self._clear_dash_area()
            return

        # Normaliza fechas (varias columnas posibles)
        df["_fecha_informe"] = pd.to_datetime(
            df.get("Fecha Informe", df.get("Fecha de informe", df.get("Fecha de ingreso", ""))),
            dayfirst=True, errors="coerce"
        )

        # Llenar combos dinámicos (servicios / responsables)
        srv_vals = sorted([s for s in df.get("Servicio", pd.Series(dtype=str)).dropna().astype(str).unique() if s.strip()])
        rsp_vals = sorted([s for s in df.get("Patologo", pd.Series(dtype=str)).dropna().astype(str).unique() if s.strip()])
        
        # Solo configurar si los componentes existen
        if self.cmb_servicio is not None:
            try:
                self.cmb_servicio.configure(values=[""] + srv_vals)
            except:
                pass
        
        if self.cmb_resp is not None:
            try:
                self.cmb_resp.configure(values=[""] + rsp_vals)
            except:
                pass

        # 2) Render de KPIs básicos
        self._render_kpis(df)

        # 3) Limpiar canvases anteriores y pintar
        self._clear_dash_area()

        # Filtros iniciales (los que estén llenos)
        dff = self._get_filtered_df(df)

        # 4) PINTAR: OVERVIEW (4 gráficos)
        self._chart_in(self.tab_overview, 0, 0, lambda: self._g_line_informes_por_mes(dff), "Informes por mes", dff)
        self._chart_in(self.tab_overview, 0, 1, lambda: self._g_pie_malignidad(dff), "Distribución de Malignidad", dff)
        self._chart_in(self.tab_overview, 1, 0, lambda: self._g_bar_top_servicio(dff), "Top Servicios", dff)
        self._chart_in(self.tab_overview, 1, 1, lambda: self._g_bar_top_organo(dff), "Top Órganos", dff)

        # 5) PINTAR: BIOMARCADORES
        self._chart_in(self.tab_biomarkers, 0, 0, lambda: self._g_hist_ki67(dff), "Ki-67 (%)", dff)
        self._chart_in(self.tab_biomarkers, 0, 1, lambda: self._g_bar_her2(dff), "HER2 (score)", dff)
        self._chart_in(self.tab_biomarkers, 1, 0, lambda: self._g_bar_re_rp(dff), "RE / RP (estado)", dff)
        self._chart_in(self.tab_biomarkers, 1, 1, lambda: self._g_bar_pdl1(dff), "PD-L1", dff)

        # 6) PINTAR: TIEMPOS
        self._chart_in(self.tab_times, 0, 0, lambda: self._g_box_tiempo_proceso(dff), "Tiempo de proceso (días)", dff)
        self._chart_in(self.tab_times, 0, 1, lambda: self._g_line_throughput_semana(dff), "Throughput semanal", dff)
        self._chart_in(self.tab_times, 1, 0, lambda: self._g_scatter_edad_ki67(dff), "Edad vs Ki-67", dff)

        # 7) PINTAR: CALIDAD
        self._chart_in(self.tab_quality, 0, 0, lambda: self._g_bar_missingness(dff), "Campos vacíos (%)", dff)
        self._chart_in(self.tab_quality, 0, 1, lambda: self._g_bar_top_responsables(dff), "Productividad por responsable", dff)
        self._chart_in(self.tab_quality, 1, 0, lambda: self._g_bar_largos_texto(dff), "Longitud del diagnóstico", dff)

        # 8) PINTAR: COMPARADOR
        self._build_comparator(self.tab_compare, dff)


    # =========================
    # Estilo de tabla (Treeview) - MÉTODO ACTUALIZADO
    # =========================
    def setup_treeview_style(self):
        style = ttk_std.Style()
        style.theme_use("clam")  # look & feel moderno y estable en Windows

        # Cuerpo de la tabla - Estilo profesional mejorado
        style.configure(
            "Custom.Treeview",
            background="#ffffff",        # Fondo blanco para mejor legibilidad
            fieldbackground="#ffffff",   # Fondo de campos blanco
            foreground="#2c3e50",        # Texto azul oscuro para mayor contraste
            rowheight=35,                # Filas más altas para mejor espaciado
            borderwidth=1,               # Borde sutil
            relief="solid",              # Borde sólido
            font=("Segoe UI", 10),       # Fuente más legible
        )

        # Configurar colores alternados para filas
        style.map(
            "Custom.Treeview",
            background=[
                ("selected", "#0078d4"),     # Azul Microsoft para selección
                ("!selected", "#ffffff")     # Blanco para filas no seleccionadas
            ],
            foreground=[
                ("selected", "white"),       # Texto blanco en selección
                ("!selected", "#2c3e50")     # Texto azul oscuro normal
            ]
        )

        # Encabezados profesionales
        style.configure(
            "Custom.Treeview.Heading",
            background="#f8f9fa",           # Gris muy claro para encabezados
            foreground="#495057",           # Gris oscuro para texto
            font=("Segoe UI", 11, "bold"),  # Fuente en negrita
            relief="ridge",                 # Relieve elevado
            borderwidth=1,                  # Borde definido
        )

        # Efectos hover en encabezados
        style.map(
            "Custom.Treeview.Heading",
            background=[("active", "#e9ecef")],  # Gris claro en hover
            foreground=[("active", "#212529")]   # Texto más oscuro en hover
        )

        return style
    # ---------- Automatización Web (modal + ejecución) ----------

    def open_web_auto_modal(self):
        top = tk.Toplevel(self)
        top.title("Interoperabilidad QHORTE - Sistema de Entrega")
        top.geometry("460x360")
        top.grab_set()

        top.transient(self)
        try:
            top.lift(); top.focus_force()
        except Exception:
            pass
        
        # Campos
        frm = ttk.Frame(top, padding=12, relief="solid", borderwidth=1)
        frm.pack(fill="both", expand=True, padx=12, pady=12)

        # Usuario / Clave
        ttk.Label(frm, text="Usuario").grid(row=0, column=0, padx=10, pady=(12,6), sticky="w")
        user_var = tk.StringVar(value="12345")
        ttk.Entry(frm, textvariable=user_var).grid(row=0, column=1, padx=10, pady=(12,6), sticky="ew")

        ttk.Label(frm, text="Contraseña").grid(row=1, column=0, padx=10, pady=6, sticky="w")
        pass_var = tk.StringVar(value="CONSULTA1")
        ttk.Entry(frm, textvariable=pass_var, show="•").grid(row=1, column=1, padx=10, pady=6, sticky="ew")

        # Criterio
        ttk.Label(frm, text="Buscar por").grid(row=2, column=0, padx=10, pady=6, sticky="w")
        criterio_var = tk.StringVar(value="Fecha de Ingreso")
        ttk.Combobox(frm, values=["Fecha de Ingreso", "Fecha de Finalizacion", "Rango de Peticion", "Datos del Paciente"], textvariable=criterio_var).grid(row=2, column=1, padx=10, pady=6, sticky="ew")

        # Fechas
        fi_var = tk.StringVar(value="")
        ff_var = tk.StringVar(value="")

        def pick_fi():
            sel = CalendarioInteligente.seleccionar_fecha(parent=top, locale='es_CO', codigo_pais_festivos='CO')
            if sel:
                fi_var.set(sel.strftime("%d/%m/%Y"))
            # RE-ADQUIRIR MODAL Y TRAER AL FRENTE
            try:
                top.deiconify()
                # truco para traer al frente en Windows
                top.attributes("-topmost", True); top.attributes("-topmost", False)
                top.lift(); top.focus_force(); top.grab_set()
            except Exception:
                pass

        def pick_ff():
            sel = CalendarioInteligente.seleccionar_fecha(parent=top, locale='es_CO', codigo_pais_festivos='CO')
            if sel:
                ff_var.set(sel.strftime("%d/%m/%Y"))
            # RE-ADQUIRIR MODAL Y TRAER AL FRENTE
            try:
                top.deiconify()
                top.attributes("-topmost", True); top.attributes("-topmost", False)
                top.lift(); top.focus_force(); top.grab_set()
            except Exception:
                pass

        ttk.Label(frm, text="Fecha inicial").grid(row=3, column=0, padx=10, pady=6, sticky="w")
        row_fi = ttk.Frame(frm); row_fi.grid(row=3, column=1, padx=10, pady=6, sticky="ew")
        ttk.Entry(row_fi, textvariable=fi_var).pack(side="left", fill="x", expand=True, padx=(0,6))
        ttk.Button(row_fi, text="Elegir…", width=10, command=pick_fi).pack(side="left")

        ttk.Label(frm, text="Fecha final").grid(row=4, column=0, padx=10, pady=6, sticky="w")
        row_ff = ttk.Frame(frm); row_ff.grid(row=4, column=1, padx=10, pady=6, sticky="ew")
        ttk.Entry(row_ff, textvariable=ff_var).pack(side="left", fill="x", expand=True, padx=(0,6))
        ttk.Button(row_ff, text="Elegir…", width=10, command=pick_ff).pack(side="left")

        # Botones
        btns = ttk.Frame(frm); btns.grid(row=5, column=0, columnspan=2, pady=(12,8), sticky="ew")
        ttk.Button(btns, text="Cancelar", command=top.destroy).pack(side="right", padx=6)
        def go():
            top.destroy()
            self._start_web_automation(
                fi_var.get().strip(), ff_var.get().strip(),
                user_var.get().strip(), pass_var.get().strip(),
                criterio_var.get().strip()
            )
        ttk.Button(btns, text="Iniciar", command=go).pack(side="right", padx=6)

        # grid conf
        frm.grid_columnconfigure(1, weight=1)

    def _start_web_automation(self, fi, ff, user, pwd, criterio):
        if not fi or not ff:
            messagebox.showwarning("Fechas requeridas", "Debe seleccionar fecha inicial y final.")
            return
        self.set_status("Automatizando Entrega de resultados…")
        t = threading.Thread(target=self._run_web_automation, args=(fi, ff, user, pwd, criterio), daemon=True)
        t.start()

    def _run_web_automation(self, fi, ff, user, pwd, criterio):
        try:
            # LAZY IMPORT: Solo importar cuando realmente se necesite
            from core.huv_web_automation import automatizar_entrega_resultados, Credenciales

            ok = automatizar_entrega_resultados(
                fecha_inicial_ddmmaa=fi,
                fecha_final_ddmmaa=ff,
                cred=Credenciales(usuario=user, clave=pwd),
                criterio=criterio,
                headless=False,
                log_cb=self._log_auto
            )
            if ok:
                self.set_status("Consulta web completada. Revise resultados en el navegador.")
                messagebox.showinfo("Automatización", "Consulta completada en el portal.")
            else:
                self.set_status("Automatización: sin resultado.")
        except Exception as e:
            self.set_status(f"Error en automatización: {e}")
            messagebox.showerror("Automatización", f"Ocurrió un error:\n{e}")

    def _log_auto(self, msg: str):
        try:
            # Si está visible el textbox de logs de Procesar, úsalo; si no, status.
            if hasattr(self, "log_textbox") and str(self.log_textbox.winfo_exists()) == "1":
                self.log_textbox.configure(state="normal")
                self.log_textbox.insert("end", f"[AUTO] {msg}\n")
                self.log_textbox.configure(state="disabled")
                self.log_textbox.see("end")
            else:
                self.set_status(msg)
        except Exception:
            self.set_status(msg)

    # =========================
    # Tema claro/oscuro
    # =========================
    # =========================
    # Métodos para el panel de procesamiento
    # =========================
    def _show_external_data_info(self):
        """Mostrar información de los datos externos en un modal"""
        top = tk.Toplevel(self)
        top.title("Información de Datos Externos")
        top.geometry("500x400")
        top.grab_set()
        top.transient(self)
        
        frame = ttk.Frame(top, padding=20)
        frame.pack(fill=BOTH, expand=True)
        
        ttk.Label(frame, text="Información de Base de Datos", font=("Segoe UI", 16, "bold")).pack(pady=(0, 20))
        
        # Obtener información de la base de datos
        try:
            from core.database_manager import get_all_records_as_dataframe
            df = get_all_records_as_dataframe()
            
            if df.empty:
                total_records = 0
                date_range = "No disponible"
                last_import = "No disponible"
                unique_services = 0
                malignant_count = 0
            else:
                total_records = len(df)
                date_range = "Disponible" 
                last_import = "Disponible"
                unique_services = df.get('Servicio', pd.Series()).nunique() if 'Servicio' in df.columns else 0
                malignant_count = (df.get('Malignidad', pd.Series()).str.contains('PRESENTE', case=False, na=False)).sum() if 'Malignidad' in df.columns else 0
            
            info_text = f"""Total de informes en BD: {total_records}

Rango de fechas: {date_range}

Última importación: {last_import}

Servicios únicos: {unique_services}

Informes con malignidad: {malignant_count}"""
            
        except Exception as e:
            info_text = f"Error al obtener información de la base de datos:\n{str(e)}"
        
        text_widget = tk.Text(frame, wrap="word", font=("Segoe UI", 11))
        text_widget.pack(fill=BOTH, expand=True, pady=(0, 20))
        text_widget.insert("1.0", info_text)
        text_widget.configure(state="disabled")
        
        ttk.Button(frame, text="Cerrar", command=top.destroy, bootstyle="primary").pack()

    def _select_pdf_file(self):
        """Seleccionar un archivo PDF individual con flujo completo de análisis"""
        file_path = filedialog.askopenfilename(
            title="Seleccionar archivo PDF",
            filetypes=[("Archivos PDF", "*.pdf")],
            initialdir=os.path.join(os.getcwd(), "pdfs_patologia") if os.path.exists("pdfs_patologia") else os.getcwd()
        )
        if file_path:
            # NUEVO: Limpiar lista de registros procesados y obtener peticiones existentes antes del procesamiento
            self._ultimos_registros_procesados = []
            peticiones_antes = set()
            try:
                import sqlite3
                from core.database_manager import DB_FILE
                conn = sqlite3.connect(DB_FILE)
                cursor = conn.cursor()
                cursor.execute('SELECT "Numero de caso" FROM informes_ihq')
                peticiones_antes = set(row[0] for row in cursor.fetchall() if row[0])
                conn.close()
            except Exception as e:
                logging.warning(f"Error obteniendo peticiones existentes: {e}")

            try:
                # V5.3.9: _process_file ahora retorna (records_count, correcciones)
                records, correcciones = self._process_file(file_path)

                # V6.9.31: usar los números de caso REALES capturados durante el
                # procesamiento (incluye reimportaciones). Antes se calculaba
                # "despues - antes" leyendo SQLite (vacío con MySQL) -> daba 0.
                try:
                    capturados = list(dict.fromkeys(
                        getattr(self, '_ultimos_registros_procesados', []) or []
                    ))
                    self._ultimos_registros_procesados = capturados
                    logging.info(f"📋 Registros procesados: {len(capturados)}")
                    if capturados:
                        logging.info(f"   IDs: {', '.join(capturados[:5])}" + (" ..." if len(capturados) > 5 else ""))
                except Exception as e:
                    logging.warning(f"⚠️ Error capturando registros procesados: {e}")
                    self._ultimos_registros_procesados = []

                # NUEVO: Analizar completitud de registros
                try:
                    from core.validation_checker import analizar_batch_registros

                    numeros_peticion_procesados = self._ultimos_registros_procesados
                    logging.info(f"🔍 Analizando completitud de {len(numeros_peticion_procesados)} registros...")
                    analisis = analizar_batch_registros(numeros_peticion_procesados)

                    logging.info(f"✅ Análisis completado:")
                    logging.info(f"   • Completos: {analisis['resumen']['completos']}")
                    logging.info(f"   • Incompletos: {analisis['resumen']['incompletos']}")

                    # Actualizar vista antes de mostrar ventana
                    try:
                        self.refresh_data_and_table()
                        self.after(500, self._delayed_refresh_after_processing)

                        if hasattr(self, 'enhanced_dashboard'):
                            self.enhanced_dashboard.refresh_all_data()
                    except Exception as e:
                        logging.warning(f"⚠️ Error en refresh: {e}")

                    # Actualizar lista de archivos
                    self._refresh_files_list()

                    # Mostrar ventana de resultados con análisis de completitud
                    from core.ventana_resultados_importacion import mostrar_ventana_resultados

                    mostrar_ventana_resultados(
                        parent=self,
                        completos=analisis['completos'],
                        incompletos=analisis['incompletos'],
                        resumen=analisis['resumen'],
                        callback_auditar=self._mostrar_selector_tipo_auditoria,
                        callback_continuar=self._nav_to_visualizar
                    )

                except Exception as e:
                    # Fallback al flujo original si hay error en el análisis
                    logging.warning(f"⚠️ Error en análisis de completitud: {e}")
                    logging.info(f"   Usando flujo de importación original")

                    messagebox.showinfo("Procesamiento", f"✅ Archivo procesado exitosamente:\n{records} registros extraídos")

                    # Actualizar la vista de datos y el dashboard
                    self.refresh_data_and_table()

                    # Actualizar el dashboard si existe
                    if hasattr(self, 'enhanced_dashboard'):
                        try:
                            self.enhanced_dashboard.refresh_all_data()
                        except Exception as e:
                            logging.error(f"Error actualizando dashboard: {e}")

                    # Actualizar lista de archivos
                    self._refresh_files_list()

                    # Redirigir a visualizar datos
                    self._nav_to_visualizar()

            except Exception as e:
                messagebox.showerror("Error", f"❌ Error procesando el archivo:\n{str(e)}")

    def _select_pdf_folder(self):
        """V6.9.25: Carga la carpeta (con subcarpetas) en el explorador NAVEGABLE
        'Archivos disponibles'. NO procesa: el usuario navega, selecciona los PDFs
        (o carpetas completas) y luego pulsa «Procesar seleccionados»."""
        folder_path = filedialog.askdirectory(
            title="Seleccionar carpeta con PDFs",
            initialdir=os.path.join(os.getcwd(), "pdfs_patologia") if os.path.exists("pdfs_patologia") else os.getcwd()
        )
        if not folder_path:
            return
        self._import_root_folder = folder_path
        n = self._build_import_tree(folder_path)
        nombre = os.path.basename(folder_path.rstrip(os.sep)) or folder_path
        if n == 0:
            messagebox.showwarning(
                "Sin archivos",
                f"No se encontraron PDFs en '{nombre}' (ni en sus subcarpetas).")
        else:
            messagebox.showinfo(
                "Carpeta cargada",
                f"Se cargaron {n} PDF(s) de '{nombre}' (incluyendo subcarpetas).\n\n"
                f"En «Archivos disponibles» navegá las subcarpetas (doble clic para abrirlas), "
                f"seleccioná los PDFs —o una carpeta completa— y pulsá «Procesar seleccionados».")
        return

    def _select_pdf_folder_LEGACY_procesar_todo(self):
        """[Obsoleto V6.9.25] Procesaba TODA la carpeta recursivamente. Conservado por
        referencia; ya no se invoca (reemplazado por el explorador navegable)."""
        folder_path = filedialog.askdirectory(
            title="Seleccionar carpeta con PDFs",
            initialdir=os.path.join(os.getcwd(), "pdfs_patologia") if os.path.exists("pdfs_patologia") else os.getcwd()
        )
        if folder_path:
            # V6.9.22: recursivo=True -> incluye PDFs en subcarpetas (cualquier nivel)
            pdf_files = ocr_helpers.obtener_pdfs_en_carpeta(folder_path, recursivo=True)
            if pdf_files:
                # NUEVO: Limpiar lista de registros procesados y obtener peticiones existentes antes del procesamiento
                self._ultimos_registros_procesados = []
                peticiones_antes = set()
                try:
                    import sqlite3
                    from core.database_manager import DB_FILE
                    conn = sqlite3.connect(DB_FILE)
                    cursor = conn.cursor()
                    cursor.execute('SELECT "Numero de caso" FROM informes_ihq')
                    peticiones_antes = set(row[0] for row in cursor.fetchall() if row[0])
                    conn.close()
                except Exception as e:
                    logging.warning(f"Error obteniendo peticiones existentes: {e}")

                processed_count = 0
                total_records = 0
                errors = []

                for pdf_path in pdf_files:
                    try:
                        # V5.3.9: _process_file ahora retorna (records_count, correcciones)
                        records, correcciones = self._process_file(pdf_path)
                        processed_count += 1
                        total_records += records
                    except Exception as e:
                        pdf_name = os.path.basename(pdf_path)
                        errors.append(f"{pdf_name}: {str(e)}")

                # V6.9.31: usar los números de caso REALES capturados durante el
                # procesamiento (incluye reimportaciones). Antes "despues - antes"
                # leía SQLite (vacío con MySQL) y daba 0.
                if processed_count > 0:
                    try:
                        capturados = list(dict.fromkeys(
                            getattr(self, '_ultimos_registros_procesados', []) or []
                        ))
                        self._ultimos_registros_procesados = capturados
                        logging.info(f"📋 Registros procesados: {len(capturados)}")
                        if capturados:
                            logging.info(f"   IDs: {', '.join(capturados[:5])}" + (" ..." if len(capturados) > 5 else ""))
                    except Exception as e:
                        logging.warning(f"⚠️ Error capturando registros procesados: {e}")
                        self._ultimos_registros_procesados = []

                    # NUEVO: Analizar completitud de registros
                    try:
                        from core.validation_checker import analizar_batch_registros

                        numeros_peticion_procesados = self._ultimos_registros_procesados
                        logging.info(f"🔍 Analizando completitud de {len(numeros_peticion_procesados)} registros...")
                        analisis = analizar_batch_registros(numeros_peticion_procesados)

                        logging.info(f"✅ Análisis completado:")
                        logging.info(f"   • Completos: {analisis['resumen']['completos']}")
                        logging.info(f"   • Incompletos: {analisis['resumen']['incompletos']}")

                        # Actualizar vista antes de mostrar ventana
                        try:
                            self.refresh_data_and_table()
                            self.after(500, self._delayed_refresh_after_processing)

                            if hasattr(self, 'enhanced_dashboard'):
                                self.enhanced_dashboard.refresh_all_data()
                        except Exception as e:
                            logging.warning(f"⚠️ Error en refresh: {e}")

                        # Actualizar lista de archivos
                        self._refresh_files_list()

                        # Mostrar ventana de resultados con análisis de completitud
                        from core.ventana_resultados_importacion import mostrar_ventana_resultados

                        mostrar_ventana_resultados(
                            parent=self,
                            completos=analisis['completos'],
                            incompletos=analisis['incompletos'],
                            resumen=analisis['resumen'],
                            callback_auditar=self._mostrar_selector_tipo_auditoria,
                            callback_continuar=self._nav_to_visualizar
                        )

                    except Exception as e:
                        # Fallback al flujo original si hay error en el análisis
                        logging.warning(f"⚠️ Error en análisis de completitud: {e}")
                        logging.info(f"   Usando flujo de importación original")

                        # Mostrar resultado tradicional
                        msg = f"✅ Procesados {processed_count} de {len(pdf_files)} archivos\n"
                        msg += f"Total de registros: {total_records}"
                        if errors:
                            msg += f"\n\n❌ Errores en {len(errors)} archivos:\n" + "\n".join(errors[:3])
                            if len(errors) > 3:
                                msg += f"\n... y {len(errors) - 3} más"

                        messagebox.showinfo("Procesamiento completado", msg)

                        # Actualizar vistas
                        self.refresh_data_and_table()
                        if hasattr(self, 'enhanced_dashboard'):
                            try:
                                self.enhanced_dashboard.refresh_all_data()
                            except Exception as e:
                                logging.error(f"Error actualizando dashboard: {e}")
                        self._refresh_files_list()
                else:
                    # No se procesó ningún archivo
                    error_msg = "❌ No se pudo procesar ningún archivo.\n\nErrores encontrados:\n"
                    error_msg += "\n".join(errors[:5])  # Mostrar los primeros 5 errores
                    messagebox.showerror("Error de procesamiento", error_msg)
            else:
                messagebox.showwarning("Sin archivos", "No se encontraron archivos PDF en la carpeta seleccionada.")

    def _build_import_tree(self, root_folder):
        """V6.9.25: Construye el árbol navegable (carpetas/subcarpetas + PDFs) en
        'Archivos disponibles'. Guarda iid->ruta en self._tree_paths. Devuelve el
        número de PDFs encontrados."""
        tree = getattr(self, 'files_tree', None)
        if tree is None:
            return 0
        try:
            tree.delete(*tree.get_children())
        except Exception:
            pass
        self._tree_paths = {}
        if not root_folder or not os.path.isdir(root_folder):
            return 0
        n_pdfs = [0]

        # V6.9.66: colorear según si el PDF YA está analizado. Con cientos de archivos no
        # había forma de saberlo salvo procesarlos. Se calcula ANTES de pintar, con UNA
        # sola consulta a la BD para todos (rápido aunque haya cientos).
        estados = {}
        try:
            from core.estado_pdfs import estado_pdfs
            todos_pdf = []
            for base, _dirs, files in os.walk(root_folder):
                todos_pdf += [os.path.join(base, f) for f in files if f.lower().endswith('.pdf')]
            if todos_pdf:
                estados = estado_pdfs(todos_pdf)
        except Exception as _e_est:
            logging.warning(f"[estado-pdfs] no se pudo calcular (se muestra sin color): {_e_est}")

        # Colores del árbol. Azul = ya analizado (el usuario pidió azul), naranja = a medias,
        # verde/negrita = sin analizar (lo que hay que procesar).
        try:
            tree.tag_configure('pdf_completo', foreground='#1565C0')
            tree.tag_configure('pdf_parcial', foreground='#E65100')
            tree.tag_configure('pdf_nuevo', foreground='#1B5E20')
            tree.tag_configure('pdf_desconocido', foreground='#616161')
        except Exception:
            pass
        _TAG = {'COMPLETO': 'pdf_completo', 'PARCIAL': 'pdf_parcial',
                'NUEVO': 'pdf_nuevo', 'DESCONOCIDO': 'pdf_desconocido'}
        _MARCA = {'COMPLETO': '✓', 'PARCIAL': '◐', 'NUEVO': '●', 'DESCONOCIDO': '?'}

        def _insert(parent, path, depth):
            try:
                entradas = sorted(os.listdir(path), key=lambda s: s.lower())
            except Exception:
                return
            dirs = [e for e in entradas if os.path.isdir(os.path.join(path, e))]
            pdfs = [e for e in entradas if e.lower().endswith('.pdf')
                    and os.path.isfile(os.path.join(path, e))]
            for d in dirs:
                full = os.path.join(path, d)
                iid = tree.insert(parent, 'end', text=f"  📁 {d}", open=(depth == 0))
                self._tree_paths[iid] = full
                _insert(iid, full, depth + 1)
            for p in pdfs:
                full = os.path.join(path, p)
                inf = estados.get(full) or {}
                est = inf.get('estado', 'DESCONOCIDO')
                # el detalle "37/50" dice de un vistazo por dónde se quedó
                det = ''
                if inf.get('total'):
                    det = f"   ({inf.get('en_bd', 0)}/{inf['total']})"
                iid = tree.insert(parent, 'end',
                                  text=f"  {_MARCA.get(est, '?')} 📄 {p}{det}",
                                  tags=(_TAG.get(est, 'pdf_desconocido'),))
                self._tree_paths[iid] = full
                n_pdfs[0] += 1

        nombre = os.path.basename(root_folder.rstrip(os.sep)) or root_folder
        root_iid = tree.insert('', 'end', text=f"📂 {nombre}", open=True)
        self._tree_paths[root_iid] = root_folder
        _insert(root_iid, root_folder, 0)
        return n_pdfs[0]

    def _seleccionar_pendientes(self):
        """V6.9.66: selecciona solos los PDFs que faltan por analizar (sin analizar +
        a medias) y deja fuera los ya hechos. Con cientos de archivos, marcarlos a mano
        era el trabajo pesado — y el que se presta a saltarse alguno por error."""
        tree = getattr(self, 'files_tree', None)
        if tree is None:
            return
        try:
            from core.estado_pdfs import estado_pdfs, COMPLETO
            rutas = [p for p in getattr(self, '_tree_paths', {}).values()
                     if p and p.lower().endswith('.pdf') and os.path.isfile(p)]
            if not rutas:
                messagebox.showinfo("Sin archivos", "No hay PDFs en la lista.")
                return
            info = estado_pdfs(rutas)
            # DESCONOCIDO también entra: si no sabemos si está hecho, mejor que el
            # usuario lo vea seleccionado que arriesgarse a dejarlo sin procesar.
            pend = {p for p, v in info.items() if v.get('estado') != COMPLETO}
            iids = [i for i, p in self._tree_paths.items() if p in pend]
            tree.selection_set(iids)
            if iids:
                tree.see(iids[0])
            hechos = len(rutas) - len(pend)
            if not iids:
                messagebox.showinfo(
                    "Todo analizado",
                    f"Los {len(rutas)} PDFs de la carpeta ya están analizados.\n"
                    "No hay nada pendiente.")
            else:
                n_medias = sum(1 for p in pend if (info.get(p) or {}).get('estado') == 'PARCIAL')
                n_nuevos = len(pend) - n_medias
                logging.info(f"[pendientes] {len(iids)} seleccionados, {hechos} ya analizados")
                messagebox.showinfo(
                    "Pendientes seleccionados",
                    f"Seleccionados {len(iids)} PDF(s) pendientes:\n"
                    f"   • {n_nuevos} sin analizar\n"
                    f"   • {n_medias} a medias\n\n"
                    f"{hechos} ya analizados quedaron FUERA de la selección.\n\n"
                    "Pulsá «Procesar seleccionados» para analizarlos.")
        except Exception as e:
            logging.error(f"[seleccionar-pendientes] {e}")
            messagebox.showerror("Error", f"No se pudieron seleccionar los pendientes:\n{e}")

    def _recolectar_pdfs_seleccionados(self):
        """V6.9.25: Rutas de PDFs según la selección del árbol (PDFs seleccionados +
        todos los PDFs bajo carpetas seleccionadas). Sin duplicados y en orden."""
        tree = getattr(self, 'files_tree', None)
        if tree is None:
            return []
        paths = getattr(self, '_tree_paths', {})
        rutas = []

        def _colectar(iid):
            p = paths.get(iid)
            if p and p.lower().endswith('.pdf') and os.path.isfile(p):
                rutas.append(p)
            for hijo in tree.get_children(iid):
                _colectar(hijo)

        for iid in tree.selection():
            _colectar(iid)
        vistos = set()
        return [p for p in rutas if not (p in vistos or vistos.add(p))]

    def _refresh_files_list(self):
        """V6.9.25: Reconstruye el árbol navegable. Usa la última carpeta cargada
        (si se eligió en «Seleccionar carpeta») o, por defecto, pdfs_patologia."""
        if not hasattr(self, 'files_tree') or self.files_tree is None:
            logging.warning("Advertencia: files_tree no está disponible")
            return
        root = getattr(self, '_import_root_folder', None)
        if not root or not os.path.isdir(root):
            root = os.path.join(os.getcwd(), "pdfs_patologia")
            if not os.path.exists(root):
                try:
                    os.makedirs(root, exist_ok=True)
                except Exception:
                    pass
        self._build_import_tree(root)

    def _peticiones_existentes_bd(self) -> set:
        """V6.9.30: set de 'Numero de caso' de la BD ACTIVA (MySQL via adapter,
        o SQLite si config=sqlite). Reemplaza las lecturas directas a la SQLite
        legacy que daban conteos erróneos ('Total Procesados' = 0 con MySQL)."""
        try:
            from core.db_adapter import cursor_ctx, quote_ident
            with cursor_ctx() as (conn, cur):
                cur.execute(
                    f'SELECT {quote_ident("Numero de caso")} FROM {quote_ident("informes_ihq")}'
                )
                return set(r[0] for r in cur.fetchall() if r[0])
        except Exception as e:
            logging.warning(f"⚠️ _peticiones_existentes_bd: {e}")
            return set()

    def _process_selected_files(self):
        """Procesar los archivos seleccionados de la lista (con barra de progreso)"""
        # V6.9.25: verificar que existe el árbol navegable
        if not hasattr(self, 'files_tree') or self.files_tree is None:
            messagebox.showerror("Error", "El explorador de archivos no está disponible.")
            return

        rutas_pdf = self._recolectar_pdfs_seleccionados()
        if not rutas_pdf:
            messagebox.showwarning(
                "Sin selección",
                "Seleccioná uno o más PDFs (o una carpeta completa) en «Archivos disponibles».")
            return

        # Pre-validar archivos y verificar duplicados (rápido, main thread)
        files_to_process = []
        # V6.9.31: permitir REIMPORTAR archivos ya importados (re-extracción).
        # reimportar_todos: None = preguntar por cada uno; True = reimportar todos.
        reimportar_todos = None
        for file_path in rutas_pdf:
            filename = os.path.basename(file_path)
            duplicado_info = self._verificar_archivo_duplicado(file_path, filename)
            if duplicado_info["es_duplicado"]:
                decision = True if reimportar_todos else None
                if reimportar_todos is None:
                    resp = messagebox.askyesnocancel(
                        "Archivo ya importado",
                        f"El archivo '{filename}' ya fue importado "
                        f"(caso {duplicado_info['numero_peticion']}, "
                        f"informe {duplicado_info['fecha_informe']}).\n\n"
                        f"¿Desea REIMPORTARLO? Se vuelve a extraer y se REEMPLAZA "
                        f"el registro existente.\n\n"
                        f"   •  Sí  = Reimportar este archivo\n"
                        f"   •  No  = Omitir este archivo\n"
                        f"   •  Cancelar = Reimportar TODOS los ya importados de la selección"
                    )
                    if resp is None:          # Cancelar -> reimportar todos
                        reimportar_todos = True
                        decision = True
                    else:
                        decision = resp        # Sí = reimportar este; No = omitir
                if not decision:
                    continue  # omitir este archivo (no se reprocesa)
                # decision True -> reimportar: cae al append; save_records hace
                # UPSERT por "Numero de caso" (reemplaza el registro existente).

            files_to_process.append((file_path, filename))

        if not files_to_process:
            messagebox.showinfo("Sin archivos", "No hay archivos nuevos para procesar.")
            return

        # Obtener peticiones existentes antes del procesamiento
        # V6.9.30 FIX: leer de la BD ACTIVA (MySQL via adapter), NO de la SQLite
        # legacy. Antes leía SQLite (no se actualiza con el guardado a MySQL),
        # por eso "Total Procesados" salía en 0 aunque sí se guardaran casos.
        peticiones_antes = self._peticiones_existentes_bd()

        # Mostrar overlay de progreso
        self._show_processing_overlay(len(files_to_process))

        # Estado compartido con el thread
        self._processing_result = {
            "done": False,
            "processed_count": 0,
            "total_records": 0,
            "errors": [],
            "correcciones": [],
            "peticiones_antes": peticiones_antes,
            "current_file": "",
            "current_index": 0,
            "total_files": len(files_to_process),
        }

        # Limpiar registros procesados
        self._ultimos_registros_procesados = []

        # Lanzar thread de procesamiento
        thread = threading.Thread(
            target=self._process_files_worker,
            args=(files_to_process,),
            daemon=True
        )
        thread.start()

        # Polling desde main thread para actualizar progreso
        self._poll_processing_progress()

    # ════════════════════════════════════════════════════════════════════
    # V6.7.0 — Procesar con IA: diagnóstico de cobertura del OCR
    # ════════════════════════════════════════════════════════════════════
    # Pipeline alternativo paralelo al tradicional. Para cada PDF:
    #   1. OCR completo (sin segmentación) → texto del PDF entero
    #   2. Texto entero → LLM con prompt de extracción de TODOS los IHQ
    #   3. LLM devuelve JSON con array de diagnósticos identificados
    #   4. Resultado se guarda en informes_ia/ para comparar con BD
    #
    # Objetivo: si el extractor tradicional procesó 588 casos pero el LLM
    # encuentra 995, el problema está en la segmentación/extractores. Si
    # el LLM también encuentra 588, el OCR sí está perdiendo casos.
    #
    # NO modifica BD, NO toca extractores existentes. Es solo lectura +
    # generación de un reporte separado.

    # V6.8.0 — Prompt expandido a 184 campos. El LLM debe rellenar TODAS
    # las columnas del informe (paciente, procedimiento, dx, biomarcadores).
    # Si un campo NO está en el informe → devolver literalmente "N/A".
    # El JSON schema fuerza la estructura; el prompt solo guía el contenido.
    _PROMPT_SYSTEM_IA_OCR = (
        "Eres analista patológico expertx en informes IHQ del HUV. Recibirás\n"
        "texto OCR de UN informe IHQ y debés extraer EXACTAMENTE 184 campos.\n"
        "El texto SIEMPRE contiene 1 informe completo — nunca está vacío.\n\n"
        "REGLA UNIVERSAL: Si un campo NO aparece en el informe, devolvé el\n"
        "string literal 'N/A' (sin comillas dentro del JSON, solo la cadena).\n"
        "NUNCA inventes datos. NUNCA dejes un campo en blanco.\n"
        "Copiá los valores TAL CUAL aparecen en el PDF (incluso typos como\n"
        "'CARICNOMA' o 'NOTHINGHAM').\n\n"
        "════════ CAMPOS A EXTRAER (agrupados por categoría) ════════\n\n"
        "▶ IDENTIFICACIÓN ADMINISTRATIVA (header del PDF):\n"
        "  numero_de_caso       → 'IHQXXXXXX' de 'N. peticion :'. Solo IHQ+dígitos.\n"
        "  hospitalizado        → 'SI'/'NO'/'N/A' (campo Hospitalizado).\n"
        "  sede                 → sede del hospital (ej: 'HUV', 'PRINCIPAL').\n"
        "  eps                  → nombre EPS del paciente.\n"
        "  servicio             → servicio clínico solicitante.\n"
        "  medico_tratante      → nombre del médico que pidió el estudio.\n"
        "  especialidad         → especialidad del médico.\n"
        "  datos_clinicos       → 'Datos Clínicos:' (texto completo del párrafo).\n"
        "  tipo_de_documento    → CC/TI/CE/RC/PA/N/A.\n"
        "  n_de_identificacion  → número de cédula/documento.\n"
        "  primer_nombre / segundo_nombre / primer_apellido / segundo_apellido.\n"
        "  edad                 → solo número (ej: '45'). N/A si no aparece.\n"
        "  genero               → 'M'/'F'/'MASCULINO'/'FEMENINO'/N/A.\n"
        "  departamento         → departamento del paciente (ej: 'VALLE DEL CAUCA').\n"
        "  municipio            → municipio (ej: 'CALI').\n"
        "  cups                 → código CUPS del procedimiento.\n\n"
        "▶ PROCEDIMIENTO (header de Estudios solicitados):\n"
        "  tipo_de_examen       → 'INMUNOHISTOQUIMICA' o lo que diga.\n"
        "  procedimiento        → procedimiento (ej: 'BIOPSIA', 'RESECCIÓN').\n"
        "  organo               → órgano anatómico LIMPIO en MAYÚSCULAS.\n"
        "    Reglas: SIN 'LESION'/'BX'/'TUMOR'/'MUCOSA DE'/procedimientos.\n"
        "    Ejemplos: 'LESION ESTOMAGO'→'ESTOMAGO', 'BX MEDULA OSEA'→'MEDULA OSEA',\n"
        "    'NEFRECTOMIA RADICAL IZQUIERDA'→'RIÑON IZQUIERDO',\n"
        "    'GANGLIO PROFUNDO METASTÁSICO'→'GANGLIO LINFATICO'.\n"
        "  fecha_de_toma_1_fecha_de_la_toma            → fecha (DD/MM/YYYY).\n"
        "  fecha_de_ingreso_2_fecha_de_la_muestra      → fecha.\n"
        "  fecha_informe                               → fecha del informe.\n"
        "  patologo                                    → nombre del patólogo firmante.\n\n"
        "▶ DIAGNÓSTICO CLÍNICO (sección DIAGNÓSTICO al final del informe):\n"
        "  malignidad           → 'MALIGNO' / 'BENIGNO' / 'PRE-MALIGNO' / 'N/A'.\n"
        "    Inferí del dx: carcinoma/sarcoma/linfoma/melanoma=MALIGNO;\n"
        "    inflamación/hiperplasia reactiva/lipoma/adenoma/fibroadenoma=BENIGNO.\n"
        "    'NEGATIVO PARA NEOPLASIA' o 'NEGATIVO PARA MALIGNIDAD'=BENIGNO.\n"
        "    'LESIÓN INTRAEPITELIAL ALTO GRADO/NIC 3' = PRE-MALIGNO.\n"
        "  descripcion_macroscopica → texto completo de DESCRIPCIÓN MACROSCÓPICA.\n"
        "  descripcion_microscopica → texto completo de DESCRIPCIÓN MICROSCÓPICA.\n"
        "  descripcion_diagnostico  → texto del párrafo final 'Diagnóstico:'\n"
        "                             si existe (alternativo a Diagnóstico Principal).\n"
        "  diagnostico_coloracion   → texto sobre la coloración usada (HE, etc).\n"
        "  diagnostico_principal    → el contenido sustantivo de la sección\n"
        "    DIAGNÓSTICO del informe. NO es solo para tumores con nombre — es\n"
        "    para CUALQUIER conclusión diagnóstica del patólogo.\n"
        "    • Devolvé la primera línea sustantiva con calificadores\n"
        "      ('WHO GRADO X', 'NOTTINGHAM 7/9', 'p40 POSITIVO', '(g2, ptc3, v0)').\n"
        "    • Descartá bullets que SON sub-items distintos (PATRÓN MICROSATELITAL,\n"
        "      HER-2: NEG, etc) — esos van en sus campos específicos.\n"
        "    • SALTAR PREÁMBULOS: 'LOS HALLAZGOS COMPATIBLES CON X'→X,\n"
        "      'SUGIEREN UN X'→X, 'FAVORECEN UN X'→X.\n"
        "      EXCEPCIÓN: 'FAVORECE CARCINOMA...' (sin 'LOS HALLAZGOS') sí mantenelo\n"
        "      cuando el patólogo no es definitivo.\n"
        "    • EXCEPCIONES — copiá frase completa: 'VER DESCRIPCIÓN MICROSCÓPICA',\n"
        "      'TEJIDO SIN REPRESENTACIÓN...', 'CÉLULAS GANGLIONARES PRESENTES'.\n"
        "    # V6.9.9 FIX: diagnósticos no-tumorales — el campo NO debe quedar\n"
        "    # en N/A cuando el dx no es un tumor con nombre de entidad.\n"
        "    • REGLA ANTI-N/A: diagnostico_principal SOLO puede ser 'N/A' si la\n"
        "      sección DIAGNÓSTICO está literalmente ausente o vacía. Si tiene\n"
        "      CUALQUIER texto sustantivo, copialo — aunque NO sea el nombre de\n"
        "      un tumor (carcinoma/linfoma/sarcoma/etc.).\n"
        "    • DIAGNÓSTICOS NEGATIVOS — son diagnósticos válidos, copialos TAL CUAL:\n"
        "      'NEGATIVO PARA MALIGNIDAD', 'NEGATIVO PARA NEOPLASIA',\n"
        "      'MUESTRA NEGATIVA PARA DISPLASIA', 'NEGATIVO PARA DISPLASIA',\n"
        "      'NEGATIVO PARA COMPROMISO POR LINFOMA DE CÉLULAS B', etc.\n"
        "    • MÉDULA ÓSEA DESCRIPTIVA — el dx puede ser una descripción de la\n"
        "      celularidad sin nombre de entidad; copiá la conclusión completa:\n"
        "      'CELULARIDAD GLOBAL DEL 60%. RELACIÓN MIELOIDE-ERITROIDE...',\n"
        "      'HIPOPLASIA (CELULARIDAD DISMINUIDA...)', 'HIPERPLASIA...',\n"
        "      'APLASIA...', 'MADURACIÓN ADECUADA DE LAS TRES LÍNEAS CELULARES'.\n"
        "    • REDIRECCIÓN — si la sección DIAGNÓSTICO dice únicamente que el dx\n"
        "      está en otra sección, copiá esa frase TAL CUAL (NO la conviertas\n"
        "      en N/A): 'VER DESCRIPCIÓN MICROSCÓPICA Y COMENTARIO'.\n"
        "    • LESIONES BENIGNAS / INFLAMATORIAS — son diagnósticos válidos aunque\n"
        "      no sean cáncer; copialos completos: dermatitis (p.ej. 'DERMATITIS\n"
        "      PERIVASCULAR Y PERIANEXIAL SUPERFICIAL Y PROFUNDA...'), gastritis,\n"
        "      colitis, fibrosis, cambios reparativos, hiperplasia reactiva,\n"
        "      'TEJIDO SIN REPRESENTACIÓN DE PARÉNQUIMA RENAL', etc.\n"
        "  factor_pronostico    → texto sobre pronóstico/grado/estadio si aparece.\n\n"
        "▶ ESTUDIOS IHQ GENERALES:\n"
        "  ihq_estudios_solicitados → lista de marcadores solicitados (raw text).\n"
        "  ihq_organo               → órgano del estudio IHQ (puede ser igual a 'organo').\n"
        "  congelaciones_otros_estudios → texto sobre congelaciones si hay.\n"
        "  liquidos_5_tipo_histologico  → líquidos analizados (raw).\n"
        "  citometria_de_flujo_5_tipo_histologico → resultado de citometría si hay.\n\n"
        "▶ BIOMARCADORES (formato esperado en cada campo):\n"
        "  Estados típicos: 'POSITIVO', 'NEGATIVO', 'EQUÍVOCO', 'POSITIVO 30%',\n"
        "  'NEGATIVO (SCORE 0)', 'PERDIDA DE EXPRESIÓN', 'CONSERVADO', etc.\n"
        "  Si el biomarcador NO se mencionó en el informe → 'N/A'.\n"
        "  Buscá menciones tipo: 'HER-2: POSITIVO (SCORE 3+)', 'Ki-67: 25%',\n"
        "  'p16 POSITIVO', 'CD20: NEGATIVO', etc.\n"
        "  Biomarcadores comunes a buscar:\n"
        "    ihq_her2, ihq_ki_67, ihq_receptor_estrogenos, ihq_receptor_progesterona,\n"
        "    ihq_pdl_1, ihq_p16_estado (POSITIVO/NEGATIVO), ihq_p16_porcentaje (% si aparece),\n"
        "    ihq_p40_estado, ihq_ck7, ihq_ck20, ihq_cdx2, ihq_gata3, ihq_p53, ihq_p63,\n"
        "    ihq_ttf1, ihq_napsin, ihq_s100, ihq_melan_a, ihq_hmb45, ihq_sox10,\n"
        "    ihq_cd3, ihq_cd5, ihq_cd10, ihq_cd20, ihq_cd23, ihq_cd30, ihq_cd45,\n"
        "    ihq_cd56, ihq_cd99, ihq_cd117, ihq_cd138, ihq_kappa, ihq_lambda,\n"
        "    ihq_bcl2, ihq_bcl6, ihq_mum1, ihq_mlh1, ihq_msh2, ihq_msh6, ihq_pms2,\n"
        "    ihq_synaptofisina, ihq_cromogranina, ihq_calretinin, ihq_dog1, ihq_alk,\n"
        "    ihq_psa, ihq_idh, ihq_atrx, ihq_gfap, ihq_olig2, ihq_e_cadherina,\n"
        "    ihq_vimentina, ihq_ema, ihq_pax8, ihq_pax5, ihq_wt1, ihq_inhibina,\n"
        "    ihq_actina_musculo_liso (SMA), ihq_actina_musculo_especifica,\n"
        "    ihq_desmin, ihq_myogenin, ihq_hcg, ihq_afp, ihq_oct4, ihq_sall4,\n"
        "    ihq_hepar, ihq_arginasa, ihq_cea, ihq_ca19_9, ihq_eber, ihq_cmv,\n"
        "    ihq_lmp1, ihq_hhv8, ihq_ber_ep4, ihq_h_caldesmon, etc.\n"
        "  Si en el OCR ves un biomarcador que NO está en este alias-list,\n"
        "  ignoralo (no inventes alias).\n\n"
        "════════ FORMATO DE SALIDA ════════\n"
        "Devolvé JSON con UNA SOLA entrada en 'diagnosticos' que contenga\n"
        "TODOS los 184 campos. El schema strict lo valida — todos los\n"
        "campos son obligatorios, los que no apliquen = 'N/A'."
    )

    def _process_selected_files_ia(self):
        """V6.7.0 — Procesa PDFs seleccionados con OCR completo + LLM para
        diagnosticar cobertura del extractor tradicional.

        Para cada PDF: OCR → texto completo → LLM extrae todos los IHQ.
        Resultado guardado en informes_ia/extraccion_ia_<pdf>.json
        """
        if not hasattr(self, 'files_tree') or self.files_tree is None:
            messagebox.showerror("Error", "El explorador de archivos no está disponible.")
            return

        rutas_pdf = self._recolectar_pdfs_seleccionados()
        if not rutas_pdf:
            messagebox.showwarning(
                "Sin selección",
                "Seleccioná uno o más PDFs (o una carpeta) para procesar con IA."
            )
            return

        # Confirmación: este pipeline es lento (1-5 min por PDF)
        n = len(rutas_pdf)
        respuesta = messagebox.askyesno(
            "Procesar con IA",
            f"Vas a procesar {n} PDF(s) con el pipeline alternativo de IA.\n\n"
            f"Para cada PDF:\n"
            f"  1. Se hace OCR completo (texto entero)\n"
            f"  2. El texto se envía al LLM (LM Studio o proveedor cloud)\n"
            f"  3. El LLM identifica TODOS los IHQ y sus diagnósticos\n"
            f"  4. Resultado se guarda en informes_ia/\n\n"
            f"Tiempo estimado: ~2-5 minutos por PDF.\n"
            f"NO modifica la BD ni los extractores. Solo genera un reporte.\n\n"
            f"¿Continuar?"
        )
        if not respuesta:
            return

        files_to_process = [(p, os.path.basename(p)) for p in rutas_pdf]
        if not files_to_process:
            messagebox.showinfo("Sin archivos", "Ningún archivo válido para procesar.")
            return

        # Crear directorio de salida
        out_dir = os.path.join(os.getcwd(), "informes_ia")
        os.makedirs(out_dir, exist_ok=True)

        # Estado compartido con worker
        self._processing_result_ia = {
            "done": False,
            "cancelled": False,
            "current_index": 0,
            "total_files": len(files_to_process),
            "current_file": "",
            "current_stage": "Inicializando...",
            "current_chunk": 0,
            "total_chunks": 0,
            "live_diagnosticos": [],   # lista compartida — worker hace append
            "results": [],
            "errors": [],
        }
        # Cuántos diagnósticos en vivo ya pintamos en el treeview
        self._ia_diagnosticos_pintados = 0

        # Mostrar ventana de progreso con tabla en vivo
        self._show_ia_progress_window(len(files_to_process))

        thread = threading.Thread(
            target=self._process_files_ia_worker,
            args=(files_to_process, out_dir),
            daemon=True
        )
        thread.start()
        self._poll_processing_progress_ia()

    def _show_ia_progress_window(self, num_files):
        """V6.7.0 — Ventana grande con barra de progreso + tabla en vivo de
        diagnósticos identificados por la IA mientras procesa los chunks."""
        win = tk.Toplevel(self)
        self._ia_progress_win = win
        win.title("🤖 Procesando con IA — análisis en vivo")
        win.transient(self)
        win.protocol("WM_DELETE_WINDOW", lambda: None)  # Solo se cierra al terminar

        # Centrar (más grande que el overlay normal)
        w, h = 920, 600
        x = self.winfo_x() + (self.winfo_width() - w) // 2
        y = self.winfo_y() + (self.winfo_height() - h) // 2
        win.geometry(f"{w}x{h}+{max(0, x)}+{max(0, y)}")
        win.minsize(700, 450)

        # Cabecera con info de PDF actual
        header = ttk.Frame(win, padding=15)
        header.pack(fill=X)

        self._ia_lbl_pdf = ttk.Label(
            header,
            text=f"Iniciando procesamiento de {num_files} PDF(s)...",
            font=("Segoe UI", 11, "bold")
        )
        self._ia_lbl_pdf.pack(anchor=W)

        self._ia_lbl_chunk = ttk.Label(
            header,
            text="",
            font=("Segoe UI", 9),
            foreground="#555"
        )
        self._ia_lbl_chunk.pack(anchor=W, pady=(3, 0))

        # Barras de progreso (PDF y chunk)
        progress_frame = ttk.Frame(win, padding=(15, 5, 15, 10))
        progress_frame.pack(fill=X)

        ttk.Label(progress_frame, text="PDFs:", font=("Segoe UI", 9)).grid(row=0, column=0, sticky=W, padx=(0, 8))
        self._ia_progress_pdfs = ttk.Progressbar(
            progress_frame, mode="determinate", maximum=num_files
        )
        self._ia_progress_pdfs.grid(row=0, column=1, sticky="ew", pady=2)
        progress_frame.grid_columnconfigure(1, weight=1)

        ttk.Label(progress_frame, text="Chunks:", font=("Segoe UI", 9)).grid(row=1, column=0, sticky=W, padx=(0, 8), pady=(4, 0))
        self._ia_progress_chunks = ttk.Progressbar(
            progress_frame, mode="determinate", maximum=1
        )
        self._ia_progress_chunks.grid(row=1, column=1, sticky="ew", pady=(4, 0))

        # Contadores: esta sesión + acumulado en BD
        counter_frame = ttk.Frame(win, padding=(15, 0, 15, 5))
        counter_frame.pack(fill=X)
        self._ia_lbl_counter = ttk.Label(
            counter_frame,
            text="Diagnósticos en esta sesión: 0",
            font=("Segoe UI", 10, "bold"),
            foreground="#0a7"
        )
        self._ia_lbl_counter.pack(anchor=W)

        # Total acumulado en BD (todas las sesiones previas + actual)
        try:
            from core.diagnosticos_ia_db import count_total as _ia_count_total
            total_acumulado = _ia_count_total()
        except Exception:
            total_acumulado = 0
        self._ia_lbl_acumulado = ttk.Label(
            counter_frame,
            text=f"Total acumulado en BD (todas las sesiones): {total_acumulado}",
            font=("Segoe UI", 9),
            foreground="#555"
        )
        self._ia_lbl_acumulado.pack(anchor=W, pady=(2, 0))

        # Tabla en vivo
        # V6.8.0 — Tabla expandida a TODAS las 184 columnas que extrae la IA
        # (mismas que la BD principal). Usa scroll horizontal + vertical.
        # Para no romper compat, las 3 primeras columnas siguen siendo
        # numero_peticion / dx / organo (alias visibles), después vienen
        # todas las columnas reales en el orden de COLUMNAS_IA.
        try:
            from core.columnas_huv_ia import COLUMNAS_IA as _COLS_IA
        except Exception:
            _COLS_IA = []
        table_frame = ttk.Frame(win, padding=(15, 5, 15, 10))
        table_frame.pack(fill=BOTH, expand=True)

        # Construir lista de columnas: 3 visibles primero + todas las BD
        # Cada column id = índice en _COLS_IA (col_0, col_1, ...)
        # Los 3 primeros son alias rápidos para visualizar el dato más relevante
        columns = ["ihq", "diagnostico", "organo"]
        col_headers = {"ihq": "N° IHQ", "diagnostico": "Diagnóstico", "organo": "Órgano"}
        col_widths = {"ihq": 110, "diagnostico": 400, "organo": 160}
        # Agregar todas las columnas de la BD
        self._ia_columnas_bd = list(_COLS_IA)
        for idx, bd_col in enumerate(_COLS_IA):
            col_id = f"c{idx}"
            columns.append(col_id)
            col_headers[col_id] = bd_col
            # Anchos por categoría (admin grandes, biomarcadores chicos)
            if idx < 19:           # Admin
                col_widths[col_id] = 130
            elif idx < 26:         # Procedimiento
                col_widths[col_id] = 130
            elif idx < 33:         # Dx clínico
                col_widths[col_id] = 200
            elif idx < 38:         # Estudios IHQ
                col_widths[col_id] = 180
            else:                  # Biomarcadores
                col_widths[col_id] = 110

        self._ia_treeview = ttk.Treeview(
            table_frame, columns=columns, show="headings", height=15
        )
        for c in columns:
            self._ia_treeview.heading(c, text=col_headers[c])
            self._ia_treeview.column(c, width=col_widths.get(c, 120), anchor=W, stretch=False)

        scroll_y = ttk.Scrollbar(table_frame, orient="vertical", command=self._ia_treeview.yview)
        scroll_x = ttk.Scrollbar(table_frame, orient="horizontal", command=self._ia_treeview.xview)
        self._ia_treeview.configure(yscrollcommand=scroll_y.set, xscrollcommand=scroll_x.set)

        # Layout con scroll en ambas direcciones
        self._ia_treeview.grid(row=0, column=0, sticky="nsew")
        scroll_y.grid(row=0, column=1, sticky="ns")
        scroll_x.grid(row=1, column=0, sticky="ew")
        table_frame.grid_rowconfigure(0, weight=1)
        table_frame.grid_columnconfigure(0, weight=1)

        # V6.7.2 — Soporte de copia: shortcuts + menú contextual derecho
        self._ia_treeview.bind("<Control-c>", lambda e: self._ia_copy_selected_to_clipboard())
        self._ia_treeview.bind("<Control-C>", lambda e: self._ia_copy_selected_to_clipboard())
        self._ia_treeview.bind("<Control-a>", lambda e: self._ia_select_all_rows())
        self._ia_treeview.bind("<Control-A>", lambda e: self._ia_select_all_rows())
        # Menú contextual al click derecho
        self._ia_context_menu = tk.Menu(win, tearoff=0)
        self._ia_context_menu.add_command(
            label="Copiar selección (Ctrl+C)",
            command=self._ia_copy_selected_to_clipboard
        )
        self._ia_context_menu.add_command(
            label="Copiar todo",
            command=self._ia_copy_all_to_clipboard
        )
        self._ia_context_menu.add_separator()
        self._ia_context_menu.add_command(
            label="Seleccionar todo (Ctrl+A)",
            command=self._ia_select_all_rows
        )
        self._ia_context_menu.add_separator()
        self._ia_context_menu.add_command(
            label="Exportar todo a CSV...",
            command=self._ia_export_to_csv
        )
        self._ia_treeview.bind("<Button-3>", self._ia_show_context_menu)

        # Footer con botones de copia + estado + cerrar
        footer = ttk.Frame(win, padding=(15, 5, 15, 15))
        footer.pack(fill=X)

        self._ia_lbl_status = ttk.Label(
            footer, text="🔄 OCR del PDF...", font=("Segoe UI", 9), foreground="#06b"
        )
        self._ia_lbl_status.pack(side=LEFT)

        self._ia_btn_cerrar = ttk.Button(
            footer, text="Cerrar", state="disabled",
            command=lambda: self._ia_progress_win.destroy()
        )
        self._ia_btn_cerrar.pack(side=RIGHT)

        ttk.Button(
            footer, text="📋 Copiar todo",
            command=self._ia_copy_all_to_clipboard,
            bootstyle="secondary-outline"
        ).pack(side=RIGHT, padx=(0, 8))

        ttk.Button(
            footer, text="💾 Exportar CSV",
            command=self._ia_export_to_csv,
            bootstyle="secondary-outline"
        ).pack(side=RIGHT, padx=(0, 8))

        win.update_idletasks()

    # V6.7.3 — Patrones de limpieza post-LLM
    # El LLM a veces devuelve dx con frases introductorias del patólogo
    # ("Muslo derecho. Lesión. Biopsia. Estudio de inmunohistoquímica:")
    # o preámbulos ("LOS HALLAZGOS COMPATIBLES CON..."), y órganos con
    # palabras-ruido ("LESION ESTOMAGO", "BX MEDULA OSEA", "TUMOR PALADAR").
    # Estos patrones limpian la salida del LLM antes de pintarla.
    # NOTA: usamos `re` (importado a nivel de módulo, línea 22) en vez de
    # un alias local porque list comprehensions en class body tienen su
    # propio scope y no ven los nombres del class body.
    _IA_PREFIJOS_ORGANO_RUIDO = re.compile(
        r'^(?:LESI[ÓO]N|LESION|BX|BIOPSIA|TUMOR(?:ACI[ÓO]N)?|MASA|PIEZA|MUESTRA|'
        r'N[ÓO]DULO|ESP[ÉE]CIMEN|FRAGMENTO|RESECCI[ÓO]N|CIRUG[ÍI]A)\s+(?:DE\s+)?',
        re.IGNORECASE
    )
    # V6.7.11 — Procedimientos quirúrgicos comunes que el patólogo escribe
    # como "tipo de pieza" en vez del órgano. Mapeo procedimiento→órgano
    # canónico para que "NEFRECTOMIA RADICAL IZQUIERDA" → "RIÑON IZQUIERDO".
    # NOTA: cada regex usa [ÍI] / [ÁA] para tolerar variantes con/sin tilde
    # del LLM (ej: "NEFRECTOMÍA" con tilde vs "NEFRECTOMIA" sin tilde).
    _IA_MAPEO_PROCEDIMIENTO_ORGANO = [
        (re.compile(r'^CUADRANTECTOM[ÍI]A(?:\s+DE)?\s*', re.IGNORECASE), 'MAMA'),
        (re.compile(r'^MASTECTOM[ÍI]A(?:\s+RADICAL)?(?:\s+DE)?\s*', re.IGNORECASE), 'MAMA'),
        (re.compile(r'^NEFRECTOM[ÍI]A(?:\s+RADICAL)?(?:\s+DE)?\s*', re.IGNORECASE), 'RIÑON'),
        (re.compile(r'^HEMICOLECTOM[ÍI]A(?:\s+DE)?\s*', re.IGNORECASE), 'COLON'),
        (re.compile(r'^SIGMOIDECTOM[ÍI]A(?:\s+DE)?\s*', re.IGNORECASE), 'SIGMOIDES'),
        (re.compile(r'^HISTERECTOM[ÍI]A(?:\s+VAGINAL|\s+TOTAL|\s+RADICAL)?(?:\s+DE)?\s*', re.IGNORECASE), 'UTERO'),
        (re.compile(r'^PROSTATECTOM[ÍI]A(?:\s+DE)?\s*', re.IGNORECASE), 'PROSTATA'),
        (re.compile(r'^TIROIDECTOM[ÍI]A(?:\s+DE)?\s*', re.IGNORECASE), 'TIROIDES'),
        (re.compile(r'^GASTRECTOM[ÍI]A(?:\s+DE)?\s*', re.IGNORECASE), 'ESTOMAGO'),
        (re.compile(r'^COLECISTECTOM[ÍI]A(?:\s+DE)?\s*', re.IGNORECASE), 'VESICULA BILIAR'),
        (re.compile(r'^APENDICECTOM[ÍI]A(?:\s+DE)?\s*', re.IGNORECASE), 'APENDICE'),
        (re.compile(r'^(?:LOBECTOM[ÍI]A|NEUMONECTOM[ÍI]A)(?:\s+DE)?\s*', re.IGNORECASE), 'PULMON'),
        (re.compile(r'^HEPATECTOM[ÍI]A(?:\s+DE)?\s*', re.IGNORECASE), 'HIGADO'),
        (re.compile(r'^POLIPECTOM[ÍI]A(?:\s+DE)?\s*', re.IGNORECASE), 'COLON'),
        (re.compile(r'^TRUCUT(?:\s+DE)?\s*', re.IGNORECASE), ''),
    ]
    # Regex auxiliar para detectar lateralidad
    _IA_LATERALIDAD = re.compile(
        r'\b(IZQUIERD[OA]|DERECH[OA]|BILATERAL)\b', re.IGNORECASE
    )
    # V6.7.11 — Modificadores clínicos al FINAL del órgano que sobran
    # (ej: "GANGLIO PROFUNDO METASTÁSICO" → "GANGLIO PROFUNDO").
    _IA_SUFIJOS_MODIFICADOR_ORGANO = re.compile(
        r'\s+(?:METAST[ÁA]SIC[OA]|METAST[ÁA]SIS|INVASIV[OA]|'
        r'INFILTRANTE|MALIGN[OA])\s*$',
        re.IGNORECASE
    )
    _IA_PREFIJO_MUCOSA = re.compile(
        r'^MUCOSA\s+(?:DE\s+)?', re.IGNORECASE
    )
    _IA_PREFIJO_TEJIDO = re.compile(
        r'^(?:TEJIDO|PIEL)\s+(?:DE\s+)?(?=\w)', re.IGNORECASE
    )
    # "Estudio de inmunohistoquímica:" como separador entre frase
    # introductoria y el diagnóstico real
    _IA_SEP_ESTUDIO = re.compile(
        r'estudios?\s+de\s+inmunohistoqu[ií]mica\s*[:.]?\s*',
        re.IGNORECASE
    )
    # Preámbulos del patólogo que preceden al dx real.
    # NOTA: HISTOL[ÓO]GIC?[OA]S? cubre el typo OCR "HISTOLOGIOS" (falta C).
    _IA_PREAMBULOS_DX = [
        re.compile(p, re.IGNORECASE) for p in (
            # V6.7.11 — IHQ250005: variante simple "LOS HALLAZGOS
            # HISTOLÓGICOS SON COMPATIBLES CON" (SIN "Y DE INMUNOHISTOQUIMICA")
            r'^LOS\s+HALLAZGOS\s+HISTOL[ÓO]GIC?[OA]S?\s+SON\s+COMPATIBLES?\s+CON\s+',
            r'^LOS\s+HALLAZGOS\s+HISTOL[ÓO]GIC?[OA]S?\s+(?:Y\s+(?:DE\s+)?INMUNOHISTOQU[ÍI]MIC[OA])?\s+(?:SON\s+)?COMPATIBLES?\s+CON\s+',
            r'^LOS\s+HALLAZGOS\s+MORFOL[ÓO]GICOS?(?:\s+E\s+INMUNOHISTOQU[ÍI]MIC[OA]S?)?(?:\s+Y\s+(?:DE\s+)?INMUNOHISTOQU[ÍI]MIC[OA])?\s+(?:FAVORECEN|SON\s+COMPATIBLES?\s+CON|FAVORECE\s+EL\s+DIAGN[ÓO]STICO\s+DE)\s+',
            r'^LOS\s+HALLAZGOS\s+SON\s+(?:COMPATIBLES?\s+CON|SUGESTIVOS\s+DE)\s+',
            r'^HALLAZGOS\s+CONSISTENTES\s+CON\s+',
            r'^HALLAZGOS\s+(?:DE\s+)?INMUNOHISTOQU[ÍI]MIC[OA]\s+COMPATIBLES?\s+CON\s+',
            # V6.7.14 IHQ250008: variante con FAVORECEN
            r'^(?:LOS\s+)?HALLAZGOS\s+(?:DE\s+)?INMUNOHISTOQU[ÍI]MIC[OA]\s+FAVORECEN\s+(?:UNA?\s+)?',
            # V6.7.6 — Variante "HALLAZGOS DE MORFOLOGIA E INMUNOHISTOQUIMICA
            # COMPATIBLES CON" (sustantivo MORFOLOGIA en vez de adjetivo)
            r'^HALLAZGOS\s+DE\s+MORFOLOG[ÍI]A\s+(?:E\s+|Y\s+(?:DE\s+)?)INMUNOHISTOQU[ÍI]MIC[OA]\s+COMPATIBLES?\s+CON\s+',
            r'^HALLAZGOS\s+MORFOL[ÓO]GICOS?\s+(?:Y\s+(?:DE\s+)?INMUNOHISTOQU[ÍI]MIC[OA])?\s+COMPATIBLES?\s+CON\s+',
            # V6.7.7 — Variante "HALLAZGOS MORFOLÓGICOS Y DE INMUNOHISTOQUÍMICA
            # QUE FAVORECEN" (combina MORFOLOGICOS + IHQ + QUE FAVORECEN)
            r'^HALLAZGOS\s+MORFOL[ÓO]GICOS?\s+Y\s+(?:DE\s+)?INMUNOHISTOQU[ÍI]MIC[OA]\s+QUE\s+FAVORECEN\s+',
            r'^HALLAZGOS\s+(?:MORFOL[ÓO]GICOS?\s+)?QUE\s+FAVORECEN\s+',
            r'^PERFIL\s+(?:DE\s+EXPRESI[ÓO]N\s+)?DE\s+INMUNOHISTOQU[ÍI]MIC[OA]\s+(?:COMPATIBLE\s+CON|QUE\s+FAVORECE)\s+',
            # V6.7.7 — Variante simple "HALLAZGOS COMPATIBLES CON"
            # (sin morfológicos / IHQ / etc. en medio)
            r'^HALLAZGOS\s+COMPATIBLES?\s+CON\s+',
            r'^COMPATIBLES?\s+CON\s+',
        )
    ]
    # Artículos que pueden quedar colgando después del preámbulo
    # (ej: "LOS HALLAZGOS SON SUGESTIVOS DE UNA NEOPLASIA..." → "UNA NEOPLASIA...")
    _IA_ARTICULOS_COLGANTES = re.compile(
        r'^(?:UNA?|EL|LA|LOS|LAS)\s+', re.IGNORECASE
    )

    # V6.7.6 — Mapping de adjetivos médicos → sustantivos canónicos para
    # órganos. El LLM a veces devuelve el órgano como adjetivo (RENAL,
    # ENDOMETRIAL, RECTAL, PULMONAR, GASTRICA, CERVICAL) — más útil tener
    # el sustantivo (RIÑON, ENDOMETRIO, RECTO, PULMON, ESTOMAGO, CERVIX).
    # NOTA: aplica solo a la palabra COMPLETA (palabra exacta o seguida de
    # lateralidad como IZQUIERDA/DERECHA). NO toca casos como "PULMON
    # IZQUIERDO" donde ya está bien.
    _IA_MAPEO_ADJETIVOS_ORGANO = [
        # (regex_palabra_completa_a_reemplazar, sustantivo_canonico)
        (re.compile(r'\bRENAL\b(?:\s+(IZQUIERD[OA]|DERECH[OA]))?', re.IGNORECASE), 'RIÑON'),
        (re.compile(r'\bENDOMETRIAL\b', re.IGNORECASE), 'ENDOMETRIO'),
        (re.compile(r'\bRECTAL\b', re.IGNORECASE), 'RECTO'),
        (re.compile(r'\bPULMONAR\b(?:\s+(IZQUIERD[OA]|DERECH[OA]))?', re.IGNORECASE), 'PULMON'),
        (re.compile(r'\bG[ÁA]STRIC[OA]\b', re.IGNORECASE), 'ESTOMAGO'),
        (re.compile(r'\bCERVICAL\b', re.IGNORECASE), 'CERVIX'),
        (re.compile(r'\bHEP[ÁA]TIC[OA]\b', re.IGNORECASE), 'HIGADO'),
        (re.compile(r'\b[OÓ]SE[OA]\b', re.IGNORECASE), 'HUESO'),
        # V6.7.14 IHQ250050: "NASOFARINGEA" (adjetivo) → "NASOFARINGE"
        (re.compile(r'\bNASOFAR[ÍI]NGE[OA]\b', re.IGNORECASE), 'NASOFARINGE'),
    ]

    # V6.7.6 — Diccionario de typos comunes del LLM detectados en producción.
    # Aplicado al órgano y al diagnóstico para corregir alucinaciones.
    _IA_CORRECCIONES_TYPOS = [
        (re.compile(r'\bABOMINAL\b', re.IGNORECASE), 'ABDOMINAL'),
        (re.compile(r'\bABOMBIAL\b', re.IGNORECASE), 'ABDOMINAL'),
        (re.compile(r'\bPADRE\s+ABDOMINAL\b', re.IGNORECASE), 'PARED ABDOMINAL'),
        (re.compile(r'\bPARITEO\b', re.IGNORECASE), 'PARIETO'),
        (re.compile(r'\bADANTIMOMATOSO\b', re.IGNORECASE), 'ADAMANTINOMATOSO'),
        (re.compile(r'\bNUEROENDOCRINO\b', re.IGNORECASE), 'NEUROENDOCRINO'),
        (re.compile(r'\bHIPOFISIARIO\b', re.IGNORECASE), 'HIPOFISARIO'),
        (re.compile(r'\bLUNGA\b', re.IGNORECASE), 'PULMON'),
        # V6.7.11 — Typos detectados en producción (LLM corta letras finales
        # o usa idioma extranjero)
        (re.compile(r'\bPULMO\b', re.IGNORECASE), 'PULMON'),
        (re.compile(r'\bRECTA\b(?!L|R)', re.IGNORECASE), 'RECTO'),  # "RECTA" pero no "RECTAL" o "RECTAR"
        (re.compile(r'\bENDOMETRIUM\b', re.IGNORECASE), 'ENDOMETRIO'),
        (re.compile(r'\bMEDULA\s+HUESO\b', re.IGNORECASE), 'MEDULA OSEA'),
        (re.compile(r'\bARQUTIECTURA\b', re.IGNORECASE), 'ARQUITECTURA'),
        # V6.7.12 — Typos detectados con nemotron-3-nano-omni
        (re.compile(r'\bHIPOFISIA\b', re.IGNORECASE), 'HIPOFISIS'),
        # V6.7.16 — Typos detectados con gpt-oss-20b (truncamientos del modelo)
        (re.compile(r'\bMEDIANTE\b', re.IGNORECASE), 'MEDIASTINO'),
        (re.compile(r'\bMEDIASTI\b', re.IGNORECASE), 'MEDIASTINO'),
    ]
    # V6.7.12 — Bullet-prefix del LLM. Algunos modelos (nemotron) preservan
    # el guión inicial cuando el dx viene como bullet en el PDF.
    # Lo aplicamos al inicio del dx para limpiar "- EXPRESIÓN DE CD117..."
    _IA_BULLET_PREFIX = re.compile(r'^\s*[-•·]\s*', re.UNICODE)

    def _aplicar_correcciones_typos(self, texto: str) -> str:
        """V6.7.6 — Aplica el diccionario de typos comunes del LLM."""
        if not texto:
            return texto
        for pat, repl in self._IA_CORRECCIONES_TYPOS:
            texto = pat.sub(repl, texto)
        return texto

    def _normalizar_organo_adjetivo(self, org: str) -> str:
        """V6.7.6 — Convierte adjetivos médicos a sustantivos canónicos."""
        if not org:
            return org
        for pat, sustantivo in self._IA_MAPEO_ADJETIVOS_ORGANO:
            m = pat.search(org)
            if m:
                # Reemplazar manteniendo lateralidad si aplica
                lateralidad = m.group(1) if m.lastindex and m.group(1) else ""
                if lateralidad:
                    repl = f"{sustantivo} {lateralidad.upper()}"
                else:
                    repl = sustantivo
                org = pat.sub(repl, org, count=1)
                break  # solo un mapping por órgano
        return org

    def _limpiar_resultado_ia(self, dx: str, organo: str) -> tuple[str, str]:
        """V6.7.5 — Post-procesa la respuesta del LLM. Quita palabras-ruido
        del órgano y preámbulos del dx.

        Estrategia para el diagnóstico:
        1. Si tiene 'Estudio de inmunohistoquímica:' → tomar lo de DESPUÉS
        2. Si tiene preámbulo 'LOS HALLAZGOS COMPATIBLES CON' → strippear
        3. Si después del cleanup queda vacío → mostrar el RAW del LLM
           con marcador '(REVISAR DX)' para transparencia. NUNCA usamos
           'NO IDENTIFICADO' a menos que el LLM no haya devuelto nada.

        Returns: (dx_para_mostrar, organo_limpio).
        """
        # ─── ÓRGANO ────────────────────────────────────────────────────
        org = (organo or "").strip()
        if org:
            # V6.7.11 — Detectar procedimiento al inicio. Si lo hay, intentar
            # extraer el órgano residual; si solo queda lateralidad, usar el
            # órgano canónico mapeado del procedimiento.
            for pat_proc, organo_canon in self._IA_MAPEO_PROCEDIMIENTO_ORGANO:
                m_proc = pat_proc.match(org)
                if m_proc:
                    residual = org[m_proc.end():].strip()
                    # Detectar si el residual es solo lateralidad o vacío
                    es_solo_lateralidad = bool(re.match(
                        r'^(?:IZQUIERD[OA]|DERECH[OA]|BILATERAL)?\s*$',
                        residual, re.IGNORECASE
                    ))
                    if not residual or es_solo_lateralidad:
                        # Usar órgano canónico + lateralidad si existe
                        if organo_canon:
                            lat = self._IA_LATERALIDAD.search(residual)
                            if lat:
                                org = f"{organo_canon} {lat.group(0).upper()}"
                            else:
                                org = organo_canon
                        else:
                            org = residual or org
                    else:
                        # Hay órgano sustantivo después del procedimiento,
                        # quedarse con eso (ej: "CUADRANTECTOMIA MAMA DERECHA"
                        # → "MAMA DERECHA")
                        org = residual
                    break

            # Strip palabras-ruido (LESION, BX, TUMOR, MUCOSA DE) en cascada
            for _ in range(3):
                prev = org
                org = self._IA_PREFIJOS_ORGANO_RUIDO.sub('', org, count=1)
                org = self._IA_PREFIJO_MUCOSA.sub('', org, count=1)
                if org == prev:
                    break
            org_test = self._IA_PREFIJO_TEJIDO.sub('', org, count=1).strip()
            if org_test and len(org_test.split()) >= 1 and org_test != org:
                org = org_test

            # V6.7.11 — Strip modificadores clínicos al final
            # (ej: "GANGLIO PROFUNDO METASTÁSICO" → "GANGLIO PROFUNDO")
            org = self._IA_SUFIJOS_MODIFICADOR_ORGANO.sub('', org).strip()

            org = org.strip().rstrip('.,;')
            # V6.7.6 — Aplicar correcciones de typos + adjetivo→sustantivo
            org = self._aplicar_correcciones_typos(org)
            org = self._normalizar_organo_adjetivo(org)

            # V6.7.11 — Normalizar a MAYÚSCULAS para consistencia
            # (el LLM a veces devuelve "Vertebra L1", "Mediastino",
            # "Pulmón izquierdo" en mixed-case)
            org = org.upper()

            # V6.7.16 — Quitar tildes en órganos (HÍGADO → HIGADO,
            # PULMÓN → PULMON, RIÑÓN → RINON pero mantener Ñ → Ñ).
            # Solo elimina acentos en vocales (Á É Í Ó Ú), preserva Ñ.
            import unicodedata as _ud
            _normalized = []
            for ch in org:
                if ch == 'Ñ':
                    _normalized.append(ch)
                else:
                    _normalized.append(
                        ''.join(c for c in _ud.normalize('NFD', ch)
                                if _ud.category(c) != 'Mn')
                    )
            org = ''.join(_normalized)

            if not org:
                org = (organo or "").strip().upper()

        # ─── DIAGNÓSTICO ───────────────────────────────────────────────
        d = (dx or "").strip()
        # V6.7.12 — Quitar bullet-prefix "- " que algunos modelos
        # (nemotron) preservan del PDF original
        d = self._IA_BULLET_PREFIX.sub('', d)
        original_dx = d

        # Si el LLM no devolvió nada, ahí sí NO IDENTIFICADO
        if not d:
            return ("NO IDENTIFICADO", org)

        # 1. Si contiene "Estudio de inmunohistoquímica:" tomar lo de DESPUÉS
        m = self._IA_SEP_ESTUDIO.search(d)
        if m:
            despues = d[m.end():].strip()
            if despues and len(despues) > 3:
                d = despues
            else:
                # No hay nada significativo después → mostrar el raw del
                # LLM con marcador para transparencia (V6.7.5)
                raw_corto = self._truncar_raw(original_dx, 120)
                return (f"(REVISAR DX) {raw_corto}", org)

        # V6.7.14 IHQ250040: si el dx empieza con frase introductoria
        # tipo "Región/Mucosa/Tumor X. Lesión. [Resección/Biopsia]" sin
        # un "Estudio IHQ:" en medio, strippear hasta encontrar HALLAZGOS
        # o la primera entidad clínica en MAYÚSCULAS.
        m_intro = re.match(
            r'^[A-Za-zÁÉÍÓÚáéíóúñÑ\s]+\.\s+[A-Za-zÁÉÍÓÚáéíóúñÑ]+\.\s+'
            r'(?:Resecci[óo]n|Biopsia|Excisional|Cuadrantectom[íi]a|Mastectom[íi]a|Nefrectom[íi]a)\s+',
            d
        )
        if m_intro:
            despues = d[m_intro.end():].strip()
            if despues and len(despues) > 3:
                d = despues

        # 2. Strippear preámbulos del patólogo
        for pat in self._IA_PREAMBULOS_DX:
            d_new = pat.sub('', d, count=1)
            if d_new != d:
                d = d_new.strip()
                d = self._IA_ARTICULOS_COLGANTES.sub('', d, count=1).strip()
                break

        # 3. Limpiar puntuación final + espacios sobrantes
        # V6.7.14: incluir ':' y "CON" colgante (cuando LLM trunca)
        d = d.strip().rstrip('.').strip()
        # Si termina con ":" o "CON:" o ", CON" colgante, quitar
        d = re.sub(r'[\s,]*(?:CON)?\s*:\s*$', '', d, flags=re.IGNORECASE).strip()
        d = re.sub(r'[,\s]+CON\s*$', '', d, flags=re.IGNORECASE).strip()

        # 4. V6.7.6 — Aplicar correcciones de typos del LLM
        d = self._aplicar_correcciones_typos(d)

        # Si después del cleanup queda muy corto, recuperar raw original
        if not d or len(d) < 3:
            if original_dx and len(original_dx) >= 5:
                raw_corto = self._truncar_raw(original_dx, 120)
                return (f"(REVISAR DX) {raw_corto}", org)
            return ("NO IDENTIFICADO", org)

        return (d, org)

    @staticmethod
    def _truncar_raw(texto: str, max_chars: int) -> str:
        """Trunca un texto raw (sin saltos de línea) preservando legibilidad."""
        t = (texto or "").replace("\n", " ").replace("\r", "").strip()
        # Colapsar múltiples espacios en uno solo
        t = re.sub(r"\s+", " ", t)
        if len(t) > max_chars:
            t = t[:max_chars].rstrip() + "…"
        return t

    def _ia_show_context_menu(self, event):
        """Muestra el menú contextual sobre el treeview en la posición del click."""
        try:
            self._ia_context_menu.tk_popup(event.x_root, event.y_root)
        finally:
            self._ia_context_menu.grab_release()

    def _ia_select_all_rows(self):
        """Selecciona todas las filas del treeview."""
        try:
            children = self._ia_treeview.get_children()
            self._ia_treeview.selection_set(children)
        except Exception:
            pass
        return "break"  # evita propagación del shortcut

    def _ia_get_table_rows(self, only_selected: bool):
        """Devuelve las filas del treeview como lista de tuplas (ihq, dx, organo).

        Args:
            only_selected: si True, solo las filas seleccionadas; si False, todas.
        """
        try:
            iids = (
                self._ia_treeview.selection() if only_selected
                else self._ia_treeview.get_children()
            )
            rows = []
            for iid in iids:
                values = self._ia_treeview.item(iid, "values")
                # values es una tupla (ihq, dx, organo)
                if values:
                    rows.append(tuple(str(v) for v in values))
            return rows
        except Exception as e:
            logging.warning(f"[IA copy] Error leyendo treeview: {e}")
            return []

    def _ia_copy_selected_to_clipboard(self):
        """Copia las filas SELECCIONADAS al clipboard en formato TSV.
        Si no hay selección, copia todo."""
        rows = self._ia_get_table_rows(only_selected=True)
        if not rows:
            # Si no hay selección, copiar todo
            rows = self._ia_get_table_rows(only_selected=False)
        if not rows:
            return
        self._ia_copy_rows_to_clipboard(rows)

    def _ia_copy_all_to_clipboard(self):
        """Copia TODAS las filas al clipboard en formato TSV."""
        rows = self._ia_get_table_rows(only_selected=False)
        if not rows:
            messagebox.showinfo(
                "Sin datos",
                "No hay diagnósticos en la tabla para copiar."
            )
            return
        self._ia_copy_rows_to_clipboard(rows, mostrar_aviso=True)

    def _ia_copy_rows_to_clipboard(self, rows, mostrar_aviso: bool = False):
        """Copia filas al clipboard en TSV con TODAS las columnas (3 alias
        + 184 BD). V6.8.0 — pega directo en Excel reproduciendo el schema
        completo de la BD principal."""
        try:
            # Header completo: 3 alias + 184 columnas BD
            cols_bd = getattr(self, "_ia_columnas_bd", [])
            header = ["N° IHQ", "Diagnóstico", "Órgano"] + list(cols_bd)
            lines = ["\t".join(header)]
            n_cols = len(header)
            for r in rows:
                clean = [
                    (c or "").replace("\t", " ").replace("\n", " ").replace("\r", "")
                    for c in r
                ]
                while len(clean) < n_cols:
                    clean.append("")
                lines.append("\t".join(clean[:n_cols]))
            tsv = "\n".join(lines)

            self._ia_progress_win.clipboard_clear()
            self._ia_progress_win.clipboard_append(tsv)
            self._ia_progress_win.update()

            if mostrar_aviso:
                messagebox.showinfo(
                    "Copiado al portapapeles",
                    f"{len(rows)} fila(s) × {n_cols} columnas copiadas.\n\n"
                    f"Pegalo en Excel — reproduce el schema completo de "
                    f"la BD del HUV."
                )
        except Exception as e:
            logging.error(f"[IA copy] Falló: {e}")
            messagebox.showerror("Error", f"No se pudo copiar al portapapeles:\n{e}")

    def _ia_export_to_csv(self):
        """Exporta todas las filas a CSV con las 184 columnas + 3 alias."""
        rows = self._ia_get_table_rows(only_selected=False)
        if not rows:
            messagebox.showinfo(
                "Sin datos", "No hay diagnósticos en la tabla para exportar."
            )
            return

        from tkinter import filedialog
        from datetime import datetime as _dt
        default_name = f"diagnosticos_ia_{_dt.now().strftime('%Y%m%d_%H%M%S')}.csv"
        path = filedialog.asksaveasfilename(
            title="Guardar CSV",
            defaultextension=".csv",
            initialfile=default_name,
            filetypes=[("CSV (separado por comas)", "*.csv"), ("Todos", "*.*")]
        )
        if not path:
            return

        import csv as _csv
        try:
            cols_bd = getattr(self, "_ia_columnas_bd", [])
            header = ["N° IHQ", "Diagnostico (alias)", "Organo (alias)"] + list(cols_bd)
            n_cols = len(header)
            with open(path, "w", newline="", encoding="utf-8-sig") as f:
                writer = _csv.writer(f)
                writer.writerow(header)
                for r in rows:
                    clean = [(c or "").replace("\n", " ") for c in r]
                    while len(clean) < n_cols:
                        clean.append("")
                    writer.writerow(clean[:n_cols])
            messagebox.showinfo(
                "Exportado",
                f"{len(rows)} fila(s) × {n_cols} columnas exportadas a:\n{path}"
            )
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo guardar el CSV:\n{e}")

    def _process_files_ia_worker(self, files_to_process, out_dir):
        """Worker thread: para cada PDF hace OCR + LLM (con chunking) y guarda."""
        import json as _json
        import re as _re
        import time as _time
        from datetime import datetime as _dt

        try:
            from core.processors.ocr_processor import pdf_to_text_enhanced
            from core.llm_client import LMStudioClient, _extraer_json_robusto
            from core.diagnosticos_ia_db import (
                init_db as _init_ia_db,
                save_caso_completo as _save_caso_ia,
            )
            from core.columnas_huv_ia import (
                COLUMNAS_IA,
                ALIAS_TO_COLUMN,
                build_json_schema,
                build_prompt_field_list,
                llm_response_to_db_dict,
            )
            # V6.8.0 — También escribir a la BD principal (informes_ihq)
            # para que el Visualizador de datos refleje los casos extraídos
            # por IA. Mismo comportamiento que "Procesar seleccionados":
            # UPSERT por "Numero de caso" (reemplaza si existe).
            from core.database_manager import (
                init_db as _init_main_db,
                save_records as _save_records_main,
            )
        except Exception as e:
            self._processing_result_ia["errors"].append(f"Imports fallaron: {e}")
            self._processing_result_ia["done"] = True
            return

        # Inicializar BD acumulativa de diagnósticos IA (idempotente)
        try:
            _init_ia_db()
        except Exception as e:
            logging.warning(f"[IA] No se pudo inicializar BD de diagnósticos IA: {e}")

        # V6.8.0 — Inicializar también la BD principal (idempotente).
        # Los casos extraídos por IA se persisten EN AMBAS BDs:
        #   • diagnosticos_ia.db → histórico de extracciones IA (debug/audit)
        #   • huv_oncologia_NUEVO.db → BD principal que ve el Visualizador
        try:
            _init_main_db()
        except Exception as e:
            logging.warning(f"[IA] No se pudo inicializar BD principal: {e}")

        # V6.9.2 — Timeout aumentado a 900s (15 min) para soportar modelos
        # de reasoning como nvidia/nemotron-3-nano que generan mucho
        # razonamiento interno antes de devolver el JSON. El default de
        # 300s era insuficiente y causaba "Timeout — Reintentando..." en
        # bucle. Modelos rápidos (qwen2.5-14b/7b) responden en <60s y
        # no se ven afectados por este timeout más largo.
        client = LMStudioClient(timeout=900)

        # V6.7.13 — Chunking 1 IHQ = 1 chunk = 1 llamada al LLM.
        # Cambio arquitectural: en vez de agrupar 5-7 IHQ por chunk,
        # cada IHQ se procesa individualmente. Más llamadas (~50 vs 10)
        # pero cada respuesta del LLM es trivial (1 dx + 1 órgano), lo
        # que elimina alucinaciones, omisiones y confusiones entre IHQs.
        # Tradeoff: ~15-25 min por PDF (vs 5-10 min antes) pero calidad
        # esperada de ~99% (vs ~95% con chunking grupal).

        def _split_by_ihq_boundaries(texto: str):
            """Divide texto OCR en 1 chunk por IHQ. Cada chunk contiene
            UN solo informe IHQ completo (header + descripciones + dx)."""
            header_pat = _re.compile(
                r'N\.?\s*petici[oó]n\s*:?\s*(IHQ\s*\d{5,7})',
                _re.IGNORECASE
            )
            matches = list(header_pat.finditer(texto))

            if not matches:
                logging.warning(
                    "[IA] No se detectaron headers 'N. peticion : IHQ...' — "
                    "fallback: texto entero como un chunk"
                )
                return [texto] if texto.strip() else []

            # Detectar boundaries por cambio de número IHQ
            # (un mismo IHQ repite header en cada página de su informe)
            def _normalizar_ihq(s):
                return _re.sub(r'\s+', '', s).upper()

            boundaries = [0]
            last_ihq = None
            for m in matches:
                ihq_num = _normalizar_ihq(m.group(1))
                if last_ihq is None:
                    last_ihq = ihq_num
                    continue
                if ihq_num != last_ihq:
                    boundaries.append(m.start())
                    last_ihq = ihq_num
            boundaries.append(len(texto))

            # 1 segmento = 1 IHQ completo = 1 chunk
            chunks = []
            for i in range(len(boundaries) - 1):
                seg = texto[boundaries[i]:boundaries[i + 1]]
                if seg.strip():
                    chunks.append(seg)

            logging.info(
                f"[IA] Chunking 1-IHQ-por-chunk: {len(matches)} headers → "
                f"{len(chunks)} chunks (cada uno = 1 IHQ)"
            )
            return chunks

        # Alias retrocompatible
        _split_in_chunks = _split_by_ihq_boundaries

        def _norm_dx_entry(d):
            """V6.8.0 — Normaliza una entrada del LLM (con 184 aliases JSON)
            a un dict con keys = nombres de columnas BD.

            Devuelve el dict completo (188 keys con N/A en faltantes) o None
            si el input no es válido.

            Mantiene compat con keys 'numero_peticion'/'diagnostico'/'organo'
            para código legacy (se inyectan tomando los nombres BD).
            """
            if not isinstance(d, dict):
                return None
            lookup = {k.lower(): v for k, v in d.items() if isinstance(k, str)}

            # Caso 1: respuesta nueva V6.8.0 — tiene los 184 aliases.
            # Detectamos por presencia de 'numero_de_caso' (nombre BD nuevo).
            if "numero_de_caso" in lookup:
                # Mapear todos los aliases → nombres BD
                db_dict = llm_response_to_db_dict(lookup)
                # Aliases legacy (compat con código viejo del worker)
                db_dict["__numero_peticion__"] = (db_dict.get("Numero de caso") or "").strip()
                db_dict["__diagnostico__"] = db_dict.get("Diagnostico Principal", "N/A")
                db_dict["__organo__"] = db_dict.get("Organo", "N/A")
                return db_dict

            # Caso 2: respuesta legacy V6.7.x — solo 3 campos.
            # Convertimos a estructura completa con N/A en el resto.
            numero = (
                lookup.get("numero_peticion")
                or lookup.get("ihq")
                or lookup.get("peticion")
                or lookup.get("numero")
                or ""
            )
            dx = (
                lookup.get("diagnostico")
                or lookup.get("dx")
                or lookup.get("diagnosis")
                or ""
            )
            organo = (
                lookup.get("organo")
                or lookup.get("organ")
                or lookup.get("sitio")
                or ""
            )
            db_dict = {col: "N/A" for col in COLUMNAS_IA}
            db_dict["Numero de caso"] = numero
            db_dict["Diagnostico Principal"] = dx
            db_dict["Organo"] = organo
            db_dict["__numero_peticion__"] = numero
            db_dict["__diagnostico__"] = dx
            db_dict["__organo__"] = organo
            return db_dict

        # Set de números de petición ya pintados en la tabla (deduplicación
        # global, evita duplicados por chunk-overlap o repeticiones del LLM)
        ihq_vistos = set()

        for idx, (pdf_path, filename) in enumerate(files_to_process):
            self._processing_result_ia["current_index"] = idx + 1
            self._processing_result_ia["current_file"] = filename
            self._processing_result_ia["current_stage"] = "OCR del PDF..."
            self._processing_result_ia["current_chunk"] = 0
            self._processing_result_ia["total_chunks"] = 0

            try:
                texto_ocr = pdf_to_text_enhanced(pdf_path)
                ocr_chars = len(texto_ocr)
                logging.info(f"[IA] {filename}: OCR completo ({ocr_chars} chars)")
            except Exception as e:
                self._processing_result_ia["errors"].append(f"{filename}: OCR falló — {e}")
                continue

            chunks = _split_in_chunks(texto_ocr)
            n_chunks = len(chunks)
            self._processing_result_ia["total_chunks"] = n_chunks

            # V6.7.7 — Contar cuántos IHQ debería haber según headers en el
            # texto OCR. Si al final faltan, lo reportamos al user.
            _header_pat_count = _re.compile(
                r'N\.?\s*petici[oó]n\s*:?\s*(IHQ\s*\d{5,7})', _re.IGNORECASE
            )
            _ihq_esperados = set()
            for m in _header_pat_count.finditer(texto_ocr):
                num = _re.sub(r'\s+', '', m.group(1)).upper()
                _ihq_esperados.add(num)
            n_ihq_esperados = len(_ihq_esperados)
            logging.info(
                f"[IA] {filename}: dividido en {n_chunks} chunks · "
                f"{n_ihq_esperados} IHQ únicos detectados en OCR"
            )

            todos_diagnosticos = []
            chunks_exitosos = 0
            errores_chunks = []

            # V6.7.7 — Helper de llamada al LLM con retry automático.
            # V6.7.13 — max_tokens reducido a 800 (antes 4000). Razón:
            # ahora cada chunk = 1 solo IHQ → output esperado ~100-200 tokens
            # (1 entrada de JSON con 3 campos). 800 da margen 4x sin que
            # el LLM gaste tiempo generando reasoning innecesario.
            # V6.7.15 — Pasar json_schema (más robusto que json_object).
            # gpt-oss-20b en LM Studio rechaza json_object pero acepta
            # json_schema, y con strict=true fuerza al modelo a generar
            # exactamente la estructura esperada (sin razonamiento previo).
            # V6.7.16 — minItems/maxItems = 1: cada chunk contiene
            # exactamente 1 IHQ. Forzar al modelo a extraer 1 entrada
            # (antes devolvía [] cuando no encontraba dx claro y se rendía).
            # V6.8.0 — Schema EXPANDIDO a 184 campos (ver
            # core/columnas_huv_ia.py). El LLM extrae el caso COMPLETO,
            # no solo dx/organo. Para campos no presentes en el informe,
            # el prompt instruye devolver "N/A".
            _IA_JSON_SCHEMA = build_json_schema()

            # V6.7.17 — max_tokens aumentado de 800 a 1200 (intento 1) y de
            # 1500 a 2000 (intento 2). Razón: dx con scoring Banff completo
            # (RECHAZO ACTIVO con clasificaciones g/ptc/i/t/v) pueden requerir
            # ~1000-1200 tokens. Caso testigo: IHQ250100.
            # V6.7.18 — Aumentado a 1500/2500 para cubrir descripciones
            # extensas de médula ósea (celularidad, relación M/E, blastos)
            # y dx prostáticos con Gleason + grupo + cores + %.
            # Casos testigo: IHQ250160 (médula), IHQ250178 (próstata).
            # V6.7.19 — Aumentado a 3000/4000 para cubrir TODOS los dx
            # extremos sin truncamiento. Trade-off: ~30% más lento por
            # chunk pero cero truncamientos. Decisión clínica: cada IHQ
            # es único, no se acepta pérdida de información.
            # Casos testigo: IHQ250275 (Banff 2022 BIOPSIA INJERTO),
            # IHQ250389 (Nottingham + molecular HER2/triple-negativo),
            # IHQ250401 (receptores hormonales completos).
            # V6.9.3 — max_tokens aumentado de 3000 a 6000 para soportar
            # modelos REASONING (nvidia/nemotron-3-nano, qwen3.6, etc.) que
            # generan 1500-2700 tokens de razonamiento interno ANTES del
            # JSON final. Con 3000 tokens, el JSON quedaba truncado a la
            # mitad y resultaba "no parseable".
            # Modelos no-reasoning (qwen2.5-instruct) usan ~2500 tokens
            # totales, así que 6000 no los afecta.
            def _llamar_llm_con_retry(chunk_text, intento_max_tokens=6000):
                last_error = None
                for intento in (1, 2):
                    try:
                        max_tok = intento_max_tokens if intento == 1 else 8000
                        resp = client.completar(
                            prompt=chunk_text,
                            system_prompt=self._PROMPT_SYSTEM_IA_OCR,
                            temperature=0.1,
                            max_tokens=max_tok,
                            formato_json=True,
                            json_schema=_IA_JSON_SCHEMA,
                        )
                        if resp.get("exito"):
                            return resp, None
                        last_error = resp.get("error", "?")[:200]
                        logging.warning(
                            f"[IA] Intento {intento} falló: {last_error[:100]}"
                        )
                    except Exception as e:
                        last_error = str(e)[:200]
                        logging.warning(f"[IA] Intento {intento} excepción: {last_error[:100]}")
                return None, last_error

            # V6.9.10 PARALELO: leer paralelo_max desde config.ini ([llm] -> paralelo_max).
            # Default 1 = secuencial (retrocompatible). Si se sube a N>1 → ThreadPoolExecutor.
            # NOTA: el cliente LM Studio (requests) es thread-safe, las llamadas HTTP son
            # independientes y los slots del backend (n_parallel=4) procesan en paralelo.
            try:
                _cfg_paralelo = configparser.ConfigParser()
                _cfg_paralelo.read(
                    os.path.join(os.getcwd(), "config", "config.ini"),
                    encoding="utf-8",
                )
                paralelo_max = max(1, int(_cfg_paralelo.get("llm", "paralelo_max", fallback="1")))
            except Exception as _e_cfg:
                logging.warning(f"[IA] paralelo_max no se pudo leer ({_e_cfg}); usando 1 (secuencial)")
                paralelo_max = 1

            # V6.9.11 FILTRO: leer filtrar_secciones desde config.ini ([llm] -> filtrar_secciones).
            # Si true, se aplica _filtrar_secciones_relevantes() al chunk_text ANTES de
            # enviarlo al LLM, reduciendo ~70% el texto (cabecera + DESCRIPCION MICROSCOPICA + DIAGNOSTICO).
            # Si false, comportamiento V6.9.10 (chunk completo).
            try:
                filtrar_secciones_habilitado = _cfg_paralelo.getboolean(
                    "llm", "filtrar_secciones", fallback=False
                )
            except Exception as _e_cfg_filt:
                logging.warning(
                    f"[IA] filtrar_secciones no se pudo leer ({_e_cfg_filt}); usando False"
                )
                filtrar_secciones_habilitado = False
            logging.info(
                f"[IA] V6.9.11 filtrar_secciones = {filtrar_secciones_habilitado}"
            )

            def _filtrar_secciones_relevantes(chunk_text: str) -> str:
                """V6.9.11 — Reduce el texto enviado al LLM a SOLO las secciones útiles:
                  1. Cabecera del paciente + "Estudios solicitados"
                  2. DESCRIPCION MICROSCOPICA (resultados de biomarcadores)
                  3. DIAGNOSTICO (conclusión del patólogo)

                Elimina ruido como: notas legales, sello ISO, DESCRIPCION MACROSCOPICA,
                tabla LABORATORIO CLINICO (otros pacientes — causa de contaminación cruzada),
                headers repetidos en pág 2, firmas, líneas separadoras.

                Si NO se detectan las secciones críticas (DIAGNOSTICO o DESCRIPCION
                MICROSCOPICA), devuelve el chunk completo como fallback seguro
                (NO rompe el flujo).

                Patrones tolerantes a:
                  - Con/sin tildes ('DIAGNÓSTICO' o 'DIAGNOSTICO')
                  - Mayúsculas mezcladas
                  - Espacios extra entre palabras
                """
                if not chunk_text or not chunk_text.strip():
                    return chunk_text

                partes = []

                # 1) CABECERA + ESTUDIOS SOLICITADOS:
                # Desde inicio hasta antes de "INFORME DE ANATOM..." o "DESCRIPCION MACROSC..."
                m_cab = _re.search(
                    r'^(.+?)(?=INFORME\s+DE\s+ANATOM|DESCRIPCI[OÓ]N\s+MACROSC)',
                    chunk_text,
                    _re.IGNORECASE | _re.DOTALL,
                )
                if m_cab:
                    cabecera = m_cab.group(1).strip()
                    if cabecera:
                        partes.append(cabecera)

                # 2) DESCRIPCION MICROSCOPICA:
                # Desde "DESCRIPCION MICROSCOPICA" hasta antes de "DIAGNOSTICO"
                # o "Todos los analisis" (nota ISO) o fin.
                m_micro = _re.search(
                    r'DESCRIPCI[OÓ]N\s+MICROSC[OÓ]PICA\b(.+?)(?=\bDIAGN[OÓ]STICO\b|Todos\s+los\s+an[aá]lisis|$)',
                    chunk_text,
                    _re.IGNORECASE | _re.DOTALL,
                )
                # 3) DIAGNOSTICO:
                # Desde la palabra DIAGNOSTICO (línea propia) hasta firma del patólogo
                # o nota ISO o LABORATORIO CLINICO o fin.
                m_dx = _re.search(
                    r'\bDIAGN[OÓ]STICO\b\s*\n(.+?)(?=Todos\s+los\s+an[aá]lisis|RM:|M[eé]dic[ao]\s+Pat[oó]log[ao]|_______|LABORATORIO\s+CLINICO|$)',
                    chunk_text,
                    _re.IGNORECASE | _re.DOTALL,
                )

                # Si NO se detectan las secciones críticas, fallback al chunk completo
                if not m_micro or not m_dx:
                    logging.warning(
                        f"[FILTRO] V6.9.11: secciones criticas no detectadas "
                        f"(micro={bool(m_micro)}, dx={bool(m_dx)}); usando chunk completo "
                        f"({len(chunk_text)} chars)"
                    )
                    return chunk_text

                # Concatenar con encabezados claros para el LLM
                micro_text = m_micro.group(1).strip()
                dx_text = m_dx.group(1).strip()
                if micro_text:
                    partes.append("DESCRIPCION MICROSCOPICA\n" + micro_text)
                if dx_text:
                    partes.append("DIAGNOSTICO\n" + dx_text)

                filtrado = "\n\n".join(partes).strip()

                # Si el filtrado quedó vacío o casi vacío, fallback al chunk completo
                if not filtrado or len(filtrado) < 100:
                    logging.warning(
                        f"[FILTRO] V6.9.11: filtrado vacio o muy corto "
                        f"({len(filtrado)} chars); usando chunk completo"
                    )
                    return chunk_text

                # Log del ratio de reducción
                try:
                    orig_len = len(chunk_text)
                    new_len = len(filtrado)
                    pct = (1 - new_len / orig_len) * 100 if orig_len > 0 else 0
                    logging.debug(
                        f"[FILTRO] {orig_len} → {new_len} chars ({pct:.1f}% reduccion)"
                    )
                except Exception:
                    pass

                return filtrado

            # Locks para estado compartido y BD (creados por cada PDF para no contaminar entre archivos)
            _ia_state_lock = threading.Lock()  # protege live_diagnosticos, errors, current_chunk, ihq_vistos
            _ia_db_lock = threading.Lock()     # serializa las 3 escrituras BD por caso (atómico por IHQ)

            # Pre-calcular constantes usadas dentro de cada chunk (read-only, thread-safe)
            _fecha_iso = _dt.now().isoformat()
            _valid_ihq_pat = _re.compile(r'^IHQ\d{4,7}$')

            # Acumulador de resultados por chunk (se llenan en threads, se consolidan al final)
            chunks_resultados = [None] * n_chunks  # cada slot: dict con 'normalizados', 'err', 'chunks_exitoso'
            chunks_completados_counter = [0]       # mutable int compartido (con lock)

            def _procesar_chunk_individual(chunk_idx, chunk_text):
                """V6.9.10 PARALELO: procesa UN chunk completo y guarda en BD.

                Encapsula la MISMA lógica que el bucle secuencial original:
                  1. Llama al LLM con retry
                  2. Parsea JSON, normaliza, deduplica
                  3. Aplica _limpiar_resultado_ia a Dx + Organo
                  4. Persiste en BD (diagnosticos_ia + informes_ihq) bajo lock
                  5. Append a live_diagnosticos bajo lock (manteniendo orden save→append)

                Devuelve un dict con resultados para que el bucle exterior
                actualice todos_diagnosticos, chunks_exitosos y errores_chunks.
                """
                result = {
                    "chunk_idx": chunk_idx,
                    "normalizados": [],
                    "err": None,
                    "exitoso": False,
                    "modelo": None,  # V6.9.10: propagar nombre del modelo al payload
                }

                # Logging detallado: qué IHQ están EN este chunk (sin lock, solo lectura local)
                ihq_en_chunk = sorted(set(
                    _re.sub(r'\s+', '', m.group(1)).upper()
                    for m in _header_pat_count.finditer(chunk_text)
                ))

                # Actualizar progreso bajo lock (current_chunk, current_stage, chunk_start_time)
                with _ia_state_lock:
                    self._processing_result_ia["current_chunk"] = chunks_completados_counter[0] + 1
                    self._processing_result_ia["chunk_start_time"] = _time.time()
                    self._processing_result_ia["current_stage"] = (
                        f"LLM analizando chunk {chunks_completados_counter[0] + 1}/{n_chunks} "
                        f"({len(chunk_text):,} chars)..."
                    )

                logging.info(
                    f"[IA] {filename}: chunk {chunk_idx + 1}/{n_chunks} "
                    f"({len(chunk_text)} chars) — IHQ esperados: "
                    f"{', '.join(ihq_en_chunk[:8])}"
                    f"{f' (+{len(ihq_en_chunk) - 8} más)' if len(ihq_en_chunk) > 8 else ''}"
                )

                # V6.9.11: filtrar secciones relevantes si está habilitado.
                # Si el filtro detecta DESCRIPCION MICROSCOPICA + DIAGNOSTICO,
                # envía solo cabecera + esas dos secciones (~70% menos texto).
                # Si NO las detecta (formato no estándar), hace fallback al chunk completo.
                if filtrar_secciones_habilitado:
                    chunk_text_para_llm = _filtrar_secciones_relevantes(chunk_text)
                else:
                    chunk_text_para_llm = chunk_text

                # === MISMA llamada al LLM que la versión secuencial ===
                resp, err = _llamar_llm_con_retry(chunk_text_para_llm)

                # Limpiar chunk_start_time SOLO si somos el último en terminar (best-effort, no crítico)
                # En modo paralelo, varios chunks pueden estar in-flight; este flag es solo cosmético.

                if resp is None:
                    err_msg = f"chunk {chunk_idx + 1}/{n_chunks}: LLM falló tras 2 intentos — {err}"
                    logging.error(f"[IA] {filename}: {err_msg}")
                    with _ia_state_lock:
                        self._processing_result_ia["errors"].append(f"{filename}: {err_msg}")
                        self._processing_result_ia["chunk_start_time"] = None
                    result["err"] = err_msg
                    return result

                contenido = resp.get("respuesta") or resp.get("contenido")
                if isinstance(contenido, dict):
                    data = contenido
                elif isinstance(contenido, str) and contenido.strip():
                    data = _extraer_json_robusto(contenido)
                else:
                    data = None

                if not data or not isinstance(data, dict):
                    err_msg = f"chunk {chunk_idx + 1}/{n_chunks}: JSON no parseable"
                    logging.error(f"[IA] {filename}: {err_msg}")
                    with _ia_state_lock:
                        self._processing_result_ia["errors"].append(f"{filename}: {err_msg}")
                        self._processing_result_ia["chunk_start_time"] = None
                    result["err"] = err_msg
                    return result

                raw_dx = data.get("diagnosticos", [])
                if not isinstance(raw_dx, list):
                    raw_dx = []
                normalizados = [d for d in (_norm_dx_entry(x) for x in raw_dx) if d]
                result["normalizados"] = normalizados
                result["exitoso"] = True

                # Logging del response: cuántos IHQ devolvió
                ihq_devueltos = sorted(set(
                    _re.sub(r'\s+', '', (nd.get("__numero_peticion__") or "")).upper()
                    for nd in normalizados
                    if nd.get("__numero_peticion__")
                ))
                ihq_omitidos_chunk = [x for x in ihq_en_chunk if x not in ihq_devueltos]
                logging.info(
                    f"[IA] {filename}: chunk {chunk_idx + 1}/{n_chunks} "
                    f"LLM devolvió {len(ihq_devueltos)}/{len(ihq_en_chunk)} IHQ"
                    + (f" — omitidos: {', '.join(ihq_omitidos_chunk[:5])}" if ihq_omitidos_chunk else "")
                )

                _modelo = resp.get("modelo", "desconocido")
                result["modelo"] = _modelo  # V6.9.10: propagar modelo al consolidador

                # === Persistencia + append a live (idéntica lógica a la versión secuencial) ===
                # Por cada IHQ del chunk: dedupe (lock), limpiar, BD bajo lock, append bajo lock.
                # Mantenemos el invariante CRÍTICO V6.9.5: save→append (polling refresca después de MySQL).
                for d in normalizados:
                    raw_key = (d.get("__numero_peticion__") or "").strip().upper()
                    key = _re.sub(r'\s+', '', raw_key)
                    if not key or not _valid_ihq_pat.match(key):
                        logging.info(f"[IA] Descartando fila con numero inválido: {raw_key!r}")
                        continue

                    # Dedupe atómico bajo lock (ihq_vistos es set compartido)
                    with _ia_state_lock:
                        if key in ihq_vistos:
                            continue
                        ihq_vistos.add(key)

                    # Limpiar Dx + Organo (las funciones helper son puras, sin estado)
                    dx_raw = d.get("Diagnostico Principal", "") or ""
                    organo_raw = d.get("Organo", "") or ""
                    dx_limpio, organo_limpio = self._limpiar_resultado_ia(dx_raw, organo_raw)
                    d["Numero de caso"] = key
                    d["Diagnostico Principal"] = dx_limpio
                    d["Organo"] = organo_limpio

                    entry = dict(d)
                    entry["numero_peticion"] = key
                    entry["diagnostico"] = dx_limpio
                    entry["organo"] = organo_limpio
                    entry["pdf_origen"] = filename

                    # PASO 1: BD bajo lock (SQLite/MySQL: las 3 escrituras del caso deben ser atómicas)
                    with _ia_db_lock:
                        try:
                            datos_para_bd = {
                                k: v for k, v in d.items()
                                if not k.startswith("__")
                            }
                            _save_caso_ia(
                                datos_columnas=datos_para_bd,
                                pdf_origen=filename,
                                fecha_procesamiento=_fecha_iso,
                                modelo_utilizado=_modelo,
                                ocr_caracteres_pdf=ocr_chars,
                            )
                            try:
                                _save_records_main([datos_para_bd])
                                logging.info(
                                    f"[IA] {key}: guardado en MySQL (informes_ihq + diagnosticos_ia)"
                                )
                            except Exception as _e_main:
                                logging.warning(
                                    f"[IA] BD principal: no se persistió {key}: {_e_main}"
                                )
                        except Exception as _e:
                            logging.warning(f"[IA] No se persistió {key}: {_e}")

                    # PASO 2: append al estado DESPUÉS del save (invariante V6.9.5)
                    with _ia_state_lock:
                        self._processing_result_ia["live_diagnosticos"].append(entry)

                # Actualizar contador de chunks completados (progreso visible)
                with _ia_state_lock:
                    chunks_completados_counter[0] += 1
                    self._processing_result_ia["current_chunk"] = chunks_completados_counter[0]
                    if chunks_completados_counter[0] >= n_chunks:
                        self._processing_result_ia["chunk_start_time"] = None

                return result

            # === BRANCHING SECUENCIAL vs PARALELO ===
            if paralelo_max == 1:
                # PATH SECUENCIAL: comportamiento original retrocompatible.
                # Reutiliza la misma función helper para que no haya dos
                # implementaciones de la lógica de procesamiento.
                logging.info(f"[IA] {filename}: procesando {n_chunks} chunks en MODO SECUENCIAL (paralelo_max=1)")
                for chunk_idx, chunk_text in enumerate(chunks):
                    res = _procesar_chunk_individual(chunk_idx, chunk_text)
                    chunks_resultados[chunk_idx] = res
            else:
                # PATH PARALELO: ThreadPoolExecutor con paralelo_max workers.
                # Aprovecha los n_parallel slots de LM Studio (n_parallel=4 por default).
                # Cada thread es independiente; las únicas zonas críticas son el lock
                # de estado compartido y el lock de BD. La calidad del output es idéntica
                # (mismo prompt, mismo schema, mismo cliente, mismo retry).
                logging.info(
                    f"[IA] {filename}: procesando {n_chunks} chunks en MODO PARALELO "
                    f"(paralelo_max={paralelo_max}) — aprovecha n_parallel de LM Studio"
                )
                with concurrent.futures.ThreadPoolExecutor(
                    max_workers=paralelo_max,
                    thread_name_prefix="ia_chunk"
                ) as executor:
                    futures = {
                        executor.submit(_procesar_chunk_individual, idx, ctxt): idx
                        for idx, ctxt in enumerate(chunks)
                    }
                    for fut in concurrent.futures.as_completed(futures):
                        cidx = futures[fut]
                        try:
                            res = fut.result()
                            chunks_resultados[cidx] = res
                        except Exception as _e_thread:
                            err_msg = f"chunk {cidx + 1}/{n_chunks}: thread excepción — {_e_thread}"
                            logging.error(f"[IA] {filename}: {err_msg}")
                            with _ia_state_lock:
                                self._processing_result_ia["errors"].append(f"{filename}: {err_msg}")
                            chunks_resultados[cidx] = {
                                "chunk_idx": cidx, "normalizados": [],
                                "err": err_msg, "exitoso": False,
                            }

            # === Consolidación de resultados (orden estable: por chunk_idx) ===
            # Equivalente al estado que tenía la versión secuencial al final del bucle.
            _modelo_pdf = "desconocido"  # V6.9.10: derivado del primer chunk exitoso
            for res in chunks_resultados:
                if res is None:
                    continue
                if res.get("exitoso"):
                    chunks_exitosos += 1
                    todos_diagnosticos.extend(res.get("normalizados", []))
                    if _modelo_pdf == "desconocido" and res.get("modelo"):
                        _modelo_pdf = res["modelo"]
                if res.get("err"):
                    errores_chunks.append(res["err"])

            # V6.9.10 COMPAT: el bloque de payload (abajo) usa `resp` para extraer
            # el modelo. Como en el path paralelo `resp` no queda en el scope local
            # del bucle externo, sintetizamos un objeto-stub mínimo que cumple el
            # contrato esperado: `.get("modelo", default)`. Si ningún chunk fue
            # exitoso, _modelo_pdf="desconocido" y el payload usará ese valor.
            class _RespStub:
                __slots__ = ("_m",)
                def __init__(self, modelo):
                    self._m = modelo
                def get(self, k, default=None):
                    return self._m if k == "modelo" else default
            resp = _RespStub(_modelo_pdf)
            _modelo = _modelo_pdf  # usado por segunda pasada (deshabilitada por default)

            # Deduplicar por numero_peticion (los chunks pueden solaparse o el
            # mismo IHQ aparecer en 2 chunks si una página repite el header)
            vistos = {}
            for d in todos_diagnosticos:
                key = (d.get("numero_peticion") or "").strip().upper()
                if not key:
                    continue
                if key not in vistos:
                    vistos[key] = d

            diagnosticos_finales = list(vistos.values())
            n_dx = len(diagnosticos_finales)
            logging.info(
                f"[IA] {filename}: LLM identificó {n_dx} dx únicos "
                f"({chunks_exitosos}/{n_chunks} chunks OK)"
            )

            # V6.7.7 — Verificar IHQ esperados vs encontrados.
            ihq_encontrados = {(d.get("numero_peticion") or "").strip().upper()
                               for d in diagnosticos_finales}
            ihq_encontrados_norm = {_re.sub(r'\s+', '', k) for k in ihq_encontrados}
            ihq_faltantes = sorted(_ihq_esperados - ihq_encontrados_norm)

            # V6.7.9 — SEGUNDA PASADA: Reintentar IHQ faltantes individualmente.
            # Cuando el LLM omite IHQ en un chunk con varios casos, mandarle
            # el IHQ AISLADO suele recuperarlo (sin distracciones de otros).
            #
            # V6.9.1 — DESHABILITADA POR DEFAULT. Razón: con el schema de 184
            # campos (V6.8.0+), cada reintento tarda ~5-7 min. Para 4 IHQ
            # faltantes son 20-30 min EXTRA después de que ya terminó la
            # primera pasada — UX inaceptable.
            #
            # Si querés recuperar los IHQ faltantes:
            #   1. Mirá los logs (se imprime la lista al final)
            #   2. Reprocesá el mismo PDF (UPSERT no duplica los ya capturados)
            #   3. O usá "Procesar seleccionados" (extractor tradicional)
            #
            # Para reactivar: cambiá HABILITAR_SEGUNDA_PASADA = True
            HABILITAR_SEGUNDA_PASADA = False
            if ihq_faltantes:
                logging.info(
                    f"[IA] {filename}: {len(ihq_faltantes)} IHQ no fueron "
                    f"capturados en primera pasada: {', '.join(ihq_faltantes[:10])}"
                    + (f' (+{len(ihq_faltantes)-10} más)' if len(ihq_faltantes) > 10 else '')
                )
            if ihq_faltantes and HABILITAR_SEGUNDA_PASADA:
                logging.info(
                    f"[IA] {filename}: 🔁 Segunda pasada — reintentando "
                    f"{len(ihq_faltantes)} IHQ faltantes individualmente..."
                )
                self._processing_result_ia["current_stage"] = (
                    f"Segunda pasada: reintentando {len(ihq_faltantes)} IHQ faltantes..."
                )

                # Construir mapping IHQ → segmento de texto del OCR
                matches_all = list(_header_pat_count.finditer(texto_ocr))
                boundaries_ihq = {}  # {ihq_num: (start, end)}
                for i, m in enumerate(matches_all):
                    ihq_num = _re.sub(r'\s+', '', m.group(1)).upper()
                    if ihq_num in boundaries_ihq:
                        continue  # ya tiene first occurrence
                    end_pos = len(texto_ocr)
                    for j in range(i + 1, len(matches_all)):
                        other = _re.sub(r'\s+', '', matches_all[j].group(1)).upper()
                        if other != ihq_num:
                            end_pos = matches_all[j].start()
                            break
                    boundaries_ihq[ihq_num] = (m.start(), end_pos)

                recuperados_segunda_pasada = 0
                for ihq_falt in ihq_faltantes:
                    if ihq_falt not in boundaries_ihq:
                        continue
                    s, e = boundaries_ihq[ihq_falt]
                    segmento = texto_ocr[s:e]
                    if len(segmento) < 50:
                        continue

                    self._processing_result_ia["current_stage"] = (
                        f"Segunda pasada: reintentando {ihq_falt}..."
                    )
                    logging.info(
                        f"[IA] Reintentando {ihq_falt} aislado ({len(segmento)} chars)"
                    )

                    resp_f, err_f = _llamar_llm_con_retry(segmento)
                    if resp_f is None:
                        logging.warning(f"[IA] {ihq_falt}: reintento individual falló")
                        continue

                    contenido_f = resp_f.get("respuesta") or resp_f.get("contenido")
                    if isinstance(contenido_f, dict):
                        data_f = contenido_f
                    elif isinstance(contenido_f, str) and contenido_f.strip():
                        data_f = _extraer_json_robusto(contenido_f)
                    else:
                        data_f = None

                    if not data_f or not isinstance(data_f, dict):
                        continue

                    raw_dx_f = data_f.get("diagnosticos", [])
                    if not isinstance(raw_dx_f, list):
                        raw_dx_f = []
                    normalizados_f = [d for d in (_norm_dx_entry(x) for x in raw_dx_f) if d]

                    for d in normalizados_f:
                        raw_key_f = (d.get("__numero_peticion__") or "").strip().upper()
                        key_f = _re.sub(r'\s+', '', raw_key_f)
                        if not key_f or not _valid_ihq_pat.match(key_f):
                            continue
                        if key_f in ihq_vistos:
                            continue
                        ihq_vistos.add(key_f)

                        dx_raw_f = d.get("Diagnostico Principal", "") or ""
                        organo_raw_f = d.get("Organo", "") or ""
                        dx_limpio_f, organo_limpio_f = self._limpiar_resultado_ia(
                            dx_raw_f, organo_raw_f
                        )
                        d["Numero de caso"] = key_f
                        d["Diagnostico Principal"] = dx_limpio_f
                        d["Organo"] = organo_limpio_f

                        # V6.9.5 — Mismo fix race condition aquí
                        entry_f = dict(d)
                        entry_f["numero_peticion"] = key_f
                        entry_f["diagnostico"] = dx_limpio_f
                        entry_f["organo"] = organo_limpio_f
                        entry_f["pdf_origen"] = filename

                        # PASO 1: Guardar en MySQL ANTES del append
                        try:
                            datos_para_bd_f = {
                                k: v for k, v in d.items()
                                if not k.startswith("__")
                            }
                            _save_caso_ia(
                                datos_columnas=datos_para_bd_f,
                                pdf_origen=filename,
                                fecha_procesamiento=_fecha_iso,
                                modelo_utilizado=resp_f.get("modelo", "desconocido"),
                                ocr_caracteres_pdf=ocr_chars,
                            )
                            try:
                                _save_records_main([datos_para_bd_f])
                                logging.info(
                                    f"[IA] {key_f}: guardado en MySQL (reintento)"
                                )
                            except Exception as _e_main:
                                logging.warning(
                                    f"[IA] BD principal: no se persistió {key_f} (reintento): {_e_main}"
                                )
                        except Exception as _e:
                            logging.warning(
                                f"[IA] No se persistió {key_f} (reintento): {_e}"
                            )

                        # PASO 2: Append DESPUÉS del save (polling refresca Visualizador)
                        self._processing_result_ia["live_diagnosticos"].append(entry_f)
                        diagnosticos_finales.append(entry_f)
                        recuperados_segunda_pasada += 1

                logging.info(
                    f"[IA] {filename}: Segunda pasada recuperó "
                    f"{recuperados_segunda_pasada}/{len(ihq_faltantes)} IHQ"
                )

                # Re-evaluar faltantes después de segunda pasada
                ihq_encontrados2 = {(d.get("numero_peticion") or "").strip().upper()
                                    for d in diagnosticos_finales}
                ihq_encontrados2_norm = {_re.sub(r'\s+', '', k) for k in ihq_encontrados2}
                ihq_faltantes_finales = sorted(_ihq_esperados - ihq_encontrados2_norm)

                if ihq_faltantes_finales:
                    msg = (
                        f"{filename}: ⚠️ Tras segunda pasada, "
                        f"{len(ihq_faltantes_finales)} IHQ siguen faltando: "
                        f"{', '.join(ihq_faltantes_finales[:10])}"
                        + (f" (+{len(ihq_faltantes_finales) - 10} más)" if len(ihq_faltantes_finales) > 10 else "")
                    )
                    self._processing_result_ia["errors"].append(msg)
                    logging.warning(f"[IA] {msg}")
                elif recuperados_segunda_pasada > 0:
                    logging.info(
                        f"[IA] {filename}: ✅ Segunda pasada recuperó TODOS los faltantes"
                    )

                # Actualizar n_dx
                n_dx = len(diagnosticos_finales)

            # Guardar resultado
            out_filename = f"extraccion_ia_{os.path.splitext(filename)[0].replace(' ', '_')}.json"
            out_path = os.path.join(out_dir, out_filename)
            # V6.9.2 — Fix bug: resp puede ser None si TODOS los chunks fallaron.
            # El check 'resp in dir()' no detecta None (solo si la variable
            # no existe). Necesitamos chequeo explícito.
            try:
                _modelo_payload = (
                    resp.get("modelo", "desconocido")
                    if 'resp' in dir() and resp is not None and hasattr(resp, 'get')
                    else "desconocido"
                )
            except Exception:
                _modelo_payload = "desconocido"
            payload = {
                "pdf_origen": filename,
                "fecha_procesamiento": _dt.now().isoformat(),
                "modelo_utilizado": _modelo_payload,
                "ocr_caracteres": ocr_chars,
                "chunks_total": n_chunks,
                "chunks_exitosos": chunks_exitosos,
                "diagnosticos_identificados": n_dx,
                "diagnosticos": diagnosticos_finales,
                "errores_chunks": errores_chunks,
            }
            try:
                with open(out_path, 'w', encoding='utf-8') as f:
                    _json.dump(payload, f, ensure_ascii=False, indent=2)
                self._processing_result_ia["results"].append({
                    "pdf": filename,
                    "ocr_chars": ocr_chars,
                    "n_dx": n_dx,
                    "n_chunks": n_chunks,
                    "chunks_ok": chunks_exitosos,
                    "out_path": out_path,
                })
            except Exception as e:
                self._processing_result_ia["errors"].append(
                    f"{filename}: guardar JSON falló — {e}"
                )

            if errores_chunks and chunks_exitosos == 0:
                self._processing_result_ia["errors"].append(
                    f"{filename}: TODOS los chunks fallaron — {errores_chunks[0]}"
                )

        self._processing_result_ia["done"] = True

        # V6.8.0 — Refrescar el Visualizador de datos al finalizar todo
        # el procesamiento IA. Los casos extraídos ya están en la BD
        # principal (informes_ihq), solo falta recargar el DataFrame
        # maestro y repintar la tabla del visualizador.
        # La llamada va por self.after() porque estamos en thread worker;
        # refresh_data_and_table() toca UI y debe correr en el main thread.
        try:
            self.after(100, self.refresh_data_and_table)
            logging.info("[IA] Refresh del Visualizador de datos disparado")
        except Exception as e:
            logging.warning(f"[IA] No se pudo disparar refresh del visualizador: {e}")

    def _poll_processing_progress_ia(self):
        """V6.7.0 — Polling: actualiza barras + tabla en vivo + counter."""
        state = self._processing_result_ia

        # Solo manejar UI si la ventana sigue existiendo
        win_alive = (
            hasattr(self, '_ia_progress_win')
            and self._ia_progress_win is not None
        )
        if win_alive:
            try:
                win_alive = bool(self._ia_progress_win.winfo_exists())
            except Exception:
                win_alive = False

        if win_alive:
            try:
                # Header — PDF actual
                idx = state["current_index"]
                total = state["total_files"]
                fname = state["current_file"] or "—"
                self._ia_lbl_pdf.config(
                    text=f"PDF {idx}/{total}: {fname}"
                )

                # Header — chunk + stage + tiempo transcurrido en chunk actual
                cidx = state.get("current_chunk", 0)
                ctot = state.get("total_chunks", 0)
                stage_text = state['current_stage']
                # V6.7.10 — Si hay timestamp activo, mostrar elapsed
                chunk_start = state.get("chunk_start_time")
                if chunk_start:
                    import time as _time_local
                    elapsed = int(_time_local.time() - chunk_start)
                    if elapsed >= 3:  # solo mostrar si lleva 3+ seg esperando
                        m, s = divmod(elapsed, 60)
                        elapsed_str = f"{m}:{s:02d}" if m > 0 else f"{s}s"
                        stage_text = f"{stage_text} ⏱ {elapsed_str}"

                if ctot > 0:
                    self._ia_lbl_chunk.config(
                        text=f"Chunk {cidx}/{ctot} — {stage_text}"
                    )
                else:
                    self._ia_lbl_chunk.config(text=stage_text)

                # Barras
                self._ia_progress_pdfs['value'] = max(0, idx - 1) if not state["done"] else total
                self._ia_progress_chunks['maximum'] = max(1, ctot)
                self._ia_progress_chunks['value'] = cidx if ctot > 0 else 0

                # Status footer
                if state["done"]:
                    self._ia_lbl_status.config(
                        text="✅ Procesamiento completo",
                        foreground="#0a7"
                    )
                else:
                    self._ia_lbl_status.config(
                        text=f"🔄 {stage_text}",
                        foreground="#06b"
                    )

                # Tabla en vivo: pintar diagnósticos nuevos
                # V6.7.20 — Mostrar dx COMPLETO en la tabla (antes truncaba
                # a 200 chars + "..."). El LLM ya extrae todo correctamente
                # y la BD guarda completo; el truncamiento visual hacía
                # parecer que la extracción estaba rota cuando no lo estaba.
                # Casos testigo recuperados visualmente: IHQ250100 (262 chars),
                # IHQ250160 (222 chars), IHQ250178 (235 chars).
                live = state.get("live_diagnosticos", [])
                already_painted = self._ia_diagnosticos_pintados
                new_count = len(live) - already_painted
                if new_count > 0:
                    cols_bd = getattr(self, "_ia_columnas_bd", [])
                    for entry in live[already_painted:]:
                        ihq = entry.get("numero_peticion", "")
                        dx = (entry.get("diagnostico", "") or "").replace("\n", " ")
                        organo = entry.get("organo", "")
                        # V6.8.0 — Construir tupla con TODAS las columnas:
                        # 3 alias visibles + 184 columnas BD en orden
                        values_full = [ihq, dx, organo]
                        for bd_col in cols_bd:
                            v = entry.get(bd_col, "N/A")
                            if isinstance(v, str):
                                v = v.replace("\n", " ")
                            else:
                                v = str(v) if v is not None else "N/A"
                            values_full.append(v)
                        self._ia_treeview.insert("", "end", values=tuple(values_full))
                    # Auto-scroll al final
                    children = self._ia_treeview.get_children()
                    if children:
                        self._ia_treeview.see(children[-1])
                    self._ia_diagnosticos_pintados = len(live)

                    # V6.9.4 — REFRESH EN TIEMPO REAL del Visualizador de Datos.
                    # Cada vez que se agrega un IHQ nuevo a la tabla del modal,
                    # también refrescamos el Visualizador (background) para que
                    # los datos aparezcan en la tabla principal sin esperar a
                    # que termine TODO el procesamiento. self.after() lo programa
                    # en el main thread (el poll ya corre ahí, pero igual usamos
                    # after para no bloquear el render del modal).
                    try:
                        if hasattr(self, 'refresh_data_and_table'):
                            self.after(50, self.refresh_data_and_table)
                    except Exception as _e_refresh:
                        logging.warning(
                            f"[IA] No se pudo refrescar Visualizador: {_e_refresh}"
                        )

                # Counter de la sesión actual
                self._ia_lbl_counter.config(
                    text=f"Diagnósticos en esta sesión: {len(live)}"
                )

                # Total acumulado en BD (refrescado cada poll para reflejar
                # los inserts del worker en tiempo real)
                try:
                    from core.diagnosticos_ia_db import count_total as _ia_count_total
                    total_acum = _ia_count_total()
                    self._ia_lbl_acumulado.config(
                        text=f"Total acumulado en BD (todas las sesiones): {total_acum}"
                    )
                except Exception:
                    pass
            except Exception as e:
                logging.warning(f"[IA UI] Error actualizando UI: {e}")

        # Si terminó, mostrar resumen y habilitar botón de cierre
        if state["done"]:
            if win_alive:
                try:
                    self._ia_btn_cerrar.config(state="normal")
                except Exception:
                    pass
            # Mostrar resumen en messagebox (puede convivir con la ventana)
            self._show_ia_results_summary()
            return

        self.after(500, self._poll_processing_progress_ia)

    def _show_ia_results_summary(self):
        """Muestra ventana con resumen de resultados del procesamiento IA."""
        state = self._processing_result_ia
        results = state.get("results", [])
        errors = state.get("errors", [])

        total_dx = sum(r["n_dx"] for r in results)
        total_pdfs = len(results)

        msg_lines = [
            f"✅ Procesamiento IA completado",
            f"",
            f"PDFs procesados: {total_pdfs}",
            f"Total de diagnósticos identificados por la IA: {total_dx}",
            f"",
        ]

        if results:
            msg_lines.append("Detalle por PDF:")
            for r in results:
                chunks_info = ""
                if "n_chunks" in r:
                    chunks_info = f", {r['chunks_ok']}/{r['n_chunks']} chunks OK"
                msg_lines.append(
                    f"  • {r['pdf']}: {r['n_dx']} diagnósticos "
                    f"(OCR: {r['ocr_chars']:,} chars{chunks_info})"
                )
            msg_lines.append("")
            msg_lines.append("Persistencia:")
            msg_lines.append("  • JSON por PDF en: informes_ia/")
            try:
                from core.diagnosticos_ia_db import count_total as _ia_count_total, count_by_pdf as _ia_count_by_pdf
                msg_lines.append(
                    f"  • BD acumulativa: data/diagnosticos_ia.db ({_ia_count_total()} dx totales)"
                )
                pdfs_db = _ia_count_by_pdf()
                if pdfs_db:
                    msg_lines.append(f"  • PDFs en BD: {len(pdfs_db)}")
            except Exception:
                pass

        if errors:
            msg_lines.append("")
            msg_lines.append(f"⚠️ Errores: {len(errors)}")
            for err in errors[:5]:
                msg_lines.append(f"  • {err[:150]}")

        messagebox.showinfo("Procesamiento con IA — resultados", "\n".join(msg_lines))

    def _show_processing_overlay(self, num_files):
        """Mostrar overlay con barra de progreso sobre la UI"""
        self._progress_overlay = tk.Toplevel(self)
        overlay = self._progress_overlay
        overlay.title("Procesando PDFs...")
        overlay.transient(self)
        overlay.grab_set()
        overlay.resizable(False, False)
        overlay.protocol("WM_DELETE_WINDOW", lambda: None)  # No se puede cerrar

        # Centrar ventana
        w, h = 480, 200
        x = self.winfo_x() + (self.winfo_width() - w) // 2
        y = self.winfo_y() + (self.winfo_height() - h) // 2
        overlay.geometry(f"{w}x{h}+{x}+{y}")

        frame = ttk.Frame(overlay, padding=30)
        frame.pack(fill=BOTH, expand=True)

        ttk.Label(
            frame,
            text="⏳ Procesando archivos PDF...",
            font=("Segoe UI", 13, "bold"),
        ).pack(pady=(0, 5))

        self._progress_status_label = ttk.Label(
            frame,
            text=f"Preparando {num_files} archivo(s)...",
            font=("Segoe UI", 10),
        )
        self._progress_status_label.pack(pady=(0, 15))

        self._progress_bar = ttk.Progressbar(
            frame,
            mode="determinate",
            length=400,
            maximum=num_files,
            bootstyle="success-striped",
        )
        self._progress_bar.pack(pady=(0, 10))

        self._progress_detail_label = ttk.Label(
            frame,
            text="El OCR puede tardar varios minutos por archivo...",
            font=("Segoe UI", 9),
            foreground="gray",
        )
        self._progress_detail_label.pack()

    def _process_files_worker(self, files_to_process):
        """Worker thread: procesa PDFs (pesado, NO toca UI)"""
        result = self._processing_result

        for i, (file_path, filename) in enumerate(files_to_process):
            result["current_file"] = filename
            result["current_index"] = i

            try:
                records_count, correcciones = self._process_file(file_path)
                result["processed_count"] += 1
                result["total_records"] += records_count
                if correcciones:
                    result["correcciones"].extend(correcciones)
            except Exception as e:
                result["errors"].append(f"{filename}: {str(e)}")
                logging.error(f"❌ Error procesando {filename}: {e}")

        # V6.9.46: reconciliar coloración<->IHQ por cédula tras importar (independiente
        # del orden de llegada: da igual si llegó antes el PDF de coloración o el IHQ).
        # Es idempotente y solo recalcula la columna derivada 'Diagnostico Coloracion 2'.
        if result["processed_count"] > 0:
            try:
                from core.coloracion_processor import reconciliar_coloraciones
                rec = reconciliar_coloraciones()
                result["reconciliacion"] = rec
                logging.info(
                    f"🔗 Reconciliación coloraciones: {rec.get('ihq_actualizados', 0)} "
                    f"filas IHQ actualizadas, {rec.get('con_varias', 0)} con varias"
                )
            except Exception as e:
                logging.error(f"⚠️ Error en reconciliación de coloraciones: {e}")

        result["done"] = True

    def _poll_processing_progress(self):
        """Polling desde main thread: actualiza barra de progreso"""
        result = self._processing_result

        if not result["done"]:
            # Actualizar progreso
            idx = result["current_index"]
            filename = result["current_file"]
            total = result["total_files"]

            if filename:
                self._progress_status_label.configure(
                    text=f"Procesando ({idx + 1}/{total}): {filename}"
                )
                self._progress_bar.configure(value=idx)

            # Seguir haciendo polling cada 500ms
            self.after(500, self._poll_processing_progress)
            return

        # Procesamiento terminado - cerrar overlay
        self._progress_bar.configure(value=result["total_files"])
        self._progress_status_label.configure(text="✅ Procesamiento completado")
        self.after(600, self._on_processing_complete)

    def _on_processing_complete(self):
        """Callback main thread: mostrar resultados después del procesamiento"""
        result = self._processing_result

        # Cerrar overlay
        if hasattr(self, '_progress_overlay') and self._progress_overlay:
            self._progress_overlay.grab_release()
            self._progress_overlay.destroy()
            self._progress_overlay = None

        processed_count = result["processed_count"]
        total_records = result["total_records"]
        errors = result["errors"]
        peticiones_antes = result["peticiones_antes"]

        if processed_count > 0:
            # V6.9.31 FIX: usar los números de caso REALES capturados durante el
            # procesamiento (incluye reimportaciones). El método anterior
            # ("despues - antes") solo detectaba casos NUEVOS y daba 0 al
            # reimportar casos ya existentes (UPSERT no cambia el set de IDs).
            capturados = []
            try:
                capturados = list(dict.fromkeys(
                    getattr(self, '_ultimos_registros_procesados', []) or []
                ))
            except Exception:
                capturados = []

            if capturados:
                numeros_peticion_procesados = capturados
                self._ultimos_registros_procesados = capturados
                logging.info(f"📋 Registros procesados (reales): {len(capturados)}")
            else:
                # Fallback: diferencia de conjuntos en la BD activa (MySQL)
                try:
                    peticiones_despues = self._peticiones_existentes_bd()
                    nuevas_peticiones = list(peticiones_despues - peticiones_antes)
                    self._ultimos_registros_procesados = nuevas_peticiones
                    numeros_peticion_procesados = nuevas_peticiones
                    logging.info(f"📋 Registros nuevos (fallback): {len(nuevas_peticiones)}")
                except Exception as e:
                    logging.warning(f"⚠️ Error capturando registros: {e}")
                    self._ultimos_registros_procesados = []
                    numeros_peticion_procesados = []

            # Analizar completitud
            try:
                from core.validation_checker import analizar_batch_registros

                logging.info(f"🔍 Analizando completitud de {len(numeros_peticion_procesados)} registros...")
                analisis = analizar_batch_registros(numeros_peticion_procesados)

                logging.info(f"✅ Análisis completado:")
                logging.info(f"   • Completos: {analisis['resumen']['completos']}")
                logging.info(f"   • Incompletos: {analisis['resumen']['incompletos']}")

                try:
                    self.refresh_data_and_table()
                    self.after(500, self._delayed_refresh_after_processing)
                    if hasattr(self, 'enhanced_dashboard'):
                        self.enhanced_dashboard.refresh_all_data()
                except Exception as e:
                    logging.warning(f"⚠️ Error en refresh: {e}")

                from core.ventana_resultados_importacion import mostrar_ventana_resultados

                mostrar_ventana_resultados(
                    parent=self,
                    completos=analisis['completos'],
                    incompletos=analisis['incompletos'],
                    resumen=analisis['resumen'],
                    callback_auditar=self._mostrar_selector_tipo_auditoria,
                    callback_continuar=self._nav_to_visualizar
                )

            except Exception as e:
                logging.warning(f"⚠️ Error en análisis de completitud: {e}")

                success_msg = f"✅ Procesamiento completado:\n"
                success_msg += f"• {processed_count} archivos procesados\n"
                success_msg += f"• {total_records} registros extraídos y guardados en BD"

                if errors:
                    success_msg += f"\n\n⚠️ Errores en {len(errors)} archivos:\n"
                    success_msg += "\n".join(errors[:3])
                    if len(errors) > 3:
                        success_msg += f"\n... y {len(errors) - 3} errores más."

                messagebox.showinfo("Procesamiento completado", success_msg)

                self.refresh_data_and_table()
                self.after(1000, self._delayed_refresh_after_processing)
                if hasattr(self, 'enhanced_dashboard'):
                    self.enhanced_dashboard.refresh_all_data()
                self._nav_to_visualizar()

        else:
            error_msg = "❌ No se pudo procesar ningún archivo.\n\nErrores encontrados:\n"
            error_msg += "\n".join(errors[:5])
            messagebox.showerror("Error de procesamiento", error_msg)

    def _process_file(self, file_path):
        """Procesar un archivo PDF individual

        Returns:
            tuple: (records_count, correcciones_list)
        """
        try:
            filename = os.path.basename(file_path)
            # V4.2.1 FIX: No llamar set_status desde worker thread (no es thread-safe en Python 3.13)
            logging.info(f"Procesando {filename}...")

            # Determinar el tipo de archivo y procesarlo adecuadamente
            # V6.9.45: detectar PRIMERO los PDFs de Coloraciones ("M … AL …") y
            # enrutarlos a su procesador aislado (NO toca el flujo IHQ).
            if self._is_coloracion_file(filename, file_path):
                records_processed = self._process_coloracion_file(file_path)
                correcciones = []
                logging.info(f"✅ {filename}: {records_processed} registros de Coloración procesados")
            elif self._is_ihq_file(filename, file_path):
                # Procesar como archivo IHQ (biomarcadores)
                records_processed = self._process_ihq_file(file_path)
                correcciones = []  # IHQ no tiene validación aún
                logging.info(f"✅ {filename}: {records_processed} registros IHQ procesados")
            else:
                # Procesar como archivo general de patología (retorna tuple)
                records_processed, correcciones = self._process_general_file(file_path)
                logging.info(f"✅ {filename}: {records_processed} registros generales procesados")

            # V5.3.9: Usar logging en lugar de print (stdout puede estar cerrado)
            logging.info(f"✅ Procesamiento completado: {file_path} - {records_processed} registros")
            return records_processed, correcciones

        except Exception as e:
            error_msg = f"Error procesando {filename}: {str(e)}"
            logging.error(f"❌ {error_msg}")
            raise Exception(error_msg)

    def _is_ihq_file(self, filename, file_path):
        """Determinar si un archivo es de IHQ basándose en el nombre y contenido"""
        # Criterio 1: Nombre del archivo
        if "ihq" in filename.lower():
            return True
        
        # Criterio 2: Revisar contenido del archivo (muestra)
        try:
            from core.processors.ocr_processor import pdf_to_text_enhanced
            # Solo leer una página para determinar el tipo
            import fitz
            doc = fitz.open(file_path)
            if len(doc) > 0:
                page_text = doc[0].get_text()
                doc.close()
                # Buscar indicadores de IHQ
                ihq_indicators = ['ihq', 'inmunohistoquimica', 'her2', 'ki-67', 'receptor estrogeno']
                return any(indicator in page_text.lower() for indicator in ihq_indicators)
        except Exception:
            pass
        
        return False

    def _process_ihq_file(self, file_path):
        """Procesar archivo IHQ - LÓGICA MOVIDA A core/ihq_processor.py"""
        from core.ihq_processor import process_ihq_file

        # Crear función de log que use el widget de UI
        def log_callback(msg):
            if hasattr(self, 'log_to_widget'):
                self.log_to_widget(msg)

        # V6.9.31: capturar los números de caso REALES procesados (incluye
        # reimportaciones). Antes el modal usaba "despues - antes" y daba 0 al
        # reimportar casos ya existentes. Acumulamos en _ultimos_registros_procesados.
        numeros = []
        count = process_ihq_file(file_path, log_callback, out_numeros=numeros)
        try:
            if getattr(self, '_ultimos_registros_procesados', None) is None:
                self._ultimos_registros_procesados = []
            self._ultimos_registros_procesados.extend(numeros)
        except Exception as e:
            logging.warning(f"⚠️ No se pudieron acumular números procesados: {e}")
        return count

    def _is_coloracion_file(self, filename, file_path):
        """V6.9.45: True si el PDF es de Coloraciones básicas (estudio M autónomo).
        Detecta por nombre ("M … AL …") o por contenido (N. peticion = M###### con
        indicios de coloración y SIN indicadores IHQ). NO clasifica IHQ como coloración."""
        import re
        # Criterio 1: nombre del lote ("M 2503754 AL 2503803" / "M2515951 AL M2516000")
        if re.match(r'^\s*M\s?\d{5,}\s+AL\s+M?\d{5,}', filename, re.IGNORECASE):
            return True
        if 'ihq' in filename.lower():
            return False
        # Criterio 2: contenido de la primera página
        try:
            import fitz
            doc = fitz.open(file_path)
            txt = doc[0].get_text() if len(doc) else ''
            doc.close()
            low = txt.lower()
            tiene_m = bool(re.search(r'petici[oó�]n\s*\n?\s*:\s*M\d{5,}', txt, re.IGNORECASE))
            ihq_ind = any(k in low for k in ('inmunohistoquimica', 'inmunohistoquímica', 'her2', 'ki-67', 'ki67'))
            color_ind = any(k in low for k in ('coloracion basica', 'coloración básica',
                                               'estudio de histologia', 'estudio de histología'))
            return tiene_m and color_ind and not ihq_ind
        except Exception:
            return False

    def _process_coloracion_file(self, file_path):
        """V6.9.45: procesa un PDF de Coloraciones - lógica en core/coloracion_processor.py."""
        from core.coloracion_processor import process_coloracion_file

        def log_callback(msg):
            if hasattr(self, 'log_to_widget'):
                self.log_to_widget(msg)

        numeros = []
        stats = process_coloracion_file(file_path, dry_run=False,
                                        log_callback=log_callback, out_numeros=numeros)
        try:
            if getattr(self, '_ultimos_registros_procesados', None) is None:
                self._ultimos_registros_procesados = []
            self._ultimos_registros_procesados.extend(numeros)
        except Exception as e:
            logging.warning(f"⚠️ No se pudieron acumular números de coloración: {e}")
        # V6.9.46: el procesador ahora reporta 'filas' (antes 'guardados').
        return stats.get('filas', stats.get('guardados', 0))

    def _process_general_file(self, file_path):
        """Procesar archivo general usando el procesador estándar

        Returns:
            tuple: (records_count, correcciones_list)
        """
        try:
            # Importar los módulos necesarios
            from core.processors.ocr_processor import pdf_to_text_enhanced
            from core import unified_extractor as ihq
            from core import database_manager

            # Extraer texto del PDF
            full_text = pdf_to_text_enhanced(file_path)
            if not isinstance(full_text, str):
                full_text = '\n'.join(full_text)

            # Segmentar por informes (usando lógica similar a IHQ pero más general)
            records = []

            # Para archivos generales, puede haber múltiples informes
            # Intentar segmentar por "N. petición" o números de orden
            segments = self._segment_general_reports(full_text)

            if not segments:
                # Si no se puede segmentar, procesar como un solo informe
                segments = [full_text]

            for segment in segments:
                # Extraer datos base usando el procesador IHQ (que maneja datos generales también)
                base_data = ihq.extract_ihq_data(segment)
                base_rows = ihq.map_to_excel_format(base_data)

                if base_rows:
                    records.extend(base_rows)

            if not records:
                raise RuntimeError("No se pudo extraer información del PDF.")

            # Guardar en base de datos
            from core.database_manager import init_db, save_records, update_incomplete_records_with_debug_data
            init_db()
            saved_count = save_records(records)

            # V5.3.9.3: VALIDACIÓN MÉDICO-SERVICIO YA APLICADA EN ihq_processor.py
            # Las correcciones médico-servicio ahora se aplican ANTES de crear debug_maps
            # en ihq_processor.py líneas 119-144, por lo que NO necesitamos volver a aplicarlas aquí.
            # Las correcciones ya están guardadas en los debug_maps y se mostrarán en la ventana.
            correcciones_aplicadas = []
            logging.info("✅ Validación médico-servicio aplicada durante extracción (ihq_processor.py)")

            # REACTIVADO: Actualización automática mejorada para completar datos faltantes
            try:
                logging.info("🔄 Ejecutando actualización automática de registros incompletos...")

                # CORREGIDO: Usar el texto OCR del archivo recién procesado para completar datos
                # Crear archivo DEBUG temporal para este archivo específico
                debug_file_path = self._create_debug_file_for_current_pdf(file_path, full_text)

                # Actualizar usando el DEBUG específico de este archivo
                updated_count = update_incomplete_records_with_debug_data(debug_file_path)
                if updated_count > 0:
                    logging.info(f"✅ Se actualizaron {updated_count} registros con datos completos del archivo actual")

                # NUEVO: Ahora aplicar mapeo específico de órganos usando parse_estudios_table_for_organo
                logging.info("🔄 Aplicando mapeo avanzado de órganos...")
                organ_updated_count = self._update_organs_with_advanced_parsing(full_text, records)
                if organ_updated_count > 0:
                    logging.info(f"✅ Se mapearon {organ_updated_count} órganos adicionales")

            except Exception as e:
                # No fallar el proceso principal si la actualización automática falla
                logging.warning(f"⚠️ Advertencia en actualización automática: {str(e)}")

            # V5.3.9: Retornar tanto el conteo como las correcciones
            return saved_count, correcciones_aplicadas

        except Exception as e:
            raise Exception(f"Error en procesamiento general: {str(e)}")

    def _create_debug_file_for_current_pdf(self, file_path, full_text):
        """Crear archivo DEBUG temporal para el PDF actual"""
        try:
            # Crear nombre de archivo DEBUG basado en el PDF actual
            pdf_name = os.path.basename(file_path)
            debug_filename = f"DEBUG_OCR_OUTPUT_{pdf_name}.txt"
            debug_path = os.path.join("EXCEL", debug_filename)
            
            # Asegurar que el directorio existe
            os.makedirs("EXCEL", exist_ok=True)
            
            # Guardar el texto OCR en formato DEBUG
            with open(debug_path, 'w', encoding='utf-8') as f:
                f.write(f"=== DEBUG OCR OUTPUT PARA {pdf_name} ===\n")
                f.write(f"Generado automáticamente el {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")
                f.write(full_text)

            logging.info(f"📄 Archivo DEBUG creado: {debug_path}")
            return debug_path

        except Exception as e:
            logging.warning(f"⚠️ Error creando archivo DEBUG: {e}")
            # Fallback al archivo DEBUG global si falla
            return None

    def _find_available_debug_files(self):
        """Buscar todos los archivos DEBUG disponibles en el directorio EXCEL"""
        debug_files = []
        excel_dir = "EXCEL"
        
        if os.path.exists(excel_dir):
            for filename in os.listdir(excel_dir):
                if filename.startswith("DEBUG_OCR_OUTPUT_") and filename.endswith(".txt"):
                    debug_path = os.path.join(excel_dir, filename)
                    debug_files.append(debug_path)
                    
        return debug_files

    def _segment_general_reports(self, text):
        """Segmentar texto en informes individuales para archivos generales"""
        segments = []
        
        # Buscar patrones de número de petición más generales
        patterns = [
            r'(?i)(?:N[°.\s]*|No\.\s*|Nº\s*|N\s*)?petici[oó]n\s*[:\-]?\s*([A-Z]?\d{6,})',
            r'(?i)(?:registro|orden|numero|no\.?)\s*[:\-]?\s*(\d{4,})',
            r'(?i)(\d{4,})\s*(?:patolog[ií]a|biopsia|citolog[ií]a)'
        ]
        
        # Intentar segmentar por cualquiera de los patrones
        for pattern in patterns:
            matches = list(re.finditer(pattern, text))
            if len(matches) > 1:  # Solo si encontramos múltiples coincidencias
                starts = [(m.start(), m.group(1)) for m in matches]
                starts.sort()
                
                for i, (start, code) in enumerate(starts):
                    if i + 1 < len(starts):
                        end = starts[i + 1][0]
                        segment = text[start:end].strip()
                    else:
                        segment = text[start:].strip()
                    
                    if len(segment) > 100:  # Solo segmentos con contenido suficiente
                        segments.append(segment)
                
                break  # Usar el primer patrón que funcione
        
        return segments

    def _export_full_database_professional(self):
        """Exportar toda la base de datos a Excel con presentación profesional y diálogo de ubicación"""
        try:
            import pandas as pd
            from datetime import datetime
            
            # Obtener todos los datos de la base de datos
            from core.database_manager import get_all_records_as_dataframe
            df_complete = get_all_records_as_dataframe()
            
            if df_complete.empty:
                messagebox.showwarning("Advertencia", "No hay datos en la base de datos para exportar.")
                return
                
            # Diálogo para guardar archivo
            file_path = filedialog.asksaveasfilename(
                title="Exportar Base de Datos Completa - ONCONOVA",
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
            )
            
            if file_path:
                self._create_professional_excel_export(df_complete, file_path, "completa")
                
                # Redirigir a pestaña de exportaciones
                if hasattr(self, 'notebook'):
                    self.notebook.select(5)  # Pestaña de Exportaciones
                
        except Exception as e:
            messagebox.showerror("Error", f"❌ Error al exportar:\n{str(e)}")
            logging.error(f"Error detallado: {e}", exc_info=True)

    def _export_full_database_direct(self):
        """Exportar toda la base de datos directamente a la carpeta de Documentos sin diálogo"""
        try:
            import pandas as pd
            from datetime import datetime
            import os
            
            # Obtener todos los datos de la base de datos
            from core.database_manager import get_all_records_as_dataframe
            df_complete = get_all_records_as_dataframe()
            
            if df_complete.empty:
                messagebox.showwarning("Advertencia", "No hay datos en la base de datos para exportar.")
                return
            
            # Generar nombre automático usando helper
            filename = export_helpers.generar_nombre_archivo_export(
                prefijo="BD_Completa_HUV",
                tipo="xlsx",
                incluir_timestamp=True
            )
            
            # Ruta automática a Documentos
            export_base_path = os.path.join(os.path.expanduser("~"), "Documents", "ONCONOVA CIRUGÍA ONCOLÓGICA", "Exportaciones Base de datos")
            excel_dir = os.path.join(export_base_path, "Excel")
            os.makedirs(excel_dir, exist_ok=True)
            
            file_path = os.path.join(excel_dir, filename)
            
            self._create_professional_excel_export(df_complete, file_path, "completa")
            
            # Redirigir a pestaña de exportaciones
            if hasattr(self, 'notebook'):
                self.notebook.select(5)  # Pestaña de Exportaciones
                
        except Exception as e:
            messagebox.showerror("Error", f"❌ Error al exportar:\n{str(e)}")
            logging.error(f"Error detallado: {e}", exc_info=True)

    def _export_selected_data_professional(self):
        """Exportar solo datos seleccionados a Excel con presentación profesional y diálogo de ubicación"""
        try:
            import pandas as pd
            from datetime import datetime
            
            # Obtener elementos seleccionados del treeview
            selected_items = self.tree.selection()
            if not selected_items:
                messagebox.showwarning("Sin Selección", "No hay elementos seleccionados para exportar")
                return

            # CORREGIDO: Obtener los datos reales de las filas seleccionadas
            selected_rows_data = []
            for item in selected_items:
                # tksheet: Usar get_row_data() en lugar de .item()
                values = self.sheet.get_row_data(item)
                if values:
                    # Buscar el registro correspondiente en master_df usando el número de petición
                    numero_peticion = values[0]  # Asumiendo que la primera columna es número de petición
                    matching_row = self.master_df[self.master_df.iloc[:, 0] == numero_peticion]
                    if not matching_row.empty:
                        selected_rows_data.append(matching_row.iloc[0])

            if not selected_rows_data:
                messagebox.showwarning("Sin Datos", "No se pudieron obtener los datos de la selección")
                return

            # Crear DataFrame con solo los registros seleccionados
            selected_df = pd.DataFrame(selected_rows_data)
            
            # Diálogo para guardar archivo
            file_path = filedialog.asksaveasfilename(
                title=f"Exportar Selección ({len(selected_df)} registros) - ONCONOVA",
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
            )
            
            if file_path:
                self._create_professional_excel_export(selected_df, file_path, "seleccion")
                
                # Redirigir a pestaña de exportaciones si el archivo está en la carpeta estándar
                documents_path = os.path.join(os.path.expanduser("~"), "Documents")
                if file_path.startswith(documents_path) and hasattr(self, 'notebook'):
                    self.notebook.select(5)  # Pestaña de Exportaciones
                
        except Exception as e:
            messagebox.showerror("Error", f"❌ Error al exportar:\n{str(e)}")
            logging.error(f"Error detallado: {e}", exc_info=True)

    def _create_professional_excel_export(self, df_data, file_path, export_type):
        """Crear archivo Excel con formato profesional (función común)"""
        try:
            import pandas as pd
            from datetime import datetime
            import os
            # Crear el archivo Excel con múltiples hojas
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                
                # ========== HOJA 1: PRESENTACIÓN ==========
                # Crear DataFrame para la presentación
                export_title = "Exportación Completa" if export_type == "completa" else "Exportación de Selección"
                
                presentation_data = [
                    ["ONCONOVA CIRUGÍA ONCOLÓGICA", ""],
                    [export_title, ""],
                    ["", ""],
                    ["Información del Reporte", ""],
                    ["━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━", ""],
                    ["Fecha y Hora de Exportación:", datetime.now().strftime("%d/%m/%Y %H:%M:%S")],
                    ["Usuario:", self.info_usuario.get("nombre", "Sistema")],
                    ["Cargo:", self.info_usuario.get("cargo", "N/A")],
                    ["", ""],
                    ["Estadísticas de la Exportación", ""],
                    ["━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━", ""],
                    ["Total de Registros:", len(df_data)],
                    ["Número de Campos:", len(df_data.columns)],
                    ["", ""],
                ]
                
                # Calcular estadísticas adicionales
                fecha_cols = [col for col in df_data.columns if 'fecha' in col.lower()]
                if fecha_cols:
                    try:
                        for col in fecha_cols:
                            fechas = pd.to_datetime(df_data[col], errors='coerce')
                            fechas_validas = fechas.dropna()
                            if not fechas_validas.empty:
                                fecha_min = fechas_validas.min().strftime("%d/%m/%Y")
                                fecha_max = fechas_validas.max().strftime("%d/%m/%Y")
                                presentation_data.extend([
                                    [f"Rango de Fechas ({col}):", f"{fecha_min} - {fecha_max}"],
                                ])
                                break
                    except:
                        pass
                
                # Servicios únicos
                servicio_cols = [col for col in df_data.columns if 'servicio' in col.lower()]
                if servicio_cols:
                    unique_services = df_data[servicio_cols[0]].nunique()
                    presentation_data.append(["Servicios Únicos:", unique_services])
                
                # Casos malignos
                malignidad_cols = [col for col in df_data.columns if 'malign' in col.lower()]
                if malignidad_cols:
                    malignant_count = (df_data[malignidad_cols[0]].str.contains('PRESENTE', case=False, na=False)).sum()
                    presentation_data.append(["Casos con Malignidad:", malignant_count])
                
                presentation_data.extend([
                    ["", ""],
                    ["Descripción", ""],
                    ["━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━", ""],
                    ["Este archivo contiene la exportación", ""],
                    ["de la base de datos del sistema ONCONOVA", ""],
                    ["Cirugía Oncológica.", ""],
                    ["", ""],
                    ["La información incluye todos los campos", ""],
                    ["almacenados para cada informe médico:", ""],
                    ["- Datos del paciente", ""],
                    ["- Información clínica", ""],
                    ["- Resultados de laboratorio", ""],
                    ["- Análisis histopatológicos", ""],
                    ["- Biomarcadores (IHQ)", ""],
                    ["- Fechas y responsables", ""],
                    ["", ""],
                    ["Hospital Universitario del Valle", ""],
                    ["Sistema ONCONOVA © 2025", ""],
                ])
                
                df_presentation = pd.DataFrame(presentation_data, columns=["Campo", "Valor"])
                df_presentation.to_excel(writer, sheet_name='Presentación', index=False)
                
                # ========== HOJA 2: DATOS COMPLETOS ==========
                sheet_name = 'Datos_Completos' if export_type == "completa" else 'Datos_Seleccionados'
                df_data.to_excel(writer, sheet_name=sheet_name, index=False)
                
                # ========== FORMATO ==========
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                
                # Formatear hoja de presentación
                ws_pres = writer.sheets['Presentación']
                
                # Título principal
                ws_pres['A1'].font = Font(size=18, bold=True, color="FFFFFF")
                ws_pres['A1'].fill = PatternFill(start_color="1B4F72", end_color="1B4F72", fill_type="solid")
                ws_pres['A1'].alignment = Alignment(horizontal="center", vertical="center")
                ws_pres.merge_cells('A1:B1')
                
                # Subtítulo
                ws_pres['A2'].font = Font(size=14, bold=True, color="FFFFFF")
                ws_pres['A2'].fill = PatternFill(start_color="2E86AB", end_color="2E86AB", fill_type="solid")
                ws_pres['A2'].alignment = Alignment(horizontal="center", vertical="center")
                ws_pres.merge_cells('A2:B2')
                
                # Secciones
                section_font = Font(size=12, bold=True, color="1B4F72")
                for row in range(1, len(presentation_data) + 2):
                    cell_value = ws_pres[f'A{row}'].value
                    if cell_value and ("Información" in str(cell_value) or "Estadísticas" in str(cell_value) or "Descripción" in str(cell_value)):
                        ws_pres[f'A{row}'].font = section_font
                        ws_pres.merge_cells(f'A{row}:B{row}')
                
                # Ajustar ancho de columnas
                ws_pres.column_dimensions['A'].width = 40
                ws_pres.column_dimensions['B'].width = 30
                
                # Formatear hoja de datos
                ws_data = writer.sheets[sheet_name]
                
                # Headers
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="2E86AB", end_color="2E86AB", fill_type="solid")
                header_alignment = Alignment(horizontal="center", vertical="center")
                
                for cell in ws_data[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                
                # Ajustar ancho de columnas automáticamente
                for column in ws_data.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)  # Máximo 50 caracteres
                    ws_data.column_dimensions[column_letter].width = adjusted_width
            
            # Mensaje de confirmación
            record_text = "completos" if export_type == "completa" else "seleccionados"
            messagebox.showinfo(
                "Exportación Exitosa", 
                f"✅ Base de datos exportada exitosamente!\n\n"
                f"📊 {len(df_data)} registros {record_text} exportados\n"
                f"📋 {len(df_data.columns)} campos por registro\n"
                f"📁 Archivo: {os.path.basename(file_path)}\n\n"
                f"El archivo incluye:\n"
                f"• Hoja de presentación con estadísticas\n"
                f"• Datos {record_text} de los registros"
            )
            
        except Exception as e:
            raise e

    def _export_current_record(self):
        """Exportar el registro actualmente seleccionado"""
        try:
            selection = self.tree.selection()
            if not selection:
                messagebox.showwarning("Advertencia", "No hay registro seleccionado.")
                return
                
            # Tomar solo el primer registro seleccionado
            item = selection[0]
            self._export_selected_data_professional()
            
        except Exception as e:
            messagebox.showerror("Error", f"Error al exportar registro: {str(e)}")

    def _copy_details(self):
        """Copiar detalles del registro al clipboard"""
        try:
            details_text = self.detail_textbox.get("1.0", tk.END)
            if details_text.strip():
                self.clipboard_clear()
                self.clipboard_append(details_text)
                messagebox.showinfo("Copiado", "Detalles copiados al portapapeles")
            else:
                messagebox.showwarning("Advertencia", "No hay detalles para copiar")
        except Exception as e:
            messagebox.showerror("Error", f"Error al copiar: {str(e)}")

    def _on_double_click(self, event):
        """Manejar doble clic en la tabla"""
        selection = self.tree.selection()
        if selection:
            # Expandir detalles o realizar acción específica
            self.mostrar_detalle_registro(event)

    def _init_treeview_style(self):
        """
        Define el estilo 'Custom.Treeview' usando los colores del tema actual de TTKBootstrap.
        """
        try:
            from tkinter import ttk as _ttk
            s = _ttk.Style()

            # Usar colores del tema TTKBootstrap actual
            bg_color = self.style.colors.bg or "#ffffff"
            fg_color = self.style.colors.fg or "#000000"  
            primary_color = self.style.colors.primary or "#0d6efd"
            secondary_color = self.style.colors.secondary or "#6c757d"

            s.configure(
                "Custom.Treeview",
                background=bg_color,
                fieldbackground=bg_color,
                foreground=fg_color,
                rowheight=26,
                borderwidth=0,
            )
            s.map(
                "Custom.Treeview",
                background=[("selected", primary_color)],
                foreground=[("selected", "white")],
            )
            s.configure(
                "Custom.Treeview.Heading",
                background=secondary_color,
                foreground="white",
                relief="flat",
                padding=6,
            )

            # V6.9.16 - Selector de secciones (Notebook) cohesivo y minimalista.
            # Tab activa en azul institucional, inactivas claras, hover sutil.
            # Afecta a todos los Notebook de la app (Base de Datos, etc.).
            s.configure("TNotebook", background=bg_color, borderwidth=0,
                        tabmargins=(0, 6, 0, 0))
            s.configure(
                "TNotebook.Tab",
                font=("Segoe UI", 10),
                padding=(20, 10),
                background="#eef1f6",
                foreground="#5a6172",
                borderwidth=0,
                focuscolor="",
            )
            s.map(
                "TNotebook.Tab",
                background=[("selected", primary_color), ("active", "#dfe4ee")],
                foreground=[("selected", "#ffffff"), ("active", primary_color)],
                padding=[("selected", (20, 11))],
            )
        except Exception as e:
            logging.warning(f"No se pudo configurar estilos (Treeview/Notebook): {e}")

def main():
    """
    Función principal que configura el entorno y lanza la aplicación.
    """
    logging.info("Iniciando ONCONOVA Gestor H.U.V...")

    # Configurar el parser de argumentos
    parser = argparse.ArgumentParser(description="ONCONOVA CIRUGÍA ONCOLÓGICA")
    
    parser.add_argument("--nombre", type=str, default="Usuario Sistema", help="Nombre del usuario logueado.")
    parser.add_argument("--cargo", type=str, default="Administrador", help="Cargo del usuario logueado.")
    parser.add_argument("--foto", type=str, default="SIN_FOTO", help="Ruta a la foto de perfil del usuario.")
    parser.add_argument("--tema", type=str, default="huv", help="Tema visual (la app usa SOLO el claro navy HUV).")
    parser.add_argument("--ruta-fotos", type=str, default="", help="Ruta al directorio base de fotos de usuarios.")
    parser.add_argument("--ruta-datos", type=str, default="", help="Ruta a datos de ONCONOVA.")
    parser.add_argument(
        "--lanzado-por-onconova",
        action='store_true',
        help="Bandera interna para verificar el lanzamiento desde la app principal."
    )
    parser.add_argument(
        "--modo-independiente",
        action='store_true',
        help="Permite ejecutar la aplicación de forma independiente (sin ONCONOVA)."
    )
    
    args = parser.parse_args()

    # 1. Configuramos Tesseract antes de que cualquier otra cosa lo necesite
    configure_tesseract()

    # Configurar información del usuario
    info_usuario_recibida = {
        "nombre": args.nombre,
        "cargo": args.cargo,
        "ruta_foto": args.foto,
        "ruta_directorio_fotos": args.ruta_fotos
    }
    
    # Configurar logging
    logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
    
    # Mapear tema del argumento a tema TTKBootstrap
    tema_ttk = THEME_MAP.get(args.tema, "huv")  # V6.9.28 - fallback claro (no oscuro)
    
    # Crear y ejecutar la aplicación
    app = App(info_usuario=info_usuario_recibida, tema=tema_ttk)
    app.mainloop()

    logging.info("Aplicacion cerrada. Hasta luego!")


# === NUEVAS FUNCIONES PARA DETECCIÓN DE DUPLICADOS ===
def _verificar_archivo_duplicado(self, file_path, filename):
    """
    Verifica si un archivo ya ha sido importado basándose en contenido y número de petición
    """
    try:
        # Extraer número de petición del archivo
        from core.processors.ocr_processor import pdf_to_text_enhanced
        from core.extractors.patient_extractor import PATIENT_PATTERNS
        import re
        
        # Obtener texto del PDF
        texto = pdf_to_text_enhanced(file_path)

        # Buscar número de petición
        match = re.search(PATIENT_PATTERNS['numero_peticion']['patrones'][0], texto, re.IGNORECASE)
        if not match:
            return {"es_duplicado": False, "razon": "No se pudo extraer número de petición"}
        
        numero_peticion = match.group(1).strip()
        
        # Verificar en base de datos
        from core.database_manager import verificar_duplicado_por_peticion
        resultado = verificar_duplicado_por_peticion(numero_peticion)
        
        if resultado.get("existe", False):
            registro = resultado.get("registro", {})
            return {
                "es_duplicado": True,
                "numero_peticion": numero_peticion,
                "fecha_informe": registro.get("Fecha Informe", "N/A"),
                "fecha_importacion": registro.get("Fecha Ingreso Base de Datos", "N/A"),  # FECHA REAL DE IMPORTACIÓN
                "registro": registro
            }
        else:
            return {"es_duplicado": False}

    except Exception as e:
        logging.error(f"Error verificando duplicado para {filename}: {e}")
        return {"es_duplicado": False, "error": str(e)}

# Agregar estas funciones como métodos de la clase App
App._verificar_archivo_duplicado = _verificar_archivo_duplicado

def _redirigir_a_visualizador_con_filtro(self, numero_peticion):
    """
    Redirige al visualizador de datos y resalta el registro específico
    """
    try:
        # Navegar a visualizador
        self._nav_to_visualizar()
        
        # Aplicar filtro para mostrar solo ese registro
        if hasattr(self, 'tree') and self.tree:
            # Limpiar filtros existentes
            self._clear_filters()
            
            # Aplicar filtro por número de petición
            for item in self.tree.get_children():
                values = self.tree.item(item)['values']
                if values and len(values) > 0:
                    # Buscar la columna del número de petición
                    if numero_peticion in str(values[0]):  # Asumiendo que está en la primera columna
                        self.tree.selection_set(item)
                        self.tree.focus(item)
                        self.tree.see(item)
                        break
                        
        messagebox.showinfo(
            "Registro encontrado",
            f"Se ha localizado y resaltado el registro con número de petición: {numero_peticion}"
        )
        
    except Exception as e:
        logging.error(f"Error redirigiendo al visualizador: {e}")

App._redirigir_a_visualizador_con_filtro = _redirigir_a_visualizador_con_filtro

def _actualizar_lista_archivos_con_estado(self):
    """V6.9.25: Compat — ahora delega en _refresh_files_list (árbol navegable).
    El marcado 🔴/🟢 por archivo se omite; los duplicados se detectan al procesar."""
    try:
        if hasattr(self, '_refresh_files_list'):
            self._refresh_files_list()
    except Exception as e:
        logging.error(f"Error actualizando lista de archivos: {e}")

App._actualizar_lista_archivos_con_estado = _actualizar_lista_archivos_con_estado

def _clear_filters(self):
    """
    Limpia todos los filtros activos en el visualizador
    """
    try:
        # Limpiar combos de filtro si existen
        if hasattr(self, 'cmb_servicio') and self.cmb_servicio:
            self.cmb_servicio.set("")
        if hasattr(self, 'cmb_malig') and self.cmb_malig:
            self.cmb_malig.set("")
        if hasattr(self, 'cmb_resp') and self.cmb_resp:
            self.cmb_resp.set("")
            
        # Refrescar datos
        self.refresh_data()

    except Exception as e:
        logging.error(f"Error limpiando filtros: {e}")

# V6.9.16: monkey-patch ELIMINADO. La version a nivel modulo de _clear_filters
# llamaba a self.refresh_data() (metodo inexistente) y rompia el boton 'Limpiar'
# del modal de filtros. Se usa el metodo correcto de la clase App._clear_filters
# (lineas ~3753): limpia todas las variables db_filters y llama _refresh_dashboard.
# (La funcion a nivel modulo de arriba queda sin uso, no se asigna a la clase.)

def _crear_footer_inteligente(self):
    """
    Crea un footer inteligente con información de rangos de fechas y distribución mensual
    """
    try:
        if self.master_df.empty:
            return "📂 Base de datos vacía - Importa archivos para comenzar"
        
        # Obtener información de fechas y distribución
        from core.database_manager import get_fecha_range_registros, get_distribucion_mensual
        # MESES_ES está definido al inicio del archivo
        from datetime import datetime
        
        fecha_info = get_fecha_range_registros()
        distribucion = get_distribucion_mensual()
        
        if fecha_info.get("error") or not fecha_info.get("fecha_min"):
            return f"💾 Base de datos cargada: {len(self.master_df)} registros (fechas no disponibles)"
        
        # Formatear fechas - Ya vienen en formato DD/MM/YYYY desde database_manager
        fecha_min = fecha_info["fecha_min"]
        fecha_max = fecha_info["fecha_max"]

        # Crear distribución legible - CORREGIDO: Mostrar TODOS los meses
        if distribucion and not distribucion.get("error"):
            dist_texto = []
            for mes_ano, cantidad in sorted(distribucion.items()):
                try:
                    año, mes = mes_ano.split('-')
                    mes_nombre = MESES_ES.get(int(mes), f"Mes {mes}")
                    # Formato compacto: Ene 25: 1
                    dist_texto.append(f"{mes_nombre[:3]} {año[2:]}: {cantidad}")
                except:
                    dist_texto.append(f"{mes_ano}: {cantidad}")

            # Mostrar TODOS los meses
            dist_resumen = " | ".join(dist_texto)
        else:
            dist_resumen = "Distribución no disponible"
        
        footer_text = f"💾 Base de datos cargada desde {fecha_min} hasta {fecha_max} | "
        footer_text += f"📊 Total: {len(self.master_df)} registros | "
        footer_text += f"📅 Distribución: {dist_resumen}"
        
        return footer_text

    except Exception as e:
        logging.error(f"Error creando footer inteligente: {e}")
        return f"💾 Base de datos cargada: {len(self.master_df)} registros"

App._crear_footer_inteligente = _crear_footer_inteligente

def _update_organs_with_advanced_parsing(self, full_text, records):
    """
    Aplica mapeo avanzado de órganos usando parse_estudios_table_for_organo
    """
    try:
        from core.unified_extractor import parse_estudios_table_for_organo
        from core.database_manager import get_connection
        
        # Extraer órganos usando el parser avanzado
        organos_mapeados = parse_estudios_table_for_organo(full_text)
        
        if not organos_mapeados:
            return 0

        logging.info(f"Organos detectados por parser avanzado: {organos_mapeados}")
        
        # Actualizar registros que no tienen órgano o tienen órgano incompleto
        updated_count = 0
        
        # Obtener números de petición de los registros recién procesados
        numeros_peticion = []
        for record in records:
            if isinstance(record, dict) and 'Numero de caso' in record:
                numeros_peticion.append(record['Numero de caso'])
        
        if numeros_peticion:
            with get_connection() as conn:
                cursor = conn.cursor()
                
                # Actualizar solo registros recién procesados que no tienen órgano completo
                for numero_peticion in numeros_peticion:
                    cursor.execute("""
                        UPDATE informes_ihq 
                        SET "Organo" = ?
                        WHERE "Numero de caso" = ? 
                        AND ("Organo" IS NULL 
                             OR "Organo" = '' 
                             OR "Organo" = 'NO ENCONTRADO')
                    """, (organos_mapeados, numero_peticion))
                    
                    if cursor.rowcount > 0:
                        updated_count += cursor.rowcount
                
                conn.commit()
        
        return updated_count

    except Exception as e:
        logging.error(f"Error en mapeo avanzado de organos: {e}")
        return 0

App._update_organs_with_advanced_parsing = _update_organs_with_advanced_parsing


if __name__ == "__main__":
    main()