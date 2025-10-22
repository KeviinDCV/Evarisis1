#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
🔍 AUDITOR DE SISTEMA - Herramienta Consolidada EVARISIS
==========================================================

Consolida 4 herramientas en una sola:
1. Auditor_bd_pdf.py - Auditoría PDF vs BD
2. leer_texto_caso.py - Lectura OCR
3. verificar_mapeo_biomarcadores.py - Validación mapeo
4. limpiar_encabezados_estudios.py - Limpieza headers

Autor: Sistema EVARISIS
Versión: 1.0.0
Fecha: 20 de octubre de 2025
"""

import sys
import os
import json
import sqlite3
import argparse
import re
from pathlib import Path
from typing import Dict, List, Tuple, Optional, Any, Set
from datetime import datetime

# Configurar salida UTF-8 en Windows
if sys.platform.startswith('win'):
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

# Agregar path del proyecto
PROJECT_ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

from core.database_manager import DB_FILE, TABLE_NAME


class AuditorSistema:
    """Auditor consolidado: Auditoría, OCR, Mapeo, Limpieza"""

    # Headers de tabla que NO son biomarcadores
    ENCABEZADOS_TABLA = {
        'N. ESTUDIO', 'N.ESTUDIO', 'N ESTUDIO', 'ESTUDIO',
        'TIPO ESTUDIO', 'TIPO DE ESTUDIO', 'ALMACENAMIENTO',
        'ORGANO', 'ÓRGANO', 'FECHA TOMA', 'BLOQUES Y LAMINAS', 'BLOQUES Y LÁMINAS',
    }

    # Mapeo completo de biomarcadores conocidos
    BIOMARCADORES = {
        'KI-67': 'IHQ_KI-67', 'KI67': 'IHQ_KI-67',
        'HER2': 'IHQ_HER2', 'HER-2': 'IHQ_HER2',
        'RECEPTOR DE ESTRÓGENO': 'IHQ_RECEPTOR_ESTROGENOS',
        'RECEPTORES DE ESTRÓGENO': 'IHQ_RECEPTOR_ESTROGENOS',
        'RE': 'IHQ_RECEPTOR_ESTROGENOS', 'ER': 'IHQ_RECEPTOR_ESTROGENOS',
        'RECEPTOR DE PROGESTERONA': 'IHQ_RECEPTOR_PROGESTERONA',
        'RECEPTORES DE PROGESTERONA': 'IHQ_RECEPTOR_PROGESTERONA',
        'RP': 'IHQ_RECEPTOR_PROGESTERONA', 'PR': 'IHQ_RECEPTOR_PROGESTERONA',
        'P53': 'IHQ_P53', 'PDL-1': 'IHQ_PDL-1', 'PDL1': 'IHQ_PDL-1',
        'P16': 'IHQ_P16_ESTADO', 'P40': 'IHQ_P40_ESTADO', 'P63': 'IHQ_P63',
        'CK7': 'IHQ_CK7', 'CK20': 'IHQ_CK20', 'CDX2': 'IHQ_CDX2',
        'TTF1': 'IHQ_TTF1', 'TTF-1': 'IHQ_TTF1',
        'CHROMOGRANINA': 'IHQ_CHROMOGRANINA', 'SYNAPTOPHYSIN': 'IHQ_SYNAPTOPHYSIN',
        'CD3': 'IHQ_CD3', 'CD5': 'IHQ_CD5', 'CD10': 'IHQ_CD10',
        'CD20': 'IHQ_CD20', 'CD23': 'IHQ_CD23', 'CD30': 'IHQ_CD30',
        'CD34': 'IHQ_CD34', 'CD45': 'IHQ_CD45', 'CD56': 'IHQ_CD56',
        'CD68': 'IHQ_CD68', 'CD117': 'IHQ_CD117', 'CD138': 'IHQ_CD138',
        'MLH1': 'IHQ_MLH1', 'MSH2': 'IHQ_MSH2',
        'MSH6': 'IHQ_MSH6', 'PMS2': 'IHQ_PMS2',
        'S100': 'IHQ_S100', 'VIMENTINA': 'IHQ_VIMENTINA',
        'EMA': 'IHQ_EMA', 'PAX5': 'IHQ_PAX5', 'PAX8': 'IHQ_PAX8',
        'GATA3': 'IHQ_GATA3', 'SOX10': 'IHQ_SOX10',
    }

    def __init__(self):
        self.db_path = DB_FILE
        self.table_name = TABLE_NAME
        self.debug_maps_dir = PROJECT_ROOT / "data" / "debug_maps"

        if not os.path.exists(self.db_path):
            raise FileNotFoundError(f"Base de datos no encontrada: {self.db_path}")
        if not self.debug_maps_dir.exists():
            raise FileNotFoundError(f"Debug_maps no encontrado: {self.debug_maps_dir}")

    # ========== AUDITORÍA PDF vs BD ==========

    def auditar_caso(self, numero_caso: str, json_export: bool = False) -> Dict:
        """Audita un caso: PDF vs BD con validación completa del flujo M → IHQ"""
        print(f"\n{'='*80}")
        print(f"🔍 AUDITANDO CASO: {numero_caso}")
        print(f"{'='*80}\n")

        debug_map = self._obtener_debug_map(numero_caso)
        if not debug_map:
            print(f"❌ No se encontró debug_map para {numero_caso}")
            return {}

        datos_bd = debug_map.get('base_datos', {}).get('datos_guardados', {})
        descripciones = self._extraer_descripciones(debug_map)

        # ═══════════════════════════════════════════════════════════════════════
        # NUEVO: Extraer contexto del flujo M → IHQ
        # ═══════════════════════════════════════════════════════════════════════
        texto_ocr = debug_map.get('ocr', {}).get('texto_consolidado', '')
        contexto_estudio = self._extraer_contexto_estudio(texto_ocr)

        # ═══════════════════════════════════════════════════════════════════════
        # VALIDACIONES CRÍTICAS: DIAGNOSTICO_PRINCIPAL y FACTOR_PRONOSTICO
        # ═══════════════════════════════════════════════════════════════════════
        resultado_diagnostico = self._validar_diagnostico_principal(datos_bd, texto_ocr)
        resultado_factor = self._validar_factor_pronostico(datos_bd, texto_ocr)

        # Identificar biomarcadores mencionados en PDF
        biomarcadores_en_pdf = self._identificar_biomarcadores_en_pdf(descripciones)

        # Validar IHQ_ESTUDIOS_SOLICITADOS
        resultado_estudios = self._validar_estudios_solicitados(datos_bd, biomarcadores_en_pdf)

        # Validar cada biomarcador individual
        resultados_biomarcadores = {}
        errores = []
        warnings = []
        correctos = []

        for columna_bd in biomarcadores_en_pdf:
            mencionado = True
            resultado = self._validar_biomarcador(columna_bd, datos_bd, mencionado)
            resultados_biomarcadores[columna_bd] = resultado

            if resultado['estado'] == 'ERROR':
                errores.append(columna_bd)
            elif resultado['estado'] == 'WARNING':
                warnings.append(columna_bd)
            elif resultado['estado'] == 'CORRECTO':
                correctos.append(columna_bd)

        # Calcular precisión
        total_biomarcadores = len(biomarcadores_en_pdf)
        total_correctos = len(correctos)
        precision = (total_correctos / total_biomarcadores * 100) if total_biomarcadores > 0 else 100.0

        # ═══════════════════════════════════════════════════════════════════════
        # MOSTRAR RESULTADOS MEJORADOS (con contexto M → IHQ)
        # ═══════════════════════════════════════════════════════════════════════
        print(f"📊 CONTEXTO DEL ESTUDIO:")
        if contexto_estudio['estudio_m_id']:
            print(f"   Estudio M: {contexto_estudio['estudio_m_id']}")
        if contexto_estudio['biomarcadores_ihq']:
            print(f"   Biomarcadores IHQ solicitados: {len(contexto_estudio['biomarcadores_ihq'])}")

        print(f"\n📊 VALIDACIÓN DE CAMPOS CRÍTICOS:")

        # Resultado DIAGNOSTICO_PRINCIPAL
        estado_diag_icono = {'OK': '✅', 'WARNING': '⚠️', 'ERROR': '❌'}.get(resultado_diagnostico['estado'], '❓')
        print(f"   {estado_diag_icono} DIAGNOSTICO_PRINCIPAL: {resultado_diagnostico['estado']}")
        print(f"      {resultado_diagnostico['mensaje']}")
        if 'sugerencia' in resultado_diagnostico:
            print(f"      💡 {resultado_diagnostico['sugerencia']}")

        # Resultado FACTOR_PRONOSTICO
        estado_factor_icono = {'OK': '✅', 'WARNING': '⚠️', 'ERROR': '❌'}.get(resultado_factor['estado'], '❓')
        print(f"   {estado_factor_icono} FACTOR_PRONOSTICO: {resultado_factor['estado']}")
        print(f"      {resultado_factor['mensaje']}")
        if 'sugerencia' in resultado_factor:
            print(f"      💡 {resultado_factor['sugerencia']}")

        print(f"\n📊 VALIDACIÓN DE BIOMARCADORES:")
        print(f"   Precisión: {precision:.1f}% ({total_correctos}/{total_biomarcadores})")
        print(f"   Completitud IHQ_ESTUDIOS: {resultado_estudios['porcentaje_captura']:.1f}%")

        if errores:
            print(f"   ❌ {len(errores)} biomarcador(es) mencionados sin valor en BD")
        if warnings:
            print(f"   ⚠️  {len(warnings)} valor(es) inferido(s)")

        resultado = {
            'numero_caso': numero_caso,
            'precision': precision,
            'biomarcadores_en_pdf': list(biomarcadores_en_pdf),
            'errores': errores,
            'warnings': warnings,
            'correctos': correctos,
            'resultado_estudios': resultado_estudios,
            'contexto_estudio': contexto_estudio,
            'validacion_diagnostico': resultado_diagnostico,
            'validacion_factor_pronostico': resultado_factor
        }

        if json_export:
            self._exportar_json(resultado, numero_caso)

        return resultado

    def auditar_todos(self, limite: Optional[int] = None, json_export: bool = False):
        """Audita todos los casos disponibles"""
        casos = self.listar_casos_disponibles()

        if limite:
            casos = casos[:limite]

        print(f"\n{'='*120}")
        print(f"🔍 AUDITORÍA INTELIGENTE COMPLETA V4.0")
        print(f"{'='*120}\n")
        print(f"📊 Total de casos a auditar: {len(casos)}\n")
        print(f"{'='*120}\n")

        resultados_globales = []
        total_biomarcadores = 0
        total_correctos = 0
        total_errores = 0
        total_warnings = 0

        for i, caso in enumerate(casos, 1):
            print(f"[{i}/{len(casos)}] Validando {caso}...")
            resultado = self.auditar_caso(caso, json_export=False)

            if resultado:
                resultados_globales.append(resultado)
                total_biomarcadores += len(resultado['biomarcadores_en_pdf'])
                total_correctos += len(resultado['correctos'])
                total_errores += len(resultado['errores'])
                total_warnings += len(resultado['warnings'])

        # Resumen global
        precision_global = (total_correctos / total_biomarcadores * 100) if total_biomarcadores > 0 else 100.0

        print(f"\n{'='*120}")
        print(f"📈 RESUMEN GLOBAL DE AUDITORÍA V4.0")
        print(f"{'='*120}")
        print(f"   Total casos en BD:                         {len(casos)}")
        print(f"   ✅ Casos validados:                         {len(resultados_globales)}")
        print(f"\n   📊 Total biomarcadores mencionados en PDFs: {total_biomarcadores}")
        print(f"   ✅ Biomarcadores con valor en BD:           {total_correctos}")
        print(f"   ❌ Biomarcadores sin valor en BD:           {total_errores}")
        print(f"   ⚠️  Valores inferidos (no en PDF):          {total_warnings}")
        print(f"\n   🎯 PRECISIÓN PROMEDIO:                     {precision_global:.1f}%")
        print(f"{'='*120}\n")

        if json_export:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"auditoria_v4_{timestamp}.json"
            self._exportar_json({'resumen': {
                'precision_global': precision_global,
                'total_casos': len(casos),
                'total_biomarcadores': total_biomarcadores,
                'total_correctos': total_correctos,
                'total_errores': total_errores
            }, 'casos': resultados_globales}, filename)

    # ========== LECTURA OCR ==========

    def leer_ocr(self, numero_caso: str, seccion: Optional[str] = None, buscar: Optional[str] = None):
        """Lee texto OCR de un caso"""
        numero_caso = self._normalizar_numero_ihq(numero_caso)
        debug_map = self._obtener_debug_map(numero_caso)

        if not debug_map or 'ocr' not in debug_map:
            print(f"❌ No se encontró OCR para {numero_caso}")
            return

        texto = debug_map['ocr'].get('texto_consolidado', '')

        print(f"\n📄 TEXTO REAL DEL CASO: {numero_caso}")
        print(f"{'='*80}")
        print(f"📊 Longitud: {len(texto):,} caracteres\n")

        if buscar:
            self._buscar_en_ocr(texto, buscar)
        elif seccion:
            self._extraer_seccion_ocr(texto, seccion)
        else:
            print(texto)
            print(f"\n{'='*80}")
            print(f"📊 Total: {len(texto):,} caracteres, {texto.count(chr(10)) + 1} líneas")

    def _buscar_en_ocr(self, texto: str, patron: str):
        """Busca patrón en OCR con contexto"""
        print(f"🔍 BÚSQUEDA: '{patron}'")
        print(f"{'='*80}\n")

        try:
            matches = list(re.finditer(patron, texto, re.IGNORECASE))
            if not matches:
                print(f"⚠️  No se encontraron ocurrencias")
                return

            print(f"✅ {len(matches)} ocurrencia(s) encontrada(s)\n")
            for i, match in enumerate(matches, 1):
                inicio = max(0, match.start() - 60)
                fin = min(len(texto), match.end() + 60)
                print(f"[{i}] ...{texto[inicio:match.start()]}[{match.group()}]{texto[match.end():fin]}...")
                print()
        except re.error as e:
            print(f"❌ Error en regex: {e}")

    def _extraer_seccion_ocr(self, texto: str, seccion: str):
        """Extrae sección específica del OCR"""
        patrones = {
            'estudios': r'ESTUDIOS?\s+SOLICITADOS?.*?(?=\n\n+|ORGANO|DIAGNOSTICO|$)',
            'diagnostico': r'DIAGNOSTICO.*?(?=\n\n+|DESCRIPCI[OÓ]N|$)',
            'microscopica': r'DESCRIPCI[OÓ]N\s+MICROSC[OÓ]PICA.*?(?=\n\n+|COMENTARIOS|$)',
            'macroscopica': r'DESCRIPCI[OÓ]N\s+MACROSC[OÓ]PICA.*?(?=\n\n+|DESCRIPCI[OÓ]N\s+MICROSC|$)',
            'comentarios': r'COMENTARIOS.*?(?=\n\n+|---|$)',
        }

        patron = patrones.get(seccion.lower())
        if not patron:
            print(f"❌ Sección '{seccion}' no reconocida")
            print(f"Secciones disponibles: {', '.join(patrones.keys())}")
            return

        match = re.search(patron, texto, re.DOTALL | re.IGNORECASE)
        if match:
            print(f"📌 SECCIÓN: {seccion.upper()}")
            print(f"{'='*80}\n")
            print(match.group(0).strip())
            print(f"\n{'='*80}")
            print(f"📊 Longitud: {len(match.group(0))} caracteres")
        else:
            print(f"⚠️  No se encontró la sección '{seccion}'")

    # ========== VALIDACIÓN DE MAPEO ==========

    def verificar_mapeo_biomarcadores(self):
        """Verifica cobertura de mapeo de biomarcadores"""
        print(f"\n{'='*80}")
        print("🔍 VERIFICACIÓN DE MAPEO DE BIOMARCADORES")
        print(f"{'='*80}\n")

        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()

        cursor.execute(f"""
            SELECT IHQ_ESTUDIOS_SOLICITADOS
            FROM {self.table_name}
            WHERE IHQ_ESTUDIOS_SOLICITADOS IS NOT NULL
              AND IHQ_ESTUDIOS_SOLICITADOS != ''
        """)

        biomarcadores_unicos = set()
        for (estudios,) in cursor.fetchall():
            for bio in estudios.split(','):
                bio = bio.strip().upper()
                if bio and bio not in self.ENCABEZADOS_TABLA:
                    biomarcadores_unicos.add(bio)

        conn.close()

        # Verificar mapeo
        mapeados = set()
        no_mapeados = set()

        for bio in biomarcadores_unicos:
            if bio in self.BIOMARCADORES:
                mapeados.add(bio)
            else:
                no_mapeados.add(bio)

        total = len(biomarcadores_unicos)
        porcentaje = (len(mapeados) / total * 100) if total > 0 else 100.0

        print(f"📊 Total biomarcadores únicos: {total}")
        print(f"✅ Correctamente mapeados: {len(mapeados)} ({porcentaje:.1f}%)")
        print(f"❌ No mapeados: {len(no_mapeados)}")

        if no_mapeados:
            print(f"\n{'='*80}")
            print("BIOMARCADORES NO MAPEADOS:")
            for bio in sorted(no_mapeados):
                print(f"   - {bio}")
            print(f"\n⚠️  Agregue estos biomarcadores a MAPEO_BIOMARCADORES")
        else:
            print(f"\n✅ Todos los biomarcadores están correctamente mapeados")

    # ========== LIMPIEZA DE HEADERS ==========

    def limpiar_headers_tabla(self):
        """Elimina headers de tabla de IHQ_ESTUDIOS_SOLICITADOS"""
        print(f"\n{'='*80}")
        print("LIMPIEZA DE ENCABEZADOS EN IHQ_ESTUDIOS_SOLICITADOS")
        print(f"{'='*80}\n")

        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()

        cursor.execute(f"""
            SELECT "Numero de caso", IHQ_ESTUDIOS_SOLICITADOS
            FROM {self.table_name}
            WHERE IHQ_ESTUDIOS_SOLICITADOS IS NOT NULL
              AND IHQ_ESTUDIOS_SOLICITADOS != ''
        """)

        registros = cursor.fetchall()
        modificados = 0

        for numero, estudios_raw in registros:
            estudios_limpios = self._limpiar_estudios(estudios_raw)

            if estudios_limpios != estudios_raw:
                cursor.execute(f"""
                    UPDATE {self.table_name}
                    SET IHQ_ESTUDIOS_SOLICITADOS = ?
                    WHERE "Numero de caso" = ?
                """, (estudios_limpios, numero))
                modificados += 1
                print(f"   {numero}: '{estudios_raw}' -> '{estudios_limpios}'")

        if modificados > 0:
            conn.commit()
            print(f"\n✅ {modificados} registro(s) limpiado(s)")
        else:
            print("✅ No se encontraron headers para limpiar")

        conn.close()

    def _limpiar_estudios(self, estudios_raw: str) -> str:
        """Limpia headers de una cadena de estudios"""
        if not estudios_raw:
            return ''

        estudios = [e.strip() for e in estudios_raw.split(',')]
        limpios = [e for e in estudios if e.upper() not in self.ENCABEZADOS_TABLA and e.strip()]
        return ', '.join(limpios)

    # ========== LISTADO ==========

    def listar_casos_disponibles(self) -> List[str]:
        """Lista todos los casos con debug_map"""
        archivos = list(self.debug_maps_dir.glob("debug_map_IHQ*.json"))
        casos = []
        for archivo in archivos:
            match = re.search(r'IHQ\d{6}', archivo.name)
            if match:
                casos.append(match.group())
        return sorted(set(casos))

    # ========== HELPERS INTERNOS ==========

    def _normalizar_numero_ihq(self, numero: str) -> str:
        """Normaliza número IHQ"""
        if not numero.upper().startswith("IHQ"):
            numero = f"IHQ{numero}"
        return numero.upper()

    def _obtener_debug_map(self, numero_caso: str) -> Optional[Dict]:
        """Obtiene el debug_map más reciente de un caso"""
        patron = f"debug_map_{numero_caso}_*.json"
        archivos = list(self.debug_maps_dir.glob(patron))

        if not archivos:
            return None

        archivo_mas_reciente = max(archivos, key=lambda p: p.stat().st_mtime)

        try:
            with open(archivo_mas_reciente, 'r', encoding='utf-8') as f:
                return json.load(f)
        except:
            return None

    def _extraer_descripciones(self, debug_map: Dict) -> Dict[str, str]:
        """Extrae descripciones del debug_map"""
        datos_bd = debug_map.get('base_datos', {}).get('datos_guardados', {})
        return {
            'macroscopica': datos_bd.get('Descripcion macroscopica', ''),
            'microscopica': datos_bd.get('Descripcion microscopica', ''),
            'diagnostico': datos_bd.get('Diagnostico Principal', '')
        }

    def _identificar_biomarcadores_en_pdf(self, descripciones: Dict) -> Set[str]:
        """Identifica biomarcadores mencionados en las descripciones"""
        texto_completo = f"{descripciones['macroscopica']}\n{descripciones['microscopica']}\n{descripciones['diagnostico']}".upper()
        biomarcadores_encontrados = set()

        for nombre_bio, columna_bd in self.BIOMARCADORES.items():
            if re.search(rf'\b{re.escape(nombre_bio)}\b', texto_completo):
                biomarcadores_encontrados.add(columna_bd)

        return biomarcadores_encontrados

    def _validar_biomarcador(self, columna_bd: str, datos_bd: Dict, mencionado: bool) -> Dict:
        """Valida un biomarcador individual"""
        valor_bd = datos_bd.get(columna_bd, '')
        tiene_valor = valor_bd and str(valor_bd).strip() and str(valor_bd) not in ['N/A', 'nan', 'None', '']

        if mencionado and tiene_valor:
            return {'estado': 'CORRECTO', 'valor_bd': valor_bd}
        elif mencionado and not tiene_valor:
            return {'estado': 'ERROR', 'valor_bd': 'VACÍO'}
        elif not mencionado and tiene_valor:
            return {'estado': 'WARNING', 'valor_bd': valor_bd}
        else:
            return {'estado': 'OK', 'valor_bd': 'VACÍO'}

    def _extraer_contexto_estudio(self, texto_ocr: str) -> Dict:
        """Extrae contexto del flujo M → IHQ del PDF.

        CONTEXTO DEL SISTEMA EVARISIS:
        ================================

        FLUJO DE ESTUDIOS ONCOLÓGICOS:

        1. ESTUDIO M (Coloración/Patología General)
           - Identificador: M25XXXXX-X (M = Coloración, 25 = año 2025)
           - DESCRIPCIÓN MACROSCÓPICA: Describe muestra física + diagnóstico completo del estudio M
           - Diagnóstico completo incluye: Tipo histológico + Grado Nottingham + Invasiones

        2. ESTUDIO IHQ (Inmunohistoquímica)
           - Se realiza SOBRE el estudio M para análisis molecular
           - DESCRIPCIÓN MICROSCÓPICA: Resultados detallados de cada biomarcador
           - DIAGNÓSTICO del IHQ:
             * Primera línea: CONFIRMACIÓN del diagnóstico del estudio M (sin grado ni invasiones)
             * Líneas siguientes: Resultados de biomarcadores IHQ

        Returns:
            Dict con:
            - estudio_m_id: Identificador del estudio M (ej: M2510488-1)
            - diagnostico_estudio_m: Diagnóstico completo del estudio de coloración
            - diagnostico_ihq_confirmacion: Primera línea del diagnóstico IHQ (solo histología)
            - biomarcadores_ihq: Lista de biomarcadores encontrados en descripción microscópica
        """
        contexto = {
            'estudio_m_id': None,
            'diagnostico_estudio_m': None,
            'diagnostico_ihq_confirmacion': None,
            'biomarcadores_ihq': []
        }

        # 1. Extraer ID del estudio M
        patron_estudio_m = r'bloque\s+(M\d{7}-\d+)'
        match_m = re.search(patron_estudio_m, texto_ocr, re.IGNORECASE)
        if match_m:
            contexto['estudio_m_id'] = match_m.group(1)

        # 2. Extraer diagnóstico completo del estudio M (en DESCRIPCIÓN MACROSCÓPICA, entre comillas)
        patron_diag_m = r'con diagn[óo]stico de\s+["\']([^"\']+)["\']'
        match_diag_m = re.search(patron_diag_m, texto_ocr, re.IGNORECASE | re.DOTALL)
        if match_diag_m:
            contexto['diagnostico_estudio_m'] = match_diag_m.group(1).strip()

        # 3. Extraer confirmación del diagnóstico IHQ (primera línea de DIAGNÓSTICO)
        patron_diagnostico = r'DIAGN[ÓO]STICO\s*\n(.*?)(?=\n\s*[A-Z]{3,}:|\Z)'
        match_diagnostico = re.search(patron_diagnostico, texto_ocr, re.DOTALL | re.IGNORECASE)
        if match_diagnostico:
            texto_diag = match_diagnostico.group(1).strip()
            lineas = texto_diag.split('\n')
            for linea in lineas:
                linea_limpia = linea.strip()
                if linea_limpia and linea_limpia.startswith('-'):
                    # Primera línea con guion = confirmación del diagnóstico
                    contexto['diagnostico_ihq_confirmacion'] = linea_limpia.lstrip('- ').strip('.').strip()
                    break

        # 4. Identificar biomarcadores de IHQ en DESCRIPCIÓN MICROSCÓPICA
        patron_microscopica = r'DESCRIPCI[ÓO]N\s+MICROSC[ÓO]PICA\s*\n(.*?)(?=\n\s*DIAGN[ÓO]STICO|\Z)'
        match_micro = re.search(patron_microscopica, texto_ocr, re.DOTALL | re.IGNORECASE)
        if match_micro:
            texto_micro = match_micro.group(1)
            # Buscar biomarcadores mencionados
            biomarcadores_comunes = [
                'KI-67', 'KI67', 'HER2', 'HER-2',
                'RECEPTOR DE ESTRÓGENO', 'RECEPTORES DE ESTRÓGENO',
                'RECEPTOR DE PROGESTERONA', 'RECEPTORES DE PROGESTERONA',
                'P53', 'TTF-1', 'TTF1', 'CK7', 'CK20',
                'SINAPTOFISINA', 'SYNAPTOPHYSIN', 'CROMOGRANINA', 'CHROMOGRANINA',
                'CD3', 'CD5', 'CD10', 'CD20', 'CD56', 'CKAE1/AE3'
            ]
            for bio in biomarcadores_comunes:
                if re.search(rf'\b{re.escape(bio)}\b', texto_micro, re.IGNORECASE):
                    if bio not in contexto['biomarcadores_ihq']:
                        contexto['biomarcadores_ihq'].append(bio)

        return contexto

    def _validar_diagnostico_principal(self, datos_bd: Dict, texto_ocr: str) -> Dict:
        """Valida DIAGNOSTICO_PRINCIPAL según flujo M → IHQ.

        El diagnóstico principal debe ser la CONFIRMACIÓN del estudio IHQ,
        que aparece en la primera línea de la sección DIAGNÓSTICO.

        NO debe incluir:
        - Grado Nottingham (es del estudio M)
        - Invasión linfovascular (es del estudio M)
        - Invasión perineural (es del estudio M)
        - Carcinoma in situ (es del estudio M)

        Returns:
            Dict con estado, mensaje, valor_bd, valor_esperado, sugerencia
        """
        diagnostico_bd = datos_bd.get('Diagnostico Principal', '').strip()

        # Extraer sección DIAGNÓSTICO del PDF
        patron_diagnostico = r'DIAGN[ÓO]STICO\s*\n(.*?)(?=\n\s*[A-Z]{3,}:|\Z)'
        match_diagnostico = re.search(patron_diagnostico, texto_ocr, re.DOTALL | re.IGNORECASE)

        if not match_diagnostico:
            return {
                'estado': 'WARNING',
                'mensaje': 'No se encontró sección DIAGNÓSTICO en el PDF',
                'valor_bd': diagnostico_bd
            }

        texto_diagnostico = match_diagnostico.group(1).strip()

        # Extraer primera línea del diagnóstico (confirmación del estudio M)
        lineas = texto_diagnostico.split('\n')
        primera_linea = ''
        for linea in lineas:
            linea_limpia = linea.strip()
            if linea_limpia and linea_limpia.startswith('-'):
                # Primera línea con guion = confirmación del diagnóstico
                primera_linea = linea_limpia.lstrip('- ').strip('.').strip()
                break

        if not primera_linea:
            return {
                'estado': 'WARNING',
                'mensaje': 'No se pudo extraer primera línea del DIAGNÓSTICO',
                'valor_bd': diagnostico_bd
            }

        # Normalizar textos para comparación
        diagnostico_bd_norm = self._normalizar_texto(diagnostico_bd)
        primera_linea_norm = self._normalizar_texto(primera_linea)

        # Validar coincidencia
        if diagnostico_bd_norm == primera_linea_norm:
            # Validar que NO contenga gradación ni invasiones (errores comunes)
            keywords_estudio_m = ['NOTTINGHAM', 'GRADO', 'INVASIÓN', 'INVASION', 'IN SITU']
            contaminacion = [kw for kw in keywords_estudio_m if kw in diagnostico_bd.upper()]

            if contaminacion:
                return {
                    'estado': 'WARNING',
                    'mensaje': f'DIAGNOSTICO_PRINCIPAL contaminado con datos del estudio M: {", ".join(contaminacion)}',
                    'valor_bd': diagnostico_bd,
                    'valor_esperado': primera_linea,
                    'sugerencia': (
                        'Eliminar gradación e invasiones del DIAGNOSTICO_PRINCIPAL.\n'
                        'Solo debe tener el diagnóstico histológico básico (confirmación IHQ).\n'
                        'Verificar extractor en medical_extractor.py::extract_principal_diagnosis()'
                    )
                }
            else:
                return {
                    'estado': 'OK',
                    'mensaje': 'DIAGNOSTICO_PRINCIPAL correctamente extraído (confirmación IHQ)',
                    'valor_bd': diagnostico_bd,
                    'ubicacion': 'Primera línea de DIAGNÓSTICO (confirmación del estudio M)'
                }
        else:
            # Verificar si es substring o match parcial
            if diagnostico_bd_norm in primera_linea_norm or primera_linea_norm in diagnostico_bd_norm:
                return {
                    'estado': 'WARNING',
                    'mensaje': 'DIAGNOSTICO_PRINCIPAL es parcial, falta información completa',
                    'valor_bd': diagnostico_bd,
                    'valor_esperado': primera_linea,
                    'sugerencia': f'Actualizar extractor para capturar: "{primera_linea}"'
                }
            else:
                return {
                    'estado': 'ERROR',
                    'mensaje': 'DIAGNOSTICO_PRINCIPAL NO coincide con primera línea del DIAGNÓSTICO IHQ',
                    'valor_bd': diagnostico_bd,
                    'valor_esperado': primera_linea,
                    'sugerencia': f'Verificar extractor. Debe capturar primera línea del DIAGNÓSTICO: "{primera_linea}"',
                    'analisis_extractor': 'El extractor extract_principal_diagnosis() está extrayendo datos incorrectos'
                }

    def _validar_factor_pronostico(self, datos_bd: Dict, texto_ocr: str) -> Dict:
        """Valida FACTOR_PRONOSTICO según flujo M → IHQ.

        CONTEXTO DEL SISTEMA:
        =====================

        1. ESTUDIO M (Coloración): Diagnóstico histológico completo
           - Contiene: Grado Nottingham, invasión linfovascular, invasión perineural
           - Ubicación: DESCRIPCIÓN MACROSCÓPICA (texto entrecomillado)
           - Ejemplo: "CARCINOMA... NOTTINGHAM GRADO 2... INVASIÓN LINFOVASCULAR PRESENTE"

        2. ESTUDIO IHQ (Inmunohistoquímica): Análisis molecular con biomarcadores
           - Contiene: Receptor de Estrógeno, HER2, Ki-67, p53, TTF-1, etc.
           - Ubicación: DESCRIPCIÓN MICROSCÓPICA + DIAGNÓSTICO (líneas 2+)
           - Ejemplo: "RECEPTORES DE ESTRÓGENO POSITIVO 80-90% / HER2 NEGATIVO / Ki-67: 51-60%"

        FACTOR_PRONOSTICO debe contener SOLO biomarcadores del estudio IHQ.
        NO debe contener información del estudio M (grado, invasiones).

        Orden estándar oncológico: Ki-67 → HER2 → ER → PR → resto

        El sistema extrae factor pronóstico en 4 PRIORIDADES (según medical_extractor.py):
        1. Ki-67 / Ki67 (índice de proliferación celular) - PRIORIDAD ALTA
        2. p53 (supresor tumoral)
        3. Líneas de inmunorreactividad (TTF-1, CK7, Napsina A, etc.)
        4. Otros biomarcadores en diagnóstico (p40, p16, HER2, etc.)

        La IA en auditoría parcial CONSTRUYE el factor pronóstico de biomarcadores encontrados.

        Esta validación:
        - Busca biomarcadores en TODAS las secciones (DIAGNÓSTICO, DESCRIPCIÓN MICROSCÓPICA, COMENTARIOS)
        - Si factor_bd == N/A, analiza si existen biomarcadores que el extractor debió capturar
        - Sugiere correcciones específicas al extractor según lo encontrado
        - DETECTA CONTAMINACIÓN con datos del estudio M (grado, invasiones)

        Returns:
            Dict con estado, mensaje, sugerencias específicas y biomarcadores encontrados
        """
        factor_bd = datos_bd.get('Factor pronostico', 'N/A').strip()

        # ═══════════════════════════════════════════════════════════════════════
        # VALIDACIÓN CRÍTICA: Detectar contaminación con datos del estudio M
        # ═══════════════════════════════════════════════════════════════════════
        keywords_estudio_m = [
            'NOTTINGHAM', 'GRADO HISTOLOGICO', 'GRADO 1', 'GRADO 2', 'GRADO 3',
            'INVASIÓN LINFOVASCULAR', 'INVASION LINFOVASCULAR',
            'INVASIÓN PERINEURAL', 'INVASION PERINEURAL',
            'CARCINOMA DUCTAL IN SITU', 'CARCINOMA IN SITU',
            'BIEN DIFERENCIADO', 'MODERADAMENTE DIFERENCIADO', 'POBREMENTE DIFERENCIADO'
        ]

        if factor_bd != 'N/A':
            contaminacion_detectada = []
            for keyword in keywords_estudio_m:
                if keyword in factor_bd.upper():
                    contaminacion_detectada.append(keyword)

            if contaminacion_detectada:
                return {
                    'estado': 'ERROR',
                    'mensaje': f'FACTOR_PRONOSTICO CONTAMINADO con datos del estudio M: {", ".join(contaminacion_detectada)}',
                    'valor_bd': factor_bd,
                    'contaminacion': contaminacion_detectada,
                    'sugerencia': (
                        'FACTOR_PRONOSTICO debe contener SOLO biomarcadores de IHQ.\n'
                        'Información del estudio M (Grado Nottingham, invasiones, diferenciación) NO pertenece aquí.\n'
                        'Esta información debe estar en otros campos específicos:\n'
                        '  - Grado Nottingham → Campo específico de gradación\n'
                        '  - Invasión linfovascular → Campo específico de invasión\n'
                        '  - Diferenciación → Campo específico de diferenciación\n'
                        '\nVerificar extractor en medical_extractor.py::extract_factor_pronostico()\n'
                        'El extractor está mezclando datos del estudio M con biomarcadores IHQ'
                    ),
                    'analisis_extractor': 'El extractor extract_factor_pronostico() está capturando información del estudio M que NO es factor pronóstico IHQ'
                }

        # ═══════════════════════════════════════════════════════════════════════
        # CONTINUAR CON VALIDACIÓN NORMAL (si no hay contaminación)
        # ═══════════════════════════════════════════════════════════════════════

        # Secciones donde buscar biomarcadores (según auditoria_parcial.py)
        secciones = {
            'diagnostico': r'DIAGNÓSTICO.*?(?=\n\s*[A-Z]{3,}|\Z)',
            'descripcion_microscopica': r'DESCRIPCI[ÓO]N\s+MICROSC[ÓO]PICA.*?(?=\n\s*[A-Z]{3,}|\Z)',
            'descripcion_macroscopica': r'DESCRIPCI[ÓO]N\s+MACROSC[ÓO]PICA.*?(?=\n\s*[A-Z]{3,}|\Z)',
            'comentarios': r'COMENTARIOS.*?(?=\n\s*[A-Z]{3,}|\Z)',
        }

        texto_secciones = {}
        for nombre, patron in secciones.items():
            match = re.search(patron, texto_ocr, re.IGNORECASE | re.DOTALL)
            if match:
                texto_secciones[nombre] = match.group(0)

        # Si no hay secciones, no podemos validar
        if not texto_secciones:
            if factor_bd == 'N/A':
                return {
                    'estado': 'WARNING',
                    'mensaje': 'No se pudieron identificar secciones médicas en el PDF',
                    'valor_bd': factor_bd,
                    'sugerencia': 'Verificar estructura del PDF - posible problema de OCR'
                }
            else:
                return {'estado': 'OK', 'valor_bd': factor_bd}

        # Buscar biomarcadores en todas las secciones
        biomarcadores_encontrados = {}
        texto_completo = ' '.join(texto_secciones.values())

        # ═══════════════════════════════════════════════════════════════════════
        # PRIORIDAD 1: Ki-67 / Ki67 (según medical_extractor.py líneas 294-327)
        # ═══════════════════════════════════════════════════════════════════════
        ki67_patterns = [
            r'[ÍI]NDICE\s+DE\s+PROLIFERACI[ÓO]N\s+C[ÉE]LULAR\s+(?:MEDIDO\s+CON\s+)?(?:\()?Ki[\s-]?67(?:\))?\s*[:\s]+([0-9]+)\s*%',
            r'Ki[\s-]?67\s+DEL\s+([0-9]+)\s*%',
            r'Ki[\s-]?67\s*[:\s]+([0-9]+)\s*%',
            r'Ki[\s-]?67\s*[:\s]+(POSITIVO|NEGATIVO|ALTO|BAJO|<\s*\d+\s*%)',
        ]

        for patron in ki67_patterns:
            match = re.search(patron, texto_completo, re.IGNORECASE)
            if match:
                valor = match.group(0).strip()
                # Validar contexto (evitar diferenciación glandular)
                match_start = match.start()
                context_before = texto_completo[max(0, match_start - 150):match_start].upper()
                if 'DIFERENCIACI' not in context_before or 'GLANDULAR' not in context_before:
                    if 'MENOR DEL' not in context_before[-30:]:
                        biomarcadores_encontrados['Ki-67'] = {
                            'valor': valor,
                            'ubicacion': self._identificar_seccion(match.start(), texto_secciones)
                        }
                        break

        # ═══════════════════════════════════════════════════════════════════════
        # PRIORIDAD 2: p53 (según medical_extractor.py líneas 337-350)
        # ═══════════════════════════════════════════════════════════════════════
        p53_patterns = [
            r'p53\s+tiene\s+expresi[ÓO]n[^.\n]+\([^)]+\)',
            r'p53\s*[:]\s*[^.\n]+',
            r'p53\s+(POSITIVO|NEGATIVO|MUTADO|NO\s+MUTADO)[^.\n]*',
        ]

        for patron in p53_patterns:
            match = re.search(patron, texto_completo, re.IGNORECASE)
            if match:
                biomarcadores_encontrados['p53'] = {
                    'valor': match.group(0).strip(),
                    'ubicacion': self._identificar_seccion(match.start(), texto_secciones)
                }
                break

        # ═══════════════════════════════════════════════════════════════════════
        # PRIORIDAD 3: Líneas de inmunorreactividad (según medical_extractor.py líneas 359-372)
        # ═══════════════════════════════════════════════════════════════════════
        inmuno_patterns = [
            r'Las\s+c[ée]lulas\s+tumorales\s+presentan\s+inmuno[^\n.]+\.',
            r'Los\s+marcadores[^\n.]+(?:negativos?|positivos?)\.',
        ]

        for patron in inmuno_patterns:
            match = re.search(patron, texto_completo, re.IGNORECASE)
            if match:
                biomarcadores_encontrados['inmunorreactividad'] = {
                    'valor': match.group(0).strip(),
                    'ubicacion': self._identificar_seccion(match.start(), texto_secciones)
                }
                break

        # ═══════════════════════════════════════════════════════════════════════
        # PRIORIDAD 4: Otros biomarcadores (HER2, ER, PR, p40, p16, etc.)
        # ═══════════════════════════════════════════════════════════════════════
        otros_marcadores_patterns = [
            r'HER[\s-]?2\s*[:\s]+(POSITIVO|NEGATIVO|\d+\+)',
            r'(?:RECEPTOR\s+DE\s+)?ESTR[ÓO]GENO[S]?\s*[:\s]+(\d+%\s+)?(?:POSITIVO|NEGATIVO)',
            r'(?:RECEPTOR\s+DE\s+)?PROGESTERONA\s*[:\s]+(\d+%\s+)?(?:POSITIVO|NEGATIVO)',
            r'p40\s+(POSITIVO|NEGATIVO)',
            r'p16\s+(POSITIVO|NEGATIVO)',
            r'TTF[\s-]?1\s+(POSITIVO|NEGATIVO)',
            r'CK7\s+(POSITIVO|NEGATIVO)',
            r'CK20\s+(POSITIVO|NEGATIVO)',
            r'(?:Sinaptofisina|Synaptophysin)\s+(POSITIVO|NEGATIVO)',
            r'Napsina\s+A\s+(POSITIVO|NEGATIVO)',
        ]

        for patron in otros_marcadores_patterns:
            matches = re.finditer(patron, texto_completo, re.IGNORECASE)
            for match in matches:
                nombre_marcador = re.sub(r'\s*[:\s]+.*', '', match.group(0), flags=re.IGNORECASE).strip()
                if nombre_marcador not in biomarcadores_encontrados:
                    biomarcadores_encontrados[nombre_marcador] = {
                        'valor': match.group(0).strip(),
                        'ubicacion': self._identificar_seccion(match.start(), texto_secciones)
                    }

        # ═══════════════════════════════════════════════════════════════════════
        # ANÁLISIS DE RESULTADOS
        # ═══════════════════════════════════════════════════════════════════════

        if factor_bd == 'N/A':
            if biomarcadores_encontrados:
                # HAY biomarcadores pero el extractor NO los capturó
                # Construir factor pronóstico como lo haría la IA
                factor_construido = ' / '.join([v['valor'] for v in biomarcadores_encontrados.values()])

                # Analizar por qué el extractor falló
                ubicaciones = {}
                for nombre, info in biomarcadores_encontrados.items():
                    ubicacion = info['ubicacion']
                    if ubicacion not in ubicaciones:
                        ubicaciones[ubicacion] = []
                    ubicaciones[ubicacion].append(nombre)

                # Generar sugerencias específicas
                sugerencias = []

                if 'Ki-67' in biomarcadores_encontrados:
                    ki67_info = biomarcadores_encontrados['Ki-67']
                    sugerencias.append(
                        f"• Ki-67 encontrado en {ki67_info['ubicacion']}: '{ki67_info['valor']}'\n"
                        f"  → Verificar patrones en medical_extractor.py líneas 294-327\n"
                        f"  → El extractor busca: 'ÍNDICE DE PROLIFERACIÓN CELULAR', 'Ki-67 DEL', 'Ki-67:'\n"
                        f"  → Posible problema: formato diferente, contexto de diferenciación glandular"
                    )

                if 'p53' in biomarcadores_encontrados:
                    p53_info = biomarcadores_encontrados['p53']
                    sugerencias.append(
                        f"• p53 encontrado en {p53_info['ubicacion']}: '{p53_info['valor']}'\n"
                        f"  → Verificar patrones en medical_extractor.py líneas 337-350\n"
                        f"  → El extractor busca: 'p53 tiene expresión', 'p53:', 'p53 POSITIVO/NEGATIVO'"
                    )

                if 'inmunorreactividad' in biomarcadores_encontrados:
                    inmuno_info = biomarcadores_encontrados['inmunorreactividad']
                    sugerencias.append(
                        f"• Línea de inmunorreactividad encontrada en {inmuno_info['ubicacion']}\n"
                        f"  → Verificar patrones en medical_extractor.py líneas 359-372\n"
                        f"  → El extractor busca: 'Las células tumorales presentan inmuno...', 'Los marcadores...'"
                    )

                # Otros biomarcadores
                otros = [k for k in biomarcadores_encontrados.keys() if k not in ['Ki-67', 'p53', 'inmunorreactividad']]
                if otros:
                    sugerencias.append(
                        f"• Otros biomarcadores encontrados: {', '.join(otros)}\n"
                        f"  → Verificar si están en última línea del diagnóstico (medical_extractor.py líneas 378-391)\n"
                        f"  → Formato esperado: '- TUMOR... p40 POSITIVO / p16 POSITIVO'"
                    )

                # Verificar si está en sección no buscada por extractor
                secciones_encontradas = list(ubicaciones.keys())
                if 'descripcion_macroscopica' in secciones_encontradas or 'comentarios' in secciones_encontradas:
                    sugerencias.append(
                        f"⚠️ IMPORTANTE: Biomarcadores encontrados en {', '.join(secciones_encontradas)}\n"
                        f"  → El extractor medical_extractor.py recibe 'diagnostico_completo' (todo el texto)\n"
                        f"  → Pero prioriza patrones específicos que pueden no coincidir\n"
                        f"  → Considerar ampliar patrones o buscar en secciones adicionales"
                    )

                return {
                    'estado': 'WARNING',
                    'mensaje': f'FACTOR_PRONOSTICO es N/A pero se encontraron {len(biomarcadores_encontrados)} biomarcadores en el PDF',
                    'valor_bd': factor_bd,
                    'biomarcadores_encontrados': biomarcadores_encontrados,
                    'factor_sugerido': factor_construido,
                    'sugerencias': '\n'.join(sugerencias),
                    'analisis_extractor': 'El extractor extract_factor_pronostico() NO capturó estos biomarcadores. Ver sugerencias para diagnóstico.'
                }
            else:
                # NO hay biomarcadores en el PDF - N/A es válido
                return {
                    'estado': 'OK',
                    'mensaje': 'N/A es apropiado - No se encontraron biomarcadores (Ki-67, p53, inmunorreactividad, etc.) en el PDF',
                    'valor_bd': factor_bd,
                    'analisis_extractor': 'El extractor funcionó correctamente - no hay datos que extraer'
                }
        else:
            # Factor pronóstico tiene valor
            # Verificar si el valor está en el PDF
            factor_normalizado = self._normalizar_texto(factor_bd)
            texto_normalizado = self._normalizar_texto(texto_completo)

            if factor_normalizado in texto_normalizado:
                return {
                    'estado': 'OK',
                    'mensaje': 'FACTOR_PRONOSTICO correctamente extraído',
                    'valor_bd': factor_bd,
                    'biomarcadores_encontrados': biomarcadores_encontrados
                }
            else:
                # Valor no encontrado en PDF - posible error
                return {
                    'estado': 'WARNING',
                    'mensaje': 'FACTOR_PRONOSTICO tiene valor pero no se encuentra explícitamente en el PDF',
                    'valor_bd': factor_bd,
                    'biomarcadores_encontrados': biomarcadores_encontrados,
                    'sugerencia': 'Verificar si el valor fue inferido incorrectamente o si está en formato diferente'
                }

    def _detectar_diagnostico_coloracion_inteligente(self, texto_ocr: str) -> Dict:
        """
        Detecta DIAGNOSTICO_COLORACION del estudio M de forma INTELIGENTE.

        Busca por CONTENIDO SEMÁNTICO, no por posición fija.

        El diagnóstico del estudio M se identifica por tener:
        - Diagnóstico histológico base
        - Grado Nottingham (keywords: NOTTINGHAM, GRADO)
        - Invasiones (keywords: INVASIÓN LINFOVASCULAR, INVASIÓN PERINEURAL)

        Returns:
            Dict con:
            - diagnostico_encontrado: str (diagnóstico completo)
            - ubicacion: str (dónde se encontró: "Descripción Macroscópica", etc.)
            - componentes: Dict (desglose semántico)
            - confianza: float (0-1, qué tan seguro está)
            - biomarcadores_solicitados: List (si los encuentra)
        """
        resultado = {
            'diagnostico_encontrado': None,
            'ubicacion': None,
            'componentes': {
                'diagnostico_base': None,
                'grado_nottingham': None,
                'invasion_linfovascular': None,
                'invasion_perineural': None,
                'carcinoma_in_situ': None,
            },
            'confianza': 0.0,
            'biomarcadores_solicitados': []
        }

        # ═══════════════════════════════════════════════════════════════════════
        # PASO 1: Extraer DESCRIPCIÓN MACROSCÓPICA (ubicación más probable)
        # ═══════════════════════════════════════════════════════════════════════

        # Buscar hasta encontrar DESCRIPCIÓN MICROSCÓPICA o DIAGNÓSTICO con líneas en blanco previas
        patron_desc_macro = r'DESCRIPCI[ÓO]N\s+MACROSC[ÓO]PICA(.*?)(?=\n\n+\s*DESCRIPCI[ÓO]N\s+MICROSC|\n\n+\s*DIAGN[ÓO]STICO|\Z)'
        match_macro = re.search(patron_desc_macro, texto_ocr, re.DOTALL | re.IGNORECASE)

        if not match_macro:
            # No hay descripción macroscópica, buscar en todo el texto
            texto_busqueda = texto_ocr
            ubicacion = "Texto completo"
        else:
            texto_busqueda = match_macro.group(1)
            ubicacion = "Descripción Macroscópica"

        # ═══════════════════════════════════════════════════════════════════════
        # PASO 2: Buscar diagnóstico del estudio M por CONTENIDO SEMÁNTICO
        # ═══════════════════════════════════════════════════════════════════════

        # Keywords que identifican el diagnóstico del estudio M (coloración)
        keywords_estudio_m = {
            'grado': ['NOTTINGHAM', 'GRADO'],
            'invasion_linfo': ['INVASIÓN LINFOVASCULAR', 'INVASIÓN VASCULAR'],
            'invasion_peri': ['INVASIÓN PERINEURAL'],
            'in_situ': ['IN SITU', 'CARCINOMA DUCTAL IN SITU', 'CARCINOMA LOBULILLAR IN SITU']
        }

        # Buscar diagnóstico entre comillas (variante 1)
        # IMPORTANTE: El diagnóstico puede tener saltos de línea dentro de las comillas
        patron_comillas = r'diagn[óo]stico\s+de\s+["\']([^"\']+?)["\']'
        match_comillas = re.search(patron_comillas, texto_busqueda, re.IGNORECASE | re.DOTALL)

        candidatos = []

        if match_comillas:
            candidato = match_comillas.group(1).strip()
            # Verificar si contiene keywords del estudio M
            score = 0
            if any(kw in candidato.upper() for kw in keywords_estudio_m['grado']):
                score += 2  # Grado es muy indicativo
            if any(kw in candidato.upper() for kw in keywords_estudio_m['invasion_linfo']):
                score += 1
            if any(kw in candidato.upper() for kw in keywords_estudio_m['invasion_peri']):
                score += 1
            if any(kw in candidato.upper() for kw in keywords_estudio_m['in_situ']):
                score += 1

            if score >= 2:  # Al menos grado + algo más
                candidatos.append({
                    'texto': candidato,
                    'score': score,
                    'variante': 'entrecomillado'
                })

        # Buscar diagnóstico sin comillas (variante 2)
        # Buscar párrafo que contenga keywords del estudio M
        parrafos = texto_busqueda.split('\n\n')
        for parrafo in parrafos:
            parrafo_limpio = parrafo.strip()
            if len(parrafo_limpio) < 50:  # Muy corto, probablemente no es diagnóstico
                continue

            score = 0
            if any(kw in parrafo_limpio.upper() for kw in keywords_estudio_m['grado']):
                score += 2
            if any(kw in parrafo_limpio.upper() for kw in keywords_estudio_m['invasion_linfo']):
                score += 1
            if any(kw in parrafo_limpio.upper() for kw in keywords_estudio_m['invasion_peri']):
                score += 1
            if any(kw in parrafo_limpio.upper() for kw in keywords_estudio_m['in_situ']):
                score += 1

            if score >= 2:
                candidatos.append({
                    'texto': parrafo_limpio,
                    'score': score,
                    'variante': 'párrafo'
                })

        # Ordenar candidatos por score
        candidatos.sort(key=lambda x: x['score'], reverse=True)

        if not candidatos:
            # No se encontró diagnóstico del estudio M
            return resultado

        # Tomar el mejor candidato
        mejor_candidato = candidatos[0]
        diagnostico_completo = mejor_candidato['texto']

        # ═══════════════════════════════════════════════════════════════════════
        # PASO 3: Extraer componentes semánticos del diagnóstico
        # ═══════════════════════════════════════════════════════════════════════

        # Extraer diagnóstico base (primera oración antes de grado/invasiones)
        primera_oracion = diagnostico_completo.split('.')[0].strip()
        resultado['componentes']['diagnostico_base'] = primera_oracion

        # Extraer grado Nottingham
        patron_grado = r'NOTTINGHAM\s+GRADO\s+(\d+)\s*(?:\(PUNTAJE\s+DE\s+(\d+)\))?'
        match_grado = re.search(patron_grado, diagnostico_completo, re.IGNORECASE)
        if match_grado:
            grado = match_grado.group(1)
            puntaje = match_grado.group(2) if match_grado.group(2) else None
            if puntaje:
                resultado['componentes']['grado_nottingham'] = f"GRADO {grado} (PUNTAJE {puntaje})"
            else:
                resultado['componentes']['grado_nottingham'] = f"GRADO {grado}"

        # Extraer invasión linfovascular
        patron_linfo = r'INVASI[ÓO]N\s+LINFOVASCULAR\s+(PRESENTE|NO\s+IDENTIFICAD[AO]|AUSENTE|NEGATIV[AO])'
        match_linfo = re.search(patron_linfo, diagnostico_completo, re.IGNORECASE)
        if match_linfo:
            estado = match_linfo.group(1).upper()
            resultado['componentes']['invasion_linfovascular'] = estado

        # Extraer invasión perineural
        patron_peri = r'INVASI[ÓO]N\s+PERINEURAL\s+(PRESENTE|NO\s+IDENTIFICAD[AO]|AUSENTE|NEGATIV[AO])'
        match_peri = re.search(patron_peri, diagnostico_completo, re.IGNORECASE)
        if match_peri:
            estado = match_peri.group(1).upper()
            resultado['componentes']['invasion_perineural'] = estado

        # Extraer carcinoma in situ
        patron_in_situ = r'CARCINOMA\s+(?:DUCTAL|LOBULILLAR)?\s*IN\s+SITU\s+(NO\s+IDENTIFICADO|PRESENTE|AUSENTE|NEGATIV[AO])'
        match_in_situ = re.search(patron_in_situ, diagnostico_completo, re.IGNORECASE)
        if match_in_situ:
            estado = match_in_situ.group(1).upper()
            resultado['componentes']['carcinoma_in_situ'] = estado

        # ═══════════════════════════════════════════════════════════════════════
        # PASO 4: Buscar biomarcadores solicitados
        # ═══════════════════════════════════════════════════════════════════════

        patrones_solicitud = [
            r'se solicita\s+([^.]+)',
            r'por lo que se solicita\s+([^.]+)',
            r'estudios?\s+solicitados?[:\s]+([^.]+)',
            r'para\s+estudios\s+de\s+inmunohistoqu[íi]mica[:\s]+([^.]+)',
        ]

        for patron in patrones_solicitud:
            match_solicitud = re.search(patron, texto_busqueda, re.IGNORECASE)
            if match_solicitud:
                biomarcadores_texto = match_solicitud.group(1).strip()
                # Separar por comas o 'y'
                biomarcadores = re.split(r',\s*|\s+y\s+', biomarcadores_texto)
                resultado['biomarcadores_solicitados'] = [b.strip() for b in biomarcadores if b.strip()]
                break

        # ═══════════════════════════════════════════════════════════════════════
        # PASO 5: Calcular confianza
        # ═══════════════════════════════════════════════════════════════════════

        confianza = 0.0

        # +0.3 si tiene grado Nottingham
        if resultado['componentes']['grado_nottingham']:
            confianza += 0.3

        # +0.2 por cada invasión encontrada
        if resultado['componentes']['invasion_linfovascular']:
            confianza += 0.2
        if resultado['componentes']['invasion_perineural']:
            confianza += 0.2

        # +0.2 si tiene carcinoma in situ
        if resultado['componentes']['carcinoma_in_situ']:
            confianza += 0.2

        # +0.1 si encontró biomarcadores solicitados
        if resultado['biomarcadores_solicitados']:
            confianza += 0.1

        resultado['diagnostico_encontrado'] = diagnostico_completo
        resultado['ubicacion'] = ubicacion
        resultado['confianza'] = min(confianza, 1.0)  # Max 1.0

        return resultado

    def _detectar_diagnostico_principal_inteligente(self, texto_ocr: str) -> Dict:
        """
        Detecta DIAGNOSTICO_PRINCIPAL (confirmación IHQ) de forma INTELIGENTE.

        El diagnóstico principal NO siempre está en la segunda línea del DIAGNÓSTICO.
        Puede estar en CUALQUIER línea.

        Se identifica por:
        - Es un diagnóstico histológico (CARCINOMA, ADENOCARCINOMA, TUMOR, etc.)
        - NO contiene grado Nottingham
        - NO contiene invasiones
        - Es una confirmación/corrección del estudio M

        Returns:
            Dict con diagnostico_encontrado, ubicacion, linea_numero, confianza
        """
        resultado = {
            'diagnostico_encontrado': None,
            'ubicacion': None,
            'linea_numero': None,
            'confianza': 0.0
        }

        # Extraer sección DIAGNÓSTICO
        patron_diagnostico = r'DIAGN[ÓO]STICO\s*\n(.*?)(?=\n\s*[A-Z]{3,}:|\Z)'
        match_diagnostico = re.search(patron_diagnostico, texto_ocr, re.DOTALL | re.IGNORECASE)

        if not match_diagnostico:
            return resultado

        texto_diagnostico = match_diagnostico.group(1).strip()
        lineas = texto_diagnostico.split('\n')

        # Keywords de diagnóstico histológico
        keywords_diagnostico = ['CARCINOMA', 'ADENOCARCINOMA', 'TUMOR', 'NEOPLASIA', 'LINFOMA',
                                'SARCOMA', 'MELANOMA', 'GLIOMA', 'METASTASIS']

        # Keywords que NO deben estar (del estudio M)
        keywords_excluir = ['NOTTINGHAM', 'GRADO', 'INVASIÓN LINFOVASCULAR', 'INVASIÓN PERINEURAL',
                            'INVASIÓN VASCULAR', 'INVASIÓN ANGIOLINFÁTICA']

        candidatos = []

        for i, linea in enumerate(lineas, 1):
            linea_limpia = linea.strip().lstrip('- ').strip()

            if len(linea_limpia) < 10:
                continue

            # Verificar si tiene keyword de diagnóstico
            tiene_diagnostico = any(kw in linea_limpia.upper() for kw in keywords_diagnostico)

            # Verificar que NO tenga keywords de estudio M
            tiene_exclusion = any(kw in linea_limpia.upper() for kw in keywords_excluir)

            # Verificar que NO sea biomarcador IHQ
            es_biomarcador = re.search(r'(RECEPTOR|HER|KI-67|P53|TTF|CK\d+|CD\d+)\s*[:\s]+(POSITIVO|NEGATIVO|\d+%)',
                                       linea_limpia, re.IGNORECASE)

            if tiene_diagnostico and not tiene_exclusion and not es_biomarcador:
                confianza = 0.5

                # +0.2 si está al inicio del diagnóstico (líneas 1-3)
                if i <= 3:
                    confianza += 0.2

                # +0.2 si empieza con guion o número
                if linea.strip().startswith('-') or linea.strip().startswith(('1', '2', '3')):
                    confianza += 0.2

                # +0.1 si termina con punto
                if linea_limpia.endswith('.'):
                    confianza += 0.1

                candidatos.append({
                    'texto': linea_limpia.rstrip('.'),
                    'linea': i,
                    'confianza': min(confianza, 1.0)
                })

        if candidatos:
            # Tomar el de mayor confianza
            mejor = max(candidatos, key=lambda x: x['confianza'])
            resultado['diagnostico_encontrado'] = mejor['texto']
            resultado['ubicacion'] = 'Diagnóstico'
            resultado['linea_numero'] = mejor['linea']
            resultado['confianza'] = mejor['confianza']

        return resultado

    def _detectar_biomarcadores_ihq_inteligente(self, texto_ocr: str) -> Dict:
        """
        Detecta biomarcadores IHQ de forma INTELIGENTE.

        Busca en TODAS las secciones posibles:
        1. DESCRIPCIÓN MICROSCÓPICA (más común)
        2. DIAGNÓSTICO (después del diagnóstico principal)
        3. COMENTARIOS (si existen)

        No asume formato fijo. Identifica por patrón: "BIOMARCADOR: RESULTADO"

        Returns:
            Dict con biomarcadores_encontrados (List), ubicaciones (Dict), confianza global
        """
        resultado = {
            'biomarcadores_encontrados': [],
            'ubicaciones': {},
            'confianza_global': 0.0
        }

        # Biomarcadores conocidos (actualizar según necesidad)
        biomarcadores_conocidos = {
            'Ki-67': [r'Ki[\s-]?67\s*[:\s]+([^.\n]+)'],
            'HER2': [r'HER[\s-]?2\s*[:\s]+([^.\n]+)'],
            'Receptor de Estrógeno': [r'RECEPTOR(?:ES)?\s+DE\s+ESTR[ÓO]GENO[S]?\s*[:\s]+([^.\n]+)'],
            'Receptor de Progesterona': [r'RECEPTOR(?:ES)?\s+DE\s+PROGESTERONA\s*[:\s]+([^.\n]+)'],
            'p53': [r'p53\s*[:\s]+([^.\n]+)'],
            'TTF-1': [r'TTF[\s-]?1\s*[:\s]+([^.\n]+)'],
            'CK7': [r'CK7\s*[:\s]+([^.\n]+)'],
            'CK20': [r'CK20\s*[:\s]+([^.\n]+)'],
            'Sinaptofisina': [r'(?:Sinaptofisina|Synaptophysin)\s*[:\s]+([^.\n]+)'],
            'Cromogranina': [r'Cromogranina\s*[AaВ]?\s*[:\s]+([^.\n]+)'],
            'CD56': [r'CD56\s*[:\s]+([^.\n]+)'],
            'CKAE1/AE3': [r'CK\s*AE1[/\s]AE3\s*[:\s]+([^.\n]+)'],
        }

        # Extraer secciones
        secciones = {
            'Descripción Microscópica': r'DESCRIPCI[ÓO]N\s+MICROSC[ÓO]PICA\s*\n(.*?)(?=\n\s*DIAGN[ÓO]STICO|\n\s*COMENTARIOS|\Z)',
            'Diagnóstico': r'DIAGN[ÓO]STICO\s*\n(.*?)(?=\n\s*COMENTARIOS|\n\s*NOTA|\Z)',
            'Comentarios': r'COMENTARIOS\s*\n(.*?)(?=\n\s*[A-Z]{3,}:|\Z)',
        }

        texto_secciones = {}
        for nombre, patron in secciones.items():
            match = re.search(patron, texto_ocr, re.DOTALL | re.IGNORECASE)
            if match:
                texto_secciones[nombre] = match.group(1).strip()

        # Buscar cada biomarcador en todas las secciones
        for nombre_bio, patrones in biomarcadores_conocidos.items():
            for nombre_seccion, texto_seccion in texto_secciones.items():
                for patron in patrones:
                    match = re.search(patron, texto_seccion, re.IGNORECASE)
                    if match:
                        valor = match.group(0).strip()
                        # Limpiar y normalizar
                        valor = re.sub(r'\s+', ' ', valor)

                        # Evitar duplicados (si ya se encontró en otra sección)
                        ya_existe = any(b['nombre'] == nombre_bio for b in resultado['biomarcadores_encontrados'])
                        if not ya_existe:
                            resultado['biomarcadores_encontrados'].append({
                                'nombre': nombre_bio,
                                'valor': valor,
                                'ubicacion': nombre_seccion
                            })
                            resultado['ubicaciones'][nombre_bio] = nombre_seccion
                        break

        # Calcular confianza global
        if resultado['biomarcadores_encontrados']:
            resultado['confianza_global'] = 1.0

        return resultado

    def _detectar_biomarcadores_solicitados_inteligente(self, texto_ocr: str) -> Dict:
        """
        Detecta biomarcadores SOLICITADOS de forma INTELIGENTE.

        Variantes posibles:
        - "se solicita receptores de estrógeno, ki67..."
        - "estudios solicitados: ER, PR, HER2"
        - "para estudios de inmunohistoquímica: Ki-67, p53"
        - Tabla con columna "ESTUDIO"

        Returns:
            Dict con biomarcadores_solicitados (List), formato detectado, ubicacion
        """
        resultado = {
            'biomarcadores_solicitados': [],
            'formato': None,
            'ubicacion': None
        }

        # Extraer DESCRIPCIÓN MACROSCÓPICA (ubicación más probable)
        patron_desc_macro = r'DESCRIPCI[ÓO]N\s+MACROSC[ÓO]PICA\s*\n(.*?)(?=\n\s*DESCRIPCI[ÓO]N\s+MICROSC|\Z)'
        match_macro = re.search(patron_desc_macro, texto_ocr, re.DOTALL | re.IGNORECASE)

        texto_busqueda = match_macro.group(1) if match_macro else texto_ocr

        # Patrones de solicitud (orden de especificidad)
        patrones_solicitud = [
            (r'se solicita\s+([^.]+)', 'se solicita'),
            (r'por lo que se solicita\s+([^.]+)', 'por lo que se solicita'),
            (r'estudios?\s+solicitados?[:\s]+([^.]+)', 'estudios solicitados'),
            (r'para\s+estudios\s+de\s+inmunohistoqu[íi]mica[:\s]*([^.]+)', 'para estudios IHQ'),
        ]

        for patron, formato in patrones_solicitud:
            match = re.search(patron, texto_busqueda, re.IGNORECASE)
            if match:
                biomarcadores_texto = match.group(1).strip()

                # Limpiar texto (remover saltos de línea, espacios extra)
                biomarcadores_texto = re.sub(r'\s+', ' ', biomarcadores_texto)

                # Separar por comas, 'y', 'e'
                biomarcadores = re.split(r',\s*|\s+y\s+|\s+e\s+', biomarcadores_texto)

                # Limpiar y normalizar cada biomarcador
                biomarcadores_limpios = []
                for bio in biomarcadores:
                    bio_limpio = bio.strip().rstrip('.,;')
                    if bio_limpio and len(bio_limpio) > 2:
                        biomarcadores_limpios.append(bio_limpio)

                resultado['biomarcadores_solicitados'] = biomarcadores_limpios
                resultado['formato'] = formato
                resultado['ubicacion'] = 'Descripción Macroscópica' if match_macro else 'Texto completo'
                break

        return resultado

    def _validar_diagnostico_coloracion_inteligente(self, datos_bd: Dict, texto_ocr: str) -> Dict:
        """
        Valida DIAGNOSTICO_COLORACION comparando SEMÁNTICAMENTE (no textualmente).

        Compara componentes individuales:
        - Diagnóstico base
        - Grado Nottingham
        - Invasión linfovascular
        - Invasión perineural
        - Carcinoma in situ

        Returns:
            Dict con estado, componentes_validos, componentes_faltantes, sugerencias
        """
        # Primero detectar diagnóstico coloración en PDF
        deteccion_pdf = self._detectar_diagnostico_coloracion_inteligente(texto_ocr)

        # Obtener valor de BD (si existe, porque aún no hay columna)
        diagnostico_coloracion_bd = datos_bd.get('DIAGNOSTICO_COLORACION', None)

        resultado = {
            'estado': 'PENDING',  # OK, WARNING, ERROR, PENDING (campo no existe en BD)
            'valor_bd': diagnostico_coloracion_bd,
            'valor_esperado_pdf': deteccion_pdf['diagnostico_encontrado'],
            'componentes_pdf': deteccion_pdf['componentes'],
            'componentes_validos': [],
            'componentes_faltantes': [],
            'confianza_deteccion': deteccion_pdf['confianza'],
            'sugerencia': None
        }

        if not deteccion_pdf['diagnostico_encontrado']:
            # No se detectó diagnóstico coloración en PDF
            if diagnostico_coloracion_bd is None:
                resultado['estado'] = 'PENDING'
                resultado['sugerencia'] = 'Campo DIAGNOSTICO_COLORACION no existe en BD (se creará en FASE 2)'
            else:
                resultado['estado'] = 'WARNING'
                resultado['sugerencia'] = 'BD tiene valor pero no se detectó en PDF. Verificar OCR.'
            return resultado

        # Validar componentes
        componentes_esperados = ['diagnostico_base', 'grado_nottingham', 'invasion_linfovascular',
                                 'invasion_perineural', 'carcinoma_in_situ']

        for componente in componentes_esperados:
            if deteccion_pdf['componentes'][componente]:
                resultado['componentes_validos'].append(componente)
            else:
                resultado['componentes_faltantes'].append(componente)

        # Determinar estado
        if len(resultado['componentes_validos']) >= 3:  # Al menos 3/5 componentes
            resultado['estado'] = 'OK'
            resultado['sugerencia'] = f"Diagnóstico coloración detectado con {len(resultado['componentes_validos'])}/5 componentes. Crear columna DIAGNOSTICO_COLORACION en BD (FASE 2)."
        elif len(resultado['componentes_validos']) >= 1:
            resultado['estado'] = 'WARNING'
            resultado['sugerencia'] = f"Diagnóstico coloración parcial ({len(resultado['componentes_validos'])}/5 componentes). Verificar PDF."
        else:
            resultado['estado'] = 'ERROR'
            resultado['sugerencia'] = 'No se pudieron extraer componentes del diagnóstico coloración. Verificar OCR o estructura del PDF.'

        return resultado

    def _validar_diagnostico_principal_inteligente(self, datos_bd: Dict, texto_ocr: str) -> Dict:
        """
        Valida DIAGNOSTICO_PRINCIPAL verificando que NO tenga grado ni invasiones.

        Independiente de posición: busca confirmación IHQ en CUALQUIER línea del DIAGNÓSTICO.

        Returns:
            Dict con estado, valor_bd, valor_esperado, tiene_contaminacion, sugerencias
        """
        diagnostico_bd = datos_bd.get('Diagnostico Principal', '').strip()

        # Detectar diagnóstico principal en PDF
        deteccion_pdf = self._detectar_diagnostico_principal_inteligente(texto_ocr)

        resultado = {
            'estado': 'PENDING',
            'valor_bd': diagnostico_bd,
            'valor_esperado': deteccion_pdf['diagnostico_encontrado'],
            'tiene_contaminacion': False,
            'contaminacion_detectada': [],
            'confianza_deteccion': deteccion_pdf['confianza'],
            'linea_correcta_pdf': deteccion_pdf['linea_numero'],
            'sugerencia': None
        }

        if not diagnostico_bd:
            resultado['estado'] = 'ERROR'
            resultado['sugerencia'] = 'DIAGNOSTICO_PRINCIPAL vacío en BD'
            return resultado

        # Verificar contaminación con datos del estudio M
        keywords_estudio_m = ['NOTTINGHAM', 'GRADO', 'INVASIÓN LINFOVASCULAR', 'INVASIÓN PERINEURAL', 'INVASIÓN VASCULAR']

        for keyword in keywords_estudio_m:
            if keyword in diagnostico_bd.upper():
                resultado['tiene_contaminacion'] = True
                resultado['contaminacion_detectada'].append(keyword)

        if resultado['tiene_contaminacion']:
            resultado['estado'] = 'ERROR'
            resultado['sugerencia'] = (
                f"DIAGNOSTICO_PRINCIPAL contaminado con datos del estudio M: {', '.join(resultado['contaminacion_detectada'])}.\n"
                f"Debe contener SOLO el diagnóstico histológico sin grado ni invasiones.\n"
                f"Valor esperado del PDF: \"{deteccion_pdf['diagnostico_encontrado']}\" (línea {deteccion_pdf['linea_numero']} del DIAGNÓSTICO)\n"
                f"Corrección: Modificar extractor extract_principal_diagnosis() en medical_extractor.py"
            )
            return resultado

        # Comparar con PDF (semánticamente)
        if not deteccion_pdf['diagnostico_encontrado']:
            resultado['estado'] = 'WARNING'
            resultado['sugerencia'] = 'No se pudo detectar diagnóstico principal en PDF. Verificar estructura o OCR.'
            return resultado

        # Normalizar textos para comparación
        diagnostico_bd_norm = self._normalizar_texto(diagnostico_bd)
        diagnostico_pdf_norm = self._normalizar_texto(deteccion_pdf['diagnostico_encontrado'])

        if diagnostico_bd_norm == diagnostico_pdf_norm:
            resultado['estado'] = 'OK'
            resultado['sugerencia'] = f"DIAGNOSTICO_PRINCIPAL correcto (línea {deteccion_pdf['linea_numero']} del DIAGNÓSTICO)"
        else:
            # Verificar si es coincidencia parcial
            if diagnostico_bd_norm in diagnostico_pdf_norm or diagnostico_pdf_norm in diagnostico_bd_norm:
                resultado['estado'] = 'WARNING'
                resultado['sugerencia'] = (
                    f"DIAGNOSTICO_PRINCIPAL parcialmente correcto.\n"
                    f"BD: \"{diagnostico_bd}\"\n"
                    f"PDF: \"{deteccion_pdf['diagnostico_encontrado']}\" (línea {deteccion_pdf['linea_numero']})\n"
                    f"Verificar si requiere ajuste."
                )
            else:
                resultado['estado'] = 'ERROR'
                resultado['sugerencia'] = (
                    f"DIAGNOSTICO_PRINCIPAL no coincide con PDF.\n"
                    f"BD: \"{diagnostico_bd}\"\n"
                    f"PDF: \"{deteccion_pdf['diagnostico_encontrado']}\" (línea {deteccion_pdf['linea_numero']} del DIAGNÓSTICO)\n"
                    f"Corrección: Modificar extractor extract_principal_diagnosis() en medical_extractor.py"
                )

        return resultado

    def _validar_factor_pronostico_inteligente(self, datos_bd: Dict, texto_ocr: str) -> Dict:
        """
        Valida FACTOR_PRONOSTICO verificando que sean SOLO biomarcadores IHQ.

        Sin importar dónde estén en el PDF (microscópica, diagnóstico, comentarios).

        Returns:
            Dict con estado, tiene_contaminacion, biomarcadores_encontrados_pdf,
                 biomarcadores_en_bd, cobertura, sugerencias
        """
        factor_bd = datos_bd.get('Factor pronostico', 'N/A').strip()

        # Detectar biomarcadores en PDF
        deteccion_biomarcadores = self._detectar_biomarcadores_ihq_inteligente(texto_ocr)
        deteccion_solicitados = self._detectar_biomarcadores_solicitados_inteligente(texto_ocr)

        resultado = {
            'estado': 'PENDING',
            'valor_bd': factor_bd,
            'tiene_contaminacion': False,
            'contaminacion_detectada': [],
            'biomarcadores_pdf': deteccion_biomarcadores['biomarcadores_encontrados'],
            'biomarcadores_solicitados': deteccion_solicitados['biomarcadores_solicitados'],
            'biomarcadores_en_bd': [],
            'cobertura': 0.0,
            'sugerencia': None
        }

        if factor_bd == 'N/A' or not factor_bd:
            # Factor pronóstico vacío
            if deteccion_biomarcadores['biomarcadores_encontrados']:
                resultado['estado'] = 'ERROR'
                resultado['sugerencia'] = (
                    f"FACTOR_PRONOSTICO vacío pero se detectaron {len(deteccion_biomarcadores['biomarcadores_encontrados'])} biomarcadores en PDF:\n"
                    f"{', '.join([b['nombre'] for b in deteccion_biomarcadores['biomarcadores_encontrados']])}\n"
                    f"Corrección: Modificar extractor extract_factor_pronostico() en medical_extractor.py"
                )
            else:
                resultado['estado'] = 'OK'
                resultado['sugerencia'] = 'FACTOR_PRONOSTICO vacío es correcto (no se detectaron biomarcadores en PDF)'
            return resultado

        # Verificar contaminación con datos del estudio M
        keywords_estudio_m = ['NOTTINGHAM', 'GRADO', 'INVASIÓN LINFOVASCULAR', 'INVASIÓN PERINEURAL',
                              'BIEN DIFERENCIADO', 'MODERADAMENTE DIFERENCIADO', 'POBREMENTE DIFERENCIADO']

        for keyword in keywords_estudio_m:
            if keyword in factor_bd.upper():
                resultado['tiene_contaminacion'] = True
                resultado['contaminacion_detectada'].append(keyword)

        if resultado['tiene_contaminacion']:
            resultado['estado'] = 'ERROR'
            resultado['sugerencia'] = (
                f"FACTOR_PRONOSTICO contaminado con datos del estudio M: {', '.join(resultado['contaminacion_detectada'])}.\n"
                f"Debe contener SOLO biomarcadores de IHQ.\n"
                f"Corrección: Modificar extractor extract_factor_pronostico() en medical_extractor.py para filtrar datos del estudio M"
            )
            return resultado

        # Extraer biomarcadores mencionados en BD
        # Patrón simple: buscar nombres de biomarcadores conocidos
        biomarcadores_conocidos = ['Ki-67', 'HER2', 'Receptor de Estrógeno', 'Receptor de Progesterona',
                                   'p53', 'TTF-1', 'CK7', 'CK20', 'Sinaptofisina', 'Cromogranina', 'CD56']

        for bio in biomarcadores_conocidos:
            if bio.upper() in factor_bd.upper() or bio.replace(' ', '').upper() in factor_bd.replace(' ', '').upper():
                resultado['biomarcadores_en_bd'].append(bio)

        # Calcular cobertura (biomarcadores en BD vs detectados en PDF)
        if deteccion_biomarcadores['biomarcadores_encontrados']:
            biomarcadores_pdf_nombres = [b['nombre'] for b in deteccion_biomarcadores['biomarcadores_encontrados']]
            biomarcadores_coincidentes = [b for b in resultado['biomarcadores_en_bd'] if b in biomarcadores_pdf_nombres]
            resultado['cobertura'] = len(biomarcadores_coincidentes) / len(biomarcadores_pdf_nombres) * 100

        # Determinar estado
        if resultado['cobertura'] >= 80:
            resultado['estado'] = 'OK'
            resultado['sugerencia'] = f"FACTOR_PRONOSTICO con buena cobertura ({resultado['cobertura']:.0f}%)"
        elif resultado['cobertura'] >= 50:
            resultado['estado'] = 'WARNING'
            resultado['sugerencia'] = (
                f"FACTOR_PRONOSTICO con cobertura media ({resultado['cobertura']:.0f}%).\n"
                f"Biomarcadores en PDF: {', '.join(biomarcadores_pdf_nombres)}\n"
                f"Biomarcadores en BD: {', '.join(resultado['biomarcadores_en_bd']) if resultado['biomarcadores_en_bd'] else 'ninguno detectado'}\n"
                f"Verificar extractor extract_factor_pronostico() en medical_extractor.py"
            )
        else:
            resultado['estado'] = 'ERROR'
            resultado['sugerencia'] = (
                f"FACTOR_PRONOSTICO con cobertura baja ({resultado['cobertura']:.0f}%).\n"
                f"Biomarcadores en PDF: {', '.join(biomarcadores_pdf_nombres)}\n"
                f"Biomarcadores en BD: {', '.join(resultado['biomarcadores_en_bd']) if resultado['biomarcadores_en_bd'] else 'ninguno detectado'}\n"
                f"Corrección urgente: Modificar extractor extract_factor_pronostico() en medical_extractor.py"
            )

        return resultado

    def _identificar_seccion(self, posicion: int, texto_secciones: Dict[str, str]) -> str:
        """Identifica en qué sección del PDF se encuentra una posición de texto"""
        texto_acumulado = 0
        for nombre_seccion, texto in texto_secciones.items():
            if posicion < texto_acumulado + len(texto):
                return nombre_seccion.replace('_', ' ').title()
            texto_acumulado += len(texto)
        return 'Desconocida'

    def _normalizar_texto(self, texto: str) -> str:
        """Normaliza texto removiendo acentos y convirtiendo a mayúsculas"""
        import unicodedata
        # Remover acentos
        texto_sin_acentos = ''.join(
            c for c in unicodedata.normalize('NFD', texto)
            if unicodedata.category(c) != 'Mn'
        )
        return texto_sin_acentos.upper()

    def _validar_estudios_solicitados(self, datos_bd: Dict, biomarcadores_en_pdf: Set[str]) -> Dict:
        """Valida completitud de IHQ_ESTUDIOS_SOLICITADOS

        Busca múltiples variantes de nomenclatura:
        - RECEPTOR-ESTROGENOS / RECEPTOR-ESTROGENO
        - RECEPTOR ESTROGENOS / RECEPTOR ESTROGENO
        - RECEPTOR DE ESTROGENOS / RECEPTOR DE ESTROGENO
        - ESTROGENOS / ESTROGENO (variante corta)

        Normaliza acentos para evitar falsos negativos (ESTRÓGENO → ESTROGENO).
        """
        estudios_bd_original = datos_bd.get('IHQ_ESTUDIOS_SOLICITADOS', '')
        estudios_bd = self._normalizar_texto(estudios_bd_original)

        capturados = []
        faltantes = []

        for columna_bd in biomarcadores_en_pdf:
            nombre_bio = columna_bd.replace('IHQ_', '').replace('_', '-')

            # Generar variantes de búsqueda
            variantes = [
                nombre_bio,  # RECEPTOR-ESTROGENOS
                nombre_bio.replace('-', ' '),  # RECEPTOR ESTROGENOS
                nombre_bio.replace('RECEPTOR-', 'RECEPTOR DE '),  # RECEPTOR DE ESTROGENOS
                nombre_bio.split('-')[-1] if '-' in nombre_bio else nombre_bio,  # ESTROGENOS (última palabra)
            ]

            # Agregar variantes singular/plural
            ultima_palabra = nombre_bio.split('-')[-1] if '-' in nombre_bio else nombre_bio
            if ultima_palabra.endswith('S'):
                # Si es plural, agregar singular
                singular = ultima_palabra[:-1]  # ESTROGENOS → ESTROGENO
                variantes.append(singular)
                variantes.append(f"RECEPTOR DE {singular}")
            else:
                # Si es singular, agregar plural
                plural = ultima_palabra + 'S'  # ESTROGENO → ESTROGENOS
                variantes.append(plural)
                variantes.append(f"RECEPTOR DE {plural}")

            encontrado = False
            for variante in variantes:
                # Búsqueda case-insensitive
                if re.search(rf'\b{re.escape(variante)}\b', estudios_bd, re.IGNORECASE):
                    encontrado = True
                    break

            if encontrado:
                capturados.append(nombre_bio)
            else:
                faltantes.append(nombre_bio)

        total = len(biomarcadores_en_pdf)
        porcentaje = (len(capturados) / total * 100) if total > 0 else 100.0

        return {
            'capturados': capturados,
            'faltantes': faltantes,
            'completitud': len(faltantes) == 0,
            'porcentaje_captura': porcentaje
        }

    def _exportar_json(self, datos: Dict, nombre_archivo: str):
        """Exporta resultados a JSON"""
        output_dir = PROJECT_ROOT / "herramientas_ia" / "resultados"
        output_dir.mkdir(exist_ok=True)

        if not nombre_archivo.endswith('.json'):
            nombre_archivo = f"{nombre_archivo}.json"

        output_path = output_dir / nombre_archivo

        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(datos, f, ensure_ascii=False, indent=2)

        print(f"\n💾 Resultados guardados en: {output_path}")

    # ========== FUNCIONALIDADES EXTENDIDAS ==========

    def auditar_con_nivel(self, numero_caso: str, nivel: str = 'basico') -> Dict:
        """Audita con diferentes niveles de detalle"""
        print(f"\n🔍 Auditoría nivel: {nivel.upper()}")

        resultado_base = self.auditar_caso(numero_caso, json_export=False)

        if nivel == 'medio':
            # Agregar análisis de campos vacíos
            debug_map = self._obtener_debug_map(numero_caso)
            if debug_map:
                datos_bd = debug_map.get('base_datos', {}).get('datos_guardados', {})
                campos_vacios = [k for k, v in datos_bd.items() if not v or str(v).strip() == '']
                resultado_base['campos_vacios'] = campos_vacios[:10]  # Top 10
                print(f"\n📊 Campos vacíos detectados: {len(campos_vacios)}")

        elif nivel == 'profundo':
            # Análisis exhaustivo + sugerencias
            debug_map = self._obtener_debug_map(numero_caso)
            if debug_map:
                # Analizar cada biomarcador en detalle
                print(f"\n🔬 Análisis profundo de biomarcadores...")
                resultado_base['sugerencias'] = self._generar_sugerencias_automaticas(resultado_base)

        return resultado_base

    def _generar_sugerencias_automaticas(self, resultado: Dict) -> List[str]:
        """Genera sugerencias automáticas basadas en errores"""
        sugerencias = []

        for error in resultado.get('errores', []):
            biomarcador = error.replace('IHQ_', '').replace('_', '-')
            sugerencias.append(f"Agregar patrón para {biomarcador} en biomarker_extractor.py")

        return sugerencias

    def comparar_precision_historica(self, fecha_inicio: str, fecha_fin: str):
        """Compara precisión entre dos periodos"""
        print(f"\n📊 COMPARACIÓN HISTÓRICA DE PRECISIÓN")
        print(f"{'='*80}\n")
        print(f"Periodo: {fecha_inicio} → {fecha_fin}\n")

        debug_maps_dir = PROJECT_ROOT / "data" / "debug_maps"
        if not debug_maps_dir.exists():
            print(f"❌ No existe directorio debug_maps")
            return

        # Recolectar todos los debug_maps en el periodo
        archivos = sorted(debug_maps_dir.glob("debug_map_*.json"))

        if not archivos:
            print(f"❌ No se encontraron debug_maps")
            return

        # Agrupar por caso
        casos_data = {}
        for archivo in archivos:
            try:
                # Extraer número de caso
                partes = archivo.stem.split('_')
                if len(partes) < 4:
                    continue

                numero_caso = partes[2]
                timestamp_str = partes[3] + (f"_{partes[4]}" if len(partes) > 4 else "")

                # Filtrar por fecha si se especifica
                if fecha_inicio and timestamp_str < fecha_inicio:
                    continue
                if fecha_fin and timestamp_str > fecha_fin:
                    continue

                with open(archivo, 'r', encoding='utf-8') as f:
                    data = json.load(f)

                bd_data = data.get('base_datos', {}).get('datos_guardados', {})

                # Calcular precisión
                total_campos = len([k for k in bd_data.keys() if not k.startswith('IHQ_')])
                campos_llenos = len([v for k, v in bd_data.items()
                                   if not k.startswith('IHQ_') and v and str(v).strip()
                                   and str(v) not in ['N/A', 'nan', 'None', '']])

                biomarcadores = len([k for k, v in bd_data.items()
                                   if k.startswith('IHQ_') and v and str(v).strip()
                                   and str(v) not in ['N/A', 'nan', 'None', '']])

                precision = (campos_llenos / total_campos * 100) if total_campos > 0 else 0

                if numero_caso not in casos_data:
                    casos_data[numero_caso] = []

                casos_data[numero_caso].append({
                    'timestamp': timestamp_str,
                    'precision': precision,
                    'biomarcadores': biomarcadores
                })

            except Exception as e:
                continue

        if not casos_data:
            print(f"❌ No se encontraron casos en el periodo especificado")
            return

        # Análisis estadístico
        todas_precisiones = []
        for caso, versiones in casos_data.items():
            for v in versiones:
                todas_precisiones.append(v['precision'])

        promedio = sum(todas_precisiones) / len(todas_precisiones) if todas_precisiones else 0
        maximo = max(todas_precisiones) if todas_precisiones else 0
        minimo = min(todas_precisiones) if todas_precisiones else 0

        print(f"📊 ESTADÍSTICAS DEL PERIODO:")
        print(f"   Casos únicos: {len(casos_data)}")
        print(f"   Total versiones: {len(todas_precisiones)}")
        print(f"   Precisión promedio: {promedio:.1f}%")
        print(f"   Rango: {minimo:.1f}% - {maximo:.1f}%")

        # Top 5 casos con mejor precisión
        print(f"\n🏆 TOP 5 CASOS CON MEJOR PRECISIÓN:")
        casos_ordenados = []
        for caso, versiones in casos_data.items():
            precision_max = max([v['precision'] for v in versiones])
            casos_ordenados.append((caso, precision_max))

        casos_ordenados.sort(key=lambda x: x[1], reverse=True)
        for i, (caso, prec) in enumerate(casos_ordenados[:5], 1):
            print(f"   {i}. {caso}: {prec:.1f}%")

        # Casos con problemas
        casos_problematicos = [(caso, prec) for caso, prec in casos_ordenados if prec < 80]
        if casos_problematicos:
            print(f"\n⚠️  CASOS CON PRECISIÓN < 80%: {len(casos_problematicos)}")
            for caso, prec in casos_problematicos[:5]:
                print(f"   • {caso}: {prec:.1f}%")

    def analizar_tendencias_errores(self, periodo: str = 'mes'):
        """Analiza tendencias de errores en el tiempo"""
        print(f"\n📈 ANÁLISIS DE TENDENCIAS DE ERRORES - {periodo.upper()}")
        print(f"{'='*80}\n")

        debug_maps_dir = PROJECT_ROOT / "data" / "debug_maps"
        if not debug_maps_dir.exists():
            print(f"❌ No existe directorio debug_maps")
            return

        archivos = sorted(debug_maps_dir.glob("debug_map_*.json"))
        if not archivos:
            print(f"❌ No se encontraron debug_maps")
            return

        # Recolectar errores por tipo
        errores_por_campo = {}
        errores_por_biomarcador = {}
        casos_analizados = 0

        for archivo in archivos:
            try:
                with open(archivo, 'r', encoding='utf-8') as f:
                    data = json.load(f)

                numero_caso = archivo.stem.split('_')[2]
                bd_data = data.get('base_datos', {}).get('datos_guardados', {})

                # Analizar campos vacíos (errores)
                for campo, valor in bd_data.items():
                    if campo.startswith('IHQ_'):
                        # Biomarcador
                        if not valor or str(valor).strip() in ['N/A', 'nan', 'None', '']:
                            if campo not in errores_por_biomarcador:
                                errores_por_biomarcador[campo] = []
                            errores_por_biomarcador[campo].append(numero_caso)
                    else:
                        # Campo normal
                        if not valor or str(valor).strip() in ['N/A', 'nan', 'None', '']:
                            if campo not in errores_por_campo:
                                errores_por_campo[campo] = []
                            errores_por_campo[campo].append(numero_caso)

                casos_analizados += 1

            except Exception as e:
                continue

        if casos_analizados == 0:
            print(f"❌ No se pudo analizar ningún caso")
            return

        print(f"📊 Casos analizados: {casos_analizados}\n")

        # TOP campos con más errores
        if errores_por_campo:
            campos_ordenados = sorted(errores_por_campo.items(), key=lambda x: len(x[1]), reverse=True)
            print(f"🔴 TOP 10 CAMPOS CON MÁS ERRORES:")
            for i, (campo, casos) in enumerate(campos_ordenados[:10], 1):
                tasa_error = (len(casos) / casos_analizados * 100)
                print(f"   {i}. {campo[:50]:<50} {len(casos):>3} casos ({tasa_error:.1f}%)")

        # TOP biomarcadores con más errores
        if errores_por_biomarcador:
            bio_ordenados = sorted(errores_por_biomarcador.items(), key=lambda x: len(x[1]), reverse=True)
            print(f"\n🔴 TOP 10 BIOMARCADORES CON MÁS ERRORES:")
            for i, (bio, casos) in enumerate(bio_ordenados[:10], 1):
                tasa_error = (len(casos) / casos_analizados * 100)
                bio_nombre = bio.replace('IHQ_', '')
                print(f"   {i}. {bio_nombre[:50]:<50} {len(casos):>3} casos ({tasa_error:.1f}%)")

        # Análisis de patrones
        print(f"\n💡 RECOMENDACIONES:")

        if errores_por_campo:
            campo_mas_problematico = campos_ordenados[0]
            tasa = (len(campo_mas_problematico[1]) / casos_analizados * 100)
            if tasa > 50:
                print(f"   ⚠️  '{campo_mas_problematico[0]}' falla en {tasa:.0f}% de casos")
                print(f"      → Revisar extractor de este campo")

        if errores_por_biomarcador:
            bio_mas_problematico = bio_ordenados[0]
            tasa = (len(bio_mas_problematico[1]) / casos_analizados * 100)
            if tasa > 30:
                print(f"   ⚠️  '{bio_mas_problematico[0]}' falla en {tasa:.0f}% de casos")
                print(f"      → Agregar más patrones de extracción para este biomarcador")

        # Estadísticas generales
        total_errores_campos = sum(len(casos) for casos in errores_por_campo.values())
        total_errores_bio = sum(len(casos) for casos in errores_por_biomarcador.values())

        print(f"\n📊 RESUMEN:")
        print(f"   Total errores en campos: {total_errores_campos}")
        print(f"   Total errores en biomarcadores: {total_errores_bio}")
        print(f"   Promedio errores/caso: {(total_errores_campos + total_errores_bio) / casos_analizados:.1f}")

    def validar_correcciones_ia(self, numero_caso: str):
        """Valida que las correcciones IA fueron aplicadas correctamente"""
        print(f"\n🤖 VALIDANDO CORRECCIONES IA: {numero_caso}")
        print(f"{'='*80}\n")

        debug_map = self._obtener_debug_map(numero_caso)
        if not debug_map:
            print(f"❌ No se encontró debug_map")
            return

        correcciones = debug_map.get('correcciones_ia', [])

        if not correcciones:
            print(f"ℹ️  No hay correcciones IA registradas para este caso")
            return

        print(f"📊 Correcciones IA aplicadas: {len(correcciones)}\n")

        for i, corr in enumerate(correcciones, 1):
            print(f"{i}. Campo: {corr.get('campo', 'N/A')}")
            print(f"   Valor original: {corr.get('valor_original', 'N/A')}")
            print(f"   Valor corregido: {corr.get('valor_corregido', 'N/A')}")
            print(f"   Confianza: {corr.get('confianza', 0):.2f}")
            print()

    def exportar_excel_formateado(self, datos: Dict, nombre_archivo: str):
        """Exporta a Excel con formato"""
        try:
            import pandas as pd
            from openpyxl import load_workbook
            from openpyxl.styles import Font, PatternFill

            output_dir = PROJECT_ROOT / "herramientas_ia" / "resultados"
            output_dir.mkdir(exist_ok=True)

            if not nombre_archivo.endswith('.xlsx'):
                nombre_archivo = f"{nombre_archivo}.xlsx"

            output_path = output_dir / nombre_archivo

            # Convertir a DataFrame
            df = pd.DataFrame([datos])

            # Guardar
            df.to_excel(output_path, index=False, engine='openpyxl')

            # Formatear
            wb = load_workbook(output_path)
            ws = wb.active

            # Header con color
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

            wb.save(output_path)
            print(f"\n💾 Excel formateado guardado en: {output_path}")

        except ImportError:
            print(f"\n⚠️  Requiere: pip install openpyxl pandas")
        except Exception as e:
            print(f"\n❌ Error exportando Excel: {e}")

    def generar_dashboard_precision(self):
        """Genera dashboard visual de precisión"""
        print(f"\n📊 DASHBOARD DE PRECISIÓN")
        print(f"{'='*80}\n")

        casos = self.listar_casos_disponibles()
        if not casos:
            print("❌ No hay casos para analizar")
            return

        print(f"Analizando {len(casos)} casos...\n")

        precisiones = []
        for caso in casos[:10]:  # Limitar para demo
            resultado = self.auditar_caso(caso, json_export=False)
            if resultado:
                precisiones.append(resultado.get('precision', 0))

        if precisiones:
            promedio = sum(precisiones) / len(precisiones)
            print(f"📊 Precisión promedio: {promedio:.1f}%")
            print(f"📊 Casos analizados: {len(precisiones)}")
            print(f"📊 Rango: {min(precisiones):.1f}% - {max(precisiones):.1f}%")

    # ========== DIAGNÓSTICO Y SUGERENCIAS (TAREAS 1.8 y 1.9) ==========

    def _diagnosticar_error_campo(self, campo: str, valor_bd: str, valor_esperado: str,
                                   texto_ocr: str, detalle_validacion: Dict) -> Dict:
        """
        Diagnostica POR QUÉ falló el extractor basándose en el PDF.

        Analiza:
        1. ¿Qué esperaba encontrar? (basado en flujo M → IHQ)
        2. ¿Qué encontró realmente? (en el PDF)
        3. ¿Por qué el extractor se equivocó? (análisis de patrones)
        4. ¿Dónde está la información correcta? (ubicación exacta en PDF)

        Returns:
            Dict con causa_error, tipo_error, ubicacion_correcta, patron_fallido
        """
        diagnostico = {
            'campo': campo,
            'valor_bd': valor_bd,
            'valor_esperado': valor_esperado,
            'causa_error': None,
            'tipo_error': None,  # CONTAMINACION, VACIO, INCORRECTO, PARCIAL, BD_SIN_COLUMNA
            'ubicacion_correcta': None,
            'patron_fallido': None,
            'contexto_pdf': None
        }

        # ═══════════════════════════════════════════════════════════════
        # DIAGNÓSTICO POR CAMPO
        # ═══════════════════════════════════════════════════════════════

        if campo == 'DIAGNOSTICO_COLORACION':
            # Analizar por qué no se detectó el diagnóstico del estudio M
            if not valor_esperado:
                diagnostico['causa_error'] = 'No se encontró diagnóstico del estudio M en DESCRIPCIÓN MACROSCÓPICA'
                diagnostico['tipo_error'] = 'VACIO'
                diagnostico['patron_fallido'] = 'Búsqueda de keywords: NOTTINGHAM, GRADO, INVASIÓN'
            else:
                diagnostico['causa_error'] = 'Diagnóstico coloración detectado en PDF pero no existe columna en BD'
                diagnostico['tipo_error'] = 'BD_SIN_COLUMNA'
                diagnostico['ubicacion_correcta'] = detalle_validacion.get('ubicacion', 'Descripción Macroscópica')
                diagnostico['contexto_pdf'] = valor_esperado[:200] + '...' if len(valor_esperado) > 200 else valor_esperado

        elif campo == 'DIAGNOSTICO_PRINCIPAL':
            # Analizar por qué el diagnóstico principal está mal
            if detalle_validacion.get('tiene_contaminacion'):
                diagnostico['causa_error'] = f"Contaminado con datos del estudio M: {', '.join(detalle_validacion['contaminacion_detectada'])}"
                diagnostico['tipo_error'] = 'CONTAMINACION'
                diagnostico['patron_fallido'] = 'Extractor no filtra grado Nottingham/invasiones'
                diagnostico['ubicacion_correcta'] = f"Línea {detalle_validacion.get('linea_correcta_pdf', '?')} del DIAGNÓSTICO"
            elif not valor_bd:
                diagnostico['causa_error'] = 'Campo vacío en BD'
                diagnostico['tipo_error'] = 'VACIO'
                diagnostico['ubicacion_correcta'] = f"Línea {detalle_validacion.get('linea_correcta_pdf', '?')} del DIAGNÓSTICO"
            elif valor_bd != valor_esperado:
                diagnostico['causa_error'] = 'Diagnóstico extraído no coincide con PDF'
                diagnostico['tipo_error'] = 'INCORRECTO'
                diagnostico['patron_fallido'] = 'Extractor busca en posición fija (segunda línea) en lugar de buscar semánticamente'
                diagnostico['ubicacion_correcta'] = f"Línea {detalle_validacion.get('linea_correcta_pdf', '?')} del DIAGNÓSTICO"

        elif campo == 'FACTOR_PRONOSTICO':
            # Analizar por qué el factor pronóstico está mal
            if detalle_validacion.get('tiene_contaminacion'):
                diagnostico['causa_error'] = f"Contaminado con datos del estudio M: {', '.join(detalle_validacion['contaminacion_detectada'])}"
                diagnostico['tipo_error'] = 'CONTAMINACION'
                diagnostico['patron_fallido'] = 'Extractor no filtra grado/invasiones/diferenciación'
            elif valor_bd == 'N/A' and detalle_validacion.get('biomarcadores_pdf'):
                diagnostico['causa_error'] = f"Factor pronóstico vacío pero se detectaron {len(detalle_validacion['biomarcadores_pdf'])} biomarcadores en PDF"
                diagnostico['tipo_error'] = 'VACIO'
                biomarcadores_nombres = [b['nombre'] for b in detalle_validacion['biomarcadores_pdf']]
                ubicaciones = [b['ubicacion'] for b in detalle_validacion['biomarcadores_pdf']]
                diagnostico['ubicacion_correcta'] = ', '.join(set(ubicaciones))
                diagnostico['contexto_pdf'] = f"Biomarcadores: {', '.join(biomarcadores_nombres)}"
                diagnostico['patron_fallido'] = 'Extractor solo busca en DIAGNÓSTICO, no en DESCRIPCIÓN MICROSCÓPICA'
            elif detalle_validacion.get('cobertura', 0) < 50:
                diagnostico['causa_error'] = f"Cobertura baja ({detalle_validacion['cobertura']:.0f}%): faltan biomarcadores"
                diagnostico['tipo_error'] = 'PARCIAL'
                diagnostico['patron_fallido'] = 'Extractor no busca todos los biomarcadores o busca en ubicación incorrecta'

        return diagnostico

    def _generar_sugerencia_correccion(self, diagnostico_error: Dict) -> Dict:
        """
        Genera sugerencia PRECISA de corrección basada en el diagnóstico de error.

        Proporciona:
        - Archivo a modificar
        - Función a corregir
        - Líneas aproximadas
        - Patrón regex sugerido (si aplica)
        - Explicación del problema
        - Comando para aplicar corrección

        Returns:
            Dict con archivo, funcion, lineas, problema, solucion, patron_sugerido, comando
        """
        campo = diagnostico_error['campo']
        tipo_error = diagnostico_error['tipo_error']

        sugerencia = {
            'archivo': None,
            'funcion': None,
            'lineas': None,
            'problema': diagnostico_error['causa_error'],
            'solucion': None,
            'patron_sugerido': None,
            'comando': None,
            'prioridad': None  # CRITICA, ALTA, MEDIA, BAJA
        }

        # ═══════════════════════════════════════════════════════════════
        # SUGERENCIAS POR CAMPO Y TIPO DE ERROR
        # ═══════════════════════════════════════════════════════════════

        if campo == 'DIAGNOSTICO_COLORACION':
            if tipo_error == 'BD_SIN_COLUMNA':
                sugerencia['archivo'] = 'FASE 2 del plan'
                sugerencia['funcion'] = 'Migración de schema BD'
                sugerencia['solucion'] = (
                    'Crear columna DIAGNOSTICO_COLORACION en la base de datos.\n'
                    'Pasos:\n'
                    '1. Migrar schema BD: agregar columna entre Descripcion Diagnostico y Diagnostico Principal\n'
                    '2. Crear extractor extract_diagnostico_coloracion() en medical_extractor.py\n'
                    '3. Modificar unified_extractor.py para incluir nuevo campo\n'
                    '4. Modificar ihq_processor.py para guardar en BD y debug_map'
                )
                sugerencia['comando'] = 'Pendiente FASE 2'
                sugerencia['prioridad'] = 'ALTA'
            elif tipo_error == 'VACIO':
                sugerencia['archivo'] = 'core/extractors/medical_extractor.py'
                sugerencia['funcion'] = 'extract_diagnostico_coloracion() (FASE 2 - crear)'
                sugerencia['solucion'] = 'Crear extractor con detección semántica de grado Nottingham + invasiones'
                sugerencia['prioridad'] = 'ALTA'

        elif campo == 'DIAGNOSTICO_PRINCIPAL':
            sugerencia['archivo'] = 'core/extractors/medical_extractor.py'
            sugerencia['funcion'] = 'extract_principal_diagnosis()'
            sugerencia['lineas'] = '~420-480 (aproximado)'

            if tipo_error == 'CONTAMINACION':
                sugerencia['solucion'] = (
                    'Filtrar datos del estudio M (grado Nottingham, invasiones).\n'
                    'Modificaciones necesarias:\n'
                    '1. Agregar validación para excluir keywords: NOTTINGHAM, GRADO, INVASIÓN\n'
                    '2. Extraer solo diagnóstico histológico base\n'
                    '3. Buscar en sección DIAGNÓSTICO (confirmación IHQ), no en DESCRIPCIÓN MACROSCÓPICA'
                )
                sugerencia['patron_sugerido'] = (
                    r'# Buscar diagnóstico sin grado ni invasiones\n'
                    r'patron_diagnostico = r\'(CARCINOMA|ADENOCARCINOMA|TUMOR)[^.]+\'\n'
                    r'# Excluir si contiene:\n'
                    r'keywords_excluir = [\'NOTTINGHAM\', \'GRADO\', \'INVASIÓN\']\n'
                    r'if not any(kw in diagnostico for kw in keywords_excluir):\n'
                    r'    return diagnostico'
                )
                sugerencia['comando'] = 'python herramientas_ia/editor_core.py --editar-extractor DIAGNOSTICO_PRINCIPAL --simular'
                sugerencia['prioridad'] = 'CRITICA'

            elif tipo_error == 'INCORRECTO':
                sugerencia['solucion'] = (
                    'NO asumir posición fija (segunda línea).\n'
                    'Modificaciones necesarias:\n'
                    '1. Buscar SEMÁNTICAMENTE en CUALQUIER línea del DIAGNÓSTICO\n'
                    '2. Identificar diagnóstico histológico sin grado ni invasiones\n'
                    '3. Usar detección inteligente similar a _detectar_diagnostico_principal_inteligente()'
                )
                sugerencia['patron_sugerido'] = (
                    r'# Buscar en TODAS las líneas, no solo la segunda\n'
                    r'for linea in lineas_diagnostico:\n'
                    r'    if tiene_diagnostico(linea) and not tiene_grado(linea):\n'
                    r'        return linea'
                )
                sugerencia['comando'] = 'python herramientas_ia/editor_core.py --editar-extractor DIAGNOSTICO_PRINCIPAL --simular'
                sugerencia['prioridad'] = 'ALTA'

            elif tipo_error == 'VACIO':
                sugerencia['solucion'] = 'Verificar que extractor busque en sección DIAGNÓSTICO correctamente'
                sugerencia['comando'] = 'python herramientas_ia/editor_core.py --validar-sintaxis medical_extractor.py'
                sugerencia['prioridad'] = 'ALTA'

        elif campo == 'FACTOR_PRONOSTICO':
            sugerencia['archivo'] = 'core/extractors/medical_extractor.py'
            sugerencia['funcion'] = 'extract_factor_pronostico()'
            sugerencia['lineas'] = '~267-430 (aproximado)'

            if tipo_error == 'CONTAMINACION':
                sugerencia['solucion'] = (
                    'Filtrar datos del estudio M (grado, invasiones, diferenciación).\n'
                    'Modificaciones necesarias:\n'
                    '1. Eliminar patrones de grado Nottingham\n'
                    '2. Eliminar patrones de invasión linfovascular/perineural\n'
                    '3. Eliminar patrones de diferenciación glandular\n'
                    '4. Mantener SOLO patrones de biomarcadores IHQ'
                )
                sugerencia['patron_sugerido'] = (
                    r'# Validar contexto antes de agregar factor\n'
                    r'keywords_excluir = [\'NOTTINGHAM\', \'GRADO\', \'INVASIÓN\', \'DIFERENCIADO\']\n'
                    r'if any(kw in context for kw in keywords_excluir):\n'
                    r'    continue  # Ignorar este patrón'
                )
                sugerencia['comando'] = 'python herramientas_ia/editor_core.py --editar-extractor FACTOR_PRONOSTICO --simular'
                sugerencia['prioridad'] = 'CRITICA'

            elif tipo_error == 'VACIO':
                ubicacion = diagnostico_error.get('ubicacion_correcta', 'Descripción Microscópica')
                sugerencia['solucion'] = (
                    f'Extractor no busca en {ubicacion}.\n'
                    'Modificaciones necesarias:\n'
                    '1. Buscar biomarcadores en DESCRIPCIÓN MICROSCÓPICA (prioridad 1)\n'
                    '2. Buscar biomarcadores en DIAGNÓSTICO (prioridad 2)\n'
                    '3. Buscar biomarcadores en COMENTARIOS (prioridad 3)\n'
                    '4. Usar búsqueda múltiple similar a _detectar_biomarcadores_ihq_inteligente()'
                )
                sugerencia['comando'] = 'python herramientas_ia/editor_core.py --editar-extractor FACTOR_PRONOSTICO --simular'
                sugerencia['prioridad'] = 'ALTA'

            elif tipo_error == 'PARCIAL':
                sugerencia['solucion'] = (
                    'Cobertura baja: extractor no encuentra todos los biomarcadores.\n'
                    'Modificaciones necesarias:\n'
                    '1. Ampliar patrones regex para cada biomarcador\n'
                    '2. Buscar en múltiples ubicaciones (microscópica, diagnóstico, comentarios)\n'
                    '3. Agregar variantes de nomenclatura (ej: ER vs Receptor de Estrógeno)'
                )
                sugerencia['patron_sugerido'] = (
                    r'# Buscar en TODAS las secciones\n'
                    r'for seccion in [desc_microscopica, diagnostico, comentarios]:\n'
                    r'    biomarcadores = extraer_biomarcadores(seccion)\n'
                    r'    factores.extend(biomarcadores)'
                )
                sugerencia['comando'] = 'python herramientas_ia/editor_core.py --editar-extractor FACTOR_PRONOSTICO --simular'
                sugerencia['prioridad'] = 'MEDIA'

        return sugerencia


def main():
    """CLI principal"""
    parser = argparse.ArgumentParser(
        description="🔍 Auditor de Sistema EVARISIS - Herramienta Consolidada",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Ejemplos de uso:

AUDITORÍA:
  python auditor_sistema.py IHQ250001                          # Auditar caso
  python auditor_sistema.py IHQ250001 --nivel profundo         # Auditoría detallada
  python auditor_sistema.py --todos                            # Auditar todos
  python auditor_sistema.py --todos --limite 10                # Primeros 10
  python auditor_sistema.py --listar                           # Listar casos

LECTURA OCR:
  python auditor_sistema.py IHQ250001 --leer-ocr               # Texto completo
  python auditor_sistema.py IHQ250001 --buscar "Ki-67"         # Buscar patrón
  python auditor_sistema.py IHQ250001 --seccion diagnostico    # Sección específica

VALIDACIÓN IA:
  python auditor_sistema.py IHQ250001 --validar-ia             # Validar correcciones IA

ANÁLISIS HISTÓRICO:
  python auditor_sistema.py --dashboard                        # Dashboard precisión
  python auditor_sistema.py --comparar-historico 20251001 20251020  # Comparar periodos
  python auditor_sistema.py --analizar-tendencias              # Tendencias de errores

VERIFICACIÓN SISTEMA:
  python auditor_sistema.py --verificar-mapeo                  # Mapeo biomarcadores
  python auditor_sistema.py --limpiar-headers                  # Limpiar headers

EXPORTACIÓN:
  python auditor_sistema.py IHQ250001 --json                   # Export JSON
  python auditor_sistema.py IHQ250001 --excel auditoria_caso   # Export Excel
  python auditor_sistema.py --todos --json                     # Export global
        """
    )

    parser.add_argument("caso", nargs='?', help="Número IHQ (ej: IHQ250001)")
    parser.add_argument("--todos", action="store_true", help="Auditar todos los casos")
    parser.add_argument("--limite", type=int, help="Limitar número de casos")
    parser.add_argument("--nivel", choices=['basico', 'medio', 'profundo'], default='basico', help="Nivel de auditoría")
    parser.add_argument("--listar", action="store_true", help="Listar casos disponibles")
    parser.add_argument("--leer-ocr", action="store_true", help="Leer texto OCR")
    parser.add_argument("--buscar", help="Buscar patrón en OCR")
    parser.add_argument("--seccion", choices=['estudios', 'diagnostico', 'microscopica', 'macroscopica', 'comentarios'], help="Extraer sección OCR")
    parser.add_argument("--verificar-mapeo", action="store_true", help="Verificar mapeo biomarcadores")
    parser.add_argument("--limpiar-headers", action="store_true", help="Limpiar headers de tabla")
    parser.add_argument("--validar-ia", action="store_true", help="Validar correcciones IA")
    parser.add_argument("--dashboard", action="store_true", help="Generar dashboard de precisión")
    parser.add_argument("--comparar-historico", nargs=2, metavar=('INICIO', 'FIN'), help="Comparar precisión entre fechas (ej: 20251001 20251020)")
    parser.add_argument("--analizar-tendencias", action="store_true", help="Analizar tendencias de errores")
    parser.add_argument("--json", action="store_true", help="Exportar a JSON")
    parser.add_argument("--excel", help="Exportar a Excel formateado")

    args = parser.parse_args()

    try:
        auditor = AuditorSistema()

        if args.listar:
            casos = auditor.listar_casos_disponibles()
            print(f"\n📋 Casos disponibles: {len(casos)}\n")
            for caso in casos:
                print(f"  - {caso}")

        elif args.todos:
            auditor.auditar_todos(limite=args.limite, json_export=args.json)

        elif args.verificar_mapeo:
            auditor.verificar_mapeo_biomarcadores()

        elif args.limpiar_headers:
            auditor.limpiar_headers_tabla()

        elif args.dashboard:
            auditor.generar_dashboard_precision()

        elif args.comparar_historico:
            fecha_inicio, fecha_fin = args.comparar_historico
            auditor.comparar_precision_historica(fecha_inicio, fecha_fin)

        elif args.analizar_tendencias:
            auditor.analizar_tendencias_errores()

        elif args.caso:
            if args.leer_ocr or args.buscar or args.seccion:
                auditor.leer_ocr(args.caso, seccion=args.seccion, buscar=args.buscar)
            elif args.validar_ia:
                auditor.validar_correcciones_ia(args.caso)
            else:
                resultado = auditor.auditar_caso(args.caso, json_export=args.json)

                # Exportar a Excel si se especifica
                if args.excel and resultado:
                    auditor.exportar_excel_formateado(resultado, args.excel)

        else:
            parser.print_help()

    except KeyboardInterrupt:
        print("\n\n⚠️  Operación cancelada")
        sys.exit(130)
    except Exception as e:
        print(f"\n❌ Error: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()
