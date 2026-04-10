#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ExportWorker - Worker asíncrono para exportación de datos
Worker QThread para exportar DataFrames a Excel sin bloquear la UI

Características:
- Exportación asíncrona a Excel
- Formateo profesional con openpyxl
- Progreso en tiempo real
- Múltiples formatos de salida
"""

from PySide6.QtCore import QThread, Signal
from pathlib import Path
from typing import Optional, Dict, Any
import pandas as pd
import traceback
from datetime import datetime


class ExportWorker(QThread):
    """
    Worker para exportación asíncrona a Excel

    Signals:
        progress(int, str): Emite (porcentaje, mensaje) durante exportación
        finished(str): Emite ruta del archivo generado cuando completa
        error(str): Emite mensaje de error si falla
    """

    # Signals
    progress = Signal(int, str)
    finished = Signal(str)
    error = Signal(str)

    def __init__(
        self,
        dataframe: pd.DataFrame,
        output_path: str,
        include_formatting: bool = True,
        parent=None
    ):
        """
        Inicializa el worker de exportación

        Args:
            dataframe: DataFrame a exportar
            output_path: Ruta del archivo de salida
            include_formatting: Si debe aplicar formato profesional
            parent: Widget padre
        """
        super().__init__(parent)
        self.dataframe = dataframe
        self.output_path = output_path
        self.include_formatting = include_formatting
        self._is_cancelled = False

    def cancel(self):
        """Cancela la exportación"""
        self._is_cancelled = True

    def run(self):
        """Ejecuta la exportación en thread separado"""
        try:
            self.progress.emit(0, "Iniciando exportación...")

            # Verificar DataFrame
            if self.dataframe is None or self.dataframe.empty:
                self.error.emit("Error: DataFrame vacío o nulo")
                return

            # Crear directorio si no existe
            output_dir = Path(self.output_path).parent
            output_dir.mkdir(parents=True, exist_ok=True)

            self.progress.emit(20, f"Exportando {len(self.dataframe)} filas...")

            # Exportar a Excel
            if self.include_formatting:
                self._export_with_formatting()
            else:
                self._export_simple()

            if self._is_cancelled:
                self.progress.emit(0, "Exportación cancelada")
                return

            self.progress.emit(100, f"Archivo generado: {self.output_path}")
            self.finished.emit(self.output_path)

        except ImportError as e:
            error_msg = (
                f"Error de importación: {str(e)}\n\n"
                "Asegúrate de que openpyxl esté instalado:\n"
                "pip install openpyxl"
            )
            self.error.emit(error_msg)

        except Exception as e:
            error_msg = (
                f"Error durante exportación:\n\n"
                f"{str(e)}\n\n"
                f"Traceback:\n{traceback.format_exc()}"
            )
            self.error.emit(error_msg)

    def _export_simple(self):
        """Exportación simple sin formateo"""
        self.dataframe.to_excel(
            self.output_path,
            index=False,
            engine='openpyxl'
        )
        self.progress.emit(90, "Datos exportados")

    def _export_with_formatting(self):
        """Exportación con formateo profesional"""
        from openpyxl import load_workbook
        from openpyxl.styles import (
            Font, PatternFill, Alignment, Border, Side
        )

        # Exportar datos base
        self.progress.emit(30, "Escribiendo datos...")
        self.dataframe.to_excel(
            self.output_path,
            index=False,
            engine='openpyxl'
        )

        if self._is_cancelled:
            return

        # Cargar workbook para formatear
        self.progress.emit(50, "Aplicando formato profesional...")
        wb = load_workbook(self.output_path)
        ws = wb.active

        # Estilos
        header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='2E5C8A', end_color='2E5C8A', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')

        cell_font = Font(name='Calibri', size=10)
        cell_alignment = Alignment(horizontal='left', vertical='center')

        border_side = Side(style='thin', color='D0D0D0')
        border = Border(
            left=border_side,
            right=border_side,
            top=border_side,
            bottom=border_side
        )

        # Formatear encabezados
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        self.progress.emit(60, "Formateando celdas...")

        # Formatear celdas de datos
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.font = cell_font
                cell.alignment = cell_alignment
                cell.border = border

                # Color coding basado en completitud (si existe columna)
                if 'completitud' in str(cell.column_letter).lower() or \
                   cell.column == self._get_completitud_column(ws):
                    try:
                        value = float(cell.value) if cell.value else 0
                        if value >= 95:
                            cell.fill = PatternFill(start_color='C6EFCE', fill_type='solid')
                        elif value >= 80:
                            cell.fill = PatternFill(start_color='FFEB9C', fill_type='solid')
                        else:
                            cell.fill = PatternFill(start_color='FFC7CE', fill_type='solid')
                    except (ValueError, TypeError):
                        pass

        if self._is_cancelled:
            return

        self.progress.emit(70, "Ajustando anchos de columna...")

        # Ajustar anchos de columna
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Congelar primera fila
        ws.freeze_panes = 'A2'

        self.progress.emit(80, "Guardando archivo...")

        # Guardar workbook formateado
        wb.save(self.output_path)
        wb.close()

        self.progress.emit(90, "Formato aplicado exitosamente")

    def _get_completitud_column(self, worksheet):
        """Obtiene el número de columna de 'Completitud'"""
        for cell in worksheet[1]:
            if 'completitud' in str(cell.value).lower():
                return cell.column
        return None


class MultiFormatExportWorker(QThread):
    """
    Worker para exportación en múltiples formatos simultáneamente

    Soporta: Excel, CSV, JSON, HTML
    """

    progress = Signal(int, str)
    file_generated = Signal(str, str)  # (formato, ruta)
    finished = Signal(dict)            # {formato: ruta}
    error = Signal(str)

    def __init__(
        self,
        dataframe: pd.DataFrame,
        base_path: str,
        formats: list = None,
        parent=None
    ):
        """
        Inicializa worker de exportación múltiple

        Args:
            dataframe: DataFrame a exportar
            base_path: Ruta base (sin extensión)
            formats: Lista de formatos ['excel', 'csv', 'json', 'html']
            parent: Widget padre
        """
        super().__init__(parent)
        self.dataframe = dataframe
        self.base_path = base_path
        self.formats = formats or ['excel', 'csv']
        self._is_cancelled = False
        self.results = {}

    def cancel(self):
        """Cancela exportación"""
        self._is_cancelled = True

    def run(self):
        """Ejecuta exportación múltiple"""
        try:
            total_formats = len(self.formats)

            for idx, fmt in enumerate(self.formats):
                if self._is_cancelled:
                    return

                progress_pct = int((idx / total_formats) * 100)
                self.progress.emit(progress_pct, f"Exportando a {fmt.upper()}...")

                # Generar ruta
                output_path = self._get_output_path(fmt)

                # Exportar según formato
                if fmt == 'excel':
                    self._export_excel(output_path)
                elif fmt == 'csv':
                    self._export_csv(output_path)
                elif fmt == 'json':
                    self._export_json(output_path)
                elif fmt == 'html':
                    self._export_html(output_path)
                else:
                    self.progress.emit(progress_pct, f"Formato {fmt} no soportado")
                    continue

                self.results[fmt] = output_path
                self.file_generated.emit(fmt, output_path)

            self.progress.emit(100, "Exportación múltiple completada")
            self.finished.emit(self.results)

        except Exception as e:
            self.error.emit(f"Error en exportación múltiple: {str(e)}")

    def _get_output_path(self, fmt: str) -> str:
        """Genera ruta de salida según formato"""
        base = Path(self.base_path)
        extensions = {
            'excel': '.xlsx',
            'csv': '.csv',
            'json': '.json',
            'html': '.html'
        }
        return str(base.with_suffix(extensions.get(fmt, '.txt')))

    def _export_excel(self, output_path: str):
        """Exporta a Excel con formato"""
        worker = ExportWorker(
            self.dataframe,
            output_path,
            include_formatting=True
        )
        worker.run()

    def _export_csv(self, output_path: str):
        """Exporta a CSV"""
        self.dataframe.to_csv(
            output_path,
            index=False,
            encoding='utf-8-sig'
        )

    def _export_json(self, output_path: str):
        """Exporta a JSON"""
        self.dataframe.to_json(
            output_path,
            orient='records',
            indent=2,
            force_ascii=False
        )

    def _export_html(self, output_path: str):
        """Exporta a HTML con estilos"""
        html_content = self.dataframe.to_html(
            index=False,
            classes='table table-striped table-hover',
            border=0
        )

        # Agregar estilos CSS
        styled_html = f"""
<!DOCTYPE html>
<html lang="es">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>EVARISIS - Datos Exportados</title>
    <style>
        body {{
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            margin: 20px;
            background-color: #f5f5f5;
        }}
        .table {{
            width: 100%;
            border-collapse: collapse;
            background-color: white;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        }}
        .table th {{
            background-color: #2E5C8A;
            color: white;
            padding: 12px;
            text-align: left;
            font-weight: bold;
        }}
        .table td {{
            padding: 10px;
            border-bottom: 1px solid #ddd;
        }}
        .table tr:hover {{
            background-color: #f8f9fa;
        }}
        h1 {{
            color: #2E5C8A;
            text-align: center;
        }}
        .export-info {{
            text-align: center;
            color: #666;
            margin-bottom: 20px;
        }}
    </style>
</head>
<body>
    <h1>EVARISIS - Datos Exportados</h1>
    <div class="export-info">
        Generado: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}<br>
        Total de registros: {len(self.dataframe)}
    </div>
    {html_content}
</body>
</html>
"""

        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(styled_html)


# Ejemplo de uso
if __name__ == "__main__":
    import sys
    from PySide6.QtWidgets import QApplication, QMainWindow, QTextEdit, QVBoxLayout, QWidget, QPushButton

    app = QApplication(sys.argv)

    class TestWindow(QMainWindow):
        def __init__(self):
            super().__init__()
            self.setWindowTitle("Test ExportWorker")
            self.resize(800, 600)

            central = QWidget()
            self.setCentralWidget(central)
            layout = QVBoxLayout(central)

            self.log = QTextEdit()
            self.log.setReadOnly(True)
            layout.addWidget(self.log)

            btn = QPushButton("Exportar Datos de Prueba")
            btn.clicked.connect(self.start_export)
            layout.addWidget(btn)

        def start_export(self):
            # Datos de prueba
            df = pd.DataFrame({
                'Numero': [f'IHQ25{i:04d}' for i in range(1, 101)],
                'Paciente': [f'Paciente {i}' for i in range(1, 101)],
                'Completitud': [95.0, 85.0, 75.0, 100.0] * 25
            })

            output_path = "test_export.xlsx"

            self.worker = ExportWorker(df, output_path, include_formatting=True)
            self.worker.progress.connect(
                lambda p, m: self.log.append(f"[{p}%] {m}")
            )
            self.worker.finished.connect(
                lambda path: self.log.append(f"\nArchivo generado: {path}")
            )
            self.worker.error.connect(
                lambda e: self.log.append(f"\nERROR: {e}")
            )

            self.worker.start()

    window = TestWindow()
    window.show()

    sys.exit(app.exec())
